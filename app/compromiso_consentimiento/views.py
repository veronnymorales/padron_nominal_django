from django.shortcuts import render

# TABLERO PAQUETE GESTANTE 
from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo, Actualizacion
from django.db.models.functions import Substr
import logging

from .queries import (obtener_avance_compromiso_consentimiento, obtener_variables_compromiso_consentimiento, obtener_avance_regional_mensual_compromiso_consentimiento,
                    obtener_cobertura_por_zona, obtener_cobertura_por_provincia, obtener_cobertura_por_distrito, 
                    obtener_seguimiento_compromiso_consentimiento,
                    obtener_avance_regional_mensual_compromiso_consentimiento, obtener_avance_cobertura_compromiso_consentimiento, 
                    obtener_cobertura_por_red, obtener_cobertura_por_microred, obtener_cobertura_por_establecimiento, 
                    obtener_seguimiento_compromiso_consentimiento_red, obtener_seguimiento_compromiso_consentimiento_microred, obtener_seguimiento_compromiso_consentimiento_establecimiento)

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models.functions import Substr

from datetime import datetime
import locale

from django.db.models import IntegerField  # Importar IntegerField
from django.db.models.functions import Cast, Substr  # Importar Cast y Substr
# linea de border 
from openpyxl.utils import column_index_from_string

# Reporte excel
from datetime import datetime
import getpass  # Para obtener el nombre del usuario
from django.contrib.auth.models import User  # O tu modelo de usuario personalizado
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

User = get_user_model()

from django.db.models import IntegerField             # Importar IntegerField
from django.db.models.functions import Cast, Substr     # Importar Cast y Substr

logger = logging.getLogger(__name__)


def BASE(request):
    actualizacion = Actualizacion.objects.all()
    return render(request,'compromiso_consentimiento/index_compromiso_consentimiento.html', {"actualizacion": actualizacion})
############################
## COMPONENTES Y GRAFICOS 
############################
def process_avance_por_region(resultados_avance_por_region):
    data = {    
        'r_numerador_resumen': [],
        'r_denominador_resumen': [],
        'r_avance_resumen': []
    }
    
    # Si no hay resultados, retornar valores por defecto
    if not resultados_avance_por_region:
        data['r_numerador_resumen'] = [0]
        data['r_denominador_resumen'] = [0]
        data['r_avance_resumen'] = [0.0]
        #print(f"[PROCESS] Sin datos de entrada, usando valores por defecto: {data}")
        return data
    
    # Procesar el primer (y único) registro
    row = resultados_avance_por_region[0]
    try:
        # Mapear las claves correctas del resultado
        numerador = row.get('num', 0)
        denominador = row.get('den', 0) 
        avance = row.get('avance', 0.0)

        # Asegurar que los valores no sean None
        numerador = numerador if numerador is not None else 0
        denominador = denominador if denominador is not None else 0
        avance = avance if avance is not None else 0.0

        data['r_numerador_resumen'].append(int(numerador))
        data['r_denominador_resumen'].append(int(denominador))
        data['r_avance_resumen'].append(float(avance))
        
        #print(f"[PROCESS] Datos procesados: Num={numerador}, Den={denominador}, Avance={avance}%")
        
    except (KeyError, ValueError, TypeError) as e:
        logger.error(f"Error procesando datos: {e}, Row: {row}")
        # Valores por defecto en caso de error
        data['r_numerador_resumen'] = [0]
        data['r_denominador_resumen'] = [0]
        data['r_avance_resumen'] = [0.0]
    
    #print(f"[PROCESS] Resultado final: {data}")
    return data

def process_avance_regional_mensual_compromiso_consentimiento(resultados_avance_regional_mensual_compromiso_consentimiento):
    """Procesa los resultados del graficos"""
    data = {
        'num_1': [],
        'den_1': [],
        'cob_1': [],
        'num_2': [],
        'den_2': [],
        'cob_2': [],
        'num_3': [],
        'den_3': [],
        'cob_3': [],
        'num_4': [],
        'den_4': [],
        'cob_4': [],
        'num_5': [],
        'den_5': [],
        'cob_5': [],
        'num_6': [],
        'den_6': [],
        'cob_6': [],
        'num_7': [],
        'den_7': [],
        'cob_7': [],
        'num_8': [],
        'den_8': [],
        'cob_8': [],                
        'num_9': [],
        'den_9': [],
        'cob_9': [],
        'num_10': [],
        'den_10': [],
        'cob_10': [],
        'num_11': [],
        'den_11': [],
        'cob_11': [],
        'num_12': [],
        'den_12': [],
        'cob_12': [],
    }
    for index, row in enumerate(resultados_avance_regional_mensual_compromiso_consentimiento):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'num_1','den_1','cob_1','num_2','den_2','cob_2','num_3','den_3','cob_3','num_4','den_4','cob_4','num_5','den_5','cob_5','num_6','den_6','cob_6','num_7','den_7','cob_7','num_8','den_8','cob_8','num_9','den_9','cob_9','num_10','den_10','cob_10','num_11','den_11','cob_11','num_12','den_12','cob_12'}
            
            if not required_keys.issubset(row.keys()):
                raise ValueError(f"La fila {index} no tiene las claves necesarias: {row}")
            # Extraer cada valor, convirtiendo a float
            num_1_value = float(row.get('num_1', 0.0))
            den_1_value = float(row.get('den_1', 0.0))
            cob_1_value = float(row.get('cob_1', 0.0))
            num_2_value = float(row.get('num_2', 0.0))
            den_2_value = float(row.get('den_2', 0.0))
            cob_2_value = float(row.get('cob_2', 0.0))
            num_3_value = float(row.get('num_3', 0.0))
            den_3_value = float(row.get('den_3', 0.0))
            cob_3_value = float(row.get('cob_3', 0.0))
            num_4_value = float(row.get('num_4', 0.0))
            den_4_value = float(row.get('den_4', 0.0))
            cob_4_value = float(row.get('cob_4', 0.0))
            num_5_value = float(row.get('num_5', 0.0))
            den_5_value = float(row.get('den_5', 0.0))
            cob_5_value = float(row.get('cob_5', 0.0))
            num_6_value = float(row.get('num_6', 0.0))
            den_6_value = float(row.get('den_6', 0.0))
            cob_6_value = float(row.get('cob_6', 0.0))
            num_7_value = float(row.get('num_7', 0.0))
            den_7_value = float(row.get('den_7', 0.0))
            cob_7_value = float(row.get('cob_7', 0.0))
            num_8_value = float(row.get('num_8', 0.0))
            den_8_value = float(row.get('den_8', 0.0))
            cob_8_value = float(row.get('cob_8', 0.0))
            num_9_value = float(row.get('num_9', 0.0))
            den_9_value = float(row.get('den_9', 0.0))
            cob_9_value = float(row.get('cob_9', 0.0))
            num_10_value = float(row.get('num_10', 0.0))
            den_10_value = float(row.get('den_10', 0.0))
            cob_10_value = float(row.get('cob_10', 0.0))
            num_11_value = float(row.get('num_11', 0.0))
            den_11_value = float(row.get('den_11', 0.0))
            cob_11_value = float(row.get('cob_11', 0.0))
            num_12_value = float(row.get('num_12', 0.0))
            den_12_value = float(row.get('den_12', 0.0))
            cob_12_value = float(row.get('cob_12', 0.0))
            
            data['num_1'].append(num_1_value)
            data['den_1'].append(den_1_value)
            data['cob_1'].append(cob_1_value)
            data['num_2'].append(num_2_value)
            data['den_2'].append(den_2_value)
            data['cob_2'].append(cob_2_value)
            data['num_3'].append(num_3_value)
            data['den_3'].append(den_3_value)
            data['cob_3'].append(cob_3_value)
            data['num_4'].append(num_4_value)
            data['den_4'].append(den_4_value)
            data['cob_4'].append(cob_4_value)
            data['num_5'].append(num_5_value)
            data['den_5'].append(den_5_value)
            data['cob_5'].append(cob_5_value)
            data['num_6'].append(num_6_value)
            data['den_6'].append(den_6_value)
            data['cob_6'].append(cob_6_value)
            data['num_7'].append(num_7_value)
            data['den_7'].append(den_7_value)
            data['cob_7'].append(cob_7_value)
            data['num_8'].append(num_8_value)
            data['den_8'].append(den_8_value)
            data['cob_8'].append(cob_8_value)
            data['num_9'].append(num_9_value)
            data['den_9'].append(den_9_value)
            data['cob_9'].append(cob_9_value)
            data['num_10'].append(num_10_value)
            data['den_10'].append(den_10_value)
            data['cob_10'].append(cob_10_value)
            data['num_11'].append(num_11_value)
            data['den_11'].append(den_11_value)
            data['cob_11'].append(cob_11_value)
            data['num_12'].append(num_12_value)
            data['den_12'].append(den_12_value)
            data['cob_12'].append(cob_12_value)

        except Exception as e:
            logger.error(f"Error procesando la fila {index}: {str(e)}")
    return data

def process_variables_por_region(resultados_variables_por_region):
    data = {    
        'c_den_variable': [],
        'c_num_apn': [],
        'c_avance_apn': [],
        'c_num_consentimiento': [],
        'c_avance_consentimiento': []
    }
    
    # Si no hay resultados, retornar valores por defecto
    if not resultados_variables_por_region:
        data['c_den_variable'] = [0]
        data['c_num_apn'] = [0]
        data['c_avance_apn'] = [0.0]
        data['c_num_consentimiento'] = [0]
        data['c_avance_consentimiento'] = [0.0]
        return data
    
    # Procesar el primer (y único) registro
    row = resultados_variables_por_region[0]
    try:
        # Mapear las claves correctas del resultado        
        den_variable = row.get('den_variable', 0)
        num_apn = row.get('num_apn', 0)
        avance_apn = row.get('avance_apn', 0.0)
        num_consentimiento = row.get('num_consentimiento', 0)
        avance_consentimiento = row.get('avance_consentimiento', 0.0)

        # Asegurar que los valores no sean None        
        den_variable = den_variable if den_variable is not None else 0
        num_apn = num_apn if num_apn is not None else 0
        avance_apn = avance_apn if avance_apn is not None else 0
        num_consentimiento = num_consentimiento if num_consentimiento is not None else 0
        avance_consentimiento = avance_consentimiento if avance_consentimiento is not None else 0.0
        

        data['c_den_variable'].append(int(den_variable))
        data['c_num_apn'].append(int(num_apn))
        data['c_avance_apn'].append(float(avance_apn))
        data['c_num_consentimiento'].append(int(num_consentimiento))
        data['c_avance_consentimiento'].append(float(avance_consentimiento))

    except (KeyError, ValueError, TypeError) as e:
        logger.error(f"Error procesando datos: {e}, Row: {row}")
        # Valores por defecto en caso de error
        data['c_den_variable'] = [0]
        data['c_num_apn'] = [0]
        data['c_avance_apn'] = [0.0]
        data['c_num_consentimiento'] = [0]
        data['c_avance_consentimiento'] = [0.0]
    #print(f"[PROCESS] Resultado final: {data}")
    return data

def process_cobertura_por_zona(resultados_cobertura_por_zona):
    """Procesa los resultados del graficos"""
    data = {
            'z_zona': [],
            'z_den': [],
            'z_num': [],
            'z_brecha': [],
            'z_cob': [],
    }   
    for row in resultados_cobertura_por_zona:
        try:
            data['z_zona'].append(row['z_zona'])

            # Cambia null (None) a 0
            data['z_den'].append(row['z_den'] if row['z_den'] is not None else 0)
            data['z_num'].append(row['z_num'] if row['z_num'] is not None else 0)
            data['z_brecha'].append(row['z_brecha'] if row['z_brecha'] is not None else 0)
            data['z_cob'].append(row['z_cob'] if row['z_cob'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_provincia(resultados_cobertura_por_provincia):
    """Procesa los resultados del graficos"""
    data = {
            'p_provincia': [],
            'p_den': [],
            'p_num': [],
            'p_brecha': [],
            'p_cob': [],
    }   
    for row in resultados_cobertura_por_provincia:
        try:
            data['p_provincia'].append(row['p_provincia'])

            # Cambia null (None) a 0
            data['p_den'].append(row['p_den'] if row['p_den'] is not None else 0)
            data['p_num'].append(row['p_num'] if row['p_num'] is not None else 0)
            data['p_brecha'].append(row['p_brecha'] if row['p_brecha'] is not None else 0)
            data['p_cob'].append(row['p_cob'] if row['p_cob'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_distrito(resultados_cobertura_por_distrito):
    """Procesa los resultados del graficos"""
    data = {
            'd_distrito': [],
            'd_den': [],
            'd_num': [],
            'd_brecha': [],
            'd_cob': [],
    }   
    for row in resultados_cobertura_por_distrito:
        try:
            data['d_distrito'].append(row['d_distrito'])

            # Cambia null (None) a 0
            data['d_den'].append(row['d_den'] if row['d_den'] is not None else 0)
            data['d_num'].append(row['d_num'] if row['d_num'] is not None else 0)
            data['d_brecha'].append(row['d_brecha'] if row['d_brecha'] is not None else 0)
            data['d_cob'].append(row['d_cob'] if row['d_cob'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

##----------------------------
def process_avance_cobertura_compromiso_consentimiento(resultados_avance_cobertura_compromiso_consentimiento):
    """Procesa los resultados del graficos"""
    data = {
            'anio': [],
            'mes': [],
            'Ubigueo_Establecimiento':[],
            'Distrito': [],
            'Provincia': [],
            'Codigo_Red': [],
            'red': [],
            'Codigo_MicroRed': [],
            'microred': [],
            'Codigo_Unico': [],
            'Nombre_Establecimiento': [],
            'grupo_edad': [],
            'total_denominador': [],
            'total_numerador': [],
            'total_brecha': [],
            'cobertura_porcentaje': [],
    }   
    for row in resultados_avance_cobertura_compromiso_consentimiento:
        try:
            data['anio'].append(row['anio'])
            data['mes'].append(row['mes'])
            data['Ubigueo_Establecimiento'].append(row['Ubigueo_Establecimiento'])
            data['Distrito'].append(row['Distrito'])
            data['Provincia'].append(row['Provincia'])
            data['Codigo_Red'].append(row['Codigo_Red'])
            data['red'].append(row['red'])
            data['Codigo_MicroRed'].append(row['Codigo_MicroRed'])
            data['microred'].append(row['microred'])
            data['Codigo_Unico'].append(row['Codigo_Unico'])
            data['Nombre_Establecimiento'].append(row['Nombre_Establecimiento'])
            data['grupo_edad'].append(row['grupo_edad'])

            # Cambia null (None) a 0
            data['total_denominador'].append(row['total_denominador'] if row['total_denominador'] is not None else 0)
            data['total_numerador'].append(row['total_numerador'] if row['total_numerador'] is not None else 0)
            data['total_brecha'].append(row['total_brecha'] if row['total_brecha'] is not None else 0)
            data['cobertura_porcentaje'].append(row['cobertura_porcentaje'] if row['cobertura_porcentaje'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_edad(resultados_cobertura_por_edad):
    """Procesa los resultados del graficos"""
    data = {    
            'c_grupo_edad': [],
            'c_denominador': [],
            'c_numerador': [],
            'c_brecha': [],
            'c_cobertura': [],
    }
    for row in resultados_cobertura_por_edad:
        try:
            data['c_grupo_edad'].append(row['c_grupo_edad'])

            # Cambia null (None) a 0
            data['c_denominador'].append(row['c_denominador'] if row['c_denominador'] is not None else 0)
            data['c_numerador'].append(row['c_numerador'] if row['c_numerador'] is not None else 0)
            data['c_brecha'].append(row['c_brecha'] if row['c_brecha'] is not None else 0)
            data['c_cobertura'].append(row['c_cobertura'] if row['c_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_red(resultados_cobertura_por_red):
    """Procesa los resultados del graficos"""
    data = {
            'r_red': [],
            'r_denominador': [],
            'r_numerador': [],
            'r_brecha': [],
            'r_cobertura': [],
    }   
    for row in resultados_cobertura_por_red:
        try:
            data['r_red'].append(row['r_red'])

            # Cambia null (None) a 0
            data['r_denominador'].append(row['r_denominador'] if row['r_denominador'] is not None else 0)
            data['r_numerador'].append(row['r_numerador'] if row['r_numerador'] is not None else 0)
            data['r_brecha'].append(row['r_brecha'] if row['r_brecha'] is not None else 0)
            data['r_cobertura'].append(row['r_cobertura'] if row['r_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_microred(resultados_cobertura_por_microred):
    """Procesa los resultados del graficos"""
    data = {
            'm_microred': [],
            'm_denominador': [],
            'm_numerador': [],
            'm_brecha': [],
            'm_cobertura': [],
    }   
    for row in resultados_cobertura_por_microred:
        try:
            data['m_microred'].append(row['m_microred'])

            # Cambia null (None) a 0
            data['m_denominador'].append(row['m_denominador'] if row['m_denominador'] is not None else 0)
            data['m_numerador'].append(row['m_numerador'] if row['m_numerador'] is not None else 0)
            data['m_brecha'].append(row['m_brecha'] if row['m_brecha'] is not None else 0)
            data['m_cobertura'].append(row['m_cobertura'] if row['m_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_establecimiento(resultados_cobertura_por_establecimiento):
    """Procesa los resultados del graficos"""
    data = {
            'e_establecimiento': [],
            'e_denominador': [],
            'e_numerador': [],
            'e_brecha': [],
            'e_cobertura': [],
    }   
    for row in resultados_cobertura_por_establecimiento:
        try:
            data['e_establecimiento'].append(row['e_establecimiento'])

            # Cambia null (None) a 0
            data['e_denominador'].append(row['e_denominador'] if row['e_denominador'] is not None else 0)
            data['e_numerador'].append(row['e_numerador'] if row['e_numerador'] is not None else 0)
            data['e_brecha'].append(row['e_brecha'] if row['e_brecha'] is not None else 0)
            data['e_cobertura'].append(row['e_cobertura'] if row['e_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data


#######################
## PANTALLA PRINCIPAL
#######################
def index_compromiso_consentimiento(request):
    actualizacion = Actualizacion.objects.all()

    # Capturamos el año que viene por GET
    anio = request.GET.get('anio', '2025')
    mes = request.GET.get('mes', '')
    provincia = request.GET.get('provincia', '')
    distrito = request.GET.get('distrito', '')

    if anio not in ['2024', '2025' , '2026']:
        anio = '2025'

    redes_h = (
        MAESTRO_HIS_ESTABLECIMIENTO
        .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL', Disa='JUNIN')
        .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
        .values('Red', 'codigo_red_filtrado')
        .distinct()
        .order_by('Red')
    )
    
    provincias_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )

    mes_seleccionado_inicio = request.GET.get('mes_inicio')
    mes_seleccionado_fin = request.GET.get('mes_fin')
    
    provincia_seleccionada = request.GET.get('provincia_h')
    distrito_seleccionado = request.GET.get('distrito_h')
    
    red_seleccionada = request.GET.get('red_h')
    microred_seleccionada = request.GET.get('p_microredes_establec_h')
    establecimiento_seleccionado = request.GET.get('p_establecimiento_h')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Unificar la lógica de obtención de datos
            #print(f"Parámetros recibidos - Año: {anio}, Mes inicio: {mes_seleccionado_inicio}, Mes fin: {mes_seleccionado_fin}, Provincia: {provincia_seleccionada}, Distrito: {distrito_seleccionado}")
            
            # Lógica para la sección de avance regional
            resultados_avance_por_region = obtener_avance_compromiso_consentimiento(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            resultados_variables_por_region = obtener_variables_compromiso_consentimiento(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)           
            resultados_avance_regional_mensual_compromiso_consentimiento = obtener_avance_regional_mensual_compromiso_consentimiento(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado) 
            
            resultados_cobertura_por_zona = obtener_cobertura_por_zona(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            resultados_cobertura_por_provincia = obtener_cobertura_por_provincia(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            resultados_cobertura_por_distrito = obtener_cobertura_por_distrito(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            
            # Lógica para la sección de cobertura
            resultados_avance_cobertura_compromiso_consentimiento = obtener_avance_cobertura_compromiso_consentimiento(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            
            resultados_cobertura_por_red = obtener_cobertura_por_red(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            resultados_cobertura_por_microred = obtener_cobertura_por_microred(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            resultados_cobertura_por_establecimiento = obtener_cobertura_por_establecimiento(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            
            # Combinar todos los resultados en un solo diccionario
            data = {
                **process_avance_por_region(resultados_avance_por_region),
                **process_variables_por_region(resultados_variables_por_region),
                **process_avance_regional_mensual_compromiso_consentimiento(resultados_avance_regional_mensual_compromiso_consentimiento),
                **process_avance_cobertura_compromiso_consentimiento(resultados_avance_cobertura_compromiso_consentimiento),
                
                # Graficos de barras
                **process_cobertura_por_zona(resultados_cobertura_por_zona),
                **process_cobertura_por_provincia(resultados_cobertura_por_provincia),
                **process_cobertura_por_distrito(resultados_cobertura_por_distrito),
                
                **process_cobertura_por_red(resultados_cobertura_por_red),
                **process_cobertura_por_microred(resultados_cobertura_por_microred),
                **process_cobertura_por_establecimiento(resultados_cobertura_por_establecimiento)
            }

            return JsonResponse(data)

        except Exception as e:
            logger.error(f"Error al obtener datos: {str(e)}")
            return JsonResponse({'error': f"Error al obtener datos: {str(e)}"}, status=500)

    # Renderizado inicial de la página
    return render(request, 'compromiso_consentimiento/index_compromiso_consentimiento.html', {
        'mes_seleccionado_inicio': mes_seleccionado_inicio,
        'mes_seleccionado_fin': mes_seleccionado_fin,
        'actualizacion': actualizacion,
        'provincia_seleccionada': provincia_seleccionada,
        'distrito_seleccionado': distrito_seleccionado,
        'provincias_h': provincias_h,
    })

#####################################
## FILTROS HORIZONTAL
#####################################
##----------------------------------
## FILTROS HORIZONTAL POR SALUD
##----------------------------------
def get_establecimientos_compromiso_consentimiento_h(request,establecimiento_id):
    redes_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
                .order_by('Red')
    )
    provincias_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes_h': redes_h,
                'provincias_h': provincias_h,
                'mes_inicio':mes_inicio
    }
    return render(request,'compromiso_consentimiento/establecimientos_h.html', context)

def p_microredes_establec_compromiso_consentimiento_h(request):
    redes_param = request.GET.get('red_h') 
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    return render(request, 'compromiso_consentimiento/partials/p_microredes_establec_h.html', context)

def p_establecimientos_compromiso_consentimiento_h(request):
    microredes = request.GET.get('p_microredes_establec_h', '')    
    codigo_red = request.GET.get('red_h', '')
    
    # Construir el filtro dinámicamente
    filtros = {
        'Descripcion_Sector': 'GOBIERNO REGIONAL',
        'Disa': 'JUNIN'
    }
    
    if microredes:
        filtros['Codigo_MicroRed'] = microredes
    if codigo_red:
        filtros['Codigo_Red__startswith'] = codigo_red
    
    establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(**filtros).values('Codigo_Unico','Nombre_Establecimiento').distinct()
    
    # Debug: contar establecimientos encontrados
    establec_list = list(establec)
    
    context= {
        'establec': establec_list
    }
    return render(request, 'compromiso_consentimiento/partials/p_establecimientos_h.html', context)

##-----------------------------------
## FILTROS HORIZONTAL POR MUNICIPIO
##-----------------------------------
def p_distritos_compromiso_consentimiento_h(request):
    provincias_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    
    provincia_param = request.GET.get('provincia_h')
    
    distritos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
        Ubigueo_Establecimiento__startswith=provincia_param,
        Descripcion_Sector='GOBIERNO REGIONAL'
    ).values('Ubigueo_Establecimiento', 'Distrito').distinct().order_by('Distrito')
    context = {
        'distritos': distritos,
        'provincias_h' :  provincias_h
    }
    return render(request, 'compromiso_consentimiento/partials/p_distritos.html', context)

###########################################
## SEGUIMIENTO NOMINAL FILTROS
##########################################

###################---------------------------
## FILTRO AMBITO DE SALUD
##################---------------------------

## SEGUIMIENTO POR REDES
def get_redes_compromiso_consentimiento(request,redes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'compromiso_consentimiento/components/salud/redes.html', context)

## SEGUIMIENTO POR MICRO-REDES
def get_microredes_compromiso_consentimiento(request, microredes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 'compromiso_consentimiento/components/salud/microredes.html', context)

def p_microredes_compromiso_consentimiento(request):
    redes_param = request.GET.get('red', '')
    

    
    # Consulta principal
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
        Codigo_Red__startswith=redes_param, 
        Descripcion_Sector='GOBIERNO REGIONAL', 
        Disa='JUNIN'
    ).values('Codigo_MicroRed','MicroRed').distinct()
    
    microredes_list = list(microredes)
    
    context = {
        'redes_param': redes_param,
        'microredes': microredes_list
    }
    
    # Si es una petición AJAX, devolver solo las opciones
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'compromiso_consentimiento/partials/p_microredes_options.html', context)
    
    return render(request, 'compromiso_consentimiento/partials/p_microredes.html', context)

## REPORTE POR ESTABLECIMIENTO
def get_establecimientos_compromiso_consentimiento(request,establecimiento_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
                .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request,'compromiso_consentimiento/components/salud/establecimientos.html', context)

def p_microredes_establec_compromiso_consentimiento(request):
    redes_param = request.GET.get('red') 
    

    
    # Usar startswith en lugar de igualdad exacta
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
        Codigo_Red__startswith=redes_param, 
        Descripcion_Sector='GOBIERNO REGIONAL',
        Disa='JUNIN'
    ).values('Codigo_MicroRed','MicroRed').distinct()
    
    microredes_list = list(microredes)
    
    context = {
        'microredes': microredes_list,
        'is_htmx': True
    }
    return render(request, 'compromiso_consentimiento/partials/p_microredes_establec.html', context)

def p_establecimientos_compromiso_consentimiento(request):
    # Limpiar parámetros - obtener solo el primer valor si hay duplicados
    microredes = request.GET.get('p_microredes_establec', '').strip()
    
    # Para el código de red, obtener todos los valores y usar el primero no vacío
    red_values = request.GET.getlist('red')
    codigo_red = ''
    for red_val in red_values:
        if red_val and red_val.strip():
            codigo_red = red_val.strip()
            break
    
    # Construir el filtro dinámicamente
    filtros = {
        'Descripcion_Sector': 'GOBIERNO REGIONAL',
        'Disa': 'JUNIN'
    }
    
    # Agregar filtro de red si está disponible
    if codigo_red:
        filtros['Codigo_Red__startswith'] = codigo_red
    
    # Agregar filtro de microred si está presente
    if microredes:
        filtros['Codigo_MicroRed'] = microredes
        
        # Si hay microred pero no hay red, intentar obtener la red de la microred
        if not codigo_red:
            try:
                red_desde_microred = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
                    Codigo_MicroRed=microredes,
                    Descripcion_Sector='GOBIERNO REGIONAL',
                    Disa='JUNIN'
                ).values('Codigo_Red').first()
                
                if red_desde_microred:
                    codigo_red_obtenido = red_desde_microred['Codigo_Red']
                    filtros['Codigo_Red'] = codigo_red_obtenido
            except Exception as e:
                pass
        
    try:
        establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(**filtros).values('Codigo_Unico','Nombre_Establecimiento').distinct()
        establec_list = list(establec)
    except Exception as e:
        establec_list = []
    
    context= {
        'establec': establec_list
    }
    
    # Si es una petición AJAX, devolver solo las opciones
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'compromiso_consentimiento/partials/p_establecimientos_options.html', context)
    
    return render(request, 'compromiso_consentimiento/partials/p_establecimientos.html', context)

######################---------------------------
## FILTRO AMBITO DE MUNICIPIO
######################-------------------------------

## SEGUIMIENTO POR PROVINCIA
def get_provincias_compromiso_consentimiento(request, provincia_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    )
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
            }
    
    return render(request, 'compromiso_consentimiento/components/municipio/provincias.html', context)

## SEGUIMIENTO POR DISTRITOS
def get_distritos_compromiso_consentimiento(request, distrito_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'compromiso_consentimiento/components/municipio/distritos.html', context)

def p_distrito_compromiso_consentimiento(request):
    provincia_param = request.GET.get('provincia', '')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 'compromiso_consentimiento/partials/p_distritos.html', context)


########################################
## SEGUIMIENTO REPORTE EXCEL 
#######################################

## REPORTE DE EXCEL
class RptconsentimientoCompromiso(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        anio = request.GET.get('anio', '2025')
        mes_inicio = request.GET.get('fecha_inicio', '')
        mes_fin = request.GET.get('fecha_fin', '')
        provincia = request.GET.get('provincia', '')
        distrito = request.GET.get('distrito', '')
        p_red = request.GET.get('red', '')
        p_microredes = request.GET.get('p_microredes', '')
        p_establecimiento = request.GET.get('p_establecimiento', '')
        p_cumple = request.GET.get('cumple', '') 

        # Creación de la consulta
        #print(f"Año: {anio}, Mes Inicio: {mes_inicio}, Mes Fin: {mes_fin}, Provincia: {provincia}, Distrito: {distrito}, Red: {p_red}, Microredes: {p_microredes}, Establecimiento: {p_establecimiento}, Cumple: {p_cumple}")
        resultado_seguimiento_compromiso_consentimiento = obtener_seguimiento_compromiso_consentimiento(anio, mes_inicio, mes_fin, provincia, distrito, p_red, p_microredes, p_establecimiento, p_cumple)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento_compromiso_consentimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_compromiso_consentimiento.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptPnPoblacionMicroRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_departamento = 'JUNIN'
        p_red = request.GET.get('red', '')
        p_microred = request.GET.get('microredes', '')
        p_establec = ''
        p_edades =  request.GET.get('edades','')
        p_cumple = request.GET.get('cumple', '') 
        # Creación de la consulta
        resultado_seguimiento_microred = obtener_seguimiento_compromiso_consentimiento_microred(p_departamento, p_red, p_microred, p_edades, p_cumple)

        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento_microred)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_compromiso_consentimiento_microred.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptPnPoblacionEstablec(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_departamento = 'JUNIN'
        p_red = request.GET.get('red','')
        p_microred = request.GET.get('p_microredes_establec','')  # Corregido
        p_establec = request.GET.get('p_establecimiento','')
        p_mes = request.GET.get('mes', '')
        p_edades = request.GET.get('edades', '')
        # Manejo seguro de fechas - usar valores por defecto si no están presentes
        p_cumple = request.GET.get('cumple', '')
        
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_compromiso_consentimiento_establecimiento(p_departamento,p_establec,p_edades,p_cumple)
                
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_compromiso_consentimiento_establecimiento.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response


def fill_worksheet(ws, results): 
# cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 12
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 30
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 1
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 9
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 5
    ws.column_dimensions['K'].width = 9
    ws.column_dimensions['L'].width = 5
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 15
    ws.column_dimensions['P'].width = 18
    ws.column_dimensions['Q'].width = 18
    ws.column_dimensions['R'].width = 5
    ws.column_dimensions['S'].width = 18
    ws.column_dimensions['T'].width = 5
    ws.column_dimensions['U'].width = 20
    ws.column_dimensions['V'].width = 10
    ws.column_dimensions['W'].width = 25
        
    # linea de division
    ws.freeze_panes = 'C10'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    # Definir el estilo de relleno celeste
    celeste_fill = PatternFill(start_color='FF87CEEB', end_color='FF87CEEB', fill_type='solid')
    # Morado más claro
    morado_claro_fill = PatternFill(start_color='FFE9D8FF', end_color='FFE9D8FF', fill_type='solid')
    # Plomo más claro
    plomo_claro_fill = PatternFill(start_color='FFEDEDED', end_color='FFEDEDED', fill_type='solid')
    # Azul más claro
    azul_claro_fill = PatternFill(start_color='FFD8EFFA', end_color='FFD8EFFA', fill_type='solid')
    # Naranja más claro
    naranja_claro_fill = PatternFill(start_color='FFFFEBD8', end_color='FFFFEBD8', fill_type='solid')
    # Verde más claro
    verde_claro_fill = PatternFill(start_color='FFBDF7BD', end_color='FFBDF7BD', fill_type='solid')
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    border_negro = Border(left=Side(style='thin', color='000000'), # negro
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000')) 
    
    # Merge cells 
    # numerador y denominador
    ws.merge_cells('B6:H6') 
    ws.merge_cells('I6:W6')
    
    # CABECERA NOMBRES
    #ws.merge_cells('B6:H6') 
    #ws.merge_cells('I6:W6')

    # Auxiliar HORIZONTAL
    ws.merge_cells('M7:M8')
    ws.merge_cells('N7:N8')
    
    # intervalo
    ws.merge_cells('B7:C7')
    ws.merge_cells('D7:H7')
    ws.merge_cells('I7:J7')
    ws.merge_cells('K7:L7')
    ws.merge_cells('O7:W7')
        
    # COD HIS
    ws.merge_cells('B8:H8')
    ws.merge_cells('I8:J8')
    ws.merge_cells('K8:L8')
    ws.merge_cells('O8:W8')

    # Combina cela
    ws['B6'] = 'META (DENOMINADOR)'
    ws['I6'] = 'AVANCE (NUMERADOR)'
    
    
    # INTERVALO
    #ws['S7'] = 'NUMERADOR PARCIAL'

    ws['D7'] = 'Datos de la madre del HIS MINSA'    
    ws['I7'] = 'Atencion prenatal de la MadreEstablecimiento de Salud del Padron Nominal'
    ws['K7'] = 'Compromiso de Consentimiento Informado'
    # CODIGO HIS
    
    ws['I8'] = 'DX = Z3491,Z3492,Z3493 '
    ws['K8'] = 'DX = 99199.25'

    
    ### numerador y denominador     
    ws['B6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B6'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['B6'].fill = gray_fill
    ws['B6'].border = border_negro

    ws['I6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I6'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['I6'].fill = naranja_claro_fill
    ws['I6'].border = border_negro
    
    
    #intervalos 
    ws['B7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B7'].font = Font(name = 'Arial', size= 7)
    ws['B7'].fill = naranja_claro_fill
    ws['B7'].border = border_negro
    
    ws['D7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D7'].font = Font(name = 'Arial', size= 7)
    ws['D7'].fill = plomo_claro_fill
    ws['D7'].border = border_negro
    
    ws['I7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I7'].font = Font(name = 'Arial', size= 7)
    ws['I7'].fill = plomo_claro_fill
    ws['I7'].border = border_negro
    
    ws['K7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K7'].font = Font(name = 'Arial', size= 7)
    ws['K7'].fill = plomo_claro_fill
    ws['K7'].border = border_negro
        
    
    # CODIGO HIS
    
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B8'].font = Font(name = 'Arial', size= 7)
    ws['B8'].fill = azul_claro_fill
    ws['B8'].border = border_negro
    
    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 7)
    ws['I8'].fill = azul_claro_fill
    ws['I8'].border = border_negro
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 7)
    ws['K8'].fill = azul_claro_fill
    ws['K8'].border = border_negro
    
    
    ws['B7'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B7'].fill = plomo_claro_fill
    ws['B7'].border = border_negro
    ws['B7'] = 'INTERVALO'
    
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B8'].fill = azul_claro_fill
    ws['B8'].border = border_negro
    ws['B8'] = 'COD HIS'
    
    ### BORDE DE CELDAS CONBINADAS
    
    # NUM y DEN
    inicio_columna = 'B'
    fin_columna = 'W'
    fila = 6
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
        
    # INTERVALO
    inicio_columna = 'B'
    fin_columna = 'W'
    fila = 7
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
        
    # CODIGO HIS 
    inicio_columna = 'B'
    fin_columna = 'W'
    fila = 8
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    ##### imprimer fecha y hora del reporte
    fecha_hora_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nombre_usuario = getpass.getuser()

    # Obtener el usuario actualmente autenticado
    try:
        user = User.objects.get(is_active=True)
    except User.DoesNotExist:
        user = None
    except User.MultipleObjectsReturned:
        # Manejar el caso donde hay múltiples usuarios activos
        user = User.objects.filter(is_active=True).first()  # Por ejemplo, obtener el primero
    
    # Asignar fecha y hora a la celda A1
    ws['V1'].value = 'Fecha y Hora:'
    ws['W1'].value = fecha_hora_actual

    # Asignar nombre de usuario a la celda A2
    ws['V2'].value = 'Usuario:'
    ws['W2'].value = nombre_usuario
    
    # Formatear las etiquetas en negrita
    etiqueta_font = Font(name='Arial', size=8)
    ws['V1'].font = etiqueta_font
    ws['W1'].font = etiqueta_font
    ws['V2'].font = etiqueta_font
    ws['W2'].font = etiqueta_font

    # Alinear el texto
    ws['V1'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W1'].alignment = Alignment(horizontal="left", vertical="center")
    ws['V2'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W2'].alignment = Alignment(horizontal="left", vertical="center")
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL CONSENTIMIENTO INFORMADO DE LA GESTANTE'
    
    ws['B3'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B3'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B3'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'
        
    ws['B9'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['B9'].fill = fill
    ws['B9'].border = border
    ws['B9'] = 'DNI'
    
    ws['C9'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['C9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['C9'].fill = fill
    ws['C9'].border = border
    ws['C9'] = 'NOMBRE DE LA GESTANTE'
    
    ws['D9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['D9'].fill = fill
    ws['D9'].border = border
    ws['D9'] = 'INICIO GEST'      
    
    ws['E9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['E9'].fill = fill
    ws['E9'].border = border
    ws['E9'] = 'SEM 14' 
    
    ws['F9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['F9'].fill = fill
    ws['F9'].border = border
    ws['F9'] = 'SEM 28'     
    
    ws['G9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['G9'].fill = fill
    ws['G9'].border = border
    ws['G9'] = 'SEM 33'    
    
    ws['H9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['H9'].fill = fill
    ws['H9'].border = border
    ws['H9'] = 'SEM 37'    
    
    ws['I9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['I9'].fill = green_fill
    ws['I9'].border = border
    ws['I9'] = 'APN'    
    
    ws['J9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['J9'].fill = green_fill
    ws['J9'].border = border
    ws['J9'] = 'VAL'  
    
    ws['K9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['K9'].fill = green_fill_2
    ws['K9'].border = border
    ws['K9'] = 'CONSENTIMIENTO'  
    
    ws['L9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['L9'].fill = green_fill_2
    ws['L9'].border = border
    ws['L9'] = 'VAL'  
    
    ws['M9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['M9'].fill = yellow_fill
    ws['M9'].border = border
    ws['M9'] = 'MES'  
    
    ws['N9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['N9'].fill = yellow_fill
    ws['N9'].border = border
    ws['N9'] = 'IND'  
    
    ws['O9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['O9'].fill = orange_fill
    ws['O9'].border = border
    ws['O9'] = 'ZONA'  
    
    ws['P9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['P9'].fill = orange_fill
    ws['P9'].border = border
    ws['P9'] = 'PROVINCIA'  
    
    ws['Q9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Q9'].fill = orange_fill
    ws['Q9'].border = border
    ws['Q9'] = 'DISTRITO'    
    
    ws['R9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['R9'].fill = orange_fill
    ws['R9'].border = border
    ws['R9'] = 'COD RED' 
    
    ws['S9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['S9'].fill = orange_fill
    ws['S9'].border = border
    ws['S9'] = 'RED' 
    
    ws['T9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['T9'].fill = orange_fill
    ws['T9'].border = border
    ws['T9'] = 'COD MICRO' 
    
    ws['U9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['U9'].fill = orange_fill   
    ws['U9'].border = border
    ws['U9'] = 'MICRORED' 
    
    ws['V9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V9'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['V9'].fill = orange_fill
    ws['V9'].border = border
    ws['V9'] = 'COD EESS' 
    
    ws['W9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W9'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['W9'].fill = orange_fill
    ws['W9'].border = border
    ws['W9'] = 'ESTABLECIMIENTO DE SALUD'   
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True)
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    
    # Definir los caracteres especiales de check y X
    check_mark = '✓'  # Unicode para check
    x_mark = '✗'  # Unicode para X
    sub_cumple = 'CUMPLE'
    sub_no_cumple = 'NO CUMPLE'
    
    # Escribir datos
    for row, record in enumerate(results, start=10):
        for col, value in enumerate(record.values(), start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas específicas
            if col in [17, 21, 23] :  # Columnas que deben estar alineadas a la izquierda
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar color plomo a las columnas D y E (columnas 4 y 5)
            if col in [4, 5]:  # Columnas D y E
                cell.fill = gray_fill  # Usar el fill plomo que ya tienes definido
                cell.font = Font(name='Arial', size=8)
            
            # Aplicar color en la columna consentimiento
            if col == 14:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=8, bold = True,color="FFFFFF")  # Letra blanca
                    elif value_upper == "CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')  # Fondo verde
                        cell.font = Font(name='Arial', size=8,  bold = True,color="FFFFFF")  # Letra blanca
                    else:
                        cell.font = Font(name='Arial', size=8, bold = True)
                else:
                    cell.font = Font(name='Arial', size=8,  bold = True)
            
            # Aplicar color de letra SUB consentimientoES
            elif col in [10,12]:
                if value == 0:
                    cell.value = sub_no_cumple  # Insertar check
                    cell.font = Font(name='Arial', size=7, color="FF0000") 
                elif value == 1:
                    cell.value = sub_cumple # Insertar check
                    cell.font = Font(name='Arial', size=7, color="00B050")
                else:
                    cell.font = Font(name='Arial', size=7)
            # Fuente normal para otras columnas
            
            # Aplicar color de letra SUB GENERALIDADES
            elif col in [39]:
                if value == 0:
                    cell.value = sub_no_cumple  # Insertar check
                    cell.font = Font(name='Arial', size=7, color="FF0000") 
                    cell.fill = PatternFill(patternType='solid', fgColor='FFEDEDED')  
                    cell.fill = gray_fill # Letra roja
                elif value == 1:
                    cell.value = sub_cumple # Insertar check
                    cell.font = Font(name='Arial', size=7, color="00B050")
                    cell.fill = PatternFill(patternType='solid', fgColor='FFEDEDED') 
                    cell.fill = gray_fill# Letra verde
                else:
                    cell.font = Font(name='Arial', size=7)
            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)  # Fuente normal para otras columnas
            
            # Aplicar caracteres especiales check y X
            if col in [10, 12]:
                if value == 1:
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == 0:
                    cell.value = x_mark  # Insertar X
                    cell.font = Font(name='Arial', size=10, color='FF0000')  # Letra roja
                else:
                    cell.font = Font(name='Arial', size=8)  # Fuente normal si no es 1 o 0
            
            cell.border = border

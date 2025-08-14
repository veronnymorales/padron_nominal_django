from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.shortcuts import render, redirect
from django.db import connection

# filtros
from base.models import DimPeriodo, DimDiscapacidadEtapa, MAESTRO_HIS_ESTABLECIMIENTO
from .models import DimDisFisicaCie,TramaBaseDiscapacidadRpt02FisicaNominal
from django.db.models import Case, When, Value, IntegerField, Sum, OuterRef, Subquery
from django.db.models.functions import Substr, Cast, Concat, Replace
from django.db.models import CharField, F

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from .utils import generar_operacional

from django.db.models.functions import Substr

from datetime import datetime
import locale

################################################
# REPORTE DE SEGUIMIENTO
################################################
@login_required
def ate_atc_discapacidad(request):
    return render(request, 'discapacidad/index_ate_atc.html')

#--- PROVINCIAS -------------------------------------------------------------
def ate_atc_get_provincias(request,provincias_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
            }
    
    return render(request, 'discapacidad/ate_atc_provincias.html', context)

#--- FUNCIONES OPERACIONALES PARTES REPORTE -----------------------------------------
def rpt_operacional_fisico(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT DISTINCT
	                SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
	                renaes,
	                Categoria,
	                SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC 
                LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s
                AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes, Categoria
        """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        resultado_prov = cursor.fetchall()
    return resultado_prov

def rpt_operacional_sensorial(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes, Categoria
            """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_sensorial = cursor.fetchall()
    return resultado_prov_sensorial

def rpt_operacional_certificado(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes, Categoria
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_certificado = cursor.fetchall()
    return resultado_prov_certificado

def rpt_operacional_rbc(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes, Categoria
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_rbc = cursor.fetchall()
    return resultado_prov_rbc

def rpt_operacional_capacitacion(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes, Categoria
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_capacitacion = cursor.fetchall()
    return resultado_prov_capacitacion

def rpt_operacional_agente(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_236) AS dis_236,
                        SUM(dis_237) AS dis_237,
                        SUM(dis_238) AS dis_238,
                        SUM(dis_239) AS dis_239,
                        SUM(dis_240) AS dis_240,
                        SUM(dis_241) AS dis_241
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 THEN 1 ELSE 0 END) 		 AS dis_236,
                            SUM(CASE WHEN Categoria = 1 THEN gedad ELSE 0 END)   AS dis_237,
                            SUM(CASE WHEN Categoria = 2 THEN 1 ELSE 0 END)     AS dis_238,
                            SUM(CASE WHEN Categoria = 2 THEN gedad ELSE 0 END)   AS dis_239,
                            SUM(CASE WHEN Categoria = 3 THEN 1 ELSE 0 END)     AS dis_240,
                            SUM(CASE WHEN Categoria = 3 THEN gedad ELSE 0 END)   AS dis_241
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s     
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_agente = cursor.fetchall()
    return resultado_prov_agente

def rpt_operacional_comite(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,
                        Actividad,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes, Actividad
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_comite = cursor.fetchall()
    return resultado_prov_comite

def get_categoria_matriz_fisico():
    return {
        '1':  'Lesiones medulares',
        '2':  'Enfermedad de Parkinson y similares',
        '3':  'Amputados de miembros superiores',
        '4':  'Amputados de miembros inferiores',
        '5':  'Enfermedades cerebrovasculares',
        '6':  'Enfermedades musculares y de la unión mioneural',
        '7':  'Lesiones de nervios periféricos',
        '8':  'Trastornos del desarrollo de la funcion motriz',
        '9':  'Enfermedad articular degenerativa',
        '10': 'Encefalopatía infantil y otras lesiones',
        '11': 'Sindrome de Down',
        '12': 'Cifosis y lordosis',
        '13': 'Espondilo artropatías',
        '14': 'Otros trastornos de los discos intervertebrales',
        '15': 'Cervicalgia, dorsalgia, lumbago',
        '16': 'Otras dorsopatías deformantes',
        '17': 'Otros trastornos articulares',
        '18': 'Defectos en la longitud de extremidades',
        '19': 'Enfermedad cardiovascular',
        '20': 'Enfermedad respiratoria',
        '21': 'Vejiga neurogénica y dolor',
        '22': 'Incontinencia',
        '23': 'Prolapso',
        '24': 'Traumatismos',
        '25': 'Enfermedades del tejido conectivo',
        '26': 'Patología articular excluida columna',
        '27': 'Lesiones infecciosas',
        '28': 'Lesión biomecánica',
        '29': 'Linfedema',
        '30': 'Sarcopenia',
        '31': 'Dolor',
        '32': 'Quemaduras, corrosiones y congelaciones',
    }

def get_categoria_matriz_sensorial():
    return {
        '1': 'Hipoacusia y sordera',
        '2': 'Baja visión y ceguera',
        '3': 'Sordomudez',
        '4': 'Enfermedad Cerebro vascular',
        '5': 'Trastornos específicos del desarrollo del habla y lenguaje',
        '6': 'Disartria',
        '7': 'Disfagia',     
        '8': 'Trastornos del aprendizaje',
        '9': 'Retardo Mental: Leve, moderado, severo',
        '10':'Trastornos del espectro autista',
        '11':'Otras alteraciones de salud mental',
    }

def get_categoria_matriz_certificado():
    return {
        '1': 'Certificación de Discapacidad (0515204) - Evaluación',
        '2': 'Certificación de Discapacidad (0515204) - Calificación',
        '3': 'Certificación de Discapacidad (0515204) - Certificación',
        '4': 'Certificación de Incapacidad (0515205)',
    }

def get_categoria_matriz_rbc():
    return {
        '1': '1º Visita',	
        '2': '2º Visita',	
        '3': '3º Visita',
        '4': '4º a Visita (trazador)',	
        '5': '5º a + Visitas',	   
    }      

def get_categoria_matriz_capacitacion():
    return {
        '1': 'Capacitación (C0009)',
    }

def get_categoria_matriz_agente():
    return {
        '1': 'Capacitación a Agentes Comunitarios (APP138)',
    }

def get_categoria_matriz_comite():
    return {
        '1': 'Actividad con Comité Multisectorial (APP96)',
    }

# validar matriz
def crear_matriz(request):    
    ubigeo = '1201'
    fecha_inicio = '20240102'# Ejemplo de ubigeo
    fecha_fin = '20240110'# Ejemplo de ubigeo
    matriz = rpt_operacional_fisico(ubigeo,fecha_inicio,fecha_fin)  
    # Puedes renderizar la matriz en una plantilla HTML o hacer cualquier otro procesamiento necesario
    return render(request, 'discapacidad/matrizes.html', {'matriz': matriz})
#########################################################
###--- PROVINCIAS EXCEL ----#############################
#########################################################
class Atc_Ate_RptOperacinalProv(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        provincia = request.GET.get('provincia')

        provincia_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Ubigueo_Establecimiento__startswith=provincia
        ).values_list('Provincia', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
        
        # Creación de la consulta
        resultado_prov = rpt_operacional_fisico(provincia, fecha_inicio, fecha_fin)
        resultado_prov_sensorial = rpt_operacional_sensorial(provincia, fecha_inicio, fecha_fin)
        resultado_prov_certificado = rpt_operacional_certificado(provincia, fecha_inicio, fecha_fin)
        resultado_prov_rbc = rpt_operacional_rbc(provincia, fecha_inicio, fecha_fin)       
        resultado_prov_capacitacion = rpt_operacional_capacitacion(provincia, fecha_inicio, fecha_fin)
        resultado_prov_agente = rpt_operacional_agente(provincia, fecha_inicio, fecha_fin)
        resultado_prov_comite = rpt_operacional_comite(provincia, fecha_inicio, fecha_fin)
        
        wb = Workbook()
        
        consultas = [
                ('Físico', resultado_prov, get_categoria_matriz_fisico),
                ('Sensorial', resultado_prov_sensorial, get_categoria_matriz_sensorial),
                ('Certificado', resultado_prov_certificado, get_categoria_matriz_certificado),
                ('RBC', resultado_prov_rbc, get_categoria_matriz_rbc),
                ('Capacitacion', resultado_prov_capacitacion, get_categoria_matriz_capacitacion),
                ('Agente', resultado_prov_agente, get_categoria_matriz_agente),
                ('Comite', resultado_prov_comite, get_categoria_matriz_comite),
        ]
        
        # Configurar locale para español
        # locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        
        # Obtener nombres de meses en español
        #meses_espanol = [datetime(2024, m, 1).strftime('%b').capitalize() for m in range(1, 13)]
        
        for index, (sheet_name, results, get_categoria_matriz) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            categoria_matriz = get_categoria_matriz()
            fill_worksheet(ws, results, categoria_matriz,provincia_codigo,fecha_inicio_codigo,fecha_fin_codigo)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_ate_atc_provincia.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response
###################################################################################
###################################################################################   
def fill_worksheet(ws, results, categoria_matriz, provincia_codigo,fecha_inicio_codigo,fecha_fin_codigo): 
    # cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[8].height = 25
    ws.row_dimensions[9].height = 25
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 6
    ws.column_dimensions['J'].width = 6
    ws.column_dimensions['K'].width = 6
    ws.column_dimensions['L'].width = 6
    ws.column_dimensions['M'].width = 6
    ws.column_dimensions['N'].width = 6
    ws.column_dimensions['O'].width = 6
    ws.column_dimensions['P'].width = 6
    ws.column_dimensions['Q'].width = 6
    ws.column_dimensions['R'].width = 6
    ws.column_dimensions['S'].width = 6
    ws.column_dimensions['T'].width = 6
    ws.column_dimensions['U'].width = 6
    ws.column_dimensions['V'].width = 6
    ws.column_dimensions['W'].width = 6
    ws.column_dimensions['X'].width = 6
    ws.column_dimensions['Y'].width = 6
    ws.column_dimensions['Z'].width = 6
    ws.column_dimensions['AA'].width = 6
    ws.column_dimensions['AB'].width = 6
    # linea de division
    ws.freeze_panes = 'E10'
    # Configuración del fondo y el borde
    fill = PatternFill(patternType='solid', fgColor='00B0F0')
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo

    # Agregar el código de provincia
    ws['D7'] = provincia_codigo[0] if provincia_codigo else ''
    ws['D7'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['D7'].font = Font(name='Arial', size=7)
    
    ws['I6'] = fecha_inicio_codigo[0] if provincia_codigo else ''
    ws['I6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['I6'].font = Font(name='Arial', size=7)
    
    ws['I7'] = fecha_fin_codigo[0] if provincia_codigo else ''
    ws['I7'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['I7'].font = Font(name='Arial', size=7)
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'REPORTE DE ATENDIDOS Y ATENCIONES DE DISCAPACIDAD'
    
    ws['B6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B6'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B6'] ='DIRESA / GERESA / DISA'
    
    ws['D6'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['D6'].font = Font(name = 'Arial', size= 7)
    ws['D6'] ='JUNIN'
    
    ws['B7'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B7'] ='PROV/ DIST/ RED/ MR/ ESTABLEC'
    
    ws['E6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['E6'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['E6'] ='PERIODO'
    
    ws['G6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['G6'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['G6'] ='MES INICIO'
    
    ws['G7'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['G7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['G7'] ='MES FIN'
    
    ws['B9'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B9'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['B9'].fill = fill
    ws['B9'].border = border
    ws['B9'] = 'UBIGEO'
    
    ws['C9'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['C9'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['C9'].fill = fill
    ws['C9'].border = border
    ws['C9'] = 'IPRESS'
    
    ws['D9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D9'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['D9'].fill = fill
    ws['D9'].border = border
    ws['D9'] = 'MORBILIDADES'      
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')

    meses_espanol = [datetime(2024, m, 1).strftime('%b').capitalize() for m in range(1, 13)]

    # Escribir encabezados con meses
    ws.merge_cells('E8:P8')
    ws.cell(row=8, column=5, value="ATENDIDOS").font = header_font
    ws.cell(row=8, column=5).alignment = centered_alignment
    ws.cell(row=8, column=5).fill = header_fill
    
    ws.merge_cells('Q8:AB8')
    ws.cell(row=8, column=17, value="ATENCIONES").font = header_font
    ws.cell(row=8, column=17).alignment = centered_alignment
    ws.cell(row=8, column=17).fill = PatternFill(patternType='solid', fgColor='A9A9A9')

    for col, mes in enumerate(meses_espanol, start=5):
        cell = ws.cell(row=9, column=col, value=mes)
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = border
        cell.fill = header_fill
        
        cell = ws.cell(row=9, column=col+12, value=mes)
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = border
        cell.fill = PatternFill(patternType='solid', fgColor='A9A9A9')

    # Escribir datos
    for row, record in enumerate(results, start=10):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name = 'Arial', size= 8)
            cell.border = border
            
            # Comparar categoría con la matriz    
            if col == 4:  # Columna de Categoría
                categoria_nombre = categoria_matriz.get(str(value), f'Categoría {value}')
                ws.cell(row=row, column=col, value=f"{value} - {categoria_nombre}")
                cell.alignment = Alignment(horizontal='left')

################################################
# REPORTE DE DISTRITO
################################################
def ate_atc_get_distritos(request, distritos_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'discapacidad/ate_atc_distritos.html', context)

def p_distritos(request):
    provincia_param = request.GET.get('provincia')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 'discapacidad/partials/p_distritos.html', context)

#--- FUNCIONES OPERACIONALES PARTES REPORTE -----------------------------------------
def rpt_operacional_fisico_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT DISTINCT
	                SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
	                renaes,
	                Categoria,
	                SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC 
                LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s
                AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes, Categoria
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_dist = cursor.fetchall()
    return resultado_dist

def rpt_operacional_sensorial_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes, Categoria
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_dist_sensorial = cursor.fetchall()
    return resultado_dist_sensorial

def rpt_operacional_certificado_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes, Categoria
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        resultado_dist_certificado = cursor.fetchall()
    return resultado_dist_certificado

def rpt_operacional_rbc_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes, Categoria
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        resultado_dist_rbc = cursor.fetchall()
    return resultado_dist_rbc

def rpt_operacional_capacitacion_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes, Categoria
                    """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_dist_capacitacion = cursor.fetchall()
    return resultado_dist_capacitacion

def rpt_operacional_agente_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes, Categoria
                    """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_dist_agente = cursor.fetchall()
    return resultado_dist_agente

def rpt_operacional_comite_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,
                        Actividad,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes, Actividad
                    """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_dist_comite = cursor.fetchall()
    return resultado_dist_comite

class Atc_Ate_RptOperacinalDist(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        distritos = request.GET.get('distritos')

        provincia_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Ubigueo_Establecimiento__startswith=distritos        
            ).values_list('Provincia', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
        
        # Creación de la consulta
        resultado_dist = rpt_operacional_fisico_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_sensorial = rpt_operacional_sensorial_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_certificado = rpt_operacional_certificado_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_rbc = rpt_operacional_rbc_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_capacitacion = rpt_operacional_capacitacion_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_agente = rpt_operacional_agente_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_comite = rpt_operacional_comite_dist(distritos, fecha_inicio, fecha_fin)
        
        wb = Workbook()
        
        consultas = [
                ('Físico', resultado_dist, get_categoria_matriz_fisico),
                ('Sensorial', resultado_dist_sensorial, get_categoria_matriz_sensorial),
                ('Certificado', resultado_dist_certificado, get_categoria_matriz_certificado),
                ('RBC', resultado_dist_rbc, get_categoria_matriz_rbc),
                ('Capacitacion', resultado_dist_capacitacion, get_categoria_matriz_capacitacion),
                ('Agente', resultado_dist_agente, get_categoria_matriz_agente),
                ('Comite', resultado_dist_comite, get_categoria_matriz_comite),
        ]
        
        # Configurar locale para español
        # locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        
        # Obtener nombres de meses en español
        #meses_espanol = [datetime(2024, m, 1).strftime('%b').capitalize() for m in range(1, 13)]
        
        for index, (sheet_name, results, get_categoria_matriz) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            categoria_matriz = get_categoria_matriz()
            fill_worksheet(ws, results, categoria_matriz,provincia_codigo,fecha_inicio_codigo,fecha_fin_codigo)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_ate_atc_distrito.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

################################################
# REPORTE POR REDES
################################################
def ate_atc_get_redes(request,redes_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 'discapacidad/ate_atc_redes.html', context)

#--- FUNCIONES OPERACIONALES PARTES REPORTE -----------------------------------------
def rpt_operacional_fisico_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                            renaes,
                            Categoria,
	                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                        SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                        SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                        SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                        SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                        SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                        SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                        SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                        SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                        SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                        SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                        SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                        SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                        SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                        SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                        SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                        SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                        SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                        SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                        SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                        SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                        SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                        SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes
                        """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_red = cursor.fetchall()
    
    return resultado_red

def rpt_operacional_sensorial_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_sensorial_red = cursor.fetchall()
    
    return resultado_sensorial_red

def rpt_operacional_certificado_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_certificado_red = cursor.fetchall()
    return resultado_certificado_red

def rpt_operacional_rbc_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY  MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_rbc_red = cursor.fetchall()

    return resultado_rbc_red

def rpt_operacional_capacitacion_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s  
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_capacitacion_red = cursor.fetchall()
    return resultado_capacitacion_red

def rpt_operacional_agente_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT DISTINCT
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        Actividad,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s  
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_agente_red = cursor.fetchall()

    return resultado_agente_red

def rpt_operacional_comite_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT DISTINCT
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        Actividad,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
	                    SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
	                    SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
	                    SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
	                    SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
	                    SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
	                    SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
	                    SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
	                    SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
	                    SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
	                    SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
	                    SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
	                    SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes, Actividad
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_comite_red = cursor.fetchall()

    return resultado_comite_red

class Atc_Ate_RptOperacinalRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        red = request.GET.get('red')

        provincia_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Codigo_Red=red
        ).values_list('Red', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
        
        # Creación de la consulta
        resultado_red = rpt_operacional_fisico_red(red, fecha_inicio, fecha_fin)
        resultado_sensorial_red = rpt_operacional_sensorial_red(red, fecha_inicio, fecha_fin)
        resultado_certificado_red = rpt_operacional_certificado_red(red, fecha_inicio, fecha_fin)
        resultado_rbc_red = rpt_operacional_rbc_red(red, fecha_inicio, fecha_fin)
        resultado_capacitacion_red = rpt_operacional_capacitacion_red(red, fecha_inicio, fecha_fin)
        resultado_agente_red = rpt_operacional_agente_red(red, fecha_inicio, fecha_fin)
        resultado_comite_red = rpt_operacional_comite_red(red, fecha_inicio, fecha_fin)
        
        wb = Workbook()
        
        consultas = [
                ('Físico', resultado_red, get_categoria_matriz_fisico),
                ('Sensorial', resultado_sensorial_red, get_categoria_matriz_sensorial),
                ('Certificado', resultado_certificado_red, get_categoria_matriz_certificado),
                ('RBC', resultado_rbc_red, get_categoria_matriz_rbc),
                ('Capacitacion', resultado_capacitacion_red, get_categoria_matriz_capacitacion),
                ('Agente', resultado_agente_red, get_categoria_matriz_agente),
                ('Comite', resultado_comite_red, get_categoria_matriz_comite),
        ]
        
        # Configurar locale para español
        # locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        
        # Obtener nombres de meses en español
        #meses_espanol = [datetime(2024, m, 1).strftime('%b').capitalize() for m in range(1, 13)]
        
        for index, (sheet_name, results, get_categoria_matriz) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            categoria_matriz = get_categoria_matriz()
            fill_worksheet(ws, results, categoria_matriz,provincia_codigo,fecha_inicio_codigo,fecha_fin_codigo)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_ate_atc_red.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

################################################
# REPORTE POR MICRO-REDES
################################################
def ate_atc_get_microredes(request, microredes_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }

    return render(request, 'discapacidad/ate_atc_microredes.html', context)

def p_microredes(request):
    redes_param = request.GET.get('redes')
    
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL').values('Codigo_MicroRed','MicroRed').distinct()
    
    context = {
        'redes_param': redes_param,
        'microredes': microredes
    }
    return render(request, 'discapacidad/partials/p_microredes.html', context)

def rpt_operacional_fisico_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,                  
                        renaes,
                        Categoria,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_microred = cursor.fetchall()
    return resultado_microred

def rpt_operacional_sensorial_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,                  
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_sensorial_microred = cursor.fetchall()
    
    return resultado_sensorial_microred

def rpt_operacional_certificado_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,                  
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_certificado_microred = cursor.fetchall()  
    return resultado_certificado_microred

def rpt_operacional_rbc_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,                  
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_rbc_microred = cursor.fetchall()

    return resultado_rbc_microred

def rpt_operacional_capacitacion_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,                  
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_capacitacion_microred = cursor.fetchall()
    return resultado_capacitacion_microred

def rpt_operacional_agente_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,                  
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_agente_microred = cursor.fetchall()

    return resultado_agente_microred

def rpt_operacional_comite_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,                  
                            renaes,
                            Actividad,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                       GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_comite_microred = cursor.fetchall()

    return resultado_comite_microred

class Atc_Ate_RptOperacinalMicroRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        red = request.GET.get('redes')
        microred = request.GET.get('microredes')        
        
        provincia_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Codigo_Red=red,Codigo_MicroRed=microred
        ).values_list('MicroRed', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())

        # Creación de la consulta
        resultado_microred = rpt_operacional_fisico_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_sensorial_microred = rpt_operacional_sensorial_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_certificado_microred = rpt_operacional_certificado_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_rbc_microred = rpt_operacional_rbc_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_capacitacion_microred = rpt_operacional_capacitacion_microred(red, microred,fecha_inicio, fecha_fin)
        resultado_agente_microred = rpt_operacional_agente_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_comite_microred = rpt_operacional_comite_microred(red, microred, fecha_inicio, fecha_fin)

        wb = Workbook()
        
        consultas = [
                ('Físico', resultado_microred, get_categoria_matriz_fisico),
                ('Sensorial', resultado_sensorial_microred, get_categoria_matriz_sensorial),
                ('Certificado', resultado_certificado_microred, get_categoria_matriz_certificado),
                ('RBC', resultado_rbc_microred, get_categoria_matriz_rbc),
                ('Capacitacion', resultado_capacitacion_microred, get_categoria_matriz_capacitacion),
                ('Agente', resultado_agente_microred, get_categoria_matriz_agente),
                ('Comite', resultado_comite_microred, get_categoria_matriz_comite),
        ]
        
        # Configurar locale para español
        # locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        
        # Obtener nombres de meses en español        
        for index, (sheet_name, results, get_categoria_matriz) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            categoria_matriz = get_categoria_matriz()
            fill_worksheet(ws, results, categoria_matriz, provincia_codigo, fecha_inicio_codigo, fecha_fin_codigo)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "ate_atc_rpt_operacional_microredes.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

################################################
# REPORTE POR ESTABLECIMIENTOS
################################################
def ate_atc_get_establecimientos(request,establecimiento_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request,'discapacidad/ate_atc_establecimientos.html', context)

def p_microredes_establec(request):
    redes_param = request.GET.get('redes') 
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    return render(request, 'discapacidad/partials/p_microredes_establec.html', context)

def p_establecimientos(request):
    microredes = request.GET.get('p_microredes_establec')    
    codigo_red = request.GET.get('redes')
    establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_MicroRed=microredes,Codigo_Red=codigo_red,Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_Unico','Nombre_Establecimiento').distinct()
    context= {
        'establec': establec
    }
    return render(request, 'discapacidad/partials/p_establecimientos.html', context)

def rpt_operacional_fisico_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL_ATE_ATC.renaes,Categoria
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])       
        resultado_establec = cursor.fetchall()
    return resultado_establec

def rpt_operacional_sensorial_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, renaes, Categoria
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_sensorial_establec = cursor.fetchall()    
    return resultado_sensorial_establec

def rpt_operacional_certificado_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, renaes, Categoria
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_certificado_establec = cursor.fetchall()
    return resultado_certificado_establec

def rpt_operacional_rbc_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, renaes, Categoria
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_rbc_establec = cursor.fetchall()
    return resultado_rbc_establec

def rpt_operacional_capacitacion_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                            renaes,
                            Categoria,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, renaes, Categoria
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_capacitacion_establec = cursor.fetchall()
    return resultado_capacitacion_establec

def rpt_operacional_agente_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                            renaes,
                            Categoria,
                           	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                           	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                           	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                           	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                           	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                           	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                           	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                           	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                           	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                           	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                           	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                           	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                           	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                           	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                           	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                           	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                           	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                           	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                           	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                           	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                           	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                           	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                           	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                           	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, renaes, Categoria
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_agente_establec = cursor.fetchall()
    return resultado_agente_establec

def rpt_operacional_comite_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                        SELECT DISTINCT
                            MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                            renaes,
                            Actividad,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATE=1) then 1 else 0 end)ATE1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATE=1) then 1 else 0 end)ATE2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATE=1) then 1 else 0 end)ATE3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATE=1) then 1 else 0 end)ATE4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATE=1) then 1 else 0 end)ATE5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATE=1) then 1 else 0 end)ATE6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATE=1) then 1 else 0 end)ATE7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATE=1) then 1 else 0 end)ATE8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATE=1) then 1 else 0 end)ATE9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATE=1) then 1 else 0 end)ATE10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATE=1) then 1 else 0 end)ATE11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATE=1) then 1 else 0 end)ATE12,
                        	SUM(CASE WHEN (periodo BETWEEN 20240101 AND 20240131 AND ATC=1) then 1 else 0 end)ATES1,
                        	SUM(CASE WHEN (periodo BETWEEN 20240201 AND 20240229 AND ATC=1) then 1 else 0 end)ATES2,
                        	SUM(CASE WHEN (periodo BETWEEN 20240301 AND 20240331 AND ATC=1) then 1 else 0 end)ATES3,
                        	SUM(CASE WHEN (periodo BETWEEN 20240401 AND 20240431 AND ATC=1) then 1 else 0 end)ATES4,
                        	SUM(CASE WHEN (periodo BETWEEN 20240501 AND 20240531 AND ATC=1) then 1 else 0 end)ATES5,
                        	SUM(CASE WHEN (periodo BETWEEN 20240601 AND 20240631 AND ATC=1) then 1 else 0 end)ATES6,
                        	SUM(CASE WHEN (periodo BETWEEN 20240701 AND 20240731 AND ATC=1) then 1 else 0 end)ATES7,
                        	SUM(CASE WHEN (periodo BETWEEN 20240801 AND 20240831 AND ATC=1) then 1 else 0 end)ATES8,
                        	SUM(CASE WHEN (periodo BETWEEN 20240901 AND 20240931 AND ATC=1) then 1 else 0 end)ATES9,
                        	SUM(CASE WHEN (periodo BETWEEN 20241001 AND 20241031 AND ATC=1) then 1 else 0 end)ATES10,
                        	SUM(CASE WHEN (periodo BETWEEN 20241101 AND 20241131 AND ATC=1) then 1 else 0 end)ATES11,
                        	SUM(CASE WHEN (periodo BETWEEN 20241201 AND 20241231 AND ATC=1) then 1 else 0 end)ATES12
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC 
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL_ATE_ATC.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, renaes, Actividad
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_comite_establec = cursor.fetchall()
    return resultado_comite_establec

class Atc_Ate_RptOperacinalEstablec(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        establec = request.GET.get('p_establecimiento')

        provincia_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Codigo_Unico=establec
        ).values_list('Nombre_Establecimiento', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
                
        # Creación de la consulta
        resultado_establec = rpt_operacional_fisico_establec(establec,fecha_inicio, fecha_fin)
        resultado_sensorial_establec = rpt_operacional_sensorial_establec(establec, fecha_inicio, fecha_fin)
        resultado_certificado_establec = rpt_operacional_certificado_establec(establec, fecha_inicio, fecha_fin)
        resultado_rbc_establec = rpt_operacional_rbc_establec(establec, fecha_inicio, fecha_fin)
        resultado_capacitacion_establec = rpt_operacional_capacitacion_establec(establec,fecha_inicio, fecha_fin)
        resultado_agente_establec = rpt_operacional_agente_establec(establec, fecha_inicio, fecha_fin)
        resultado_comite_establec = rpt_operacional_comite_establec(establec, fecha_inicio, fecha_fin)
        

        wb = Workbook()
        
        consultas = [
                ('Físico', resultado_establec, get_categoria_matriz_fisico),
                ('Sensorial', resultado_sensorial_establec, get_categoria_matriz_sensorial),
                ('Certificado', resultado_certificado_establec, get_categoria_matriz_certificado),
                ('RBC', resultado_rbc_establec, get_categoria_matriz_rbc),
                ('Capacitacion', resultado_capacitacion_establec, get_categoria_matriz_capacitacion),
                ('Agente', resultado_agente_establec, get_categoria_matriz_agente),
                ('Comite', resultado_comite_establec, get_categoria_matriz_comite),
        ]
        
        # Configurar locale para español
        # locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
        
        # Obtener nombres de meses en español
        
        for index, (sheet_name, results, get_categoria_matriz) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            categoria_matriz = get_categoria_matriz()
            fill_worksheet(ws, results, categoria_matriz,provincia_codigo,fecha_inicio_codigo,fecha_fin_codigo)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "ate_atc_rpt_operacional_establec.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response
    
    
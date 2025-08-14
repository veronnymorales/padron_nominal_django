from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.shortcuts import render, redirect
from django.db import connection

# filtros
from base.models import DimPeriodo, DimDiscapacidadEtapa, MAESTRO_HIS_ESTABLECIMIENTO
from .models import TramaBaseDiscapacidadRpt02FisicaNominal
from django.db.models import Case, When, Value, IntegerField, Sum, OuterRef, Subquery
from django.db.models.functions import Substr, Cast, Concat, Replace
from django.db.models import CharField, F

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from .utils import generar_operacional

from django.db.models.functions import Substr



# Create your views here.
@login_required
def operacional(request):
    return render(request, 'discapacidad/index.html')

################################################
# REPORTE DE SEGUIMIENTO
################################################
#--- PROVINCIAS -------------------------------------------------------------
def get_provincias(request,provincias_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    context = {
                'provincias': provincias,
            }
    
    return render(request, 'discapacidad/provincias.html', context)

#--- FUNCIONES OPERACIONALES PARTES REPORTE -----------------------------------------
def rpt_operacional_fisico(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                    renaes,                   
                    SUM(dis_1) AS dis_1,
                    SUM(dis_2) AS dis_2,
                    SUM(dis_3) AS dis_3,
                    SUM(dis_4) AS dis_4,
                    SUM(dis_5) AS dis_5,
                    SUM(dis_6) AS dis_6,
                    SUM(dis_7) AS dis_7,
                    SUM(dis_8) AS dis_8,
                    SUM(dis_9) AS dis_9,
                    SUM(dis_10) AS dis_10,
                    SUM(dis_11) AS dis_11,
                    SUM(dis_12) AS dis_12,
                    SUM(dis_13) AS dis_13,
                    SUM(dis_14) AS dis_14,
                    SUM(dis_15) AS dis_15,
                    SUM(dis_16) AS dis_16,
                    SUM(dis_17) AS dis_17,
                    SUM(dis_18) AS dis_18,
                    SUM(dis_19) AS dis_19,
                    SUM(dis_20) AS dis_20,
                    SUM(dis_21) AS dis_21,
                    SUM(dis_22) AS dis_22,
                    SUM(dis_23) AS dis_23,
                    SUM(dis_24) AS dis_24,
                    SUM(dis_25) AS dis_25,
                    SUM(dis_26) AS dis_26,
                    SUM(dis_27) AS dis_27,
                    SUM(dis_28) AS dis_28,
                    SUM(dis_29) AS dis_29,
                    SUM(dis_30) AS dis_30,
                    SUM(dis_31) AS dis_31,
                    SUM(dis_32) AS dis_32,
                    SUM(dis_33) AS dis_33,
                    SUM(dis_34) AS dis_34,
                    SUM(dis_35) AS dis_35,
                    SUM(dis_36) AS dis_36,
                    SUM(dis_37) AS dis_37,
                    SUM(dis_38) AS dis_38,
                    SUM(dis_39) AS dis_39,
                    SUM(dis_40) AS dis_40,
                    SUM(dis_41) AS dis_41,
                    SUM(dis_42) AS dis_42,
                    SUM(dis_43) AS dis_43,
                    SUM(dis_44) AS dis_44,
                    SUM(dis_45) AS dis_45,
                    SUM(dis_46) AS dis_46,
                    SUM(dis_47) AS dis_47,
                    SUM(dis_48) AS dis_48,
                    SUM(dis_49) AS dis_49,
                    SUM(dis_50) AS dis_50,
                    SUM(dis_51) AS dis_51,
                    SUM(dis_52) AS dis_52,
                    SUM(dis_53) AS dis_53,
                    SUM(dis_54) AS dis_54,
                    SUM(dis_55) AS dis_55,
                    SUM(dis_56) AS dis_56,
                    SUM(dis_57) AS dis_57,
                    SUM(dis_58) AS dis_58,
                    SUM(dis_59) AS dis_59,
                    SUM(dis_60) AS dis_60,
                    SUM(dis_61) AS dis_61,
                    SUM(dis_62) AS dis_62,
                    SUM(dis_63) AS dis_63,
                    SUM(dis_64) AS dis_64,
                    SUM(dis_65) AS dis_65,
                    SUM(dis_66) AS dis_66,
                    SUM(dis_67) AS dis_67,
                    SUM(dis_68) AS dis_68,
                    SUM(dis_69) AS dis_69,
                    SUM(dis_70) AS dis_70,
                    SUM(dis_71) AS dis_71,
                    SUM(dis_72) AS dis_72,
                    SUM(dis_73) AS dis_73,
                    SUM(dis_74) AS dis_74,
                    SUM(dis_75) AS dis_75,
                    SUM(dis_76) AS dis_76,
                    SUM(dis_77) AS dis_77,
                    SUM(dis_78) AS dis_78,
                    SUM(dis_79) AS dis_79,
                    SUM(dis_80) AS dis_80,
                    SUM(dis_81) AS dis_81,
                    SUM(dis_82) AS dis_82,
                    SUM(dis_83) AS dis_83,
                    SUM(dis_84) AS dis_84,
                    SUM(dis_85) AS dis_85,
                    SUM(dis_86) AS dis_86,
                    SUM(dis_87) AS dis_87,
                    SUM(dis_88) AS dis_88,
                    SUM(dis_89) AS dis_89,
                    SUM(dis_90) AS dis_90,
                    SUM(dis_91) AS dis_91,
                    SUM(dis_92) AS dis_92,
                    SUM(dis_93) AS dis_93,
                    SUM(dis_94) AS dis_94,
                    SUM(dis_95) AS dis_95,
                    SUM(dis_96) AS dis_96,
                    SUM(dis_97) AS dis_97,
                    SUM(dis_98) AS dis_98,
                    SUM(dis_99) AS dis_99,
                    SUM(dis_100) AS dis_100,
                    SUM(dis_101) AS dis_101,
                    SUM(dis_102) AS dis_102,
                    SUM(dis_103) AS dis_103,
                    SUM(dis_104) AS dis_104,
                    SUM(dis_105) AS dis_105,
                    SUM(dis_106) AS dis_106,
                    SUM(dis_107) AS dis_107,
                    SUM(dis_108) AS dis_108,
                    SUM(dis_109) AS dis_109,
                    SUM(dis_110) AS dis_110,
                    SUM(dis_111) AS dis_111,
                    SUM(dis_112) AS dis_112,
                    SUM(dis_113) AS dis_113,
                    SUM(dis_114) AS dis_114,
                    SUM(dis_115) AS dis_115,
                    SUM(dis_116) AS dis_116,
                    SUM(dis_117) AS dis_117,
                    SUM(dis_118) AS dis_118,
                    SUM(dis_119) AS dis_119,
                    SUM(dis_120) AS dis_120,
                    SUM(dis_121) AS dis_121,
                    SUM(dis_122) AS dis_122,
                    SUM(dis_123) AS dis_123,
                    SUM(dis_124) AS dis_124,
                    SUM(dis_125) AS dis_125,
                    SUM(dis_126) AS dis_126,
                    SUM(dis_127) AS dis_127,
                    SUM(dis_128) AS dis_128,
                    SUM(dis_129) AS dis_129,
                    SUM(dis_130) AS dis_130, 
                    SUM(dis_131) AS dis_131,
                    SUM(dis_132) AS dis_132,
                    SUM(dis_133) AS dis_133,
                    SUM(dis_134) AS dis_134,
                    SUM(dis_135) AS dis_135,
                    SUM(dis_136) AS dis_136,
                    SUM(dis_137) AS dis_137,
                    SUM(dis_138) AS dis_138,
                    SUM(dis_139) AS dis_139,
                    SUM(dis_140) AS dis_140, 
                    SUM(dis_141) AS dis_141,
                    SUM(dis_142) AS dis_142,
                    SUM(dis_143) AS dis_143,
                    SUM(dis_144) AS dis_144,
                    SUM(dis_145) AS dis_145,
                    SUM(dis_146) AS dis_146,
                    SUM(dis_147) AS dis_147,
                    SUM(dis_148) AS dis_148,
                    SUM(dis_149) AS dis_149,
                    SUM(dis_150) AS dis_150,
                    SUM(dis_151) AS dis_151,
                    SUM(dis_152) AS dis_152,
                    SUM(dis_153) AS dis_153,
                    SUM(dis_154) AS dis_154,
                    SUM(dis_155) AS dis_155,
                    SUM(dis_156) AS dis_156,
                    SUM(dis_157) AS dis_157,
                    SUM(dis_158) AS dis_158,
                    SUM(dis_159) AS dis_159,
                    SUM(dis_160) AS dis_160
                FROM (
                    SELECT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_1,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_2,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_3,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_4,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_5,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_6,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_7,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_8,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_9,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_10,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_11,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_12,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_13,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_14,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_15,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_16,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_17,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_18,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_19,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_20,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_21,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_22,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_23,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_24,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_25,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_26,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_27,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_28,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_29,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_30,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_31,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_32,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_33,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_34,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_35,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_36,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_37,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_38,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_39,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_40,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_41,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_42,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_43,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_44,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_45,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_46,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_47,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_48,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_49,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_50,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_51,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_52,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_53,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_54,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_55,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_56,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_57,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_58,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_59,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_60,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_61,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_62,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_63,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_64,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_65,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_66,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_67,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_68,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_69,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_70,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_71,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_72,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_73,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_74,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_75,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_76,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_77,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_78,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_79,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_80,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_81,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_82,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_83,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_84,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_85,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_86,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_87,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_88,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_89,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_90,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_91,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_92,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_93,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_94,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_95,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_96,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_97,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_98,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_99,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_100,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_101,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_102,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_103,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_104,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_105,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_106,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_107,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_108,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_109,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_110,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_111,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_112,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_113,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_114,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_115,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_116,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_117,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_118,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_119,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_120,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_121,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_122,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_123,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_124,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_125,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_126,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_127,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_128,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_129,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_130,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_131,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_132,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_133,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_134,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_135,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_136,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_137,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_138,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_139,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_140,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_141,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_142,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_143,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_144,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_145,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_146,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_147,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_148,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_149,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_150,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_151,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_152,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_153,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_154,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_155,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_156,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_157,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_158,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_159,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_160
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                ) subquery
                GROUP BY renaes, ubigeo_filtrado
        """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        resultado_prov = cursor.fetchall()
    return resultado_prov

def rpt_operacional_sensorial(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,   
                        SUM(dis_161) AS dis_161,
                        SUM(dis_162) AS dis_162,
                        SUM(dis_163) AS dis_163,
                        SUM(dis_164) AS dis_164,
                        SUM(dis_165) AS dis_165,
                        SUM(dis_166) AS dis_166,
                        SUM(dis_167) AS dis_167,
                        SUM(dis_168) AS dis_168,
                        SUM(dis_169) AS dis_169,
                        SUM(dis_170) AS dis_170,
                        SUM(dis_171) AS dis_171,
                        SUM(dis_172) AS dis_172,
                        SUM(dis_173) AS dis_173,
                        SUM(dis_174) AS dis_174,
                        SUM(dis_175) AS dis_175,
                        SUM(dis_176) AS dis_176,
                        SUM(dis_177) AS dis_177,
                        SUM(dis_178) AS dis_178,
                        SUM(dis_179) AS dis_179,
                        SUM(dis_180) AS dis_180,
                        SUM(dis_181) AS dis_181,
                        SUM(dis_182) AS dis_182,
                        SUM(dis_183) AS dis_183,
                        SUM(dis_184) AS dis_184,
                        SUM(dis_185) AS dis_185,
                        SUM(dis_186) AS dis_186,
                        SUM(dis_187) AS dis_187,
                        SUM(dis_188) AS dis_188,
                        SUM(dis_189) AS dis_189,
                        SUM(dis_190) AS dis_190,
                        SUM(dis_191) AS dis_191,
                        SUM(dis_192) AS dis_192,
                        SUM(dis_193) AS dis_193,
                        SUM(dis_194) AS dis_194,
                        SUM(dis_195) AS dis_195
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_161,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_162,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_163,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_164,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_165,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_166,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_167,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_168,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_169,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_170,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_171,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_172,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_173,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_174,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_175,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_176,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_177,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_178,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_179,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_180,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_181,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_182,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_183,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_184,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_185,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_186,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_187,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_188,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_189,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_190,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_191,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_192,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_193,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_194,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_195
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
            """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_sensorial = cursor.fetchall()
    return resultado_prov_sensorial

def rpt_operacional_certificado(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,   
                        SUM(dis_216) AS dis_216,
                        SUM(dis_217) AS dis_217,
                        SUM(dis_218) AS dis_218,
                        SUM(dis_219) AS dis_219,
                        SUM(dis_220) AS dis_220,
                        SUM(dis_221) AS dis_221,
                        SUM(dis_222) AS dis_222,
                        SUM(dis_223) AS dis_223,
                        SUM(dis_224) AS dis_224,
                        SUM(dis_225) AS dis_225,
                        SUM(dis_226) AS dis_226,
                        SUM(dis_227) AS dis_227,
                        SUM(dis_228) AS dis_228,
                        SUM(dis_229) AS dis_229,
                        SUM(dis_230) AS dis_230,
                        SUM(dis_231) AS dis_231,
                        SUM(dis_232) AS dis_232,
                        SUM(dis_233) AS dis_233,
                        SUM(dis_234) AS dis_234,
                        SUM(dis_235) AS dis_235
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_216,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_217,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_218,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_219,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_220,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_221,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_222,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_223,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_224,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_225,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_226,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_227,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_228,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_229,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_230,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_231,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_232,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_233,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_234,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_235
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_certificado = cursor.fetchall()
    return resultado_prov_certificado

def rpt_operacional_rbc(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_242) AS dis_242,
                        SUM(dis_243) AS dis_243,
                        SUM(dis_244) AS dis_244,
                        SUM(dis_245) AS dis_245,
                        SUM(dis_246) AS dis_246,
                        SUM(dis_247) AS dis_247,
                        SUM(dis_248) AS dis_248,
                        SUM(dis_249) AS dis_249,
                        SUM(dis_250) AS dis_250,
                        SUM(dis_251) AS dis_251,
                        SUM(dis_252) AS dis_252,
                        SUM(dis_253) AS dis_253,
                        SUM(dis_254) AS dis_254,
                        SUM(dis_255) AS dis_255,
                        SUM(dis_256) AS dis_256,
                        SUM(dis_257) AS dis_257,
                        SUM(dis_258) AS dis_258,
                        SUM(dis_259) AS dis_259,
                        SUM(dis_260) AS dis_260, 
                        SUM(dis_261) AS dis_261, 
                        SUM(dis_262) AS dis_262, 
                        SUM(dis_263) AS dis_263, 
                        SUM(dis_264) AS dis_264, 
                        SUM(dis_265) AS dis_265, 
                        SUM(dis_266) AS dis_266
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_242,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_243,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_244,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_245,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_246,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_247,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_248,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_249,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_250,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_251,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_252,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_253,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_254,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_255,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_256,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_257,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_258,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_259,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_260,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_261,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_262,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_263,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_264,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_265,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_266
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s     
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_rbc = cursor.fetchall()
    return resultado_prov_rbc

def rpt_operacional_mental(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_196) AS dis_196,
                        SUM(dis_197) AS dis_197,
                        SUM(dis_198) AS dis_198,
                        SUM(dis_199) AS dis_199,
                        SUM(dis_200) AS dis_200,
                        SUM(dis_201) AS dis_201,
                        SUM(dis_202) AS dis_202,
                        SUM(dis_203) AS dis_203,
                        SUM(dis_204) AS dis_204,
                        SUM(dis_205) AS dis_205,
                        SUM(dis_206) AS dis_206,
                        SUM(dis_207) AS dis_207,
                        SUM(dis_208) AS dis_208,
                        SUM(dis_209) AS dis_209,
                        SUM(dis_210) AS dis_210, 
                        SUM(dis_211) AS dis_211, 
                        SUM(dis_212) AS dis_212, 
                        SUM(dis_213) AS dis_213, 
                        SUM(dis_214) AS dis_214, 
                        SUM(dis_215) AS dis_215
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_196,
                            SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_197,
                            SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_198,
                            SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_199,
                            SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_200,
                            SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_201,
                            SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_202,
                            SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_203,
                            SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_204,
                            SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_205,
                            SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_206,
                            SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_207,
                            SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_208,
                            SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_209,
                            SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_210,
                            SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_211,
                            SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_212,
                            SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_213,
                            SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_214,
                            SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_215
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s     
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_mental = cursor.fetchall()
    return resultado_prov_mental

def rpt_operacional_capacitacion(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_273) AS dis_273,
                        SUM(dis_274) AS dis_274
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            COUNT(Categoria) AS dis_273,
                            SUM(gedad) AS dis_274
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s     
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_capacitacion = cursor.fetchall()
    return resultado_prov_capacitacion

def rpt_operacional_agente(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_236) AS dis_236,
                        SUM(dis_237) AS dis_237,
                        SUM(dis_238) AS dis_238,
                        SUM(dis_239) AS dis_239,
                        SUM(dis_240) AS dis_240,
                        SUM(dis_241) AS dis_241
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 THEN 1 ELSE 0 END) 		 AS dis_236,
                            SUM(CASE WHEN Categoria = 1 THEN gedad ELSE 0 END)   AS dis_237,
                            SUM(CASE WHEN Categoria = 2 THEN 1 ELSE 0 END)     AS dis_238,
                            SUM(CASE WHEN Categoria = 2 THEN gedad ELSE 0 END)   AS dis_239,
                            SUM(CASE WHEN Categoria = 3 THEN 1 ELSE 0 END)     AS dis_240,
                            SUM(CASE WHEN Categoria = 3 THEN gedad ELSE 0 END)   AS dis_241
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s     
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_agente = cursor.fetchall()
    return resultado_prov_agente

def rpt_operacional_comite(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_267) AS dis_267,
                        SUM(dis_268) AS dis_268,
                        SUM(dis_269) AS dis_269,
                        SUM(dis_270) AS dis_270,
                        SUM(dis_271) AS dis_271,
                        SUM(dis_272) AS dis_272
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Actividad = 1 THEN 1 ELSE 0 END) 		AS dis_267,
                            SUM(CASE WHEN Actividad = 1 THEN Partic ELSE 0 END) AS dis_268,
                            SUM(CASE WHEN Actividad = 2 THEN 1 ELSE 0 END)      AS dis_269,
                            SUM(CASE WHEN Actividad = 2 THEN Partic ELSE 0 END) AS dis_270,
                            SUM(CASE WHEN Actividad = 3 THEN 1 ELSE 0 END)      AS dis_271,
                            SUM(CASE WHEN Actividad = 3 THEN Partic ELSE 0 END) AS dis_272
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) = %s     
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:4], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_prov_comite = cursor.fetchall()
    return resultado_prov_comite

# validar matriz
def crear_matriz(request):    
    ubigeo = '1201'
    fecha_inicio = '20240102'# Ejemplo de ubigeo
    fecha_fin = '20240110'# Ejemplo de ubigeo
    matriz = rpt_operacional_fisico(ubigeo,fecha_inicio,fecha_fin)  
    # Puedes renderizar la matriz en una plantilla HTML o hacer cualquier otro procesamiento necesario
    return render(request, 'discapacidad/matrizes.html', {'matriz': matriz})
###
###--- PROVINCIAS EXCEL ---------------------------------##
##
class RptOperacinalProv(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        provincia = request.GET.get('provincia')

        # Creación de la consulta
        resultado_prov = rpt_operacional_fisico(provincia, fecha_inicio, fecha_fin)
        resultado_prov_sensorial = rpt_operacional_sensorial(provincia, fecha_inicio, fecha_fin)
        resultado_prov_certificado = rpt_operacional_certificado(provincia, fecha_inicio, fecha_fin)
        resultado_prov_rbc = rpt_operacional_rbc(provincia, fecha_inicio, fecha_fin)       
        resultado_prov_mental = rpt_operacional_mental(provincia, fecha_inicio, fecha_fin)
        resultado_prov_capacitacion = rpt_operacional_capacitacion(provincia, fecha_inicio, fecha_fin)
        resultado_prov_agente = rpt_operacional_agente(provincia, fecha_inicio, fecha_fin)
        resultado_prov_comite = rpt_operacional_comite(provincia, fecha_inicio, fecha_fin)

        provincia_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Ubigueo_Establecimiento__startswith=provincia
        ).values_list('Provincia', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())

        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambia el alto de la columna
        sheet.row_dimensions[1].height = 14
        sheet.row_dimensions[2].height = 14
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[15].height = 25
        # cambia el ancho de la columna
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 28
        sheet.column_dimensions['C'].width = 28
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 9
        sheet.column_dimensions['F'].width = 9
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 9
        sheet.column_dimensions['L'].width = 9
        # linea de division
        sheet.freeze_panes = 'AL8'
        
        # Configuración del fondo y el borde
        fill = PatternFill(patternType='solid', fgColor='00B0F0')
        border = Border(left=Side(style='thin', color='00B0F0'),
                        right=Side(style='thin', color='00B0F0'),
                        top=Side(style='thin', color='00B0F0'),
                        bottom=Side(style='thin', color='00B0F0'))

        borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                        right=Side(style='thin', color='A9A9A9'), # Plomo
                        top=Side(style='thin', color='A9A9A9'), # Plomo
                        bottom=Side(style='thin', color='A9A9A9')) # Plomo

        # crea titulo del reporte
        sheet['B1'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B1'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
        
        sheet['B2'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B2'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
        
        sheet['B4'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B4'].font = Font(name = 'Arial', size= 12, bold = True)
        sheet['B4'] = 'REPORTE DE ACTIVIDADES DEL COMPONENTE DE DISCAPACIDAD'
        
        sheet['B6'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B6'] ='DIRESA / GERESA / DISA'
        
        sheet['C6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C6'].font = Font(name = 'Arial', size= 7)
        sheet['C6'] ='JUNIN'

        sheet['B7'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B7'] ='PROV/ DIST/ RED/ MR/ ESTABLEC'
        
        sheet['C7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C7'].font = Font(name = 'Arial', size= 7)
        sheet['C7'] = provincia_codigo[0]
        
        sheet['E6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['E6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['E6'] ='PERIODO'
        
        sheet['F6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F6'] ='MES INICIO'
        
        sheet['F7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F7'] ='MES FIN'
        
        sheet['G6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G6'].font = Font(name = 'Arial', size= 8)
        sheet['G6'] = fecha_inicio_codigo[0]
        
        sheet['G7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G7'].font = Font(name = 'Arial', size= 8)
        sheet['G7'] = fecha_fin_codigo[0]
        
        sheet['B9'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B9'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B9'] ='PERSONAS CON DISCAPACIDAD RECIBEN ATENCION DE REHABILITACION EN ESTABLECIMIENTOS DE SALUD (3000688)'
        
        sheet['B10'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B10'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B10'] ='Capacitación en medicina de rehabilitación integral (5004449)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=12, max_row=12, min_col=3, max_col=5):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['C12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C12'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['C12'] ='Capacitación  (C0009)' 
        
        sheet['D11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D11'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D11'].fill = fill
        sheet['D11'].border = border
        sheet['D11'] = 'N°'
                
        sheet['E11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E11'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['E11'].fill = fill
        sheet['E11'].border = border
        sheet['E11'] = 'Capacitados'
        #######################################################
        ########## DISCAPACIDAD FISICA ########################
        #######################################################
        sheet['B14'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B14'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B14'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Física (5005150)' 
                
        sheet['B15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B15'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B15'].fill = fill
        sheet['B15'].border = border
        sheet['B15'] = 'Atenciones'
        
        sheet['D15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D15'].fill = fill
        sheet['D15'].border = border
        sheet['D15'] = 'Total'
        
        sheet['E15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E15'].fill = fill
        sheet['E15'].border = border
        sheet['E15'] = 'Niños         (1d - 11a)'
        
        sheet['F15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F15'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F15'].fill = fill
        sheet['F15'].border = border
        sheet['F15'] = 'Adolescentes (12a - 17a)'
        
        sheet['G15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G15'].fill = fill
        sheet['G15'].border = border
        sheet['G15'] = 'Jóvenes (18a - 29a)'
        
        sheet['H15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H15'].fill = fill
        sheet['H15'].border = border
        sheet['H15'] = 'Adultos (30a - 59a)'
        
        sheet['I15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I15'].fill = fill
        sheet['I15'].border = border
        sheet['I15'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=16, max_row=47, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B16'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B16'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B16'] ='LESIONES MEDULARES' 
                
        sheet['B17'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B17'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B17'] ='ENFERMEDAD DE PARKINSON Y SIMILARES' 
        
        sheet['B18'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B18'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B18'] ='REHABILITACIÓN EN PACIENTES AMPUTADOS' 
                
        sheet['B20'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B20'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B20'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA NEUROLÓGICA' 
        
        sheet['B23'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B23'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B23'] ='TRASTORNOS DEL DESARROLLO DE LA FUNCIÓN MOTRIZ' 
        
        sheet['B24'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B24'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B24'] ='ATENCIÓN DE REHABILITACIÓN DE ENFERMEDAD ARTICULAR DEGENERATIVA' 
        
        sheet['B25'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B25'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B25'] ='ENCEFALOPATÍA INFANTIL' 
                
        sheet['B26'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B26'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B26'] ='SÍNDROME DOWN' 
        
        sheet['B27'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B27'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B27'] ='REHABILITACIÓN EN PATOLOGÍA DE LA COLUMNA VERTEBRAL Y OTROS TRASTORNOS POSTURALES' 
        
        sheet['B34'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B34'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B34'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD CARDIOVASCULAR' 
        
        sheet['B35'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B35'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B35'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD RESPIRATORIA' 
        
        sheet['B36'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B36'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B36'] ='ATENCIÓN DE REHABILITACIÓN EN ALTERACIONES DEL PISO PÉLVICO' 
        
        sheet['B37'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B37'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B37'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA TRAUMATOLÓGICA Y REUMATOLÓGICA' 
        
        sheet['B44'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B44'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B44'] ='ATENCIÓN DE REHABILITACIÓN ONCOLÓGICA' 
        
        sheet['B46'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B46'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B46'] ='ATENCIÓN DE REHABILITACIÓN EN DOLOR' 
        
        sheet['B47'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B47'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B47'] ='ATENCIÓN DE REHABILITACIÓN EN PACIENTES QUEMADOS' 
        ####     
        sheet['C16'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C16'].font = Font(name = 'Arial', size= 7)
        sheet['C16'] ='Lesiones medulares' 
    
        sheet['C17'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C17'].font = Font(name = 'Arial', size= 7)
        sheet['C17'] ='Enfermedad de Parkinson y similares' 
        
        sheet['C18'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C18'].font = Font(name = 'Arial', size= 7)
        sheet['C18'] ='Amputados de miembros superiores' 
        
        sheet['C19'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C19'].font = Font(name = 'Arial', size= 7)
        sheet['C19'] ='Amputados de miembros inferiores' 
        
        sheet['C20'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C20'].font = Font(name = 'Arial', size= 7)
        sheet['C20'] ='Enfermedades cerebrovasculares'
        
        sheet['C21'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C21'].font = Font(name = 'Arial', size= 7)
        sheet['C21'] ='Enfermedades musculares y de la unión mioneural'
        
        sheet['C22'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C22'].font = Font(name = 'Arial', size= 7)
        sheet['C22'] ='Lesiones de nervios periféricos'
        
        sheet['C23'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C23'].font = Font(name = 'Arial', size= 7)
        sheet['C23'] ='Trastornos del desarrollo de la funcion motriz'
        
        sheet['C24'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C24'].font = Font(name = 'Arial', size= 7)
        sheet['C24'] ='Enfermedad articular degenerativa'
        
        sheet['C25'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C25'].font = Font(name = 'Arial', size= 7)
        sheet['C25'] ='Encefalopatía infantil y otras lesiones'
        
        sheet['C26'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C26'].font = Font(name = 'Arial', size= 7)
        sheet['C26'] ='Sindrome de Down'
        
        sheet['C27'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C27'].font = Font(name = 'Arial', size= 7)
        sheet['C27'] ='Cifosis y lordosis'
        
        sheet['C28'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C28'].font = Font(name = 'Arial', size= 7)
        sheet['C28'] ='Espondilo artropatías'
        
        sheet['C29'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C29'].font = Font(name = 'Arial', size= 7)
        sheet['C29'] ='Otros trastornos de los discos intervertebrales'
        
        sheet['C30'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C30'].font = Font(name = 'Arial', size= 7)
        sheet['C30'] ='Cervicalgia, dorsalgia, lumbago'
        
        sheet['C31'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C31'].font = Font(name = 'Arial', size= 7)
        sheet['C31'] ='Otras dorsopatías deformantes'
        
        sheet['C32'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C32'].font = Font(name = 'Arial', size= 7)
        sheet['C32'] ='Otros trastornos articulares'
        
        sheet['C33'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C33'].font = Font(name = 'Arial', size= 7)
        sheet['C33'] ='Defectos en la longitud de extremidades'
        
        sheet['C34'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C34'].font = Font(name = 'Arial', size= 7)
        sheet['C34'] ='Enfermedad cardiovascular'
        
        sheet['C35'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C35'].font = Font(name = 'Arial', size= 7)
        sheet['C35'] ='Enfermedad respiratoria'
        
        sheet['C36'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C36'].font = Font(name = 'Arial', size= 7)
        sheet['C36'] ='Vejiga neurogénica y dolor'
        
        sheet['C37'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C37'].font = Font(name = 'Arial', size= 7)
        sheet['C37'] ='Incontinencia'
        
        sheet['C38'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C38'].font = Font(name = 'Arial', size= 7)
        sheet['C38'] ='Prolapso'
        
        sheet['C39'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C39'].font = Font(name = 'Arial', size= 7)
        sheet['C39'] ='Traumatismos'
        
        sheet['C40'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C40'].font = Font(name = 'Arial', size= 7)
        sheet['C40'] ='Enfermedades del tejido conectivo'
        
        sheet['C41'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C41'].font = Font(name = 'Arial', size= 7)
        sheet['C41'] ='Patología articular excluida columna'
        
        sheet['C42'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C42'].font = Font(name = 'Arial', size= 7)
        sheet['C42'] ='Lesiones infecciosas'
        
        sheet['C43'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C43'].font = Font(name = 'Arial', size= 7)
        sheet['C43'] ='Lesión biomecánica'
        
        sheet['C44'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C44'].font = Font(name = 'Arial', size= 7)
        sheet['C44'] ='Linfedema'
        
        sheet['C45'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C45'].font = Font(name = 'Arial', size= 7)
        sheet['C45'] ='Sarcopenia'
        
        sheet['C46'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C46'].font = Font(name = 'Arial', size= 7)
        sheet['C46'] ='Dolor'
        
        sheet['C47'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C47'].font = Font(name = 'Arial', size= 7)
        sheet['C47'] ='Quemaduras, corrosiones y congelaciones'
        
        ##########################################################    
        ########## DISCAPACIDAD SENSORIAL ########################
        ##########################################################
        sheet['B50'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B50'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B50'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Sensorial (5005151)' 
                
        sheet['B51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B51'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B51'].fill = fill
        sheet['B51'].border = border
        sheet['B51'] = 'Atenciones'
        
        sheet['D51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D51'].fill = fill
        sheet['D51'].border = border
        sheet['D51'] = 'Total'
        
        sheet['E51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E51'].fill = fill
        sheet['E51'].border = border
        sheet['E51'] = 'Niños         (1d - 11a)'
        
        sheet['F51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F51'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F51'].fill = fill
        sheet['F51'].border = border
        sheet['F51'] = 'Adolescentes (12a - 17a)'
        
        sheet['G51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G51'].fill = fill
        sheet['G51'].border = border
        sheet['G51'] = 'Jóvenes (18a - 29a)'
        
        sheet['H51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H51'].fill = fill
        sheet['H51'].border = border
        sheet['H51'] = 'Adultos (30a - 59a)'
        
        sheet['I51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I51'].fill = fill
        sheet['I51'].border = border
        sheet['I51'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=52, max_row=58, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B52'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B52'] ='HIPOACUSIA Y/O SORDERA' 
        
        sheet['B53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B53'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B53'] ='BAJA VISION Y/O CEGUERA' 
        
        sheet['B54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B54'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B54'] ='SORDOMUDEZ' 
        
        sheet['B55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B55'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B55'] ='ENFERMEDAD CEREBRO VASCULAR' 
        
        sheet['B56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B56'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B56'] ='TRASTORNOS ESPECIFICOS DEL DESARROLLO DEL HABLA Y LENGUAJE' 
        
        sheet['B57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B57'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B57'] ='DISARTRIA Y DISFAGIA' 
        
        sheet['B59'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B59'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B59'] ='SUB TOTAL' 
        
        ########               
        sheet['C52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C52'].font = Font(name = 'Arial', size= 7)
        sheet['C52'] ='Hipoacusia y sordera' 
        
        sheet['C53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C53'].font = Font(name = 'Arial', size= 7)
        sheet['C53'] ='Baja visión y ceguera' 
        
        sheet['C54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C54'].font = Font(name = 'Arial', size= 7)
        sheet['C54'] ='Sordomudez' 
        
        sheet['C55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C55'].font = Font(name = 'Arial', size= 7)
        sheet['C55'] ='Enfermedad Cerebro vascular' 
        
        sheet['C56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C56'].font = Font(name = 'Arial', size= 7)
        sheet['C56'] ='Trastornos específicos del desarrollo del habla y lenguaje' 
        
        sheet['C57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C57'].font = Font(name = 'Arial', size= 7)
        sheet['C57'] ='Disartria' 
        
        sheet['C58'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C58'].font = Font(name = 'Arial', size= 7)
        sheet['C58'] ='Disfagia' 
        
        ########################################################
        ########## DISCAPACIDAD MENTAL #########################
        ########################################################
        sheet['B61'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B61'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B61'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Mental (5005152)' 
                
        sheet['B62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B62'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B62'].fill = fill
        sheet['B62'].border = border
        sheet['B62'] = 'Atenciones'
        
        sheet['D62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D62'].fill = fill
        sheet['D62'].border = border
        sheet['D62'] = 'Total'
        
        sheet['E62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E62'].fill = fill
        sheet['E62'].border = border
        sheet['E62'] = 'Niños         (1d - 11a)'
        
        sheet['F62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F62'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F62'].fill = fill
        sheet['F62'].border = border
        sheet['F62'] = 'Adolescentes (12a - 17a)'
        
        sheet['G62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G62'].fill = fill
        sheet['G62'].border = border
        sheet['G62'] = 'Jóvenes (18a - 29a)'
        
        sheet['H62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H62'].fill = fill
        sheet['H62'].border = border
        sheet['H62'] = 'Adultos (30a - 59a)'
        
        sheet['I62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I62'].fill = fill
        sheet['I62'].border = border
        sheet['I62'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=63, max_row=66, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B63'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B63'] ='TRASTORNOS DE APRENDIZAJE' 
        
        sheet['B64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B64'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B64'] ='RETRASO MENTAL LEVE, MODERADO, SEVERO' 
        
        sheet['B65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B65'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B65'] ='TRASTORNOS DEL ESPECTRO AUTISTA' 
        
        sheet['B66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B66'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B66'] ='OTROS TRASTORNOS DE SALUD MENTAL' 
        
        sheet['B67'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B67'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B67'] ='SUB TOTAL' 
        
        ##########
        
        sheet['C63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C63'].font = Font(name = 'Arial', size= 7)
        sheet['C63'] ='Trastornos del aprendizaje' 
        
        sheet['C64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C64'].font = Font(name = 'Arial', size= 7)
        sheet['C64'] ='Retardo Mental: Leve, moderado, severo' 
        
        sheet['C65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C65'].font = Font(name = 'Arial', size= 7)
        sheet['C65'] ='Trastornos del espectro autista' 
        
        sheet['C66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C66'].font = Font(name = 'Arial', size= 7)
        sheet['C66'] ='Otras alteraciones de salud mental' 
                
        ##################################################
        ########## CERTIFICACION #########################
        ##################################################
        sheet['B69'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B69'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B69'] ='PERSONAS CON DISCAPACIDAD CERTIFICADAS EN ESTABLECIMIENTOS DE SALUD (3000689)' 
                
        sheet['B70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B70'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B70'].fill = fill
        sheet['B70'].border = border
        sheet['B70'] = 'Atenciones'
        
        sheet['D70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D70'].fill = fill
        sheet['D70'].border = border
        sheet['D70'] = 'Total'
        
        sheet['E70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E70'].fill = fill
        sheet['E70'].border = border
        sheet['E70'] = 'Niños         (1d - 11a)'
        
        sheet['F70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F70'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F70'].fill = fill
        sheet['F70'].border = border
        sheet['F70'] = 'Adolescentes (12a - 17a)'
        
        sheet['G70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G70'].fill = fill
        sheet['G70'].border = border
        sheet['G70'] = 'Jóvenes (18a - 29a)'
        
        sheet['H70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H70'].fill = fill
        sheet['H70'].border = border
        sheet['H70'] = 'Adultos (30a - 59a)'
        
        sheet['I70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I70'].fill = fill
        sheet['I70'].border = border
        sheet['I70'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=71, max_row=74, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B71'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B71'] ='Certificación de Discapacidad (0515204)' 
        
        sheet['B74'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B74'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B74'] ='Certificación de Incapacidad (0515205)' 
        
        sheet['B75'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B75'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B75'] ='SUB TOTAL' 
        
        sheet['C71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C71'].font = Font(name = 'Arial', size= 7)
        sheet['C71'] ='Evaluación' 
        
        sheet['C72'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C72'].font = Font(name = 'Arial', size= 7)
        sheet['C72'] ='Calificación' 
        
        sheet['C73'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C73'].font = Font(name = 'Arial', size= 7)
        sheet['C73'] ='Certificación' 

        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B77'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B77'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B77'] ='PERSONAS CON DISCAPACIDAD RECIBEN SERVICIOS DE REHABILITACIÓN BASADA EN LA COMUNIDAD (3000690)' 
        
        sheet['B78'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B78'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B78'] ='CAPACITACIÓN A AGENTES COMUNITARIOS EN REHABILITACIÓN BASADA EN LA COMUNIDAD (5005155)' 
        
        sheet['B82'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B82'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B82'] ='Capacitación a Agentes Comunitarios  (APP138)' 
        
        sheet['D80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D80'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D80'].fill = fill
        sheet['D80'].border = border
        sheet['D80'] = 'Taller'
        
        sheet['F80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F80'].fill = fill
        sheet['F80'].border = border
        sheet['F80'] = 'Sesion Educativa'
        
        sheet['H80'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H80'].fill = fill
        sheet['H80'].border = border
        sheet['H80'] = 'Sesion Demostrativa'
        
        sheet['D81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D81'].fill = fill
        sheet['D81'].border = border
        sheet['D81'] = 'N°'
        
        sheet['E81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E81'].fill = fill
        sheet['E81'].border = border
        sheet['E81'] = 'Capacitados'
        
        sheet['F81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F81'].fill = fill
        sheet['F81'].border = border
        sheet['F81'] = 'N°'
        
        sheet['G81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G81'].fill = fill
        sheet['G81'].border = border
        sheet['G81'] = 'Capacitados'
        
        sheet['H81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H81'].fill = fill
        sheet['H81'].border = border
        sheet['H81'] = 'N° '
        
        sheet['I81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I81'].fill = fill
        sheet['I81'].border = border
        sheet['I81'] = 'Capacitados'
        
        
        #borde plomo
        for row in sheet.iter_rows(min_row=82, max_row=82, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        ################################################
        ########## VISITAS RBC #########################
        ################################################
        sheet['B84'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B84'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B84'] ='Vistas a alas familias Rehabilitacion Basada en la Comunidad' 
                
        sheet['B85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B85'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B85'].fill = fill
        sheet['B85'].border = border
        sheet['B85'] = 'Visitas'
        
        sheet['D85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D85'].fill = fill
        sheet['D85'].border = border
        sheet['D85'] = 'Total'
        
        sheet['E85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E85'].fill = fill
        sheet['E85'].border = border
        sheet['E85'] = 'Niños         (1d - 11a)'
        
        sheet['F85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F85'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F85'].fill = fill
        sheet['F85'].border = border
        sheet['F85'] = 'Adolescentes (12a - 17a)'
        
        sheet['G85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G85'].fill = fill
        sheet['G85'].border = border
        sheet['G85'] = 'Jóvenes (18a - 29a)'
        
        sheet['H85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H85'].fill = fill
        sheet['H85'].border = border
        sheet['H85'] = 'Adultos (30a - 59a)'
        
        sheet['I85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I85'].fill = fill
        sheet['I85'].border = border
        sheet['I85'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=86, max_row=90, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = borde_plomo
        
        sheet['B86'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B86'].font = Font(name = 'Arial', size= 8)
        sheet['B86'] ='1º Visita' 
        
        sheet['B87'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B87'].font = Font(name = 'Arial', size= 8)
        sheet['B87'] ='2º Visita' 
        
        sheet['B88'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B88'].font = Font(name = 'Arial', size= 8)
        sheet['B88'] ='3º Visita' 
        
        sheet['B89'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B89'].font = Font(name = 'Arial', size= 8)
        sheet['B89'] ='4º a Visita (trazador)' 
        
        sheet['B90'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B90'].font = Font(name = 'Arial', size= 8)
        sheet['B90'] ='5º a + Visitas' 
        
        sheet['B91'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B91'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B91'] ='SUB TOTAL' 
        
        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B93'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B93'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B93'] ='Capacitación a Actores Sociales para la aplicación de la estrategia de Rehabilitación Basada en la Comunidad' 
                
        sheet['B94'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B94'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B94'] ='Actividades con Gobiernos Locales:' 
        
        sheet['B97'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B97'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B97'] ='Actividad con Comité Multisectorial (APP96)' 
        
        sheet['D95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D95'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D95'].fill = fill
        sheet['D95'].border = border
        sheet['D95'] = 'Taller'
        
        sheet['F95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F95'].fill = fill
        sheet['F95'].border = border
        sheet['F95'] = 'Sesion Educativa'
        
        sheet['H95'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H95'].fill = fill
        sheet['H95'].border = border
        sheet['H95'] = 'Sesion Demostrativa'
        
        sheet['D96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D96'].fill = fill
        sheet['D96'].border = border
        sheet['D96'] = 'N°'
        
        sheet['E96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E96'].fill = fill
        sheet['E96'].border = border
        sheet['E96'] = 'Capacitados'
        
        sheet['F96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F96'].fill = fill
        sheet['F96'].border = border
        sheet['F96'] = 'N°'
        
        sheet['G96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G96'].fill = fill
        sheet['G96'].border = border
        sheet['G96'] = 'Capacitados'
        
        sheet['H96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H96'].fill = fill
        sheet['H96'].border = border
        sheet['H96'] = 'N° '
        
        sheet['I96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I96'].fill = fill
        sheet['I96'].border = border
        sheet['I96'] = 'Capacitados'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=97, max_row=97, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        #############################################################################
        #############################################################################                
        # cambina celdas
        sheet.merge_cells('C6:D6')
        sheet.merge_cells('C7:E7')
        
        sheet.merge_cells('B18:B19')
        sheet.merge_cells('B20:B22')
        sheet.merge_cells('B27:B33')
        sheet.merge_cells('B37:B43')
        sheet.merge_cells('B44:B45')
        
        # sensorial
        sheet.merge_cells('B57:B58')
        
        sheet.merge_cells('B15:C15')
        sheet.merge_cells('B51:C51')
        
        # mental
        sheet.merge_cells('B62:C62')
        
        #certificado
        sheet.merge_cells('B70:C70')
        
        sheet.merge_cells('B71:B73')
        
        #RBC
        sheet.merge_cells('B85:C85')
        
        #capa
        sheet.merge_cells('D80:E80')
        sheet.merge_cells('F80:G80')
        sheet.merge_cells('H80:I80')

        sheet.merge_cells('D95:E95')
        sheet.merge_cells('F95:G95')
        sheet.merge_cells('H95:I95')
        
        #capacitacion
        sheet.merge_cells('B82:C82')
        sheet.merge_cells('B97:C97')
        
        #visita
        sheet.merge_cells('B86:C86')
        sheet.merge_cells('B87:C87')
        sheet.merge_cells('B88:C88')
        sheet.merge_cells('B89:C89')
        sheet.merge_cells('B90:C90')
        
        # Definir ubicaciones específicas para cada columna y su suma total
        columnas_ubicaciones = {
            'PROVINCIA': 'D10',
            'DIS_1': 'E16', 
            'DIS_2': 'F16',
            'DIS_3': 'G16',
            'DIS_4': 'H16',
            'DIS_5': 'I16',
            'DIS_6': 'E17',
            'DIS_7': 'F17',
            'DIS_8': 'G17',
            'DIS_9': 'H17',
            'DIS_10': 'I17',
            'DIS_11': 'E18',
            'DIS_12': 'F18',
            'DIS_13': 'G18',
            'DIS_14': 'H18',
            'DIS_15': 'I18',
            'DIS_16': 'E19',
            'DIS_17': 'F19',
            'DIS_18': 'G19',
            'DIS_19': 'H19',
            'DIS_20': 'I19',
            'DIS_21': 'E20',
            'DIS_22': 'F20',
            'DIS_23': 'G20',
            'DIS_24': 'H20',
            'DIS_25': 'I20',
            'DIS_26': 'E21',
            'DIS_27': 'F21',
            'DIS_28': 'G21',
            'DIS_29': 'H21',
            'DIS_30': 'I21',
            'DIS_31': 'E22',
            'DIS_32': 'F22',
            'DIS_33': 'G22',
            'DIS_34': 'H22',
            'DIS_35': 'I22',
            'DIS_36': 'E23',
            'DIS_37': 'F23',
            'DIS_38': 'G23',
            'DIS_39': 'H23',
            'DIS_40': 'I23',
            'DIS_41': 'E24',
            'DIS_42': 'F24',
            'DIS_43': 'G24',
            'DIS_44': 'H24',
            'DIS_45': 'I24',
            'DIS_46': 'E25',
            'DIS_47': 'F25',
            'DIS_48': 'G25',
            'DIS_49': 'H25',
            'DIS_50': 'I25',
            'DIS_51': 'E26',
            'DIS_52': 'F26',
            'DIS_53': 'G26',
            'DIS_54': 'H26',
            'DIS_55': 'I26',
            'DIS_56': 'E27',
            'DIS_57': 'F27',
            'DIS_58': 'G27',
            'DIS_59': 'H27',
            'DIS_60': 'I27',
            'DIS_61': 'E28',
            'DIS_62': 'F28',
            'DIS_63': 'G28',
            'DIS_64': 'H28',
            'DIS_65': 'I28',
            'DIS_66': 'E29',
            'DIS_67': 'F29',
            'DIS_68': 'G29',
            'DIS_69': 'H29',
            'DIS_70': 'I29',
            'DIS_71': 'E30',
            'DIS_72': 'F30',
            'DIS_73': 'G30',
            'DIS_74': 'H30',
            'DIS_75': 'I30',
            'DIS_76': 'E31',
            'DIS_77': 'F31',
            'DIS_78': 'G31',
            'DIS_79': 'H31',
            'DIS_80': 'I31',
            'DIS_81': 'E32',
            'DIS_82': 'F32',
            'DIS_83': 'G32',
            'DIS_84': 'H32',
            'DIS_85': 'I32',
            'DIS_86': 'E33',
            'DIS_87': 'F33',
            'DIS_88': 'G33',
            'DIS_89': 'H33',
            'DIS_90': 'I33',
            'DIS_91': 'E34',
            'DIS_92': 'F34',
            'DIS_93': 'G34',
            'DIS_94': 'H34',
            'DIS_95': 'I34',
            'DIS_96': 'E35',
            'DIS_97': 'F35',
            'DIS_98': 'G35',
            'DIS_99': 'H35',
            'DIS_100': 'I35',
            'DIS_101': 'E36',
            'DIS_102': 'F36',
            'DIS_103': 'G36',
            'DIS_104': 'H36',
            'DIS_105': 'I36',
            'DIS_106': 'E37',
            'DIS_107': 'F37',
            'DIS_108': 'G37',
            'DIS_109': 'H37',
            'DIS_110': 'I37',
            'DIS_111': 'E38',
            'DIS_112': 'F38',
            'DIS_113': 'G38',
            'DIS_114': 'H38',
            'DIS_115': 'I38',
            'DIS_116': 'E39',
            'DIS_117': 'F39',
            'DIS_118': 'G39',
            'DIS_119': 'H39',
            'DIS_120': 'I39',
            'DIS_121': 'E40',
            'DIS_122': 'F40',
            'DIS_123': 'G40',
            'DIS_124': 'H40',
            'DIS_125': 'I40',
            'DIS_126': 'E41',
            'DIS_127': 'F41',
            'DIS_128': 'G41',
            'DIS_129': 'H41',
            'DIS_130': 'I41', 
            'DIS_131': 'E42',
            'DIS_132': 'F42',
            'DIS_133': 'G42',
            'DIS_134': 'H42',
            'DIS_135': 'I42', 
            'DIS_136': 'E43',
            'DIS_137': 'F43',
            'DIS_138': 'G43',
            'DIS_139': 'H43',
            'DIS_140': 'I43', 
            'DIS_141': 'E44',
            'DIS_142': 'F44',
            'DIS_143': 'G44',
            'DIS_144': 'H44',
            'DIS_145': 'I44', 
            'DIS_146': 'E45',
            'DIS_147': 'F45',
            'DIS_148': 'G45',
            'DIS_149': 'H45',
            'DIS_150': 'I45', 
            'DIS_151': 'E46',
            'DIS_152': 'F46',
            'DIS_153': 'G46',
            'DIS_154': 'H46',
            'DIS_155': 'I46', 
            'DIS_156': 'E47',
            'DIS_157': 'F47',
            'DIS_158': 'G47',
            'DIS_159': 'H47',
            'DIS_160': 'I47',            
        }
        
        col_ubi_sensorial = {    
            'PROVINCIA': 'D10',
            'DIS_161': 'E52',
            'DIS_162': 'F52',
            'DIS_163': 'G52',
            'DIS_164': 'H52',
            'DIS_165': 'I52',
            'DIS_166': 'E53',
            'DIS_167': 'F53',
            'DIS_168': 'G53',
            'DIS_169': 'H53',
            'DIS_170': 'I53',
            'DIS_171': 'E54',
            'DIS_172': 'F54',
            'DIS_173': 'G54',
            'DIS_174': 'H54',
            'DIS_175': 'I54',
            'DIS_176': 'E55',
            'DIS_177': 'F55',
            'DIS_178': 'G55',
            'DIS_179': 'H55',
            'DIS_180': 'I55',
            'DIS_181': 'E56',
            'DIS_182': 'F56',
            'DIS_183': 'G56',
            'DIS_184': 'H56',
            'DIS_185': 'I56',
            'DIS_186': 'E57',
            'DIS_187': 'F57',
            'DIS_188': 'G57',
            'DIS_189': 'H57',
            'DIS_190': 'I57',
            'DIS_191': 'E58',
            'DIS_192': 'F58',
            'DIS_193': 'G58',
            'DIS_194': 'H58',
            'DIS_195': 'I58',
        }
        
        col_ubi_mental = {    
            'PROVINCIA': 'D10',
            'DIS_196': 'E63',
            'DIS_197': 'F63',
            'DIS_198': 'G63',
            'DIS_199': 'H63',
            'DIS_200': 'I63',
            'DIS_201': 'E64',
            'DIS_202': 'F64',
            'DIS_203': 'G64',
            'DIS_204': 'H64',
            'DIS_205': 'I64',
            'DIS_206': 'E65',
            'DIS_207': 'F65',
            'DIS_208': 'G65',
            'DIS_209': 'H65',
            'DIS_210': 'I65',
            'DIS_211': 'E66',
            'DIS_212': 'F66',
            'DIS_213': 'G66',
            'DIS_214': 'H66',
            'DIS_215': 'I66',
        }
        
        col_ubi_certificado = {    
            'PROVINCIA': 'D10',
            'DIS_216': 'E71',
            'DIS_217': 'F71',
            'DIS_218': 'G71',
            'DIS_219': 'H71',
            'DIS_220': 'I71',
            'DIS_221': 'E72',
            'DIS_222': 'F72',
            'DIS_223': 'G72',
            'DIS_224': 'H72',
            'DIS_225': 'I72',
            'DIS_226': 'E73',
            'DIS_227': 'F73',
            'DIS_228': 'G73',
            'DIS_229': 'H73',
            'DIS_230': 'I73',
            'DIS_231': 'E74',
            'DIS_232': 'F74',
            'DIS_233': 'G74',
            'DIS_234': 'H74',
            'DIS_235': 'I74',
        }
        
        col_ubi_capacitacion = {    
            'PROVINCIA': 'D10',
            'DIS_273': 'D12',
            'DIS_274': 'E12',
        }
        
        col_ubi_agente = {    
            'PROVINCIA': 'D10',
            'DIS_236': 'D82',
            'DIS_237': 'E82',
            'DIS_238': 'F82',
            'DIS_239': 'G82',
            'DIS_240': 'H82',
            'DIS_241': 'I82',
        }      
        
        col_ubi_rbc = {    
            'PROVINCIA': 'D10',
            'DIS_242': 'E86',
            'DIS_243': 'F86',
            'DIS_244': 'G86',
            'DIS_245': 'H86',
            'DIS_246': 'I86',
            'DIS_247': 'E87',
            'DIS_248': 'F87',
            'DIS_249': 'G87',
            'DIS_250': 'H87',
            'DIS_251': 'I87',
            'DIS_252': 'E88',
            'DIS_253': 'F88',
            'DIS_254': 'G88',
            'DIS_255': 'H88',
            'DIS_256': 'I88',
            'DIS_257': 'E89',
            'DIS_258': 'F89',
            'DIS_259': 'G89',
            'DIS_260': 'H89',
            'DIS_261': 'I89',
            'DIS_262': 'E90',
            'DIS_263': 'F90',
            'DIS_264': 'G90',
            'DIS_265': 'H90',
            'DIS_266': 'I90'
        }
        
        col_ubi_comite = {    
            'PROVINCIA': 'D10',
            'DIS_267': 'D97',
            'DIS_268': 'E97',
            'DIS_269': 'F97',
            'DIS_270': 'G97',
            'DIS_271': 'H97',
            'DIS_272': 'I97',
        }
        
        # Inicializar diccionario para almacenar sumas por columna
        column_sums = {
            'DIS_1': 0,
            'DIS_2': 0,
            'DIS_3': 0,
            'DIS_4': 0,
            'DIS_5': 0,
            'DIS_6': 0,
            'DIS_7': 0,
            'DIS_8': 0,
            'DIS_9': 0,
            'DIS_10': 0,
            'DIS_11': 0,
            'DIS_12': 0,
            'DIS_13': 0,
            'DIS_14': 0,
            'DIS_15': 0,
            'DIS_16': 0,
            'DIS_17': 0,
            'DIS_18': 0,
            'DIS_19': 0,
            'DIS_20': 0,
            'DIS_21': 0,
            'DIS_22': 0,
            'DIS_23': 0,
            'DIS_24': 0,
            'DIS_25': 0,
            'DIS_26': 0,
            'DIS_27': 0,
            'DIS_28': 0,
            'DIS_29': 0,
            'DIS_30': 0,
            'DIS_31': 0,
            'DIS_32': 0,
            'DIS_33': 0,
            'DIS_34': 0,
            'DIS_35': 0,
            'DIS_36': 0,
            'DIS_37': 0,
            'DIS_38': 0,
            'DIS_39': 0,
            'DIS_40': 0,
            'DIS_41': 0,
            'DIS_42': 0,
            'DIS_43': 0,
            'DIS_44': 0,
            'DIS_45': 0,
            'DIS_46': 0,
            'DIS_47': 0,
            'DIS_48': 0,
            'DIS_49': 0,
            'DIS_50': 0,
            'DIS_51': 0,
            'DIS_52': 0,
            'DIS_53': 0,
            'DIS_54': 0,
            'DIS_55': 0,
            'DIS_56': 0,
            'DIS_57': 0,
            'DIS_58': 0,
            'DIS_59': 0,
            'DIS_60': 0,
            'DIS_61': 0,
            'DIS_62': 0,
            'DIS_63': 0,
            'DIS_64': 0,
            'DIS_65': 0,
            'DIS_66': 0,
            'DIS_67': 0,
            'DIS_68': 0,
            'DIS_69': 0,
            'DIS_70': 0,
            'DIS_71': 0,
            'DIS_72': 0,
            'DIS_73': 0,
            'DIS_74': 0,
            'DIS_75': 0,
            'DIS_76': 0,
            'DIS_77': 0,
            'DIS_78': 0,
            'DIS_79': 0,
            'DIS_80': 0,
            'DIS_81': 0,
            'DIS_82': 0,
            'DIS_83': 0,
            'DIS_84': 0,
            'DIS_85': 0,
            'DIS_86': 0,
            'DIS_87': 0,
            'DIS_88': 0,
            'DIS_89': 0,
            'DIS_90': 0,
            'DIS_91': 0,
            'DIS_92': 0,
            'DIS_93': 0,
            'DIS_94': 0,
            'DIS_95': 0,
            'DIS_96': 0,
            'DIS_97': 0,
            'DIS_98': 0,
            'DIS_99': 0,
            'DIS_100': 0,
            'DIS_101': 0,
            'DIS_102': 0,
            'DIS_103': 0,
            'DIS_104': 0,
            'DIS_105': 0,
            'DIS_106': 0,
            'DIS_107': 0,
            'DIS_108': 0,
            'DIS_109': 0,
            'DIS_110': 0,
            'DIS_111': 0,
            'DIS_112': 0,
            'DIS_113': 0,
            'DIS_114': 0,
            'DIS_115': 0,
            'DIS_116': 0,
            'DIS_117': 0,
            'DIS_118': 0,
            'DIS_119': 0,
            'DIS_120': 0,
            'DIS_121': 0,
            'DIS_122': 0,
            'DIS_123': 0,
            'DIS_124': 0,
            'DIS_125': 0,
            'DIS_126': 0,
            'DIS_127': 0,
            'DIS_128': 0,
            'DIS_129': 0,
            'DIS_130': 0, 
            'DIS_131': 0,
            'DIS_132': 0,
            'DIS_133': 0,
            'DIS_134': 0,
            'DIS_135': 0, 
            'DIS_136': 0,
            'DIS_137': 0,
            'DIS_138': 0,
            'DIS_139': 0,
            'DIS_140': 0, 
            'DIS_141': 0,
            'DIS_142': 0,
            'DIS_143': 0,
            'DIS_144': 0,
            'DIS_145': 0, 
            'DIS_146': 0,
            'DIS_147': 0,
            'DIS_148': 0,
            'DIS_149': 0,
            'DIS_150': 0, 
            'DIS_151': 0,
            'DIS_152': 0,
            'DIS_153': 0,
            'DIS_154': 0,
            'DIS_155': 0, 
            'DIS_156': 0,
            'DIS_157': 0,
            'DIS_158': 0,
            'DIS_159': 0,
            'DIS_160': 0,    
        }
        
        col_sum_sensorial = {       
            'DIS_161': 0,
            'DIS_162': 0,
            'DIS_163': 0,
            'DIS_164': 0,
            'DIS_165': 0,
            'DIS_166': 0,
            'DIS_167': 0,
            'DIS_168': 0,
            'DIS_169': 0,
            'DIS_170': 0,
            'DIS_171': 0,
            'DIS_172': 0,
            'DIS_173': 0,
            'DIS_174': 0,
            'DIS_175': 0,
            'DIS_176': 0,
            'DIS_177': 0,
            'DIS_178': 0,
            'DIS_179': 0,
            'DIS_180': 0,
            'DIS_181': 0,
            'DIS_182': 0,
            'DIS_183': 0,
            'DIS_184': 0,
            'DIS_185': 0,
            'DIS_186': 0,
            'DIS_187': 0,
            'DIS_188': 0,
            'DIS_189': 0,
            'DIS_190': 0,
            'DIS_191': 0,
            'DIS_192': 0,
            'DIS_193': 0,
            'DIS_194': 0,
            'DIS_195': 0,
        } 

        col_sum_mental = {    
            'DIS_196': 0,
            'DIS_197': 0,
            'DIS_198': 0,
            'DIS_199': 0,
            'DIS_200': 0,
            'DIS_201': 0,
            'DIS_202': 0,
            'DIS_203': 0,
            'DIS_204': 0,
            'DIS_205': 0,
            'DIS_206': 0,
            'DIS_207': 0,
            'DIS_208': 0,
            'DIS_209': 0,
            'DIS_210': 0,
            'DIS_211': 0,
            'DIS_212': 0,
            'DIS_213': 0,
            'DIS_214': 0,
            'DIS_215': 0,
        }
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_certificado = {       
            'DIS_216': 0,
            'DIS_217': 0,
            'DIS_218': 0,
            'DIS_219': 0,
            'DIS_220': 0,
            'DIS_221': 0,
            'DIS_222': 0,
            'DIS_223': 0,
            'DIS_224': 0,
            'DIS_225': 0,
            'DIS_226': 0,
            'DIS_227': 0,
            'DIS_228': 0,
            'DIS_229': 0,
            'DIS_230': 0,
            'DIS_231': 0,
            'DIS_232': 0,
            'DIS_233': 0,
            'DIS_234': 0,
            'DIS_235': 0,
        }  
        
        col_sum_capacitacion = {    
            'DIS_273': 0,
            'DIS_274': 0,
        }
        
        col_sum_agente = {    
            'DIS_236': 0,
            'DIS_237': 0,
            'DIS_238': 0,
            'DIS_239': 0,
            'DIS_240': 0,
            'DIS_241': 0,
        }      
        
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_rbc = {       
            'DIS_242': 0,
            'DIS_243': 0,
            'DIS_244': 0,
            'DIS_245': 0,
            'DIS_246': 0,
            'DIS_247': 0,
            'DIS_248': 0,
            'DIS_249': 0,
            'DIS_250': 0,
            'DIS_251': 0,
            'DIS_252': 0,
            'DIS_253': 0,
            'DIS_254': 0,
            'DIS_255': 0,
            'DIS_256': 0,
            'DIS_257': 0,
            'DIS_258': 0,
            'DIS_259': 0,
            'DIS_260': 0,
            'DIS_261': 0,
            'DIS_262': 0,
            'DIS_263': 0,
            'DIS_264': 0,
            'DIS_265': 0,
            'DIS_266': 0,
        } 
        
        col_sum_comite = {    
            'DIS_267': 0,
            'DIS_268': 0,
            'DIS_269': 0,
            'DIS_270': 0,
            'DIS_271': 0,
            'DIS_272': 0,
        }
                    
        ############################
        ###  DISCAPACIDAD FISICA ###
        ############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov:
            for col_name in column_sums:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(columnas_ubicaciones.keys()).index(col_name) + 1
                    column_sums[col_name] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila: {row}")                        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_name, total_cell in columnas_ubicaciones.items():
            if col_name in column_sums:
                # Obtener la celda correspondiente según la ubicación
                cell = sheet[total_cell]
                # Asignar el valor de la suma a la celda
                cell.value = column_sums[col_name]
                # Aplicar formato a la celda
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        
        # Sumar los valores del diccionario      
        total_sum_cat_1 =  sum([column_sums['DIS_1'], column_sums['DIS_2'], column_sums['DIS_3'],column_sums['DIS_4'],column_sums['DIS_5']])
        total_sum_cat_2 =  sum([column_sums['DIS_6'], column_sums['DIS_7'], column_sums['DIS_8'],column_sums['DIS_9'],column_sums['DIS_10']])
        total_sum_cat_3 =  sum([column_sums['DIS_11'], column_sums['DIS_12'], column_sums['DIS_13'],column_sums['DIS_14'],column_sums['DIS_15']])
        total_sum_cat_4 =  sum([column_sums['DIS_16'], column_sums['DIS_17'], column_sums['DIS_18'],column_sums['DIS_19'],column_sums['DIS_20']])
        total_sum_cat_5 =  sum([column_sums['DIS_21'], column_sums['DIS_22'], column_sums['DIS_23'],column_sums['DIS_24'],column_sums['DIS_25']])
        total_sum_cat_6 =  sum([column_sums['DIS_26'], column_sums['DIS_27'], column_sums['DIS_28'],column_sums['DIS_29'],column_sums['DIS_30']])
        total_sum_cat_7 =  sum([column_sums['DIS_31'], column_sums['DIS_32'], column_sums['DIS_33'],column_sums['DIS_34'],column_sums['DIS_35']])
        total_sum_cat_8 =  sum([column_sums['DIS_36'], column_sums['DIS_37'], column_sums['DIS_38'],column_sums['DIS_39'],column_sums['DIS_40']])
        total_sum_cat_9 =  sum([column_sums['DIS_41'], column_sums['DIS_42'], column_sums['DIS_43'],column_sums['DIS_44'],column_sums['DIS_45']])
        total_sum_cat_10 =  sum([column_sums['DIS_46'], column_sums['DIS_47'], column_sums['DIS_48'],column_sums['DIS_49'],column_sums['DIS_50']])
        total_sum_cat_11 =  sum([column_sums['DIS_51'], column_sums['DIS_52'], column_sums['DIS_53'],column_sums['DIS_54'],column_sums['DIS_55']])
        total_sum_cat_12 =  sum([column_sums['DIS_56'], column_sums['DIS_57'], column_sums['DIS_58'],column_sums['DIS_59'],column_sums['DIS_60']])
        total_sum_cat_13 =  sum([column_sums['DIS_61'], column_sums['DIS_62'], column_sums['DIS_63'],column_sums['DIS_64'],column_sums['DIS_65']])
        total_sum_cat_14 =  sum([column_sums['DIS_66'], column_sums['DIS_67'], column_sums['DIS_68'],column_sums['DIS_69'],column_sums['DIS_70']])
        total_sum_cat_15 =  sum([column_sums['DIS_71'], column_sums['DIS_72'], column_sums['DIS_73'],column_sums['DIS_74'],column_sums['DIS_75']])
        total_sum_cat_16 =  sum([column_sums['DIS_76'], column_sums['DIS_77'], column_sums['DIS_78'],column_sums['DIS_79'],column_sums['DIS_80']])   
        total_sum_cat_17 =  sum([column_sums['DIS_81'], column_sums['DIS_82'], column_sums['DIS_83'],column_sums['DIS_84'],column_sums['DIS_85']])
        total_sum_cat_18 =  sum([column_sums['DIS_86'], column_sums['DIS_87'], column_sums['DIS_88'],column_sums['DIS_89'],column_sums['DIS_90']])
        total_sum_cat_19 =  sum([column_sums['DIS_91'], column_sums['DIS_92'], column_sums['DIS_93'],column_sums['DIS_94'],column_sums['DIS_95']])
        total_sum_cat_20 =  sum([column_sums['DIS_96'], column_sums['DIS_97'], column_sums['DIS_98'],column_sums['DIS_99'],column_sums['DIS_100']])
        total_sum_cat_21 =  sum([column_sums['DIS_101'], column_sums['DIS_102'], column_sums['DIS_103'],column_sums['DIS_104'],column_sums['DIS_105']])
        total_sum_cat_22 =  sum([column_sums['DIS_106'], column_sums['DIS_107'], column_sums['DIS_108'],column_sums['DIS_109'],column_sums['DIS_110']])
        total_sum_cat_23 =  sum([column_sums['DIS_111'], column_sums['DIS_112'], column_sums['DIS_113'],column_sums['DIS_114'],column_sums['DIS_115']])
        total_sum_cat_24 =  sum([column_sums['DIS_116'], column_sums['DIS_117'], column_sums['DIS_118'],column_sums['DIS_119'],column_sums['DIS_120']])
        total_sum_cat_25 =  sum([column_sums['DIS_121'], column_sums['DIS_122'], column_sums['DIS_123'],column_sums['DIS_124'],column_sums['DIS_125']])
        total_sum_cat_26 =  sum([column_sums['DIS_126'], column_sums['DIS_127'], column_sums['DIS_128'],column_sums['DIS_129'],column_sums['DIS_130']])
        total_sum_cat_27 =  sum([column_sums['DIS_131'], column_sums['DIS_132'], column_sums['DIS_133'],column_sums['DIS_134'],column_sums['DIS_135']])
        total_sum_cat_28 =  sum([column_sums['DIS_136'], column_sums['DIS_137'], column_sums['DIS_138'],column_sums['DIS_139'],column_sums['DIS_140']])
        total_sum_cat_29 =  sum([column_sums['DIS_141'], column_sums['DIS_142'], column_sums['DIS_143'],column_sums['DIS_144'],column_sums['DIS_145']])
        total_sum_cat_30 =  sum([column_sums['DIS_146'], column_sums['DIS_147'], column_sums['DIS_148'],column_sums['DIS_149'],column_sums['DIS_150']])
        total_sum_cat_31 =  sum([column_sums['DIS_151'], column_sums['DIS_152'], column_sums['DIS_153'],column_sums['DIS_154'],column_sums['DIS_155']])
        total_sum_cat_32 =  sum([column_sums['DIS_156'], column_sums['DIS_157'], column_sums['DIS_158'],column_sums['DIS_159'],column_sums['DIS_160']])

        sheet['D16'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D16'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D16'] = total_sum_cat_1     
        
        sheet['D17'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D17'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D17'] = total_sum_cat_2 
        
        sheet['D18'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D18'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D18'] = total_sum_cat_3    
        
        sheet['D19'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D19'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D19'] = total_sum_cat_4    
        
        sheet['D20'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D20'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D20'] = total_sum_cat_5    
        
        sheet['D21'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D21'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D21'] = total_sum_cat_6    
        
        sheet['D22'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D22'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D22'] = total_sum_cat_7    
        
        sheet['D23'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D23'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D23'] = total_sum_cat_8    
        
        sheet['D24'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D24'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D24'] = total_sum_cat_9    
        
        sheet['D25'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D25'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D25'] = total_sum_cat_10 
        
        sheet['D26'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D26'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D26'] = total_sum_cat_11
                
        sheet['D27'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D27'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D27'] = total_sum_cat_12    
        
        sheet['D28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D28'] = total_sum_cat_13   
        
        sheet['D29'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D29'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D29'] = total_sum_cat_14   
        
        sheet['D30'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D30'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D30'] = total_sum_cat_15   
        
        sheet['D31'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D31'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D31'] = total_sum_cat_16   
        
        sheet['D32'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D32'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D32'] = total_sum_cat_17         
        
        sheet['D33'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D33'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D33'] = total_sum_cat_18   
        
        sheet['D34'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D34'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D34'] = total_sum_cat_19   
        
        sheet['D35'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D35'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D35'] = total_sum_cat_20   
        
        sheet['D36'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D36'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D36'] = total_sum_cat_21   
        
        sheet['D37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D37'] = total_sum_cat_22   
        
        sheet['D38'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D38'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D38'] = total_sum_cat_23   
        
        sheet['D39'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D39'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D39'] = total_sum_cat_24   
        
        sheet['D40'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D40'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D40'] = total_sum_cat_25  
        
        sheet['D41'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D41'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D41'] = total_sum_cat_26 
        
        sheet['D42'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D42'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D42'] = total_sum_cat_27   
        
        sheet['D43'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D43'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D43'] = total_sum_cat_28   
        
        sheet['D44'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D44'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D44'] = total_sum_cat_29  
        
        sheet['D45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D45'] = total_sum_cat_30  
        
        sheet['D46'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D46'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D46'] = total_sum_cat_31
        
        sheet['D47'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D47'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D47'] = total_sum_cat_32
        
        # Sumar los valores del VERTICAL      
        total_sum_cat_vertical_1 =  sum([column_sums['DIS_1'],column_sums['DIS_6'], column_sums['DIS_11'],column_sums['DIS_16'],column_sums['DIS_21'],column_sums['DIS_26'],column_sums['DIS_31'],column_sums['DIS_36'],column_sums['DIS_41'],column_sums['DIS_46'],column_sums['DIS_51'],column_sums['DIS_56'],column_sums['DIS_61'],column_sums['DIS_66'],column_sums['DIS_71'],column_sums['DIS_76'],column_sums['DIS_81'],column_sums['DIS_86'],column_sums['DIS_91'],column_sums['DIS_96'],column_sums['DIS_101'],column_sums['DIS_106'] ,column_sums['DIS_111'],column_sums['DIS_116'],column_sums['DIS_121'],column_sums['DIS_126'],column_sums['DIS_131'],column_sums['DIS_136'],column_sums['DIS_141'],column_sums['DIS_146'],column_sums['DIS_151'],column_sums['DIS_156']])
        total_sum_cat_vertical_2 =  sum([column_sums['DIS_2'],column_sums['DIS_7'], column_sums['DIS_12'],column_sums['DIS_17'],column_sums['DIS_22'],column_sums['DIS_27'],column_sums['DIS_32'],column_sums['DIS_37'],column_sums['DIS_42'],column_sums['DIS_47'],column_sums['DIS_52'],column_sums['DIS_57'],column_sums['DIS_62'],column_sums['DIS_67'],column_sums['DIS_72'],column_sums['DIS_77'],column_sums['DIS_82'],column_sums['DIS_87'],column_sums['DIS_92'],column_sums['DIS_97'],column_sums['DIS_102'],column_sums['DIS_107'] ,column_sums['DIS_112'],column_sums['DIS_117'],column_sums['DIS_122'],column_sums['DIS_127'],column_sums['DIS_132'],column_sums['DIS_137'],column_sums['DIS_142'],column_sums['DIS_147'],column_sums['DIS_152'],column_sums['DIS_157']])
        total_sum_cat_vertical_3 =  sum([column_sums['DIS_3'],column_sums['DIS_8'], column_sums['DIS_13'],column_sums['DIS_18'],column_sums['DIS_23'],column_sums['DIS_28'],column_sums['DIS_33'],column_sums['DIS_38'],column_sums['DIS_43'],column_sums['DIS_48'],column_sums['DIS_53'],column_sums['DIS_58'],column_sums['DIS_63'],column_sums['DIS_68'],column_sums['DIS_73'],column_sums['DIS_78'],column_sums['DIS_83'],column_sums['DIS_88'],column_sums['DIS_93'],column_sums['DIS_98'],column_sums['DIS_103'],column_sums['DIS_108'] ,column_sums['DIS_113'],column_sums['DIS_118'],column_sums['DIS_123'],column_sums['DIS_128'],column_sums['DIS_133'],column_sums['DIS_138'],column_sums['DIS_143'],column_sums['DIS_148'],column_sums['DIS_153'],column_sums['DIS_158']])
        total_sum_cat_vertical_4 =  sum([column_sums['DIS_4'],column_sums['DIS_9'], column_sums['DIS_14'],column_sums['DIS_19'],column_sums['DIS_24'],column_sums['DIS_29'],column_sums['DIS_34'],column_sums['DIS_39'],column_sums['DIS_44'],column_sums['DIS_49'],column_sums['DIS_54'],column_sums['DIS_59'],column_sums['DIS_64'],column_sums['DIS_69'],column_sums['DIS_74'],column_sums['DIS_79'],column_sums['DIS_84'],column_sums['DIS_89'],column_sums['DIS_94'],column_sums['DIS_99'],column_sums['DIS_104'],column_sums['DIS_109'] ,column_sums['DIS_114'],column_sums['DIS_119'],column_sums['DIS_124'],column_sums['DIS_129'],column_sums['DIS_134'],column_sums['DIS_139'],column_sums['DIS_144'],column_sums['DIS_149'],column_sums['DIS_154'],column_sums['DIS_159']])
        total_sum_cat_vertical_5 =  sum([column_sums['DIS_5'],column_sums['DIS_10'],column_sums['DIS_15'],column_sums['DIS_20'],column_sums['DIS_25'],column_sums['DIS_30'],column_sums['DIS_35'],column_sums['DIS_40'],column_sums['DIS_45'],column_sums['DIS_50'],column_sums['DIS_55'],column_sums['DIS_60'],column_sums['DIS_65'],column_sums['DIS_70'],column_sums['DIS_75'],column_sums['DIS_80'],column_sums['DIS_85'],column_sums['DIS_90'],column_sums['DIS_95'],column_sums['DIS_100'],column_sums['DIS_105'],column_sums['DIS_110'],column_sums['DIS_115'],column_sums['DIS_120'],column_sums['DIS_125'],column_sums['DIS_130'],column_sums['DIS_135'],column_sums['DIS_140'],column_sums['DIS_145'],column_sums['DIS_150'],column_sums['DIS_155'],column_sums['DIS_160']])

        sheet['E48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E48'] = total_sum_cat_vertical_1     
        
        sheet['F48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F48'] = total_sum_cat_vertical_2 
        
        sheet['G48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G48'] = total_sum_cat_vertical_3    
        
        sheet['H48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H48'] = total_sum_cat_vertical_4    
        
        sheet['I48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I48'] = total_sum_cat_vertical_5    
        ##########################################################################
        
        ###############################
        ###  DISCAPACIDAD SENSORIAL ###
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_sensorial:
            for col_sensorial in col_sum_sensorial:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_sensorial.keys()).index(col_sensorial) + 1
                    col_sum_sensorial[col_sensorial] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_sensorial, total_cell_sensorial in col_ubi_sensorial.items():
            if col_sensorial in col_sum_sensorial:
                # Obtener la celda correspondiente según la ubicación
                cell_sensorial = sheet[total_cell_sensorial]
                # Asignar el valor de la suma a la celda
                cell_sensorial.value = col_sum_sensorial[col_sensorial]
                # Aplicar formato a la celda
                cell_sensorial.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_sensorial.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_sensorial.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_sensorial['DIS_161'], col_sum_sensorial['DIS_162'], col_sum_sensorial['DIS_163'], col_sum_sensorial['DIS_164'], col_sum_sensorial['DIS_165']])
        t_sum_cat_2 =  sum([col_sum_sensorial['DIS_166'], col_sum_sensorial['DIS_167'], col_sum_sensorial['DIS_168'], col_sum_sensorial['DIS_169'], col_sum_sensorial['DIS_170']])
        t_sum_cat_3 =  sum([col_sum_sensorial['DIS_171'], col_sum_sensorial['DIS_172'], col_sum_sensorial['DIS_173'], col_sum_sensorial['DIS_174'], col_sum_sensorial['DIS_175']])
        t_sum_cat_4 =  sum([col_sum_sensorial['DIS_176'], col_sum_sensorial['DIS_177'], col_sum_sensorial['DIS_178'], col_sum_sensorial['DIS_179'], col_sum_sensorial['DIS_180']])
        t_sum_cat_5 =  sum([col_sum_sensorial['DIS_181'], col_sum_sensorial['DIS_182'], col_sum_sensorial['DIS_183'], col_sum_sensorial['DIS_184'], col_sum_sensorial['DIS_185']])
        t_sum_cat_6 =  sum([col_sum_sensorial['DIS_186'], col_sum_sensorial['DIS_187'], col_sum_sensorial['DIS_188'], col_sum_sensorial['DIS_189'], col_sum_sensorial['DIS_190']])
        t_sum_cat_7 =  sum([col_sum_sensorial['DIS_191'], col_sum_sensorial['DIS_192'], col_sum_sensorial['DIS_193'], col_sum_sensorial['DIS_194'], col_sum_sensorial['DIS_195']])
        
        sheet['D52'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D52'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D52'] = t_sum_cat_1     
        
        sheet['D53'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D53'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D53'] = t_sum_cat_2 
        
        sheet['D54'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D54'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D54'] = t_sum_cat_3    
        
        sheet['D55'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D55'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D55'] = t_sum_cat_4    
        
        sheet['D56'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D56'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D56'] = t_sum_cat_5    
        
        sheet['D57'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D57'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D57'] = t_sum_cat_6    
        
        sheet['D58'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D58'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D58'] = t_sum_cat_7    
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_sensorial['DIS_161'],col_sum_sensorial['DIS_166'],col_sum_sensorial['DIS_171'],col_sum_sensorial['DIS_176'],col_sum_sensorial['DIS_181'],col_sum_sensorial['DIS_186'],col_sum_sensorial['DIS_191']])
        t_sum_cat_vertical_2 =  sum([col_sum_sensorial['DIS_162'],col_sum_sensorial['DIS_167'],col_sum_sensorial['DIS_172'],col_sum_sensorial['DIS_177'],col_sum_sensorial['DIS_182'],col_sum_sensorial['DIS_187'],col_sum_sensorial['DIS_192']])
        t_sum_cat_vertical_3 =  sum([col_sum_sensorial['DIS_163'],col_sum_sensorial['DIS_168'],col_sum_sensorial['DIS_173'],col_sum_sensorial['DIS_178'],col_sum_sensorial['DIS_183'],col_sum_sensorial['DIS_188'],col_sum_sensorial['DIS_193']])
        t_sum_cat_vertical_4 =  sum([col_sum_sensorial['DIS_164'],col_sum_sensorial['DIS_169'],col_sum_sensorial['DIS_174'],col_sum_sensorial['DIS_179'],col_sum_sensorial['DIS_184'],col_sum_sensorial['DIS_189'],col_sum_sensorial['DIS_194']])
        t_sum_cat_vertical_5 =  sum([col_sum_sensorial['DIS_165'],col_sum_sensorial['DIS_170'],col_sum_sensorial['DIS_175'],col_sum_sensorial['DIS_180'],col_sum_sensorial['DIS_185'],col_sum_sensorial['DIS_190'],col_sum_sensorial['DIS_195']])
        
        sheet['E59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E59'] = t_sum_cat_vertical_1     
        
        sheet['F59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F59'] = t_sum_cat_vertical_2 
        
        sheet['G59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G59'] = t_sum_cat_vertical_3    
        
        sheet['H59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H59'] = t_sum_cat_vertical_4    
        
        sheet['I59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I59'] = t_sum_cat_vertical_5    
        ##########################################################################
                
        ###############################
        ###  DISCAPACIDAD MENTAL ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_mental:
            for col_mental in col_sum_mental:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_mental.keys()).index(col_mental) + 1
                    col_sum_mental[col_mental] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_mental, total_cell_mental in col_ubi_mental.items():
            if col_mental in col_sum_mental:
                # Obtener la celda correspondiente según la ubicación
                cell_mental = sheet[total_cell_mental]
                # Asignar el valor de la suma a la celda
                cell_mental.value = col_sum_mental[col_mental]
                # Aplicar formato a la celda
                cell_mental.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_mental.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_mental.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_mental['DIS_196'], col_sum_mental['DIS_197'], col_sum_mental['DIS_198'], col_sum_mental['DIS_199'], col_sum_mental['DIS_200']])
        t_sum_cat_2 =  sum([col_sum_mental['DIS_201'], col_sum_mental['DIS_202'], col_sum_mental['DIS_203'], col_sum_mental['DIS_204'], col_sum_mental['DIS_205']])
        t_sum_cat_3 =  sum([col_sum_mental['DIS_206'], col_sum_mental['DIS_207'], col_sum_mental['DIS_208'], col_sum_mental['DIS_209'], col_sum_mental['DIS_210']])
        t_sum_cat_4 =  sum([col_sum_mental['DIS_211'], col_sum_mental['DIS_212'], col_sum_mental['DIS_213'], col_sum_mental['DIS_214'], col_sum_mental['DIS_215']])
        
        sheet['D63'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D63'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D63'] = t_sum_cat_1     
        
        sheet['D64'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D64'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D64'] = t_sum_cat_2 
        
        sheet['D65'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D65'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D65'] = t_sum_cat_3    
        
        sheet['D66'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D66'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D66'] = t_sum_cat_4    

        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_mental['DIS_196'],col_sum_mental['DIS_201'],col_sum_mental['DIS_206'],col_sum_mental['DIS_211']])
        t_sum_cat_vertical_2 =  sum([col_sum_mental['DIS_197'],col_sum_mental['DIS_202'],col_sum_mental['DIS_207'],col_sum_mental['DIS_212']])
        t_sum_cat_vertical_3 =  sum([col_sum_mental['DIS_198'],col_sum_mental['DIS_203'],col_sum_mental['DIS_208'],col_sum_mental['DIS_213']])
        t_sum_cat_vertical_4 =  sum([col_sum_mental['DIS_199'],col_sum_mental['DIS_204'],col_sum_mental['DIS_209'],col_sum_mental['DIS_214']])
        t_sum_cat_vertical_5 =  sum([col_sum_mental['DIS_200'],col_sum_mental['DIS_205'],col_sum_mental['DIS_210'],col_sum_mental['DIS_215']])
        
        sheet['E67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E67'] = t_sum_cat_vertical_1     
        
        sheet['F67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F67'] = t_sum_cat_vertical_2 
        
        sheet['G67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G67'] = t_sum_cat_vertical_3    
        
        sheet['H67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H67'] = t_sum_cat_vertical_4    
        
        sheet['I67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I67'] = t_sum_cat_vertical_5    
        ##########################################################################
        
        
        #################################
        ###  DISCAPACIDAD CERTIFICADO ###
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_certificado:
            for col_certificado in col_sum_certificado:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_certificado.keys()).index(col_certificado) + 1
                    col_sum_certificado[col_certificado] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_certificado, total_cell_certificado in col_ubi_certificado.items():
            if col_certificado in col_sum_certificado:
                # Obtener la celda correspondiente según la ubicación
                cell_certificado = sheet[total_cell_certificado]
                # Asignar el valor de la suma a la celda
                cell_certificado.value = col_sum_certificado[col_certificado]
                # Aplicar formato a la celda
                cell_certificado.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_certificado.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_certificado.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        # Sumar los valores del diccionario      
        t_sum_cat_cert_1 =  sum([col_sum_certificado['DIS_216'], col_sum_certificado['DIS_217'], col_sum_certificado['DIS_218'], col_sum_certificado['DIS_219'], col_sum_certificado['DIS_220']])
        t_sum_cat_cert_2 =  sum([col_sum_certificado['DIS_221'], col_sum_certificado['DIS_222'], col_sum_certificado['DIS_223'], col_sum_certificado['DIS_224'], col_sum_certificado['DIS_225']])
        t_sum_cat_cert_3 =  sum([col_sum_certificado['DIS_226'], col_sum_certificado['DIS_227'], col_sum_certificado['DIS_228'], col_sum_certificado['DIS_229'], col_sum_certificado['DIS_230']])
        t_sum_cat_cert_4 =  sum([col_sum_certificado['DIS_231'], col_sum_certificado['DIS_232'], col_sum_certificado['DIS_233'], col_sum_certificado['DIS_234'], col_sum_certificado['DIS_235']])

        sheet['D71'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D71'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D71'] = t_sum_cat_cert_1     
        
        sheet['D72'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D72'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D72'] = t_sum_cat_cert_2 
        
        sheet['D73'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D73'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D73'] = t_sum_cat_cert_3 
        
        sheet['D74'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D74'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D74'] = t_sum_cat_cert_4 
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vert_1 =  sum([col_sum_certificado['DIS_216'],col_sum_certificado['DIS_221'],col_sum_certificado['DIS_226'],col_sum_certificado['DIS_231']])
        t_sum_cat_vert_2 =  sum([col_sum_certificado['DIS_217'],col_sum_certificado['DIS_222'],col_sum_certificado['DIS_227'],col_sum_certificado['DIS_232']])
        t_sum_cat_vert_3 =  sum([col_sum_certificado['DIS_218'],col_sum_certificado['DIS_223'],col_sum_certificado['DIS_228'],col_sum_certificado['DIS_233']])
        t_sum_cat_vert_4 =  sum([col_sum_certificado['DIS_219'],col_sum_certificado['DIS_224'],col_sum_certificado['DIS_229'],col_sum_certificado['DIS_234']])
        t_sum_cat_vert_5 =  sum([col_sum_certificado['DIS_220'],col_sum_certificado['DIS_225'],col_sum_certificado['DIS_230'],col_sum_certificado['DIS_235']])
        
        sheet['E75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E75'] = t_sum_cat_vert_1     
        
        sheet['F75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F75'] = t_sum_cat_vert_2 
        
        sheet['G75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G75'] = t_sum_cat_vert_3    
        
        sheet['H75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H75'] = t_sum_cat_vert_4    
        
        sheet['I75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I75'] = t_sum_cat_vert_5    
        
        #################################
        ###  DISCAPACIDAD RBC ###########
        #################################       
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_rbc:
            for col_rbc in col_sum_rbc:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_rbc.keys()).index(col_rbc) + 1
                    col_sum_rbc[col_rbc] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
                    
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_rbc, total_cell_rbc in col_ubi_rbc.items():
            if col_rbc in col_sum_rbc:
                # Obtener la celda correspondiente según la ubicación
                cell_rbc = sheet[total_cell_rbc]
                # Asignar el valor de la suma a la celda
                cell_rbc.value = col_sum_rbc[col_rbc]
                # Aplicar formato a la celda
                cell_rbc.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_rbc.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_rbc.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        ## Sumar los valores del diccionario      
        t_sum_cat_rbc_1 =  sum([col_sum_rbc['DIS_242'], col_sum_rbc['DIS_243'], col_sum_rbc['DIS_244'], col_sum_rbc['DIS_245'], col_sum_rbc['DIS_246']])
        t_sum_cat_rbc_2 =  sum([col_sum_rbc['DIS_247'], col_sum_rbc['DIS_248'], col_sum_rbc['DIS_249'], col_sum_rbc['DIS_250'], col_sum_rbc['DIS_251']])
        t_sum_cat_rbc_3 =  sum([col_sum_rbc['DIS_252'], col_sum_rbc['DIS_253'], col_sum_rbc['DIS_254'], col_sum_rbc['DIS_255'], col_sum_rbc['DIS_256']])
        t_sum_cat_rbc_4 =  sum([col_sum_rbc['DIS_257'], col_sum_rbc['DIS_258'], col_sum_rbc['DIS_259'], col_sum_rbc['DIS_260'], col_sum_rbc['DIS_261']])
        t_sum_cat_rbc_5 =  sum([col_sum_rbc['DIS_262'], col_sum_rbc['DIS_263'], col_sum_rbc['DIS_264'], col_sum_rbc['DIS_265'], col_sum_rbc['DIS_266']])

        sheet['D86'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D86'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D86'] = t_sum_cat_rbc_1     
        
        sheet['D87'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D87'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D87'] = t_sum_cat_rbc_2 
        
        sheet['D88'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D88'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D88'] = t_sum_cat_rbc_3     
        
        sheet['D89'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D89'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D89'] = t_sum_cat_rbc_4 
        
        sheet['D90'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D90'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D90'] = t_sum_cat_rbc_5 
        
        # Sumar los valores del VERTICAL      
        t_sum_vert_rbc_1 =  sum([col_sum_rbc['DIS_242'],col_sum_rbc['DIS_247'],col_sum_rbc['DIS_252'],col_sum_rbc['DIS_257'],col_sum_rbc['DIS_262']])
        t_sum_vert_rbc_2 =  sum([col_sum_rbc['DIS_243'],col_sum_rbc['DIS_248'],col_sum_rbc['DIS_253'],col_sum_rbc['DIS_258'],col_sum_rbc['DIS_263']])
        t_sum_vert_rbc_3 =  sum([col_sum_rbc['DIS_244'],col_sum_rbc['DIS_249'],col_sum_rbc['DIS_254'],col_sum_rbc['DIS_259'],col_sum_rbc['DIS_264']])
        t_sum_vert_rbc_4 =  sum([col_sum_rbc['DIS_245'],col_sum_rbc['DIS_250'],col_sum_rbc['DIS_255'],col_sum_rbc['DIS_260'],col_sum_rbc['DIS_265']])
        t_sum_vert_rbc_5 =  sum([col_sum_rbc['DIS_246'],col_sum_rbc['DIS_251'],col_sum_rbc['DIS_256'],col_sum_rbc['DIS_261'],col_sum_rbc['DIS_266']])
        
        sheet['E91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E91'] = t_sum_vert_rbc_1
        
        sheet['F91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F91'] = t_sum_vert_rbc_2 
        
        sheet['G91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G91'] = t_sum_vert_rbc_3    
        
        sheet['H91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H91'] = t_sum_vert_rbc_4    
        
        sheet['I91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I91'] = t_sum_vert_rbc_5   
        
        #################################
        ###  CAPACITACION PERSONAL ######
        #################################
        
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_capacitacion:
            for col_capacitacion in col_sum_capacitacion:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_capacitacion.keys()).index(col_capacitacion) + 1
                    col_sum_capacitacion[col_capacitacion] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_capacitacion, total_cell_capacitacion in col_ubi_capacitacion.items():
            if col_capacitacion in col_sum_capacitacion:
                # Obtener la celda correspondiente según la ubicación
                cell_capacitacion = sheet[total_cell_capacitacion]
                # Asignar el valor de la suma a la celda
                cell_capacitacion.value = col_sum_capacitacion[col_capacitacion]
                # Aplicar formato a la celda
                cell_capacitacion.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_capacitacion.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_capacitacion.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_capacitacion['DIS_273']])
        t_sum_cat_2 = sum([col_sum_capacitacion['DIS_274']])
        
        sheet['D12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D12'] = t_sum_cat_1     
        
        sheet['E12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E12'] = t_sum_cat_2 
        
        ###############################
        ###  CAPACITACION AGENTE ######
        ###############################
                
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_agente:
            for col_agente in col_sum_agente:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_agente.keys()).index(col_agente) + 1
                    col_sum_agente[col_agente] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_agente, total_cell_agente in col_ubi_agente.items():
            if col_agente in col_sum_agente:
                # Obtener la celda correspondiente según la ubicación
                cell_agente = sheet[total_cell_agente]
                # Asignar el valor de la suma a la celda
                cell_agente.value = col_sum_agente[col_agente]
                # Aplicar formato a la celda
                cell_agente.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_agente.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_agente.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_agente['DIS_236']])
        t_sum_cat_2 = sum([col_sum_agente['DIS_237']])
        t_sum_cat_3 = sum([col_sum_agente['DIS_238']])
        t_sum_cat_4 = sum([col_sum_agente['DIS_239']])
        t_sum_cat_5 = sum([col_sum_agente['DIS_240']])
        t_sum_cat_6 = sum([col_sum_agente['DIS_241']])
        
        sheet['D82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D82'] = t_sum_cat_1     
        
        sheet['E82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E82'] = t_sum_cat_2 
        
        sheet['F82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F82'] = t_sum_cat_3
        
        sheet['G82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G82'] = t_sum_cat_4 
        
        sheet['H82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H82'] = t_sum_cat_5
        
        sheet['I82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I82'] = t_sum_cat_6 
        
        ############################
        ###  CAPACITACION COMITE ###
        #############################
        
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_prov_comite:
            for col_comite in col_sum_comite:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_comite.keys()).index(col_comite) + 1
                    col_sum_comite[col_comite] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_comite, total_cell_comite in col_ubi_comite.items():
            if col_comite in col_sum_comite:
                # Obtener la celda correspondiente según la ubicación
                cell_comite = sheet[total_cell_comite]
                # Asignar el valor de la suma a la celda
                cell_comite.value = col_sum_comite[col_comite]
                # Aplicar formato a la celda
                cell_comite.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_comite.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_comite.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_comite['DIS_267']])
        t_sum_cat_2 = sum([col_sum_comite['DIS_268']])
        t_sum_cat_3 = sum([col_sum_comite['DIS_269']])
        t_sum_cat_4 = sum([col_sum_comite['DIS_270']])
        t_sum_cat_5 = sum([col_sum_comite['DIS_271']])
        t_sum_cat_6 = sum([col_sum_comite['DIS_272']])
        
        sheet['D97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D97'] = t_sum_cat_1     
        
        sheet['E97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E97'] = t_sum_cat_2 
        
        sheet['F97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F97'] = t_sum_cat_3
        
        sheet['G97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G97'] = t_sum_cat_4 
        
        sheet['H97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H97'] = t_sum_cat_5
        
        sheet['I97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I97'] = t_sum_cat_6 
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_operacional_provincia.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        workbook.save(response)

        return response

################################################
# REPORTE DE DISTRITO
################################################
def get_distritos(request, distritos_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'discapacidad/distritos.html', context)

def p_distritos(request):
    provincia_param = request.GET.get('provincia')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 'discapacidad/partials/p_distritos.html', context)

#--- FUNCIONES OPERACIONALES PARTES REPORTE -----------------------------------------
def rpt_operacional_fisico_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                    renaes,                   
                    SUM(dis_1) AS dis_1,
                    SUM(dis_2) AS dis_2,
                    SUM(dis_3) AS dis_3,
                    SUM(dis_4) AS dis_4,
                    SUM(dis_5) AS dis_5,
                    SUM(dis_6) AS dis_6,
                    SUM(dis_7) AS dis_7,
                    SUM(dis_8) AS dis_8,
                    SUM(dis_9) AS dis_9,
                    SUM(dis_10) AS dis_10,
                    SUM(dis_11) AS dis_11,
                    SUM(dis_12) AS dis_12,
                    SUM(dis_13) AS dis_13,
                    SUM(dis_14) AS dis_14,
                    SUM(dis_15) AS dis_15,
                    SUM(dis_16) AS dis_16,
                    SUM(dis_17) AS dis_17,
                    SUM(dis_18) AS dis_18,
                    SUM(dis_19) AS dis_19,
                    SUM(dis_20) AS dis_20,
                    SUM(dis_21) AS dis_21,
                    SUM(dis_22) AS dis_22,
                    SUM(dis_23) AS dis_23,
                    SUM(dis_24) AS dis_24,
                    SUM(dis_25) AS dis_25,
                    SUM(dis_26) AS dis_26,
                    SUM(dis_27) AS dis_27,
                    SUM(dis_28) AS dis_28,
                    SUM(dis_29) AS dis_29,
                    SUM(dis_30) AS dis_30,
                    SUM(dis_31) AS dis_31,
                    SUM(dis_32) AS dis_32,
                    SUM(dis_33) AS dis_33,
                    SUM(dis_34) AS dis_34,
                    SUM(dis_35) AS dis_35,
                    SUM(dis_36) AS dis_36,
                    SUM(dis_37) AS dis_37,
                    SUM(dis_38) AS dis_38,
                    SUM(dis_39) AS dis_39,
                    SUM(dis_40) AS dis_40,
                    SUM(dis_41) AS dis_41,
                    SUM(dis_42) AS dis_42,
                    SUM(dis_43) AS dis_43,
                    SUM(dis_44) AS dis_44,
                    SUM(dis_45) AS dis_45,
                    SUM(dis_46) AS dis_46,
                    SUM(dis_47) AS dis_47,
                    SUM(dis_48) AS dis_48,
                    SUM(dis_49) AS dis_49,
                    SUM(dis_50) AS dis_50,
                    SUM(dis_51) AS dis_51,
                    SUM(dis_52) AS dis_52,
                    SUM(dis_53) AS dis_53,
                    SUM(dis_54) AS dis_54,
                    SUM(dis_55) AS dis_55,
                    SUM(dis_56) AS dis_56,
                    SUM(dis_57) AS dis_57,
                    SUM(dis_58) AS dis_58,
                    SUM(dis_59) AS dis_59,
                    SUM(dis_60) AS dis_60,
                    SUM(dis_61) AS dis_61,
                    SUM(dis_62) AS dis_62,
                    SUM(dis_63) AS dis_63,
                    SUM(dis_64) AS dis_64,
                    SUM(dis_65) AS dis_65,
                    SUM(dis_66) AS dis_66,
                    SUM(dis_67) AS dis_67,
                    SUM(dis_68) AS dis_68,
                    SUM(dis_69) AS dis_69,
                    SUM(dis_70) AS dis_70,
                    SUM(dis_71) AS dis_71,
                    SUM(dis_72) AS dis_72,
                    SUM(dis_73) AS dis_73,
                    SUM(dis_74) AS dis_74,
                    SUM(dis_75) AS dis_75,
                    SUM(dis_76) AS dis_76,
                    SUM(dis_77) AS dis_77,
                    SUM(dis_78) AS dis_78,
                    SUM(dis_79) AS dis_79,
                    SUM(dis_80) AS dis_80,
                    SUM(dis_81) AS dis_81,
                    SUM(dis_82) AS dis_82,
                    SUM(dis_83) AS dis_83,
                    SUM(dis_84) AS dis_84,
                    SUM(dis_85) AS dis_85,
                    SUM(dis_86) AS dis_86,
                    SUM(dis_87) AS dis_87,
                    SUM(dis_88) AS dis_88,
                    SUM(dis_89) AS dis_89,
                    SUM(dis_90) AS dis_90,
                    SUM(dis_91) AS dis_91,
                    SUM(dis_92) AS dis_92,
                    SUM(dis_93) AS dis_93,
                    SUM(dis_94) AS dis_94,
                    SUM(dis_95) AS dis_95,
                    SUM(dis_96) AS dis_96,
                    SUM(dis_97) AS dis_97,
                    SUM(dis_98) AS dis_98,
                    SUM(dis_99) AS dis_99,
                    SUM(dis_100) AS dis_100,
                    SUM(dis_101) AS dis_101,
                    SUM(dis_102) AS dis_102,
                    SUM(dis_103) AS dis_103,
                    SUM(dis_104) AS dis_104,
                    SUM(dis_105) AS dis_105,
                    SUM(dis_106) AS dis_106,
                    SUM(dis_107) AS dis_107,
                    SUM(dis_108) AS dis_108,
                    SUM(dis_109) AS dis_109,
                    SUM(dis_110) AS dis_110,
                    SUM(dis_111) AS dis_111,
                    SUM(dis_112) AS dis_112,
                    SUM(dis_113) AS dis_113,
                    SUM(dis_114) AS dis_114,
                    SUM(dis_115) AS dis_115,
                    SUM(dis_116) AS dis_116,
                    SUM(dis_117) AS dis_117,
                    SUM(dis_118) AS dis_118,
                    SUM(dis_119) AS dis_119,
                    SUM(dis_120) AS dis_120,
                    SUM(dis_121) AS dis_121,
                    SUM(dis_122) AS dis_122,
                    SUM(dis_123) AS dis_123,
                    SUM(dis_124) AS dis_124,
                    SUM(dis_125) AS dis_125,
                    SUM(dis_126) AS dis_126,
                    SUM(dis_127) AS dis_127,
                    SUM(dis_128) AS dis_128,
                    SUM(dis_129) AS dis_129,
                    SUM(dis_130) AS dis_130, 
                    SUM(dis_131) AS dis_131,
                    SUM(dis_132) AS dis_132,
                    SUM(dis_133) AS dis_133,
                    SUM(dis_134) AS dis_134,
                    SUM(dis_135) AS dis_135,
                    SUM(dis_136) AS dis_136,
                    SUM(dis_137) AS dis_137,
                    SUM(dis_138) AS dis_138,
                    SUM(dis_139) AS dis_139,
                    SUM(dis_140) AS dis_140, 
                    SUM(dis_141) AS dis_141,
                    SUM(dis_142) AS dis_142,
                    SUM(dis_143) AS dis_143,
                    SUM(dis_144) AS dis_144,
                    SUM(dis_145) AS dis_145,
                    SUM(dis_146) AS dis_146,
                    SUM(dis_147) AS dis_147,
                    SUM(dis_148) AS dis_148,
                    SUM(dis_149) AS dis_149,
                    SUM(dis_150) AS dis_150,
                    SUM(dis_151) AS dis_151,
                    SUM(dis_152) AS dis_152,
                    SUM(dis_153) AS dis_153,
                    SUM(dis_154) AS dis_154,
                    SUM(dis_155) AS dis_155,
                    SUM(dis_156) AS dis_156,
                    SUM(dis_157) AS dis_157,
                    SUM(dis_158) AS dis_158,
                    SUM(dis_159) AS dis_159,
                    SUM(dis_160) AS dis_160
                FROM (
                    SELECT
                        SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_1,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_2,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_3,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_4,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_5,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_6,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_7,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_8,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_9,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_10,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_11,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_12,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_13,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_14,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_15,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_16,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_17,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_18,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_19,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_20,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_21,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_22,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_23,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_24,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_25,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_26,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_27,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_28,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_29,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_30,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_31,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_32,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_33,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_34,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_35,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_36,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_37,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_38,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_39,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_40,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_41,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_42,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_43,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_44,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_45,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_46,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_47,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_48,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_49,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_50,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_51,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_52,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_53,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_54,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_55,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_56,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_57,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_58,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_59,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_60,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_61,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_62,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_63,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_64,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_65,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_66,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_67,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_68,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_69,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_70,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_71,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_72,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_73,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_74,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_75,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_76,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_77,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_78,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_79,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_80,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_81,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_82,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_83,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_84,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_85,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_86,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_87,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_88,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_89,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_90,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_91,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_92,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_93,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_94,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_95,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_96,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_97,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_98,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_99,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_100,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_101,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_102,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_103,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_104,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_105,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_106,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_107,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_108,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_109,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_110,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_111,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_112,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_113,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_114,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_115,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_116,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_117,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_118,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_119,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_120,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_121,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_122,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_123,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_124,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_125,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_126,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_127,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_128,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_129,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_130,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_131,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_132,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_133,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_134,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_135,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_136,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_137,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_138,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_139,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_140,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_141,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_142,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_143,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_144,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_145,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_146,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_147,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_148,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_149,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_150,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_151,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_152,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_153,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_154,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_155,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_156,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_157,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_158,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_159,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_160
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                ) subquery
                GROUP BY renaes, ubigeo_filtrado
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_dist = cursor.fetchall()
    return resultado_dist

def rpt_operacional_sensorial_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,   
                        SUM(dis_161) AS dis_161,
                        SUM(dis_162) AS dis_162,
                        SUM(dis_163) AS dis_163,
                        SUM(dis_164) AS dis_164,
                        SUM(dis_165) AS dis_165,
                        SUM(dis_166) AS dis_166,
                        SUM(dis_167) AS dis_167,
                        SUM(dis_168) AS dis_168,
                        SUM(dis_169) AS dis_169,
                        SUM(dis_170) AS dis_170,
                        SUM(dis_171) AS dis_171,
                        SUM(dis_172) AS dis_172,
                        SUM(dis_173) AS dis_173,
                        SUM(dis_174) AS dis_174,
                        SUM(dis_175) AS dis_175,
                        SUM(dis_176) AS dis_176,
                        SUM(dis_177) AS dis_177,
                        SUM(dis_178) AS dis_178,
                        SUM(dis_179) AS dis_179,
                        SUM(dis_180) AS dis_180,
                        SUM(dis_181) AS dis_181,
                        SUM(dis_182) AS dis_182,
                        SUM(dis_183) AS dis_183,
                        SUM(dis_184) AS dis_184,
                        SUM(dis_185) AS dis_185,
                        SUM(dis_186) AS dis_186,
                        SUM(dis_187) AS dis_187,
                        SUM(dis_188) AS dis_188,
                        SUM(dis_189) AS dis_189,
                        SUM(dis_190) AS dis_190,
                        SUM(dis_191) AS dis_191,
                        SUM(dis_192) AS dis_192,
                        SUM(dis_193) AS dis_193,
                        SUM(dis_194) AS dis_194,
                        SUM(dis_195) AS dis_195
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_161,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_162,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_163,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_164,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_165,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_166,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_167,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_168,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_169,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_170,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_171,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_172,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_173,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_174,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_175,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_176,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_177,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_178,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_179,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_180,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_181,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_182,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_183,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_184,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_185,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_186,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_187,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_188,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_189,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_190,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_191,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_192,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_193,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_194,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_195
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_dist_sensorial = cursor.fetchall()
    return resultado_dist_sensorial

def rpt_operacional_certificado_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,   
                        SUM(dis_216) AS dis_216,
                        SUM(dis_217) AS dis_217,
                        SUM(dis_218) AS dis_218,
                        SUM(dis_219) AS dis_219,
                        SUM(dis_220) AS dis_220,
                        SUM(dis_221) AS dis_221,
                        SUM(dis_222) AS dis_222,
                        SUM(dis_223) AS dis_223,
                        SUM(dis_224) AS dis_224,
                        SUM(dis_225) AS dis_225,
                        SUM(dis_226) AS dis_226,
                        SUM(dis_227) AS dis_227,
                        SUM(dis_228) AS dis_228,
                        SUM(dis_229) AS dis_229,
                        SUM(dis_230) AS dis_230,
                        SUM(dis_231) AS dis_231,
                        SUM(dis_232) AS dis_232,
                        SUM(dis_233) AS dis_233,
                        SUM(dis_234) AS dis_234,
                        SUM(dis_235) AS dis_235
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_216,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_217,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_218,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_219,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_220,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_221,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_222,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_223,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_224,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_225,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_226,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_227,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_228,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_229,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_230,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_231,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_232,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_233,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_234,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_235
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        resultado_dist_certificado = cursor.fetchall()
    return resultado_dist_certificado

def rpt_operacional_rbc_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_242) AS dis_242,
                        SUM(dis_243) AS dis_243,
                        SUM(dis_244) AS dis_244,
                        SUM(dis_245) AS dis_245,
                        SUM(dis_246) AS dis_246,
                        SUM(dis_247) AS dis_247,
                        SUM(dis_248) AS dis_248,
                        SUM(dis_249) AS dis_249,
                        SUM(dis_250) AS dis_250,
                        SUM(dis_251) AS dis_251,
                        SUM(dis_252) AS dis_252,
                        SUM(dis_253) AS dis_253,
                        SUM(dis_254) AS dis_254,
                        SUM(dis_255) AS dis_255,
                        SUM(dis_256) AS dis_256,
                        SUM(dis_257) AS dis_257,
                        SUM(dis_258) AS dis_258,
                        SUM(dis_259) AS dis_259,
                        SUM(dis_260) AS dis_260, 
                        SUM(dis_261) AS dis_261, 
                        SUM(dis_262) AS dis_262, 
                        SUM(dis_263) AS dis_263, 
                        SUM(dis_264) AS dis_264, 
                        SUM(dis_265) AS dis_265, 
                        SUM(dis_266) AS dis_266
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_242,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_243,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_244,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_245,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_246,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_247,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_248,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_249,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_250,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_251,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_252,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_253,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_254,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_255,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_256,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_257,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_258,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_259,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_260,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_261,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_262,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_263,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_264,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_265,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_266
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s 
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                ) subquery
                GROUP BY renaes, ubigeo_filtrado
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        resultado_dist_rbc = cursor.fetchall()
    return resultado_dist_rbc

def rpt_operacional_mental_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_242) AS dis_242,
                        SUM(dis_243) AS dis_243,
                        SUM(dis_244) AS dis_244,
                        SUM(dis_245) AS dis_245,
                        SUM(dis_246) AS dis_246,
                        SUM(dis_247) AS dis_247,
                        SUM(dis_248) AS dis_248,
                        SUM(dis_249) AS dis_249,
                        SUM(dis_250) AS dis_250,
                        SUM(dis_251) AS dis_251,
                        SUM(dis_252) AS dis_252,
                        SUM(dis_253) AS dis_253,
                        SUM(dis_254) AS dis_254,
                        SUM(dis_255) AS dis_255,
                        SUM(dis_256) AS dis_256,
                        SUM(dis_257) AS dis_257,
                        SUM(dis_258) AS dis_258,
                        SUM(dis_259) AS dis_259,
                        SUM(dis_260) AS dis_260, 
                        SUM(dis_261) AS dis_261, 
                        SUM(dis_262) AS dis_262, 
                        SUM(dis_263) AS dis_263, 
                        SUM(dis_264) AS dis_264, 
                        SUM(dis_265) AS dis_265, 
                        SUM(dis_266) AS dis_266
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_242,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_243,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_244,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_245,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_246,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_247,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_248,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_249,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_250,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_251,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_252,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_253,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_254,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_255,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_256,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_257,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_258,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_259,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_260,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_261,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_262,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_263,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_264,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_265,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_266
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s 
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                ) subquery
                GROUP BY renaes, ubigeo_filtrado
        """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        # Consultar los resultados finales desde la tabla temporal
        resultado_dist_mental = cursor.fetchall()
    return resultado_dist_mental

def rpt_operacional_capacitacion_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_273) AS dis_273,
                        SUM(dis_274) AS dis_274
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            COUNT(Categoria) AS dis_273,
                            SUM(gedad) AS dis_274
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_dist_capacitacion = cursor.fetchall()
    return resultado_dist_capacitacion

def rpt_operacional_agente_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_236) AS dis_236,
                        SUM(dis_237) AS dis_237,
                        SUM(dis_238) AS dis_238,
                        SUM(dis_239) AS dis_239,
                        SUM(dis_240) AS dis_240,
                        SUM(dis_241) AS dis_241
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Categoria = 1 THEN 1 ELSE 0 END) 		 AS dis_236,
                            SUM(CASE WHEN Categoria = 1 THEN gedad ELSE 0 END)   AS dis_237,
                            SUM(CASE WHEN Categoria = 2 THEN 1 ELSE 0 END)     AS dis_238,
                            SUM(CASE WHEN Categoria = 2 THEN gedad ELSE 0 END)   AS dis_239,
                            SUM(CASE WHEN Categoria = 3 THEN 1 ELSE 0 END)     AS dis_240,
                            SUM(CASE WHEN Categoria = 3 THEN gedad ELSE 0 END)   AS dis_241
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_dist_agente = cursor.fetchall()
    return resultado_dist_agente

def rpt_operacional_comite_dist(ubigeo, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                    SELECT
                        SUBSTRING(CAST(ubigeo_filtrado AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                        renaes,     
                        SUM(dis_267) AS dis_267,
                        SUM(dis_268) AS dis_268,
                        SUM(dis_269) AS dis_269,
                        SUM(dis_270) AS dis_270,
                        SUM(dis_271) AS dis_271,
                        SUM(dis_272) AS dis_272
                    FROM (
                        SELECT
                            SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 4) AS ubigeo_filtrado,
                            renaes,
                            SUM(CASE WHEN Actividad = 1 THEN 1 ELSE 0 END) 		AS dis_267,
                            SUM(CASE WHEN Actividad = 1 THEN Partic ELSE 0 END) AS dis_268,
                            SUM(CASE WHEN Actividad = 2 THEN 1 ELSE 0 END)      AS dis_269,
                            SUM(CASE WHEN Actividad = 2 THEN Partic ELSE 0 END) AS dis_270,
                            SUM(CASE WHEN Actividad = 3 THEN 1 ELSE 0 END)      AS dis_271,
                            SUM(CASE WHEN Actividad = 3 THEN Partic ELSE 0 END) AS dis_272
                        FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL
                        LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                        WHERE SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6) = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                        GROUP BY SUBSTRING(CAST(MAESTRO_HIS_ESTABLECIMIENTO.Ubigueo_Establecimiento AS VARCHAR(10)), 1, 6), renaes
                    ) subquery
                    GROUP BY renaes, ubigeo_filtrado
                    """, [str(ubigeo)[:6], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_dist_comite = cursor.fetchall()
    return resultado_dist_comite

class RptOperacinalDist(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        distritos = request.GET.get('distritos')

        # Creación de la consulta
        resultado_dist = rpt_operacional_fisico_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_sensorial = rpt_operacional_sensorial_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_certificado = rpt_operacional_certificado_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_rbc = rpt_operacional_rbc_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_mental = rpt_operacional_mental_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_capacitacion = rpt_operacional_capacitacion_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_agente = rpt_operacional_agente_dist(distritos, fecha_inicio, fecha_fin)
        resultado_dist_comite = rpt_operacional_comite_dist(distritos, fecha_inicio, fecha_fin)
        

        distrito_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Ubigueo_Establecimiento__startswith=distritos
        ).values_list('Distrito', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
        
        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambia el alto de la columna
        sheet.row_dimensions[1].height = 14
        sheet.row_dimensions[2].height = 14
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[15].height = 25

        # cambia el ancho de la columna
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 28
        sheet.column_dimensions['C'].width = 28
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 9
        sheet.column_dimensions['F'].width = 9
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 9
        sheet.column_dimensions['L'].width = 9
        # linea de division
        sheet.freeze_panes = 'AL8'
        
        # Configuración del fondo y el borde
        fill = PatternFill(patternType='solid', fgColor='00B0F0')
        border = Border(left=Side(style='thin', color='00B0F0'),
                right=Side(style='thin', color='00B0F0'),
                top=Side(style='thin', color='00B0F0'),
                bottom=Side(style='thin', color='00B0F0'))

        borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                right=Side(style='thin', color='A9A9A9'), # Plomo
                top=Side(style='thin', color='A9A9A9'), # Plomo
                bottom=Side(style='thin', color='A9A9A9')) # Plomo

        # crea titulo del reporte
        sheet['B1'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B1'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
        
        sheet['B2'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B2'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
        
        sheet['B4'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B4'].font = Font(name = 'Arial', size= 12, bold = True)
        sheet['B4'] = 'REPORTE DE ACTIVIDADES DEL COMPONENTE DE DISCAPACIDAD'
        
        sheet['B6'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B6'] ='DIRESA / GERESA / DISA'
        
        sheet['C6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C6'].font = Font(name = 'Arial', size= 7)
        sheet['C6'] ='JUNIN'

        sheet['B7'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B7'] ='PROV/ DIST/ RED/ MR/ ESTABLEC'
        
        sheet['C7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C7'].font = Font(name = 'Arial', size= 7)
        sheet['C7'] = distrito_codigo[0]
        
        sheet['E6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['E6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['E6'] ='PERIODO'
        
        sheet['F6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F6'] ='MES INICIO'
        
        sheet['F7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F7'] ='MES FIN'
        
        sheet['G6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G6'].font = Font(name = 'Arial', size= 8)
        sheet['G6'] = fecha_inicio_codigo[0]
        
        sheet['G7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G7'].font = Font(name = 'Arial', size= 8)
        sheet['G7'] = fecha_fin_codigo[0]
        
        sheet['B9'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B9'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B9'] ='PERSONAS CON DISCAPACIDAD RECIBEN ATENCION DE REHABILITACION EN ESTABLECIMIENTOS DE SALUD (3000688)'
        
        sheet['B10'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B10'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B10'] ='Capacitación en medicina de rehabilitación integral (5004449)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=12, max_row=12, min_col=3, max_col=5):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['C12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C12'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['C12'] ='Capacitación  (C0009)' 
        
        sheet['D11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D11'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D11'].fill = fill
        sheet['D11'].border = border
        sheet['D11'] = 'N°'
                
        sheet['E11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E11'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['E11'].fill = fill
        sheet['E11'].border = border
        sheet['E11'] = 'Capacitados'
        #######################################################
        ########## DISCAPACIDAD FISICA ########################
        #######################################################
        sheet['B14'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B14'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B14'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Física (5005150)' 
                
        sheet['B15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B15'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B15'].fill = fill
        sheet['B15'].border = border
        sheet['B15'] = 'Atenciones'
        
        sheet['D15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D15'].fill = fill
        sheet['D15'].border = border
        sheet['D15'] = 'Total'
        
        sheet['E15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E15'].fill = fill
        sheet['E15'].border = border
        sheet['E15'] = 'Niños         (1d - 11a)'
        
        sheet['F15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F15'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F15'].fill = fill
        sheet['F15'].border = border
        sheet['F15'] = 'Adolescentes (12a - 17a)'
        
        sheet['G15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G15'].fill = fill
        sheet['G15'].border = border
        sheet['G15'] = 'Jóvenes (18a - 29a)'
        
        sheet['H15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H15'].fill = fill
        sheet['H15'].border = border
        sheet['H15'] = 'Adultos (30a - 59a)'
        
        sheet['I15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I15'].fill = fill
        sheet['I15'].border = border
        sheet['I15'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=16, max_row=47, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B16'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B16'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B16'] ='LESIONES MEDULARES' 
                
        sheet['B17'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B17'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B17'] ='ENFERMEDAD DE PARKINSON Y SIMILARES' 
        
        sheet['B18'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B18'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B18'] ='REHABILITACIÓN EN PACIENTES AMPUTADOS' 
                
        sheet['B20'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B20'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B20'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA NEUROLÓGICA' 
        
        sheet['B23'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B23'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B23'] ='TRASTORNOS DEL DESARROLLO DE LA FUNCIÓN MOTRIZ' 
        
        sheet['B24'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B24'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B24'] ='ATENCIÓN DE REHABILITACIÓN DE ENFERMEDAD ARTICULAR DEGENERATIVA' 
        
        sheet['B25'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B25'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B25'] ='ENCEFALOPATÍA INFANTIL' 
                
        sheet['B26'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B26'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B26'] ='SÍNDROME DOWN' 
        
        sheet['B27'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B27'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B27'] ='REHABILITACIÓN EN PATOLOGÍA DE LA COLUMNA VERTEBRAL Y OTROS TRASTORNOS POSTURALES' 
        
        sheet['B34'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B34'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B34'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD CARDIOVASCULAR' 
        
        sheet['B35'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B35'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B35'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD RESPIRATORIA' 
        
        sheet['B36'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B36'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B36'] ='ATENCIÓN DE REHABILITACIÓN EN ALTERACIONES DEL PISO PÉLVICO' 
        
        sheet['B37'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B37'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B37'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA TRAUMATOLÓGICA Y REUMATOLÓGICA' 
        
        sheet['B44'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B44'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B44'] ='ATENCIÓN DE REHABILITACIÓN ONCOLÓGICA' 
        
        sheet['B46'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B46'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B46'] ='ATENCIÓN DE REHABILITACIÓN EN DOLOR' 
        
        sheet['B47'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B47'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B47'] ='ATENCIÓN DE REHABILITACIÓN EN PACIENTES QUEMADOS' 
        ####     
        sheet['C16'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C16'].font = Font(name = 'Arial', size= 7)
        sheet['C16'] ='Lesiones medulares' 
    
        sheet['C17'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C17'].font = Font(name = 'Arial', size= 7)
        sheet['C17'] ='Enfermedad de Parkinson y similares' 
        
        sheet['C18'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C18'].font = Font(name = 'Arial', size= 7)
        sheet['C18'] ='Amputados de miembros superiores' 
        
        sheet['C19'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C19'].font = Font(name = 'Arial', size= 7)
        sheet['C19'] ='Amputados de miembros inferiores' 
        
        sheet['C20'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C20'].font = Font(name = 'Arial', size= 7)
        sheet['C20'] ='Enfermedades cerebrovasculares'
        
        sheet['C21'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C21'].font = Font(name = 'Arial', size= 7)
        sheet['C21'] ='Enfermedades musculares y de la unión mioneural'
        
        sheet['C22'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C22'].font = Font(name = 'Arial', size= 7)
        sheet['C22'] ='Lesiones de nervios periféricos'
        
        sheet['C23'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C23'].font = Font(name = 'Arial', size= 7)
        sheet['C23'] ='Trastornos del desarrollo de la funcion motriz'
        
        sheet['C24'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C24'].font = Font(name = 'Arial', size= 7)
        sheet['C24'] ='Enfermedad articular degenerativa'
        
        sheet['C25'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C25'].font = Font(name = 'Arial', size= 7)
        sheet['C25'] ='Encefalopatía infantil y otras lesiones'
        
        sheet['C26'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C26'].font = Font(name = 'Arial', size= 7)
        sheet['C26'] ='Sindrome de Down'
        
        sheet['C27'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C27'].font = Font(name = 'Arial', size= 7)
        sheet['C27'] ='Cifosis y lordosis'
        
        sheet['C28'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C28'].font = Font(name = 'Arial', size= 7)
        sheet['C28'] ='Espondilo artropatías'
        
        sheet['C29'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C29'].font = Font(name = 'Arial', size= 7)
        sheet['C29'] ='Otros trastornos de los discos intervertebrales'
        
        sheet['C30'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C30'].font = Font(name = 'Arial', size= 7)
        sheet['C30'] ='Cervicalgia, dorsalgia, lumbago'
        
        sheet['C31'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C31'].font = Font(name = 'Arial', size= 7)
        sheet['C31'] ='Otras dorsopatías deformantes'
        
        sheet['C32'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C32'].font = Font(name = 'Arial', size= 7)
        sheet['C32'] ='Otros trastornos articulares'
        
        sheet['C33'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C33'].font = Font(name = 'Arial', size= 7)
        sheet['C33'] ='Defectos en la longitud de extremidades'
        
        sheet['C34'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C34'].font = Font(name = 'Arial', size= 7)
        sheet['C34'] ='Enfermedad cardiovascular'
        
        sheet['C35'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C35'].font = Font(name = 'Arial', size= 7)
        sheet['C35'] ='Enfermedad respiratoria'
        
        sheet['C36'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C36'].font = Font(name = 'Arial', size= 7)
        sheet['C36'] ='Vejiga neurogénica y dolor'
        
        sheet['C37'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C37'].font = Font(name = 'Arial', size= 7)
        sheet['C37'] ='Incontinencia'
        
        sheet['C38'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C38'].font = Font(name = 'Arial', size= 7)
        sheet['C38'] ='Prolapso'
        
        sheet['C39'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C39'].font = Font(name = 'Arial', size= 7)
        sheet['C39'] ='Traumatismos'
        
        sheet['C40'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C40'].font = Font(name = 'Arial', size= 7)
        sheet['C40'] ='Enfermedades del tejido conectivo'
        
        sheet['C41'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C41'].font = Font(name = 'Arial', size= 7)
        sheet['C41'] ='Patología articular excluida columna'
        
        sheet['C42'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C42'].font = Font(name = 'Arial', size= 7)
        sheet['C42'] ='Lesiones infecciosas'
        
        sheet['C43'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C43'].font = Font(name = 'Arial', size= 7)
        sheet['C43'] ='Lesión biomecánica'
        
        sheet['C44'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C44'].font = Font(name = 'Arial', size= 7)
        sheet['C44'] ='Linfedema'
        
        sheet['C45'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C45'].font = Font(name = 'Arial', size= 7)
        sheet['C45'] ='Sarcopenia'
        
        sheet['C46'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C46'].font = Font(name = 'Arial', size= 7)
        sheet['C46'] ='Dolor'
        
        sheet['C47'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C47'].font = Font(name = 'Arial', size= 7)
        sheet['C47'] ='Quemaduras, corrosiones y congelaciones'
        ##########################################################    
        ########## DISCAPACIDAD SENSORIAL ########################
        ##########################################################
        sheet['B50'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B50'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B50'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Sensorial (5005151)' 
                
        sheet['B51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B51'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B51'].fill = fill
        sheet['B51'].border = border
        sheet['B51'] = 'Atenciones'
        
        sheet['D51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D51'].fill = fill
        sheet['D51'].border = border
        sheet['D51'] = 'Total'
        
        sheet['E51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E51'].fill = fill
        sheet['E51'].border = border
        sheet['E51'] = 'Niños         (1d - 11a)'
        
        sheet['F51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F51'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F51'].fill = fill
        sheet['F51'].border = border
        sheet['F51'] = 'Adolescentes (12a - 17a)'
        
        sheet['G51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G51'].fill = fill
        sheet['G51'].border = border
        sheet['G51'] = 'Jóvenes (18a - 29a)'
        
        sheet['H51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H51'].fill = fill
        sheet['H51'].border = border
        sheet['H51'] = 'Adultos (30a - 59a)'
        
        sheet['I51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I51'].fill = fill
        sheet['I51'].border = border
        sheet['I51'] = 'A Mayores (60a +)'
        #borde plomo
        for row in sheet.iter_rows(min_row=52, max_row=58, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B52'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B52'] ='HIPOACUSIA Y/O SORDERA' 
        
        sheet['B53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B53'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B53'] ='BAJA VISION Y/O CEGUERA' 
        
        sheet['B54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B54'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B54'] ='SORDOMUDEZ' 
        
        sheet['B55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B55'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B55'] ='ENFERMEDAD CEREBRO VASCULAR' 
        
        sheet['B56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B56'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B56'] ='TRASTORNOS ESPECIFICOS DEL DESARROLLO DEL HABLA Y LENGUAJE' 
        
        sheet['B57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B57'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B57'] ='DISARTRIA Y DISFAGIA' 
        
        sheet['B59'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B59'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B59'] ='SUB TOTAL' 
        ########               
        sheet['C52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C52'].font = Font(name = 'Arial', size= 7)
        sheet['C52'] ='Hipoacusia y sordera' 
        
        sheet['C53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C53'].font = Font(name = 'Arial', size= 7)
        sheet['C53'] ='Baja visión y ceguera' 
        
        sheet['C54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C54'].font = Font(name = 'Arial', size= 7)
        sheet['C54'] ='Sordomudez' 
        
        sheet['C55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C55'].font = Font(name = 'Arial', size= 7)
        sheet['C55'] ='Enfermedad Cerebro vascular' 
        
        sheet['C56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C56'].font = Font(name = 'Arial', size= 7)
        sheet['C56'] ='Trastornos específicos del desarrollo del habla y lenguaje' 
        
        sheet['C57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C57'].font = Font(name = 'Arial', size= 7)
        sheet['C57'] ='Disartria' 
        
        sheet['C58'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C58'].font = Font(name = 'Arial', size= 7)
        sheet['C58'] ='Disfagia' 
        ########################################################
        ########## DISCAPACIDAD MENTAL #########################
        ########################################################
        sheet['B61'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B61'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B61'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Mental (5005152)' 
                
        sheet['B62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B62'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B62'].fill = fill
        sheet['B62'].border = border
        sheet['B62'] = 'Atenciones'
        
        sheet['D62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D62'].fill = fill
        sheet['D62'].border = border
        sheet['D62'] = 'Total'
        
        sheet['E62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E62'].fill = fill
        sheet['E62'].border = border
        sheet['E62'] = 'Niños         (1d - 11a)'
        
        sheet['F62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F62'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F62'].fill = fill
        sheet['F62'].border = border
        sheet['F62'] = 'Adolescentes (12a - 17a)'
        
        sheet['G62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G62'].fill = fill
        sheet['G62'].border = border
        sheet['G62'] = 'Jóvenes (18a - 29a)'
        
        sheet['H62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H62'].fill = fill
        sheet['H62'].border = border
        sheet['H62'] = 'Adultos (30a - 59a)'
        
        sheet['I62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I62'].fill = fill
        sheet['I62'].border = border
        sheet['I62'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=63, max_row=66, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B63'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B63'] ='TRASTORNOS DE APRENDIZAJE' 
        
        sheet['B64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B64'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B64'] ='RETRASO MENTAL LEVE, MODERADO, SEVERO' 
        
        sheet['B65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B65'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B65'] ='TRASTORNOS DEL ESPECTRO AUTISTA' 
        
        sheet['B66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B66'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B66'] ='OTROS TRASTORNOS DE SALUD MENTAL' 
        
        sheet['B67'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B67'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B67'] ='SUB TOTAL' 
        
        ##########
        sheet['C63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C63'].font = Font(name = 'Arial', size= 7)
        sheet['C63'] ='Trastornos del aprendizaje' 
        
        sheet['C64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C64'].font = Font(name = 'Arial', size= 7)
        sheet['C64'] ='Retardo Mental: Leve, moderado, severo' 
        
        sheet['C65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C65'].font = Font(name = 'Arial', size= 7)
        sheet['C65'] ='Trastornos del espectro autista' 
        
        sheet['C66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C66'].font = Font(name = 'Arial', size= 7)
        sheet['C66'] ='Otras alteraciones de salud mental' 
        ##################################################
        ########## CERTIFICACION #########################
        ##################################################
        sheet['B69'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B69'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B69'] ='PERSONAS CON DISCAPACIDAD CERTIFICADAS EN ESTABLECIMIENTOS DE SALUD (3000689)' 
                
        sheet['B70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B70'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B70'].fill = fill
        sheet['B70'].border = border
        sheet['B70'] = 'Atenciones'
        
        sheet['D70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D70'].fill = fill
        sheet['D70'].border = border
        sheet['D70'] = 'Total'
        
        sheet['E70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E70'].fill = fill
        sheet['E70'].border = border
        sheet['E70'] = 'Niños         (1d - 11a)'
        
        sheet['F70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F70'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F70'].fill = fill
        sheet['F70'].border = border
        sheet['F70'] = 'Adolescentes (12a - 17a)'
        
        sheet['G70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G70'].fill = fill
        sheet['G70'].border = border
        sheet['G70'] = 'Jóvenes (18a - 29a)'
        
        sheet['H70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H70'].fill = fill
        sheet['H70'].border = border
        sheet['H70'] = 'Adultos (30a - 59a)'
        
        sheet['I70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I70'].fill = fill
        sheet['I70'].border = border
        sheet['I70'] = 'A. Mayores (60a +)'
        #borde plomo
        for row in sheet.iter_rows(min_row=71, max_row=74, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B71'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B71'] ='Certificación de Discapacidad (0515204)' 
        
        sheet['B74'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B74'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B74'] ='Certificación de Incapacidad (0515205)' 
        
        sheet['B75'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B75'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B75'] ='SUB TOTAL' 
        
        sheet['C71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C71'].font = Font(name = 'Arial', size= 7)
        sheet['C71'] ='Evaluación' 
        
        sheet['C72'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C72'].font = Font(name = 'Arial', size= 7)
        sheet['C72'] ='Calificación' 
        
        sheet['C73'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C73'].font = Font(name = 'Arial', size= 7)
        sheet['C73'] ='Certificación' 
        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B77'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B77'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B77'] ='PERSONAS CON DISCAPACIDAD RECIBEN SERVICIOS DE REHABILITACIÓN BASADA EN LA COMUNIDAD (3000690)' 
        
        sheet['B78'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B78'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B78'] ='CAPACITACIÓN A AGENTES COMUNITARIOS EN REHABILITACIÓN BASADA EN LA COMUNIDAD (5005155)' 
        
        sheet['B82'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B82'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B82'] ='Capacitación a Agentes Comunitarios  (APP138)' 
        
        sheet['D80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D80'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D80'].fill = fill
        sheet['D80'].border = border
        sheet['D80'] = 'Taller'
        
        sheet['F80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F80'].fill = fill
        sheet['F80'].border = border
        sheet['F80'] = 'Sesion Educativa'
        
        sheet['H80'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H80'].fill = fill
        sheet['H80'].border = border
        sheet['H80'] = 'Sesion Demostrativa'
        
        sheet['D81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D81'].fill = fill
        sheet['D81'].border = border
        sheet['D81'] = 'N°'
        
        sheet['E81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E81'].fill = fill
        sheet['E81'].border = border
        sheet['E81'] = 'Capacitados'
        
        sheet['F81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F81'].fill = fill
        sheet['F81'].border = border
        sheet['F81'] = 'N°'
        
        sheet['G81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G81'].fill = fill
        sheet['G81'].border = border
        sheet['G81'] = 'Capacitados'
        
        sheet['H81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H81'].fill = fill
        sheet['H81'].border = border
        sheet['H81'] = 'N° '
        
        sheet['I81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I81'].fill = fill
        sheet['I81'].border = border
        sheet['I81'] = 'Capacitados'
        #borde plomo
        for row in sheet.iter_rows(min_row=82, max_row=82, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        ################################################
        ########## VISITAS RBC #########################
        ################################################
        sheet['B84'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B84'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B84'] ='Vistas a alas familias Rehabilitacion Basada en la Comunidad' 
                
        sheet['B85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B85'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B85'].fill = fill
        sheet['B85'].border = border
        sheet['B85'] = 'Visitas'
        
        sheet['D85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D85'].fill = fill
        sheet['D85'].border = border
        sheet['D85'] = 'Total'
        
        sheet['E85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E85'].fill = fill
        sheet['E85'].border = border
        sheet['E85'] = 'Niños         (1d - 11a)'
        
        sheet['F85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F85'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F85'].fill = fill
        sheet['F85'].border = border
        sheet['F85'] = 'Adolescentes (12a - 17a)'
        
        sheet['G85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G85'].fill = fill
        sheet['G85'].border = border
        sheet['G85'] = 'Jóvenes (18a - 29a)'
        
        sheet['H85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H85'].fill = fill
        sheet['H85'].border = border
        sheet['H85'] = 'Adultos (30a - 59a)'
        
        sheet['I85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I85'].fill = fill
        sheet['I85'].border = border
        sheet['I85'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=86, max_row=90, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = borde_plomo
        
        sheet['B86'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B86'].font = Font(name = 'Arial', size= 8)
        sheet['B86'] ='1º Visita' 
        
        sheet['B87'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B87'].font = Font(name = 'Arial', size= 8)
        sheet['B87'] ='2º Visita' 
        
        sheet['B88'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B88'].font = Font(name = 'Arial', size= 8)
        sheet['B88'] ='3º Visita' 
        
        sheet['B89'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B89'].font = Font(name = 'Arial', size= 8)
        sheet['B89'] ='4º a Visita (trazador)' 
        
        sheet['B90'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B90'].font = Font(name = 'Arial', size= 8)
        sheet['B90'] ='5º a + Visitas' 
        
        sheet['B91'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B91'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B91'] ='SUB TOTAL' 
        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B93'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B93'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B93'] ='Capacitación a Actores Sociales para la aplicación de la estrategia de Rehabilitación Basada en la Comunidad' 
                
        sheet['B94'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B94'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B94'] ='Actividades con Gobiernos Locales:' 
        
        sheet['B97'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B97'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B97'] ='Actividad con Comité Multisectorial (APP96)' 
        
        sheet['D95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D95'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D95'].fill = fill
        sheet['D95'].border = border
        sheet['D95'] = 'Taller'
        
        sheet['F95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F95'].fill = fill
        sheet['F95'].border = border
        sheet['F95'] = 'Sesion Educativa'
        
        sheet['H95'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H95'].fill = fill
        sheet['H95'].border = border
        sheet['H95'] = 'Sesion Demostrativa'
        
        sheet['D96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D96'].fill = fill
        sheet['D96'].border = border
        sheet['D96'] = 'N°'
        
        sheet['E96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E96'].fill = fill
        sheet['E96'].border = border
        sheet['E96'] = 'Capacitados'
        
        sheet['F96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F96'].fill = fill
        sheet['F96'].border = border
        sheet['F96'] = 'N°'
        
        sheet['G96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G96'].fill = fill
        sheet['G96'].border = border
        sheet['G96'] = 'Capacitados'
        
        sheet['H96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H96'].fill = fill
        sheet['H96'].border = border
        sheet['H96'] = 'N° '
        
        sheet['I96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I96'].fill = fill
        sheet['I96'].border = border
        sheet['I96'] = 'Capacitados'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=97, max_row=97, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        #############################################################################
        #############################################################################                
        # cambina celdas
        sheet.merge_cells('C6:D6')
        sheet.merge_cells('C7:E7')
        
        sheet.merge_cells('B18:B19')
        sheet.merge_cells('B20:B22')
        sheet.merge_cells('B27:B33')
        sheet.merge_cells('B37:B43')
        sheet.merge_cells('B44:B45')
        
        # sensorial
        sheet.merge_cells('B57:B58')
        
        sheet.merge_cells('B15:C15')
        sheet.merge_cells('B51:C51')
        
        # mental
        sheet.merge_cells('B62:C62')
        
        #certificado
        sheet.merge_cells('B70:C70')
        
        sheet.merge_cells('B71:B73')
        
        #RBC
        sheet.merge_cells('B85:C85')
        
        #capa
        sheet.merge_cells('D80:E80')
        sheet.merge_cells('F80:G80')
        sheet.merge_cells('H80:I80')

        sheet.merge_cells('D95:E95')
        sheet.merge_cells('F95:G95')
        sheet.merge_cells('H95:I95')
        
        #capacitacion
        sheet.merge_cells('B82:C82')
        sheet.merge_cells('B97:C97')
        
        #visita
        sheet.merge_cells('B86:C86')
        sheet.merge_cells('B87:C87')
        sheet.merge_cells('B88:C88')
        sheet.merge_cells('B89:C89')
        sheet.merge_cells('B90:C90')
        
        # Definir ubicaciones específicas para cada columna y su suma total
        columnas_ubicaciones = {
            'PROVINCIA': 'D10',
            'DIS_1': 'E16', 
            'DIS_2': 'F16',
            'DIS_3': 'G16',
            'DIS_4': 'H16',
            'DIS_5': 'I16',
            'DIS_6': 'E17',
            'DIS_7': 'F17',
            'DIS_8': 'G17',
            'DIS_9': 'H17',
            'DIS_10': 'I17',
            'DIS_11': 'E18',
            'DIS_12': 'F18',
            'DIS_13': 'G18',
            'DIS_14': 'H18',
            'DIS_15': 'I18',
            'DIS_16': 'E19',
            'DIS_17': 'F19',
            'DIS_18': 'G19',
            'DIS_19': 'H19',
            'DIS_20': 'I19',
            'DIS_21': 'E20',
            'DIS_22': 'F20',
            'DIS_23': 'G20',
            'DIS_24': 'H20',
            'DIS_25': 'I20',
            'DIS_26': 'E21',
            'DIS_27': 'F21',
            'DIS_28': 'G21',
            'DIS_29': 'H21',
            'DIS_30': 'I21',
            'DIS_31': 'E22',
            'DIS_32': 'F22',
            'DIS_33': 'G22',
            'DIS_34': 'H22',
            'DIS_35': 'I22',
            'DIS_36': 'E23',
            'DIS_37': 'F23',
            'DIS_38': 'G23',
            'DIS_39': 'H23',
            'DIS_40': 'I23',
            'DIS_41': 'E24',
            'DIS_42': 'F24',
            'DIS_43': 'G24',
            'DIS_44': 'H24',
            'DIS_45': 'I24',
            'DIS_46': 'E25',
            'DIS_47': 'F25',
            'DIS_48': 'G25',
            'DIS_49': 'H25',
            'DIS_50': 'I25',
            'DIS_51': 'E26',
            'DIS_52': 'F26',
            'DIS_53': 'G26',
            'DIS_54': 'H26',
            'DIS_55': 'I26',
            'DIS_56': 'E27',
            'DIS_57': 'F27',
            'DIS_58': 'G27',
            'DIS_59': 'H27',
            'DIS_60': 'I27',
            'DIS_61': 'E28',
            'DIS_62': 'F28',
            'DIS_63': 'G28',
            'DIS_64': 'H28',
            'DIS_65': 'I28',
            'DIS_66': 'E29',
            'DIS_67': 'F29',
            'DIS_68': 'G29',
            'DIS_69': 'H29',
            'DIS_70': 'I29',
            'DIS_71': 'E30',
            'DIS_72': 'F30',
            'DIS_73': 'G30',
            'DIS_74': 'H30',
            'DIS_75': 'I30',
            'DIS_76': 'E31',
            'DIS_77': 'F31',
            'DIS_78': 'G31',
            'DIS_79': 'H31',
            'DIS_80': 'I31',
            'DIS_81': 'E32',
            'DIS_82': 'F32',
            'DIS_83': 'G32',
            'DIS_84': 'H32',
            'DIS_85': 'I32',
            'DIS_86': 'E33',
            'DIS_87': 'F33',
            'DIS_88': 'G33',
            'DIS_89': 'H33',
            'DIS_90': 'I33',
            'DIS_91': 'E34',
            'DIS_92': 'F34',
            'DIS_93': 'G34',
            'DIS_94': 'H34',
            'DIS_95': 'I34',
            'DIS_96': 'E35',
            'DIS_97': 'F35',
            'DIS_98': 'G35',
            'DIS_99': 'H35',
            'DIS_100': 'I35',
            'DIS_101': 'E36',
            'DIS_102': 'F36',
            'DIS_103': 'G36',
            'DIS_104': 'H36',
            'DIS_105': 'I36',
            'DIS_106': 'E37',
            'DIS_107': 'F37',
            'DIS_108': 'G37',
            'DIS_109': 'H37',
            'DIS_110': 'I37',
            'DIS_111': 'E38',
            'DIS_112': 'F38',
            'DIS_113': 'G38',
            'DIS_114': 'H38',
            'DIS_115': 'I38',
            'DIS_116': 'E39',
            'DIS_117': 'F39',
            'DIS_118': 'G39',
            'DIS_119': 'H39',
            'DIS_120': 'I39',
            'DIS_121': 'E40',
            'DIS_122': 'F40',
            'DIS_123': 'G40',
            'DIS_124': 'H40',
            'DIS_125': 'I40',
            'DIS_126': 'E41',
            'DIS_127': 'F41',
            'DIS_128': 'G41',
            'DIS_129': 'H41',
            'DIS_130': 'I41', 
            'DIS_131': 'E42',
            'DIS_132': 'F42',
            'DIS_133': 'G42',
            'DIS_134': 'H42',
            'DIS_135': 'I42', 
            'DIS_136': 'E43',
            'DIS_137': 'F43',
            'DIS_138': 'G43',
            'DIS_139': 'H43',
            'DIS_140': 'I43', 
            'DIS_141': 'E44',
            'DIS_142': 'F44',
            'DIS_143': 'G44',
            'DIS_144': 'H44',
            'DIS_145': 'I44', 
            'DIS_146': 'E45',
            'DIS_147': 'F45',
            'DIS_148': 'G45',
            'DIS_149': 'H45',
            'DIS_150': 'I45', 
            'DIS_151': 'E46',
            'DIS_152': 'F46',
            'DIS_153': 'G46',
            'DIS_154': 'H46',
            'DIS_155': 'I46', 
            'DIS_156': 'E47',
            'DIS_157': 'F47',
            'DIS_158': 'G47',
            'DIS_159': 'H47',
            'DIS_160': 'I47',            
        }
        
        col_ubi_sensorial = {    
            'PROVINCIA': 'D10',
            'DIS_161': 'E52',
            'DIS_162': 'F52',
            'DIS_163': 'G52',
            'DIS_164': 'H52',
            'DIS_165': 'I52',
            'DIS_166': 'E53',
            'DIS_167': 'F53',
            'DIS_168': 'G53',
            'DIS_169': 'H53',
            'DIS_170': 'I53',
            'DIS_171': 'E54',
            'DIS_172': 'F54',
            'DIS_173': 'G54',
            'DIS_174': 'H54',
            'DIS_175': 'I54',
            'DIS_176': 'E55',
            'DIS_177': 'F55',
            'DIS_178': 'G55',
            'DIS_179': 'H55',
            'DIS_180': 'I55',
            'DIS_181': 'E56',
            'DIS_182': 'F56',
            'DIS_183': 'G56',
            'DIS_184': 'H56',
            'DIS_185': 'I56',
            'DIS_186': 'E57',
            'DIS_187': 'F57',
            'DIS_188': 'G57',
            'DIS_189': 'H57',
            'DIS_190': 'I57',
            'DIS_191': 'E58',
            'DIS_192': 'F58',
            'DIS_193': 'G58',
            'DIS_194': 'H58',
            'DIS_195': 'I58',
        }
        
        col_ubi_mental = {    
            'PROVINCIA': 'D10',
            'DIS_196': 'E63',
            'DIS_197': 'F63',
            'DIS_198': 'G63',
            'DIS_199': 'H63',
            'DIS_200': 'I63',
            'DIS_201': 'E64',
            'DIS_202': 'F64',
            'DIS_203': 'G64',
            'DIS_204': 'H64',
            'DIS_205': 'I64',
            'DIS_206': 'E65',
            'DIS_207': 'F65',
            'DIS_208': 'G65',
            'DIS_209': 'H65',
            'DIS_210': 'I65',
            'DIS_211': 'E66',
            'DIS_212': 'F66',
            'DIS_213': 'G66',
            'DIS_214': 'H66',
            'DIS_215': 'I66',
        }
        
        col_ubi_certificado = {    
            'PROVINCIA': 'D10',
            'DIS_216': 'E71',
            'DIS_217': 'F71',
            'DIS_218': 'G71',
            'DIS_219': 'H71',
            'DIS_220': 'I71',
            'DIS_221': 'E72',
            'DIS_222': 'F72',
            'DIS_223': 'G72',
            'DIS_224': 'H72',
            'DIS_225': 'I72',
            'DIS_226': 'E73',
            'DIS_227': 'F73',
            'DIS_228': 'G73',
            'DIS_229': 'H73',
            'DIS_230': 'I73',
            'DIS_231': 'E74',
            'DIS_232': 'F74',
            'DIS_233': 'G74',
            'DIS_234': 'H74',
            'DIS_235': 'I74',
        }
        
        col_ubi_capacitacion = {    
            'PROVINCIA': 'D10',
            'DIS_273': 'D12',
            'DIS_274': 'E12',
        }
        
        col_ubi_agente = {    
            'PROVINCIA': 'D10',
            'DIS_236': 'D82',
            'DIS_237': 'E82',
            'DIS_238': 'F82',
            'DIS_239': 'G82',
            'DIS_240': 'H82',
            'DIS_241': 'I82',
        }      
        
        col_ubi_rbc = {    
            'PROVINCIA': 'D10',
            'DIS_242': 'E86',
            'DIS_243': 'F86',
            'DIS_244': 'G86',
            'DIS_245': 'H86',
            'DIS_246': 'I86',
            'DIS_247': 'E87',
            'DIS_248': 'F87',
            'DIS_249': 'G87',
            'DIS_250': 'H87',
            'DIS_251': 'I87',
            'DIS_252': 'E88',
            'DIS_253': 'F88',
            'DIS_254': 'G88',
            'DIS_255': 'H88',
            'DIS_256': 'I88',
            'DIS_257': 'E89',
            'DIS_258': 'F89',
            'DIS_259': 'G89',
            'DIS_260': 'H89',
            'DIS_261': 'I89',
            'DIS_262': 'E90',
            'DIS_263': 'F90',
            'DIS_264': 'G90',
            'DIS_265': 'H90',
            'DIS_266': 'I90'
        }
        
        col_ubi_comite = {    
            'PROVINCIA': 'D10',
            'DIS_267': 'D97',
            'DIS_268': 'E97',
            'DIS_269': 'F97',
            'DIS_270': 'G97',
            'DIS_271': 'H97',
            'DIS_272': 'I97',
        }
        
        # Inicializar diccionario para almacenar sumas por columna
        column_sums = {
            'DIS_1': 0,
            'DIS_2': 0,
            'DIS_3': 0,
            'DIS_4': 0,
            'DIS_5': 0,
            'DIS_6': 0,
            'DIS_7': 0,
            'DIS_8': 0,
            'DIS_9': 0,
            'DIS_10': 0,
            'DIS_11': 0,
            'DIS_12': 0,
            'DIS_13': 0,
            'DIS_14': 0,
            'DIS_15': 0,
            'DIS_16': 0,
            'DIS_17': 0,
            'DIS_18': 0,
            'DIS_19': 0,
            'DIS_20': 0,
            'DIS_21': 0,
            'DIS_22': 0,
            'DIS_23': 0,
            'DIS_24': 0,
            'DIS_25': 0,
            'DIS_26': 0,
            'DIS_27': 0,
            'DIS_28': 0,
            'DIS_29': 0,
            'DIS_30': 0,
            'DIS_31': 0,
            'DIS_32': 0,
            'DIS_33': 0,
            'DIS_34': 0,
            'DIS_35': 0,
            'DIS_36': 0,
            'DIS_37': 0,
            'DIS_38': 0,
            'DIS_39': 0,
            'DIS_40': 0,
            'DIS_41': 0,
            'DIS_42': 0,
            'DIS_43': 0,
            'DIS_44': 0,
            'DIS_45': 0,
            'DIS_46': 0,
            'DIS_47': 0,
            'DIS_48': 0,
            'DIS_49': 0,
            'DIS_50': 0,
            'DIS_51': 0,
            'DIS_52': 0,
            'DIS_53': 0,
            'DIS_54': 0,
            'DIS_55': 0,
            'DIS_56': 0,
            'DIS_57': 0,
            'DIS_58': 0,
            'DIS_59': 0,
            'DIS_60': 0,
            'DIS_61': 0,
            'DIS_62': 0,
            'DIS_63': 0,
            'DIS_64': 0,
            'DIS_65': 0,
            'DIS_66': 0,
            'DIS_67': 0,
            'DIS_68': 0,
            'DIS_69': 0,
            'DIS_70': 0,
            'DIS_71': 0,
            'DIS_72': 0,
            'DIS_73': 0,
            'DIS_74': 0,
            'DIS_75': 0,
            'DIS_76': 0,
            'DIS_77': 0,
            'DIS_78': 0,
            'DIS_79': 0,
            'DIS_80': 0,
            'DIS_81': 0,
            'DIS_82': 0,
            'DIS_83': 0,
            'DIS_84': 0,
            'DIS_85': 0,
            'DIS_86': 0,
            'DIS_87': 0,
            'DIS_88': 0,
            'DIS_89': 0,
            'DIS_90': 0,
            'DIS_91': 0,
            'DIS_92': 0,
            'DIS_93': 0,
            'DIS_94': 0,
            'DIS_95': 0,
            'DIS_96': 0,
            'DIS_97': 0,
            'DIS_98': 0,
            'DIS_99': 0,
            'DIS_100': 0,
            'DIS_101': 0,
            'DIS_102': 0,
            'DIS_103': 0,
            'DIS_104': 0,
            'DIS_105': 0,
            'DIS_106': 0,
            'DIS_107': 0,
            'DIS_108': 0,
            'DIS_109': 0,
            'DIS_110': 0,
            'DIS_111': 0,
            'DIS_112': 0,
            'DIS_113': 0,
            'DIS_114': 0,
            'DIS_115': 0,
            'DIS_116': 0,
            'DIS_117': 0,
            'DIS_118': 0,
            'DIS_119': 0,
            'DIS_120': 0,
            'DIS_121': 0,
            'DIS_122': 0,
            'DIS_123': 0,
            'DIS_124': 0,
            'DIS_125': 0,
            'DIS_126': 0,
            'DIS_127': 0,
            'DIS_128': 0,
            'DIS_129': 0,
            'DIS_130': 0, 
            'DIS_131': 0,
            'DIS_132': 0,
            'DIS_133': 0,
            'DIS_134': 0,
            'DIS_135': 0, 
            'DIS_136': 0,
            'DIS_137': 0,
            'DIS_138': 0,
            'DIS_139': 0,
            'DIS_140': 0, 
            'DIS_141': 0,
            'DIS_142': 0,
            'DIS_143': 0,
            'DIS_144': 0,
            'DIS_145': 0, 
            'DIS_146': 0,
            'DIS_147': 0,
            'DIS_148': 0,
            'DIS_149': 0,
            'DIS_150': 0, 
            'DIS_151': 0,
            'DIS_152': 0,
            'DIS_153': 0,
            'DIS_154': 0,
            'DIS_155': 0, 
            'DIS_156': 0,
            'DIS_157': 0,
            'DIS_158': 0,
            'DIS_159': 0,
            'DIS_160': 0,    
        }
        
        col_sum_sensorial = {       
            'DIS_161': 0,
            'DIS_162': 0,
            'DIS_163': 0,
            'DIS_164': 0,
            'DIS_165': 0,
            'DIS_166': 0,
            'DIS_167': 0,
            'DIS_168': 0,
            'DIS_169': 0,
            'DIS_170': 0,
            'DIS_171': 0,
            'DIS_172': 0,
            'DIS_173': 0,
            'DIS_174': 0,
            'DIS_175': 0,
            'DIS_176': 0,
            'DIS_177': 0,
            'DIS_178': 0,
            'DIS_179': 0,
            'DIS_180': 0,
            'DIS_181': 0,
            'DIS_182': 0,
            'DIS_183': 0,
            'DIS_184': 0,
            'DIS_185': 0,
            'DIS_186': 0,
            'DIS_187': 0,
            'DIS_188': 0,
            'DIS_189': 0,
            'DIS_190': 0,
            'DIS_191': 0,
            'DIS_192': 0,
            'DIS_193': 0,
            'DIS_194': 0,
            'DIS_195': 0,
        } 

        col_sum_mental = {    
            'DIS_196': 0,
            'DIS_197': 0,
            'DIS_198': 0,
            'DIS_199': 0,
            'DIS_200': 0,
            'DIS_201': 0,
            'DIS_202': 0,
            'DIS_203': 0,
            'DIS_204': 0,
            'DIS_205': 0,
            'DIS_206': 0,
            'DIS_207': 0,
            'DIS_208': 0,
            'DIS_209': 0,
            'DIS_210': 0,
            'DIS_211': 0,
            'DIS_212': 0,
            'DIS_213': 0,
            'DIS_214': 0,
            'DIS_215': 0,
        }
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_certificado = {       
            'DIS_216': 0,
            'DIS_217': 0,
            'DIS_218': 0,
            'DIS_219': 0,
            'DIS_220': 0,
            'DIS_221': 0,
            'DIS_222': 0,
            'DIS_223': 0,
            'DIS_224': 0,
            'DIS_225': 0,
            'DIS_226': 0,
            'DIS_227': 0,
            'DIS_228': 0,
            'DIS_229': 0,
            'DIS_230': 0,
            'DIS_231': 0,
            'DIS_232': 0,
            'DIS_233': 0,
            'DIS_234': 0,
            'DIS_235': 0,
        }  
        
        col_sum_capacitacion = {    
            'DIS_273': 0,
            'DIS_274': 0,
        }
        
        col_sum_agente = {    
            'DIS_236': 0,
            'DIS_237': 0,
            'DIS_238': 0,
            'DIS_239': 0,
            'DIS_240': 0,
            'DIS_241': 0,
        }      
        
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_rbc = {       
            'DIS_242': 0,
            'DIS_243': 0,
            'DIS_244': 0,
            'DIS_245': 0,
            'DIS_246': 0,
            'DIS_247': 0,
            'DIS_248': 0,
            'DIS_249': 0,
            'DIS_250': 0,
            'DIS_251': 0,
            'DIS_252': 0,
            'DIS_253': 0,
            'DIS_254': 0,
            'DIS_255': 0,
            'DIS_256': 0,
            'DIS_257': 0,
            'DIS_258': 0,
            'DIS_259': 0,
            'DIS_260': 0,
            'DIS_261': 0,
            'DIS_262': 0,
            'DIS_263': 0,
            'DIS_264': 0,
            'DIS_265': 0,
            'DIS_266': 0,
        } 
        
        col_sum_comite = {    
            'DIS_267': 0,
            'DIS_268': 0,
            'DIS_269': 0,
            'DIS_270': 0,
            'DIS_271': 0,
            'DIS_272': 0,
        }
                    
        ############################
        ###  DISCAPACIDAD FISICA ###
        ############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist:
            for col_name in column_sums:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(columnas_ubicaciones.keys()).index(col_name) + 1
                    column_sums[col_name] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila: {row}")                        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_name, total_cell in columnas_ubicaciones.items():
            if col_name in column_sums:
                # Obtener la celda correspondiente según la ubicación
                cell = sheet[total_cell]
                # Asignar el valor de la suma a la celda
                cell.value = column_sums[col_name]
                # Aplicar formato a la celda
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        
        # Sumar los valores del diccionario      
        total_sum_cat_1 =  sum([column_sums['DIS_1'], column_sums['DIS_2'], column_sums['DIS_3'],column_sums['DIS_4'],column_sums['DIS_5']])
        total_sum_cat_2 =  sum([column_sums['DIS_6'], column_sums['DIS_7'], column_sums['DIS_8'],column_sums['DIS_9'],column_sums['DIS_10']])
        total_sum_cat_3 =  sum([column_sums['DIS_11'], column_sums['DIS_12'], column_sums['DIS_13'],column_sums['DIS_14'],column_sums['DIS_15']])
        total_sum_cat_4 =  sum([column_sums['DIS_16'], column_sums['DIS_17'], column_sums['DIS_18'],column_sums['DIS_19'],column_sums['DIS_20']])
        total_sum_cat_5 =  sum([column_sums['DIS_21'], column_sums['DIS_22'], column_sums['DIS_23'],column_sums['DIS_24'],column_sums['DIS_25']])
        total_sum_cat_6 =  sum([column_sums['DIS_26'], column_sums['DIS_27'], column_sums['DIS_28'],column_sums['DIS_29'],column_sums['DIS_30']])
        total_sum_cat_7 =  sum([column_sums['DIS_31'], column_sums['DIS_32'], column_sums['DIS_33'],column_sums['DIS_34'],column_sums['DIS_35']])
        total_sum_cat_8 =  sum([column_sums['DIS_36'], column_sums['DIS_37'], column_sums['DIS_38'],column_sums['DIS_39'],column_sums['DIS_40']])
        total_sum_cat_9 =  sum([column_sums['DIS_41'], column_sums['DIS_42'], column_sums['DIS_43'],column_sums['DIS_44'],column_sums['DIS_45']])
        total_sum_cat_10 =  sum([column_sums['DIS_46'], column_sums['DIS_47'], column_sums['DIS_48'],column_sums['DIS_49'],column_sums['DIS_50']])
        total_sum_cat_11 =  sum([column_sums['DIS_51'], column_sums['DIS_52'], column_sums['DIS_53'],column_sums['DIS_54'],column_sums['DIS_55']])
        total_sum_cat_12 =  sum([column_sums['DIS_56'], column_sums['DIS_57'], column_sums['DIS_58'],column_sums['DIS_59'],column_sums['DIS_60']])
        total_sum_cat_13 =  sum([column_sums['DIS_61'], column_sums['DIS_62'], column_sums['DIS_63'],column_sums['DIS_64'],column_sums['DIS_65']])
        total_sum_cat_14 =  sum([column_sums['DIS_66'], column_sums['DIS_67'], column_sums['DIS_68'],column_sums['DIS_69'],column_sums['DIS_70']])
        total_sum_cat_15 =  sum([column_sums['DIS_71'], column_sums['DIS_72'], column_sums['DIS_73'],column_sums['DIS_74'],column_sums['DIS_75']])
        total_sum_cat_16 =  sum([column_sums['DIS_76'], column_sums['DIS_77'], column_sums['DIS_78'],column_sums['DIS_79'],column_sums['DIS_80']])   
        total_sum_cat_17 =  sum([column_sums['DIS_81'], column_sums['DIS_82'], column_sums['DIS_83'],column_sums['DIS_84'],column_sums['DIS_85']])
        total_sum_cat_18 =  sum([column_sums['DIS_86'], column_sums['DIS_87'], column_sums['DIS_88'],column_sums['DIS_89'],column_sums['DIS_90']])
        total_sum_cat_19 =  sum([column_sums['DIS_91'], column_sums['DIS_92'], column_sums['DIS_93'],column_sums['DIS_94'],column_sums['DIS_95']])
        total_sum_cat_20 =  sum([column_sums['DIS_96'], column_sums['DIS_97'], column_sums['DIS_98'],column_sums['DIS_99'],column_sums['DIS_100']])
        total_sum_cat_21 =  sum([column_sums['DIS_101'], column_sums['DIS_102'], column_sums['DIS_103'],column_sums['DIS_104'],column_sums['DIS_105']])
        total_sum_cat_22 =  sum([column_sums['DIS_106'], column_sums['DIS_107'], column_sums['DIS_108'],column_sums['DIS_109'],column_sums['DIS_110']])
        total_sum_cat_23 =  sum([column_sums['DIS_111'], column_sums['DIS_112'], column_sums['DIS_113'],column_sums['DIS_114'],column_sums['DIS_115']])
        total_sum_cat_24 =  sum([column_sums['DIS_116'], column_sums['DIS_117'], column_sums['DIS_118'],column_sums['DIS_119'],column_sums['DIS_120']])
        total_sum_cat_25 =  sum([column_sums['DIS_121'], column_sums['DIS_122'], column_sums['DIS_123'],column_sums['DIS_124'],column_sums['DIS_125']])
        total_sum_cat_26 =  sum([column_sums['DIS_126'], column_sums['DIS_127'], column_sums['DIS_128'],column_sums['DIS_129'],column_sums['DIS_130']])
        total_sum_cat_27 =  sum([column_sums['DIS_131'], column_sums['DIS_132'], column_sums['DIS_133'],column_sums['DIS_134'],column_sums['DIS_135']])
        total_sum_cat_28 =  sum([column_sums['DIS_136'], column_sums['DIS_137'], column_sums['DIS_138'],column_sums['DIS_139'],column_sums['DIS_140']])
        total_sum_cat_29 =  sum([column_sums['DIS_141'], column_sums['DIS_142'], column_sums['DIS_143'],column_sums['DIS_144'],column_sums['DIS_145']])
        total_sum_cat_30 =  sum([column_sums['DIS_146'], column_sums['DIS_147'], column_sums['DIS_148'],column_sums['DIS_149'],column_sums['DIS_150']])
        total_sum_cat_31 =  sum([column_sums['DIS_151'], column_sums['DIS_152'], column_sums['DIS_153'],column_sums['DIS_154'],column_sums['DIS_155']])
        total_sum_cat_32 =  sum([column_sums['DIS_156'], column_sums['DIS_157'], column_sums['DIS_158'],column_sums['DIS_159'],column_sums['DIS_160']])

        sheet['D16'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D16'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D16'] = total_sum_cat_1     
        
        sheet['D17'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D17'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D17'] = total_sum_cat_2 
        
        sheet['D18'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D18'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D18'] = total_sum_cat_3    
        
        sheet['D19'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D19'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D19'] = total_sum_cat_4    
        
        sheet['D20'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D20'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D20'] = total_sum_cat_5    
        
        sheet['D21'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D21'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D21'] = total_sum_cat_6    
        
        sheet['D22'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D22'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D22'] = total_sum_cat_7    
        
        sheet['D23'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D23'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D23'] = total_sum_cat_8    
        
        sheet['D24'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D24'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D24'] = total_sum_cat_9    
        
        sheet['D25'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D25'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D25'] = total_sum_cat_10 
        
        sheet['D26'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D26'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D26'] = total_sum_cat_11
                
        sheet['D27'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D27'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D27'] = total_sum_cat_12    
        
        sheet['D28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D28'] = total_sum_cat_13   
        
        sheet['D29'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D29'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D29'] = total_sum_cat_14   
        
        sheet['D30'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D30'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D30'] = total_sum_cat_15   
        
        sheet['D31'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D31'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D31'] = total_sum_cat_16   
        
        sheet['D32'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D32'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D32'] = total_sum_cat_17         
        
        sheet['D33'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D33'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D33'] = total_sum_cat_18   
        
        sheet['D34'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D34'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D34'] = total_sum_cat_19   
        
        sheet['D35'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D35'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D35'] = total_sum_cat_20   
        
        sheet['D36'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D36'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D36'] = total_sum_cat_21   
        
        sheet['D37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D37'] = total_sum_cat_22   
        
        sheet['D38'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D38'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D38'] = total_sum_cat_23   
        
        sheet['D39'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D39'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D39'] = total_sum_cat_24   
        
        sheet['D40'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D40'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D40'] = total_sum_cat_25  
        
        sheet['D41'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D41'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D41'] = total_sum_cat_26 
        
        sheet['D42'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D42'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D42'] = total_sum_cat_27   
        
        sheet['D43'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D43'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D43'] = total_sum_cat_28   
        
        sheet['D44'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D44'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D44'] = total_sum_cat_29  
        
        sheet['D45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D45'] = total_sum_cat_30  
        
        sheet['D46'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D46'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D46'] = total_sum_cat_31
        
        sheet['D47'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D47'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D47'] = total_sum_cat_32
        
        # Sumar los valores del VERTICAL      
        total_sum_cat_vertical_1 =  sum([column_sums['DIS_1'],column_sums['DIS_6'], column_sums['DIS_11'],column_sums['DIS_16'],column_sums['DIS_21'],column_sums['DIS_26'],column_sums['DIS_31'],column_sums['DIS_36'],column_sums['DIS_41'],column_sums['DIS_46'],column_sums['DIS_51'],column_sums['DIS_56'],column_sums['DIS_61'],column_sums['DIS_66'],column_sums['DIS_71'],column_sums['DIS_76'],column_sums['DIS_81'],column_sums['DIS_86'],column_sums['DIS_91'],column_sums['DIS_96'],column_sums['DIS_101'],column_sums['DIS_106'] ,column_sums['DIS_111'],column_sums['DIS_116'],column_sums['DIS_121'],column_sums['DIS_126'],column_sums['DIS_131'],column_sums['DIS_136'],column_sums['DIS_141'],column_sums['DIS_146'],column_sums['DIS_151'],column_sums['DIS_156']])
        total_sum_cat_vertical_2 =  sum([column_sums['DIS_2'],column_sums['DIS_7'], column_sums['DIS_12'],column_sums['DIS_17'],column_sums['DIS_22'],column_sums['DIS_27'],column_sums['DIS_32'],column_sums['DIS_37'],column_sums['DIS_42'],column_sums['DIS_47'],column_sums['DIS_52'],column_sums['DIS_57'],column_sums['DIS_62'],column_sums['DIS_67'],column_sums['DIS_72'],column_sums['DIS_77'],column_sums['DIS_82'],column_sums['DIS_87'],column_sums['DIS_92'],column_sums['DIS_97'],column_sums['DIS_102'],column_sums['DIS_107'] ,column_sums['DIS_112'],column_sums['DIS_117'],column_sums['DIS_122'],column_sums['DIS_127'],column_sums['DIS_132'],column_sums['DIS_137'],column_sums['DIS_142'],column_sums['DIS_147'],column_sums['DIS_152'],column_sums['DIS_157']])
        total_sum_cat_vertical_3 =  sum([column_sums['DIS_3'],column_sums['DIS_8'], column_sums['DIS_13'],column_sums['DIS_18'],column_sums['DIS_23'],column_sums['DIS_28'],column_sums['DIS_33'],column_sums['DIS_38'],column_sums['DIS_43'],column_sums['DIS_48'],column_sums['DIS_53'],column_sums['DIS_58'],column_sums['DIS_63'],column_sums['DIS_68'],column_sums['DIS_73'],column_sums['DIS_78'],column_sums['DIS_83'],column_sums['DIS_88'],column_sums['DIS_93'],column_sums['DIS_98'],column_sums['DIS_103'],column_sums['DIS_108'] ,column_sums['DIS_113'],column_sums['DIS_118'],column_sums['DIS_123'],column_sums['DIS_128'],column_sums['DIS_133'],column_sums['DIS_138'],column_sums['DIS_143'],column_sums['DIS_148'],column_sums['DIS_153'],column_sums['DIS_158']])
        total_sum_cat_vertical_4 =  sum([column_sums['DIS_4'],column_sums['DIS_9'], column_sums['DIS_14'],column_sums['DIS_19'],column_sums['DIS_24'],column_sums['DIS_29'],column_sums['DIS_34'],column_sums['DIS_39'],column_sums['DIS_44'],column_sums['DIS_49'],column_sums['DIS_54'],column_sums['DIS_59'],column_sums['DIS_64'],column_sums['DIS_69'],column_sums['DIS_74'],column_sums['DIS_79'],column_sums['DIS_84'],column_sums['DIS_89'],column_sums['DIS_94'],column_sums['DIS_99'],column_sums['DIS_104'],column_sums['DIS_109'] ,column_sums['DIS_114'],column_sums['DIS_119'],column_sums['DIS_124'],column_sums['DIS_129'],column_sums['DIS_134'],column_sums['DIS_139'],column_sums['DIS_144'],column_sums['DIS_149'],column_sums['DIS_154'],column_sums['DIS_159']])
        total_sum_cat_vertical_5 =  sum([column_sums['DIS_5'],column_sums['DIS_10'],column_sums['DIS_15'],column_sums['DIS_20'],column_sums['DIS_25'],column_sums['DIS_30'],column_sums['DIS_35'],column_sums['DIS_40'],column_sums['DIS_45'],column_sums['DIS_50'],column_sums['DIS_55'],column_sums['DIS_60'],column_sums['DIS_65'],column_sums['DIS_70'],column_sums['DIS_75'],column_sums['DIS_80'],column_sums['DIS_85'],column_sums['DIS_90'],column_sums['DIS_95'],column_sums['DIS_100'],column_sums['DIS_105'],column_sums['DIS_110'],column_sums['DIS_115'],column_sums['DIS_120'],column_sums['DIS_125'],column_sums['DIS_130'],column_sums['DIS_135'],column_sums['DIS_140'],column_sums['DIS_145'],column_sums['DIS_150'],column_sums['DIS_155'],column_sums['DIS_160']])

        sheet['E48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E48'] = total_sum_cat_vertical_1     
        
        sheet['F48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F48'] = total_sum_cat_vertical_2 
        
        sheet['G48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G48'] = total_sum_cat_vertical_3    
        
        sheet['H48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H48'] = total_sum_cat_vertical_4    
        
        sheet['I48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I48'] = total_sum_cat_vertical_5    
        ##########################################################################
        
        ###############################
        ###  DISCAPACIDAD SENSORIAL ###
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist_sensorial:
            for col_sensorial in col_sum_sensorial:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_sensorial.keys()).index(col_sensorial) + 1
                    col_sum_sensorial[col_sensorial] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_sensorial, total_cell_sensorial in col_ubi_sensorial.items():
            if col_sensorial in col_sum_sensorial:
                # Obtener la celda correspondiente según la ubicación
                cell_sensorial = sheet[total_cell_sensorial]
                # Asignar el valor de la suma a la celda
                cell_sensorial.value = col_sum_sensorial[col_sensorial]
                # Aplicar formato a la celda
                cell_sensorial.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_sensorial.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_sensorial.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_sensorial['DIS_161'], col_sum_sensorial['DIS_162'], col_sum_sensorial['DIS_163'], col_sum_sensorial['DIS_164'], col_sum_sensorial['DIS_165']])
        t_sum_cat_2 =  sum([col_sum_sensorial['DIS_166'], col_sum_sensorial['DIS_167'], col_sum_sensorial['DIS_168'], col_sum_sensorial['DIS_169'], col_sum_sensorial['DIS_170']])
        t_sum_cat_3 =  sum([col_sum_sensorial['DIS_171'], col_sum_sensorial['DIS_172'], col_sum_sensorial['DIS_173'], col_sum_sensorial['DIS_174'], col_sum_sensorial['DIS_175']])
        t_sum_cat_4 =  sum([col_sum_sensorial['DIS_176'], col_sum_sensorial['DIS_177'], col_sum_sensorial['DIS_178'], col_sum_sensorial['DIS_179'], col_sum_sensorial['DIS_180']])
        t_sum_cat_5 =  sum([col_sum_sensorial['DIS_181'], col_sum_sensorial['DIS_182'], col_sum_sensorial['DIS_183'], col_sum_sensorial['DIS_184'], col_sum_sensorial['DIS_185']])
        t_sum_cat_6 =  sum([col_sum_sensorial['DIS_186'], col_sum_sensorial['DIS_187'], col_sum_sensorial['DIS_188'], col_sum_sensorial['DIS_189'], col_sum_sensorial['DIS_190']])
        t_sum_cat_7 =  sum([col_sum_sensorial['DIS_191'], col_sum_sensorial['DIS_192'], col_sum_sensorial['DIS_193'], col_sum_sensorial['DIS_194'], col_sum_sensorial['DIS_195']])
        
        sheet['D52'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D52'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D52'] = t_sum_cat_1     
        
        sheet['D53'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D53'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D53'] = t_sum_cat_2 
        
        sheet['D54'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D54'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D54'] = t_sum_cat_3    
        
        sheet['D55'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D55'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D55'] = t_sum_cat_4    
        
        sheet['D56'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D56'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D56'] = t_sum_cat_5    
        
        sheet['D57'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D57'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D57'] = t_sum_cat_6    
        
        sheet['D58'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D58'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D58'] = t_sum_cat_7    
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_sensorial['DIS_161'],col_sum_sensorial['DIS_166'],col_sum_sensorial['DIS_171'],col_sum_sensorial['DIS_176'],col_sum_sensorial['DIS_181'],col_sum_sensorial['DIS_186'],col_sum_sensorial['DIS_191']])
        t_sum_cat_vertical_2 =  sum([col_sum_sensorial['DIS_162'],col_sum_sensorial['DIS_167'],col_sum_sensorial['DIS_172'],col_sum_sensorial['DIS_177'],col_sum_sensorial['DIS_182'],col_sum_sensorial['DIS_187'],col_sum_sensorial['DIS_192']])
        t_sum_cat_vertical_3 =  sum([col_sum_sensorial['DIS_163'],col_sum_sensorial['DIS_168'],col_sum_sensorial['DIS_173'],col_sum_sensorial['DIS_178'],col_sum_sensorial['DIS_183'],col_sum_sensorial['DIS_188'],col_sum_sensorial['DIS_193']])
        t_sum_cat_vertical_4 =  sum([col_sum_sensorial['DIS_164'],col_sum_sensorial['DIS_169'],col_sum_sensorial['DIS_174'],col_sum_sensorial['DIS_179'],col_sum_sensorial['DIS_184'],col_sum_sensorial['DIS_189'],col_sum_sensorial['DIS_194']])
        t_sum_cat_vertical_5 =  sum([col_sum_sensorial['DIS_165'],col_sum_sensorial['DIS_170'],col_sum_sensorial['DIS_175'],col_sum_sensorial['DIS_180'],col_sum_sensorial['DIS_185'],col_sum_sensorial['DIS_190'],col_sum_sensorial['DIS_195']])
        
        sheet['E59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E59'] = t_sum_cat_vertical_1     
        
        sheet['F59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F59'] = t_sum_cat_vertical_2 
        
        sheet['G59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G59'] = t_sum_cat_vertical_3    
        
        sheet['H59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H59'] = t_sum_cat_vertical_4    
        
        sheet['I59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I59'] = t_sum_cat_vertical_5    
        ##########################################################################
        
        ###############################
        ###  DISCAPACIDAD MENTAL ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist_mental:
            for col_mental in col_sum_mental:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_mental.keys()).index(col_mental) + 1
                    col_sum_mental[col_mental] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_mental, total_cell_mental in col_ubi_mental.items():
            if col_mental in col_sum_mental:
                # Obtener la celda correspondiente según la ubicación
                cell_mental = sheet[total_cell_mental]
                # Asignar el valor de la suma a la celda
                cell_mental.value = col_sum_mental[col_mental]
                # Aplicar formato a la celda
                cell_mental.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_mental.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_mental.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_mental['DIS_196'], col_sum_mental['DIS_197'], col_sum_mental['DIS_198'], col_sum_mental['DIS_199'], col_sum_mental['DIS_200']])
        t_sum_cat_2 =  sum([col_sum_mental['DIS_201'], col_sum_mental['DIS_202'], col_sum_mental['DIS_203'], col_sum_mental['DIS_204'], col_sum_mental['DIS_205']])
        t_sum_cat_3 =  sum([col_sum_mental['DIS_206'], col_sum_mental['DIS_207'], col_sum_mental['DIS_208'], col_sum_mental['DIS_209'], col_sum_mental['DIS_210']])
        t_sum_cat_4 =  sum([col_sum_mental['DIS_211'], col_sum_mental['DIS_212'], col_sum_mental['DIS_213'], col_sum_mental['DIS_214'], col_sum_mental['DIS_215']])
        
        sheet['D63'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D63'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D63'] = t_sum_cat_1     
        
        sheet['D64'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D64'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D64'] = t_sum_cat_2 
        
        sheet['D65'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D65'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D65'] = t_sum_cat_3    
        
        sheet['D66'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D66'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D66'] = t_sum_cat_4    

        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_mental['DIS_196'],col_sum_mental['DIS_201'],col_sum_mental['DIS_206'],col_sum_mental['DIS_211']])
        t_sum_cat_vertical_2 =  sum([col_sum_mental['DIS_197'],col_sum_mental['DIS_202'],col_sum_mental['DIS_207'],col_sum_mental['DIS_212']])
        t_sum_cat_vertical_3 =  sum([col_sum_mental['DIS_198'],col_sum_mental['DIS_203'],col_sum_mental['DIS_208'],col_sum_mental['DIS_213']])
        t_sum_cat_vertical_4 =  sum([col_sum_mental['DIS_199'],col_sum_mental['DIS_204'],col_sum_mental['DIS_209'],col_sum_mental['DIS_214']])
        t_sum_cat_vertical_5 =  sum([col_sum_mental['DIS_200'],col_sum_mental['DIS_205'],col_sum_mental['DIS_210'],col_sum_mental['DIS_215']])
        
        sheet['E67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E67'] = t_sum_cat_vertical_1     
        
        sheet['F67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F67'] = t_sum_cat_vertical_2 
        
        sheet['G67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G67'] = t_sum_cat_vertical_3    
        
        sheet['H67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H67'] = t_sum_cat_vertical_4    
        
        sheet['I67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I67'] = t_sum_cat_vertical_5    
        ##########################################################################
        
        
        #################################
        ###  DISCAPACIDAD CERTIFICADO ###
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist_certificado:
            for col_certificado in col_sum_certificado:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_certificado.keys()).index(col_certificado) + 1
                    col_sum_certificado[col_certificado] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_certificado, total_cell_certificado in col_ubi_certificado.items():
            if col_certificado in col_sum_certificado:
                # Obtener la celda correspondiente según la ubicación
                cell_certificado = sheet[total_cell_certificado]
                # Asignar el valor de la suma a la celda
                cell_certificado.value = col_sum_certificado[col_certificado]
                # Aplicar formato a la celda
                cell_certificado.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_certificado.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_certificado.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        # Sumar los valores del diccionario      
        t_sum_cat_cert_1 =  sum([col_sum_certificado['DIS_216'], col_sum_certificado['DIS_217'], col_sum_certificado['DIS_218'], col_sum_certificado['DIS_219'], col_sum_certificado['DIS_220']])
        t_sum_cat_cert_2 =  sum([col_sum_certificado['DIS_221'], col_sum_certificado['DIS_222'], col_sum_certificado['DIS_223'], col_sum_certificado['DIS_224'], col_sum_certificado['DIS_225']])
        t_sum_cat_cert_3 =  sum([col_sum_certificado['DIS_226'], col_sum_certificado['DIS_227'], col_sum_certificado['DIS_228'], col_sum_certificado['DIS_229'], col_sum_certificado['DIS_230']])
        t_sum_cat_cert_4 =  sum([col_sum_certificado['DIS_231'], col_sum_certificado['DIS_232'], col_sum_certificado['DIS_233'], col_sum_certificado['DIS_234'], col_sum_certificado['DIS_235']])

        sheet['D71'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D71'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D71'] = t_sum_cat_cert_1     
        
        sheet['D72'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D72'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D72'] = t_sum_cat_cert_2 
        
        sheet['D73'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D73'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D73'] = t_sum_cat_cert_3 
        
        sheet['D74'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D74'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D74'] = t_sum_cat_cert_4 
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vert_1 =  sum([col_sum_certificado['DIS_216'],col_sum_certificado['DIS_221'],col_sum_certificado['DIS_226'],col_sum_certificado['DIS_231']])
        t_sum_cat_vert_2 =  sum([col_sum_certificado['DIS_217'],col_sum_certificado['DIS_222'],col_sum_certificado['DIS_227'],col_sum_certificado['DIS_232']])
        t_sum_cat_vert_3 =  sum([col_sum_certificado['DIS_218'],col_sum_certificado['DIS_223'],col_sum_certificado['DIS_228'],col_sum_certificado['DIS_233']])
        t_sum_cat_vert_4 =  sum([col_sum_certificado['DIS_219'],col_sum_certificado['DIS_224'],col_sum_certificado['DIS_229'],col_sum_certificado['DIS_234']])
        t_sum_cat_vert_5 =  sum([col_sum_certificado['DIS_220'],col_sum_certificado['DIS_225'],col_sum_certificado['DIS_230'],col_sum_certificado['DIS_235']])
        
        sheet['E75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E75'] = t_sum_cat_vert_1     
        
        sheet['F75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F75'] = t_sum_cat_vert_2 
        
        sheet['G75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G75'] = t_sum_cat_vert_3    
        
        sheet['H75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H75'] = t_sum_cat_vert_4    
        
        sheet['I75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I75'] = t_sum_cat_vert_5    
        
        #################################
        ###  DISCAPACIDAD RBC ###########
        #################################       
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist_rbc:
            for col_rbc in col_sum_rbc:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_rbc.keys()).index(col_rbc) + 1
                    col_sum_rbc[col_rbc] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
                    
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_rbc, total_cell_rbc in col_ubi_rbc.items():
            if col_rbc in col_sum_rbc:
                # Obtener la celda correspondiente según la ubicación
                cell_rbc = sheet[total_cell_rbc]
                # Asignar el valor de la suma a la celda
                cell_rbc.value = col_sum_rbc[col_rbc]
                # Aplicar formato a la celda
                cell_rbc.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_rbc.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_rbc.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        ## Sumar los valores del diccionario      
        t_sum_cat_rbc_1 =  sum([col_sum_rbc['DIS_242'], col_sum_rbc['DIS_243'], col_sum_rbc['DIS_244'], col_sum_rbc['DIS_245'], col_sum_rbc['DIS_246']])
        t_sum_cat_rbc_2 =  sum([col_sum_rbc['DIS_247'], col_sum_rbc['DIS_248'], col_sum_rbc['DIS_249'], col_sum_rbc['DIS_250'], col_sum_rbc['DIS_251']])
        t_sum_cat_rbc_3 =  sum([col_sum_rbc['DIS_252'], col_sum_rbc['DIS_253'], col_sum_rbc['DIS_254'], col_sum_rbc['DIS_255'], col_sum_rbc['DIS_256']])
        t_sum_cat_rbc_4 =  sum([col_sum_rbc['DIS_257'], col_sum_rbc['DIS_258'], col_sum_rbc['DIS_259'], col_sum_rbc['DIS_260'], col_sum_rbc['DIS_261']])
        t_sum_cat_rbc_5 =  sum([col_sum_rbc['DIS_262'], col_sum_rbc['DIS_263'], col_sum_rbc['DIS_264'], col_sum_rbc['DIS_265'], col_sum_rbc['DIS_266']])

        sheet['D86'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D86'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D86'] = t_sum_cat_rbc_1     
        
        sheet['D87'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D87'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D87'] = t_sum_cat_rbc_2 
        
        sheet['D88'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D88'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D88'] = t_sum_cat_rbc_3     
        
        sheet['D89'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D89'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D89'] = t_sum_cat_rbc_4 
        
        sheet['D90'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D90'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D90'] = t_sum_cat_rbc_5 
        
        # Sumar los valores del VERTICAL      
        t_sum_vert_rbc_1 =  sum([col_sum_rbc['DIS_242'],col_sum_rbc['DIS_247'],col_sum_rbc['DIS_252'],col_sum_rbc['DIS_257'],col_sum_rbc['DIS_262']])
        t_sum_vert_rbc_2 =  sum([col_sum_rbc['DIS_243'],col_sum_rbc['DIS_248'],col_sum_rbc['DIS_253'],col_sum_rbc['DIS_258'],col_sum_rbc['DIS_263']])
        t_sum_vert_rbc_3 =  sum([col_sum_rbc['DIS_244'],col_sum_rbc['DIS_249'],col_sum_rbc['DIS_254'],col_sum_rbc['DIS_259'],col_sum_rbc['DIS_264']])
        t_sum_vert_rbc_4 =  sum([col_sum_rbc['DIS_245'],col_sum_rbc['DIS_250'],col_sum_rbc['DIS_255'],col_sum_rbc['DIS_260'],col_sum_rbc['DIS_265']])
        t_sum_vert_rbc_5 =  sum([col_sum_rbc['DIS_246'],col_sum_rbc['DIS_251'],col_sum_rbc['DIS_256'],col_sum_rbc['DIS_261'],col_sum_rbc['DIS_266']])
        
        sheet['E91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E91'] = t_sum_vert_rbc_1
        
        sheet['F91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F91'] = t_sum_vert_rbc_2 
        
        sheet['G91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G91'] = t_sum_vert_rbc_3    
        
        sheet['H91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H91'] = t_sum_vert_rbc_4    
        
        sheet['I91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I91'] = t_sum_vert_rbc_5   
        
        #################################
        ###  CAPACITACION PERSONAL ######
        #################################
        
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist_capacitacion:
            for col_capacitacion in col_sum_capacitacion:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_capacitacion.keys()).index(col_capacitacion) + 1
                    col_sum_capacitacion[col_capacitacion] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_capacitacion, total_cell_capacitacion in col_ubi_capacitacion.items():
            if col_capacitacion in col_sum_capacitacion:
                # Obtener la celda correspondiente según la ubicación
                cell_capacitacion = sheet[total_cell_capacitacion]
                # Asignar el valor de la suma a la celda
                cell_capacitacion.value = col_sum_capacitacion[col_capacitacion]
                # Aplicar formato a la celda
                cell_capacitacion.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_capacitacion.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_capacitacion.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_capacitacion['DIS_273']])
        t_sum_cat_2 = sum([col_sum_capacitacion['DIS_274']])
        
        sheet['D12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D12'] = t_sum_cat_1     
        
        sheet['E12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E12'] = t_sum_cat_2 
        
        ###############################
        ###  CAPACITACION AGENTE ######
        ###############################
                
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist_agente:
            for col_agente in col_sum_agente:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_agente.keys()).index(col_agente) + 1
                    col_sum_agente[col_agente] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_agente, total_cell_agente in col_ubi_agente.items():
            if col_agente in col_sum_agente:
                # Obtener la celda correspondiente según la ubicación
                cell_agente = sheet[total_cell_agente]
                # Asignar el valor de la suma a la celda
                cell_agente.value = col_sum_agente[col_agente]
                # Aplicar formato a la celda
                cell_agente.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_agente.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_agente.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_agente['DIS_236']])
        t_sum_cat_2 = sum([col_sum_agente['DIS_237']])
        t_sum_cat_3 = sum([col_sum_agente['DIS_238']])
        t_sum_cat_4 = sum([col_sum_agente['DIS_239']])
        t_sum_cat_5 = sum([col_sum_agente['DIS_240']])
        t_sum_cat_6 = sum([col_sum_agente['DIS_241']])
        
        sheet['D82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D82'] = t_sum_cat_1     
        
        sheet['E82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E82'] = t_sum_cat_2 
        
        sheet['F82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F82'] = t_sum_cat_3
        
        sheet['G82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G82'] = t_sum_cat_4 
        
        sheet['H82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H82'] = t_sum_cat_5
        
        sheet['I82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I82'] = t_sum_cat_6 
        
        ############################
        ###  CAPACITACION COMITE ###
        #############################
        
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_dist_comite:
            for col_comite in col_sum_comite:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_comite.keys()).index(col_comite) + 1
                    col_sum_comite[col_comite] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_comite, total_cell_comite in col_ubi_comite.items():
            if col_comite in col_sum_comite:
                # Obtener la celda correspondiente según la ubicación
                cell_comite = sheet[total_cell_comite]
                # Asignar el valor de la suma a la celda
                cell_comite.value = col_sum_comite[col_comite]
                # Aplicar formato a la celda
                cell_comite.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_comite.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_comite.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_comite['DIS_267']])
        t_sum_cat_2 = sum([col_sum_comite['DIS_268']])
        t_sum_cat_3 = sum([col_sum_comite['DIS_269']])
        t_sum_cat_4 = sum([col_sum_comite['DIS_270']])
        t_sum_cat_5 = sum([col_sum_comite['DIS_271']])
        t_sum_cat_6 = sum([col_sum_comite['DIS_272']])
        
        sheet['D97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D97'] = t_sum_cat_1     
        
        sheet['E97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E97'] = t_sum_cat_2 
        
        sheet['F97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F97'] = t_sum_cat_3
        
        sheet['G97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G97'] = t_sum_cat_4 
        
        sheet['H97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H97'] = t_sum_cat_5
        
        sheet['I97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I97'] = t_sum_cat_6 
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_operacional_distrito.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        workbook.save(response)

        return response

################################################
# REPORTE POR REDES
################################################
def get_redes(request,redes_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 'discapacidad/redes.html', context)

#--- FUNCIONES OPERACIONALES PARTES REPORTE -----------------------------------------
def rpt_operacional_fisico_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,                    
                    SUM(dis_1) AS dis_1,
                    SUM(dis_2) AS dis_2,
                    SUM(dis_3) AS dis_3,
                    SUM(dis_4) AS dis_4,
                    SUM(dis_5) AS dis_5,
                    SUM(dis_6) AS dis_6,
                    SUM(dis_7) AS dis_7,
                    SUM(dis_8) AS dis_8,
                    SUM(dis_9) AS dis_9,
                    SUM(dis_10) AS dis_10,
                    SUM(dis_11) AS dis_11,
                    SUM(dis_12) AS dis_12,
                    SUM(dis_13) AS dis_13,
                    SUM(dis_14) AS dis_14,
                    SUM(dis_15) AS dis_15,
                    SUM(dis_16) AS dis_16,
                    SUM(dis_17) AS dis_17,
                    SUM(dis_18) AS dis_18,
                    SUM(dis_19) AS dis_19,
                    SUM(dis_20) AS dis_20,
                    SUM(dis_21) AS dis_21,
                    SUM(dis_22) AS dis_22,
                    SUM(dis_23) AS dis_23,
                    SUM(dis_24) AS dis_24,
                    SUM(dis_25) AS dis_25,
                    SUM(dis_26) AS dis_26,
                    SUM(dis_27) AS dis_27,
                    SUM(dis_28) AS dis_28,
                    SUM(dis_29) AS dis_29,
                    SUM(dis_30) AS dis_30,
                    SUM(dis_31) AS dis_31,
                    SUM(dis_32) AS dis_32,
                    SUM(dis_33) AS dis_33,
                    SUM(dis_34) AS dis_34,
                    SUM(dis_35) AS dis_35,
                    SUM(dis_36) AS dis_36,
                    SUM(dis_37) AS dis_37,
                    SUM(dis_38) AS dis_38,
                    SUM(dis_39) AS dis_39,
                    SUM(dis_40) AS dis_40,
                    SUM(dis_41) AS dis_41,
                    SUM(dis_42) AS dis_42,
                    SUM(dis_43) AS dis_43,
                    SUM(dis_44) AS dis_44,
                    SUM(dis_45) AS dis_45,
                    SUM(dis_46) AS dis_46,
                    SUM(dis_47) AS dis_47,
                    SUM(dis_48) AS dis_48,
                    SUM(dis_49) AS dis_49,
                    SUM(dis_50) AS dis_50,
                    SUM(dis_51) AS dis_51,
                    SUM(dis_52) AS dis_52,
                    SUM(dis_53) AS dis_53,
                    SUM(dis_54) AS dis_54,
                    SUM(dis_55) AS dis_55,
                    SUM(dis_56) AS dis_56,
                    SUM(dis_57) AS dis_57,
                    SUM(dis_58) AS dis_58,
                    SUM(dis_59) AS dis_59,
                    SUM(dis_60) AS dis_60,
                    SUM(dis_61) AS dis_61,
                    SUM(dis_62) AS dis_62,
                    SUM(dis_63) AS dis_63,
                    SUM(dis_64) AS dis_64,
                    SUM(dis_65) AS dis_65,
                    SUM(dis_66) AS dis_66,
                    SUM(dis_67) AS dis_67,
                    SUM(dis_68) AS dis_68,
                    SUM(dis_69) AS dis_69,
                    SUM(dis_70) AS dis_70,
                    SUM(dis_71) AS dis_71,
                    SUM(dis_72) AS dis_72,
                    SUM(dis_73) AS dis_73,
                    SUM(dis_74) AS dis_74,
                    SUM(dis_75) AS dis_75,
                    SUM(dis_76) AS dis_76,
                    SUM(dis_77) AS dis_77,
                    SUM(dis_78) AS dis_78,
                    SUM(dis_79) AS dis_79,
                    SUM(dis_80) AS dis_80,
                    SUM(dis_81) AS dis_81,
                    SUM(dis_82) AS dis_82,
                    SUM(dis_83) AS dis_83,
                    SUM(dis_84) AS dis_84,
                    SUM(dis_85) AS dis_85,
                    SUM(dis_86) AS dis_86,
                    SUM(dis_87) AS dis_87,
                    SUM(dis_88) AS dis_88,
                    SUM(dis_89) AS dis_89,
                    SUM(dis_90) AS dis_90,
                    SUM(dis_91) AS dis_91,
                    SUM(dis_92) AS dis_92,
                    SUM(dis_93) AS dis_93,
                    SUM(dis_94) AS dis_94,
                    SUM(dis_95) AS dis_95,
                    SUM(dis_96) AS dis_96,
                    SUM(dis_97) AS dis_97,
                    SUM(dis_98) AS dis_98,
                    SUM(dis_99) AS dis_99,
                    SUM(dis_100) AS dis_100,
                    SUM(dis_101) AS dis_101,
                    SUM(dis_102) AS dis_102,
                    SUM(dis_103) AS dis_103,
                    SUM(dis_104) AS dis_104,
                    SUM(dis_105) AS dis_105,
                    SUM(dis_106) AS dis_106,
                    SUM(dis_107) AS dis_107,
                    SUM(dis_108) AS dis_108,
                    SUM(dis_109) AS dis_109,
                    SUM(dis_110) AS dis_110,
                    SUM(dis_111) AS dis_111,
                    SUM(dis_112) AS dis_112,
                    SUM(dis_113) AS dis_113,
                    SUM(dis_114) AS dis_114,
                    SUM(dis_115) AS dis_115,
                    SUM(dis_116) AS dis_116,
                    SUM(dis_117) AS dis_117,
                    SUM(dis_118) AS dis_118,
                    SUM(dis_119) AS dis_119,
                    SUM(dis_120) AS dis_120,
                    SUM(dis_121) AS dis_121,
                    SUM(dis_122) AS dis_122,
                    SUM(dis_123) AS dis_123,
                    SUM(dis_124) AS dis_124,
                    SUM(dis_125) AS dis_125,
                    SUM(dis_126) AS dis_126,
                    SUM(dis_127) AS dis_127,
                    SUM(dis_128) AS dis_128,
                    SUM(dis_129) AS dis_129,
                    SUM(dis_130) AS dis_130, 
                    SUM(dis_131) AS dis_131,
                    SUM(dis_132) AS dis_132,
                    SUM(dis_133) AS dis_133,
                    SUM(dis_134) AS dis_134,
                    SUM(dis_135) AS dis_135,
                    SUM(dis_136) AS dis_136,
                    SUM(dis_137) AS dis_137,
                    SUM(dis_138) AS dis_138,
                    SUM(dis_139) AS dis_139,
                    SUM(dis_140) AS dis_140, 
                    SUM(dis_141) AS dis_141,
                    SUM(dis_142) AS dis_142,
                    SUM(dis_143) AS dis_143,
                    SUM(dis_144) AS dis_144,
                    SUM(dis_145) AS dis_145,
                    SUM(dis_146) AS dis_146,
                    SUM(dis_147) AS dis_147,
                    SUM(dis_148) AS dis_148,
                    SUM(dis_149) AS dis_149,
                    SUM(dis_150) AS dis_150,
                    SUM(dis_151) AS dis_151,
                    SUM(dis_152) AS dis_152,
                    SUM(dis_153) AS dis_153,
                    SUM(dis_154) AS dis_154,
                    SUM(dis_155) AS dis_155,
                    SUM(dis_156) AS dis_156,
                    SUM(dis_157) AS dis_157,
                    SUM(dis_158) AS dis_158,
                    SUM(dis_159) AS dis_159,
                    SUM(dis_160) AS dis_160
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_1,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_2,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_3,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_4,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_5,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_6,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_7,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_8,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_9,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_10,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_11,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_12,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_13,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_14,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_15,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_16,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_17,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_18,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_19,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_20,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_21,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_22,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_23,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_24,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_25,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_26,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_27,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_28,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_29,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_30,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_31,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_32,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_33,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_34,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_35,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_36,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_37,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_38,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_39,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_40,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_41,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_42,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_43,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_44,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_45,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_46,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_47,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_48,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_49,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_50,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_51,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_52,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_53,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_54,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_55,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_56,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_57,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_58,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_59,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_60,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_61,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_62,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_63,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_64,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_65,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_66,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_67,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_68,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_69,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_70,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_71,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_72,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_73,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_74,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_75,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_76,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_77,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_78,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_79,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_80,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_81,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_82,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_83,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_84,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_85,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_86,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_87,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_88,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_89,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_90,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_91,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_92,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_93,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_94,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_95,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_96,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_97,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_98,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_99,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_100,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_101,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_102,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_103,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_104,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_105,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_106,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_107,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_108,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_109,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_110,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_111,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_112,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_113,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_114,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_115,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_116,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_117,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_118,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_119,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_120,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_121,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_122,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_123,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_124,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_125,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_126,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_127,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_128,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_129,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_130,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_131,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_132,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_133,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_134,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_135,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_136,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_137,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_138,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_139,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_140,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_141,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_142,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_143,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_144,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_145,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_146,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_147,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_148,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_149,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_150,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_151,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_152,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_153,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_154,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_155,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_156,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_157,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_158,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_159,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_160
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_red = cursor.fetchall()
    
    return resultado_red

def rpt_operacional_sensorial_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,
                    SUM(dis_161) AS dis_161,
                    SUM(dis_162) AS dis_162,
                    SUM(dis_163) AS dis_163,
                    SUM(dis_164) AS dis_164,
                    SUM(dis_165) AS dis_165,
                    SUM(dis_166) AS dis_166,
                    SUM(dis_167) AS dis_167,
                    SUM(dis_168) AS dis_168,
                    SUM(dis_169) AS dis_169,
                    SUM(dis_170) AS dis_170,
                    SUM(dis_171) AS dis_171,
                    SUM(dis_172) AS dis_172,
                    SUM(dis_173) AS dis_173,
                    SUM(dis_174) AS dis_174,
                    SUM(dis_175) AS dis_175,
                    SUM(dis_176) AS dis_176,
                    SUM(dis_177) AS dis_177,
                    SUM(dis_178) AS dis_178,
                    SUM(dis_179) AS dis_179,
                    SUM(dis_180) AS dis_180,
                    SUM(dis_181) AS dis_181,
                    SUM(dis_182) AS dis_182,
                    SUM(dis_183) AS dis_183,
                    SUM(dis_184) AS dis_184,
                    SUM(dis_185) AS dis_185,
                    SUM(dis_186) AS dis_186,
                    SUM(dis_187) AS dis_187,
                    SUM(dis_188) AS dis_188,
                    SUM(dis_189) AS dis_189,
                    SUM(dis_190) AS dis_190,
                    SUM(dis_191) AS dis_191,
                    SUM(dis_192) AS dis_192,
                    SUM(dis_193) AS dis_193,
                    SUM(dis_194) AS dis_194,
                    SUM(dis_195) AS dis_195
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_161,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_162,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_163,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_164,
                            SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_165,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_166,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_167,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_168,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_169,
                            SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_170,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_171,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_172,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_173,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_174,
                            SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_175,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_176,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_177,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_178,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_179,
                            SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_180,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_181,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_182,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_183,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_184,
                            SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_185,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_186,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_187,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_188,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_189,
                            SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_190,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_191,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_192,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_193,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_194,
                            SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_195
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_sensorial_red = cursor.fetchall()
    
    return resultado_sensorial_red

def rpt_operacional_certificado_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,
                    SUM(dis_216) AS dis_216,
                    SUM(dis_217) AS dis_217,
                    SUM(dis_218) AS dis_218,
                    SUM(dis_219) AS dis_219,
                    SUM(dis_220) AS dis_220,
                    SUM(dis_221) AS dis_221,
                    SUM(dis_222) AS dis_222,
                    SUM(dis_223) AS dis_223,
                    SUM(dis_224) AS dis_224,
                    SUM(dis_225) AS dis_225,
                    SUM(dis_226) AS dis_226,
                    SUM(dis_227) AS dis_227,
                    SUM(dis_228) AS dis_228,
                    SUM(dis_229) AS dis_229,
                    SUM(dis_230) AS dis_230,
                    SUM(dis_231) AS dis_231,
                    SUM(dis_232) AS dis_232,
                    SUM(dis_233) AS dis_233,
                    SUM(dis_234) AS dis_234,
                    SUM(dis_235) AS dis_235
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_216,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_217,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_218,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_219,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_220,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_221,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_222,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_223,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_224,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_225,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_226,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_227,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_228,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_229,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_230,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_231,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_232,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_233,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_234,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_235
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_certificado_red = cursor.fetchall()
    
    return resultado_certificado_red

def rpt_operacional_rbc_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,
                    SUM(dis_242) AS dis_242,
                    SUM(dis_243) AS dis_243,
                    SUM(dis_244) AS dis_244,
                    SUM(dis_245) AS dis_245,
                    SUM(dis_246) AS dis_246,
                    SUM(dis_247) AS dis_247,
                    SUM(dis_248) AS dis_248,
                    SUM(dis_249) AS dis_249,
                    SUM(dis_250) AS dis_250,
                    SUM(dis_251) AS dis_251,
                    SUM(dis_252) AS dis_252,
                    SUM(dis_253) AS dis_253,
                    SUM(dis_254) AS dis_254,
                    SUM(dis_255) AS dis_255,
                    SUM(dis_256) AS dis_256,
                    SUM(dis_257) AS dis_257,
                    SUM(dis_258) AS dis_258,
                    SUM(dis_259) AS dis_259,
                    SUM(dis_260) AS dis_260, 
                    SUM(dis_261) AS dis_261, 
                    SUM(dis_262) AS dis_262, 
                    SUM(dis_263) AS dis_263, 
                    SUM(dis_264) AS dis_264, 
                    SUM(dis_265) AS dis_265, 
                    SUM(dis_266) AS dis_266
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_242,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_243,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_244,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_245,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_246,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_247,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_248,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_249,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_250,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_251,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_252,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_253,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_254,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_255,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_256,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_257,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_258,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_259,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_260,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_261,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_262,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_263,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_264,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_265,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_266
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_rbc_red = cursor.fetchall()

    return resultado_rbc_red

def rpt_operacional_mental_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,
                    SUM(dis_196) AS dis_196,
                    SUM(dis_197) AS dis_197,
                    SUM(dis_198) AS dis_198,
                    SUM(dis_199) AS dis_199,
                    SUM(dis_200) AS dis_200,
                    SUM(dis_201) AS dis_201,
                    SUM(dis_202) AS dis_202,
                    SUM(dis_203) AS dis_203,
                    SUM(dis_204) AS dis_204,
                    SUM(dis_205) AS dis_205,
                    SUM(dis_206) AS dis_206,
                    SUM(dis_207) AS dis_207,
                    SUM(dis_208) AS dis_208,
                    SUM(dis_209) AS dis_209,
                    SUM(dis_210) AS dis_210, 
                    SUM(dis_211) AS dis_211, 
                    SUM(dis_212) AS dis_212, 
                    SUM(dis_213) AS dis_213, 
                    SUM(dis_214) AS dis_214, 
                    SUM(dis_215) AS dis_215
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_196,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_197,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_198,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_199,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_200,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_201,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_202,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_203,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_204,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_205,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_206,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_207,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_208,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_209,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_210,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_211,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_212,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_213,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_214,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_215
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_mental_red = cursor.fetchall()

    return resultado_mental_red

def rpt_operacional_capacitacion_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,
                    SUM(dis_273) AS dis_273,
                    SUM(dis_274) AS dis_274
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        COUNT(Categoria) AS dis_273,
                        SUM(gedad) AS dis_274
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_capacitacion_red = cursor.fetchall()

    return resultado_capacitacion_red

def rpt_operacional_agente_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,
                    SUM(dis_236) AS dis_236,
                    SUM(dis_237) AS dis_237,
                    SUM(dis_238) AS dis_238,
                    SUM(dis_239) AS dis_239,
                    SUM(dis_240) AS dis_240,
                    SUM(dis_241) AS dis_241
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 THEN 1 ELSE 0 END) 	   AS dis_236,
                        SUM(CASE WHEN Categoria = 1 THEN gedad ELSE 0 END) AS dis_237,
                        SUM(CASE WHEN Categoria = 2 THEN 1 ELSE 0 END)     AS dis_238,
                        SUM(CASE WHEN Categoria = 2 THEN gedad ELSE 0 END) AS dis_239,
                        SUM(CASE WHEN Categoria = 3 THEN 1 ELSE 0 END)     AS dis_240,
                        SUM(CASE WHEN Categoria = 3 THEN gedad ELSE 0 END) AS dis_241
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_agente_red = cursor.fetchall()

    return resultado_agente_red

def rpt_operacional_comite_red(codigo_red, fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,
                    SUM(dis_267) AS dis_267,
                    SUM(dis_268) AS dis_268,
                    SUM(dis_269) AS dis_269,
                    SUM(dis_270) AS dis_270,
                    SUM(dis_271) AS dis_271,
                    SUM(dis_272) AS dis_272
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        renaes,
                        SUM(CASE WHEN Actividad = 1 THEN 1 ELSE 0 END) 		AS dis_267,
                        SUM(CASE WHEN Actividad = 1 THEN Partic ELSE 0 END) AS dis_268,
                        SUM(CASE WHEN Actividad = 2 THEN 1 ELSE 0 END)      AS dis_269,
                        SUM(CASE WHEN Actividad = 2 THEN Partic ELSE 0 END) AS dis_270,
                        SUM(CASE WHEN Actividad = 3 THEN 1 ELSE 0 END)      AS dis_271,
                        SUM(CASE WHEN Actividad = 3 THEN Partic ELSE 0 END) AS dis_272
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_Red, red
                """, [str(codigo_red)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_comite_red = cursor.fetchall()

    return resultado_comite_red

class RptOperacinalRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        red = request.GET.get('red')

        # Creación de la consulta
        resultado_red = rpt_operacional_fisico_red(red, fecha_inicio, fecha_fin)
        resultado_sensorial_red = rpt_operacional_sensorial_red(red, fecha_inicio, fecha_fin)
        resultado_certificado_red = rpt_operacional_certificado_red(red, fecha_inicio, fecha_fin)
        resultado_rbc_red = rpt_operacional_rbc_red(red, fecha_inicio, fecha_fin)
        resultado_mental_red = rpt_operacional_mental_red(red, fecha_inicio, fecha_fin)
        resultado_capacitacion_red = rpt_operacional_capacitacion_red(red, fecha_inicio, fecha_fin)
        resultado_agente_red = rpt_operacional_agente_red(red, fecha_inicio, fecha_fin)
        resultado_comite_red = rpt_operacional_comite_red(red, fecha_inicio, fecha_fin)
        
        red_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Codigo_Red=red
        ).values_list('Red', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
        
        
        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambia el alto de la columna
        sheet.row_dimensions[1].height = 14
        sheet.row_dimensions[2].height = 14
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[15].height = 25
        # cambia el ancho de la columna
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 28
        sheet.column_dimensions['C'].width = 28
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 9
        sheet.column_dimensions['F'].width = 9
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 9
        sheet.column_dimensions['L'].width = 9
        # linea de division
        sheet.freeze_panes = 'AL8'
        
        # Configuración del fondo y el borde
        fill = PatternFill(patternType='solid', fgColor='00B0F0')
        border = Border(left=Side(style='thin', color='00B0F0'),
                        right=Side(style='thin', color='00B0F0'),
                        top=Side(style='thin', color='00B0F0'),
                        bottom=Side(style='thin', color='00B0F0'))

        borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                        right=Side(style='thin', color='A9A9A9'), # Plomo
                        top=Side(style='thin', color='A9A9A9'), # Plomo
                        bottom=Side(style='thin', color='A9A9A9')) # Plomo

        # crea titulo del reporte
        sheet['B1'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B1'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
        
        sheet['B2'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B2'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
        
        sheet['B4'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B4'].font = Font(name = 'Arial', size= 12, bold = True)
        sheet['B4'] = 'REPORTE DE ACTIVIDADES DEL COMPONENTE DE DISCAPACIDAD'
        
        sheet['B6'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B6'] ='DIRESA / GERESA / DISA'
        
        sheet['C6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C6'].font = Font(name = 'Arial', size= 7)
        sheet['C6'] ='JUNIN'

        sheet['B7'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B7'] ='PROV/ DIST/ RED/ MR/ ESTABLEC'
        
        sheet['C7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C7'].font = Font(name = 'Arial', size= 7)
        sheet['C7'] = red_codigo[0]
        
        sheet['E6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['E6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['E6'] ='PERIODO'
        
        sheet['F6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F6'] ='MES INICIO'
        
        sheet['F7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F7'] ='MES FIN'
        
        sheet['G6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G6'].font = Font(name = 'Arial', size= 8)
        sheet['G6'] = fecha_inicio_codigo[0]
        
        sheet['G7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G7'].font = Font(name = 'Arial', size= 8)
        sheet['G7'] = fecha_fin_codigo[0]
        
        sheet['B9'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B9'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B9'] ='PERSONAS CON DISCAPACIDAD RECIBEN ATENCION DE REHABILITACION EN ESTABLECIMIENTOS DE SALUD (3000688)'
        
        sheet['B10'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B10'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B10'] ='Capacitación en medicina de rehabilitación integral (5004449)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=12, max_row=12, min_col=3, max_col=5):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['C12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C12'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['C12'] ='Capacitación  (C0009)' 
        
        sheet['D11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D11'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D11'].fill = fill
        sheet['D11'].border = border
        sheet['D11'] = 'N°'
                
        sheet['E11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E11'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['E11'].fill = fill
        sheet['E11'].border = border
        sheet['E11'] = 'Capacitados'
        #######################################################
        ########## DISCAPACIDAD FISICA ########################
        #######################################################
        sheet['B14'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B14'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B14'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Física (5005150)' 
                
        sheet['B15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B15'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B15'].fill = fill
        sheet['B15'].border = border
        sheet['B15'] = 'Atenciones'
        
        sheet['D15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D15'].fill = fill
        sheet['D15'].border = border
        sheet['D15'] = 'Total'
        
        sheet['E15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E15'].fill = fill
        sheet['E15'].border = border
        sheet['E15'] = 'Niños         (1d - 11a)'
        
        sheet['F15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F15'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F15'].fill = fill
        sheet['F15'].border = border
        sheet['F15'] = 'Adolescentes (12a - 17a)'
        
        sheet['G15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G15'].fill = fill
        sheet['G15'].border = border
        sheet['G15'] = 'Jóvenes (18a - 29a)'
        
        sheet['H15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H15'].fill = fill
        sheet['H15'].border = border
        sheet['H15'] = 'Adultos (30a - 59a)'
        
        sheet['I15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I15'].fill = fill
        sheet['I15'].border = border
        sheet['I15'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=16, max_row=47, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B16'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B16'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B16'] ='LESIONES MEDULARES' 
                
        sheet['B17'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B17'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B17'] ='ENFERMEDAD DE PARKINSON Y SIMILARES' 
        
        sheet['B18'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B18'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B18'] ='REHABILITACIÓN EN PACIENTES AMPUTADOS' 
                
        sheet['B20'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B20'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B20'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA NEUROLÓGICA' 
        
        sheet['B23'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B23'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B23'] ='TRASTORNOS DEL DESARROLLO DE LA FUNCIÓN MOTRIZ' 
        
        sheet['B24'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B24'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B24'] ='ATENCIÓN DE REHABILITACIÓN DE ENFERMEDAD ARTICULAR DEGENERATIVA' 
        
        sheet['B25'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B25'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B25'] ='ENCEFALOPATÍA INFANTIL' 
                
        sheet['B26'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B26'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B26'] ='SÍNDROME DOWN' 
        
        sheet['B27'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B27'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B27'] ='REHABILITACIÓN EN PATOLOGÍA DE LA COLUMNA VERTEBRAL Y OTROS TRASTORNOS POSTURALES' 
        
        sheet['B34'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B34'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B34'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD CARDIOVASCULAR' 
        
        sheet['B35'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B35'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B35'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD RESPIRATORIA' 
        
        sheet['B36'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B36'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B36'] ='ATENCIÓN DE REHABILITACIÓN EN ALTERACIONES DEL PISO PÉLVICO' 
        
        sheet['B37'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B37'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B37'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA TRAUMATOLÓGICA Y REUMATOLÓGICA' 
        
        sheet['B44'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B44'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B44'] ='ATENCIÓN DE REHABILITACIÓN ONCOLÓGICA' 
        
        sheet['B46'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B46'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B46'] ='ATENCIÓN DE REHABILITACIÓN EN DOLOR' 
        
        sheet['B47'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B47'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B47'] ='ATENCIÓN DE REHABILITACIÓN EN PACIENTES QUEMADOS' 
        ####     
        sheet['C16'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C16'].font = Font(name = 'Arial', size= 7)
        sheet['C16'] ='Lesiones medulares' 
    
        sheet['C17'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C17'].font = Font(name = 'Arial', size= 7)
        sheet['C17'] ='Enfermedad de Parkinson y similares' 
        
        sheet['C18'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C18'].font = Font(name = 'Arial', size= 7)
        sheet['C18'] ='Amputados de miembros superiores' 
        
        sheet['C19'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C19'].font = Font(name = 'Arial', size= 7)
        sheet['C19'] ='Amputados de miembros inferiores' 
        
        sheet['C20'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C20'].font = Font(name = 'Arial', size= 7)
        sheet['C20'] ='Enfermedades cerebrovasculares'
        
        sheet['C21'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C21'].font = Font(name = 'Arial', size= 7)
        sheet['C21'] ='Enfermedades musculares y de la unión mioneural'
        
        sheet['C22'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C22'].font = Font(name = 'Arial', size= 7)
        sheet['C22'] ='Lesiones de nervios periféricos'
        
        sheet['C23'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C23'].font = Font(name = 'Arial', size= 7)
        sheet['C23'] ='Trastornos del desarrollo de la funcion motriz'
        
        sheet['C24'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C24'].font = Font(name = 'Arial', size= 7)
        sheet['C24'] ='Enfermedad articular degenerativa'
        
        sheet['C25'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C25'].font = Font(name = 'Arial', size= 7)
        sheet['C25'] ='Encefalopatía infantil y otras lesiones'
        
        sheet['C26'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C26'].font = Font(name = 'Arial', size= 7)
        sheet['C26'] ='Sindrome de Down'
        
        sheet['C27'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C27'].font = Font(name = 'Arial', size= 7)
        sheet['C27'] ='Cifosis y lordosis'
        
        sheet['C28'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C28'].font = Font(name = 'Arial', size= 7)
        sheet['C28'] ='Espondilo artropatías'
        
        sheet['C29'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C29'].font = Font(name = 'Arial', size= 7)
        sheet['C29'] ='Otros trastornos de los discos intervertebrales'
        
        sheet['C30'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C30'].font = Font(name = 'Arial', size= 7)
        sheet['C30'] ='Cervicalgia, dorsalgia, lumbago'
        
        sheet['C31'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C31'].font = Font(name = 'Arial', size= 7)
        sheet['C31'] ='Otras dorsopatías deformantes'
        
        sheet['C32'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C32'].font = Font(name = 'Arial', size= 7)
        sheet['C32'] ='Otros trastornos articulares'
        
        sheet['C33'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C33'].font = Font(name = 'Arial', size= 7)
        sheet['C33'] ='Defectos en la longitud de extremidades'
        
        sheet['C34'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C34'].font = Font(name = 'Arial', size= 7)
        sheet['C34'] ='Enfermedad cardiovascular'
        
        sheet['C35'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C35'].font = Font(name = 'Arial', size= 7)
        sheet['C35'] ='Enfermedad respiratoria'
        
        sheet['C36'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C36'].font = Font(name = 'Arial', size= 7)
        sheet['C36'] ='Vejiga neurogénica y dolor'
        
        sheet['C37'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C37'].font = Font(name = 'Arial', size= 7)
        sheet['C37'] ='Incontinencia'
        
        sheet['C38'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C38'].font = Font(name = 'Arial', size= 7)
        sheet['C38'] ='Prolapso'
        
        sheet['C39'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C39'].font = Font(name = 'Arial', size= 7)
        sheet['C39'] ='Traumatismos'
        
        sheet['C40'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C40'].font = Font(name = 'Arial', size= 7)
        sheet['C40'] ='Enfermedades del tejido conectivo'
        
        sheet['C41'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C41'].font = Font(name = 'Arial', size= 7)
        sheet['C41'] ='Patología articular excluida columna'
        
        sheet['C42'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C42'].font = Font(name = 'Arial', size= 7)
        sheet['C42'] ='Lesiones infecciosas'
        
        sheet['C43'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C43'].font = Font(name = 'Arial', size= 7)
        sheet['C43'] ='Lesión biomecánica'
        
        sheet['C44'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C44'].font = Font(name = 'Arial', size= 7)
        sheet['C44'] ='Linfedema'
        
        sheet['C45'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C45'].font = Font(name = 'Arial', size= 7)
        sheet['C45'] ='Sarcopenia'
        
        sheet['C46'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C46'].font = Font(name = 'Arial', size= 7)
        sheet['C46'] ='Dolor'
        
        sheet['C47'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C47'].font = Font(name = 'Arial', size= 7)
        sheet['C47'] ='Quemaduras, corrosiones y congelaciones'
        
        ##########################################################    
        ########## DISCAPACIDAD SENSORIAL ########################
        ##########################################################
        sheet['B50'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B50'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B50'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Sensorial (5005151)' 
                
        sheet['B51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B51'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B51'].fill = fill
        sheet['B51'].border = border
        sheet['B51'] = 'Atenciones'
        
        sheet['D51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D51'].fill = fill
        sheet['D51'].border = border
        sheet['D51'] = 'Total'
        
        sheet['E51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E51'].fill = fill
        sheet['E51'].border = border
        sheet['E51'] = 'Niños         (1d - 11a)'
        
        sheet['F51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F51'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F51'].fill = fill
        sheet['F51'].border = border
        sheet['F51'] = 'Adolescentes (12a - 17a)'
        
        sheet['G51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G51'].fill = fill
        sheet['G51'].border = border
        sheet['G51'] = 'Jóvenes (18a - 29a)'
        
        sheet['H51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H51'].fill = fill
        sheet['H51'].border = border
        sheet['H51'] = 'Adultos (30a - 59a)'
        
        sheet['I51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I51'].fill = fill
        sheet['I51'].border = border
        sheet['I51'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=52, max_row=58, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B52'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B52'] ='HIPOACUSIA Y/O SORDERA' 
        
        sheet['B53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B53'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B53'] ='BAJA VISION Y/O CEGUERA' 
        
        sheet['B54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B54'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B54'] ='SORDOMUDEZ' 
        
        sheet['B55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B55'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B55'] ='ENFERMEDAD CEREBRO VASCULAR' 
        
        sheet['B56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B56'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B56'] ='TRASTORNOS ESPECIFICOS DEL DESARROLLO DEL HABLA Y LENGUAJE' 
        
        sheet['B57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B57'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B57'] ='DISARTRIA Y DISFAGIA' 
        
        sheet['B59'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B59'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B59'] ='SUB TOTAL' 
        
        ########               
        sheet['C52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C52'].font = Font(name = 'Arial', size= 7)
        sheet['C52'] ='Hipoacusia y sordera' 
        
        sheet['C53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C53'].font = Font(name = 'Arial', size= 7)
        sheet['C53'] ='Baja visión y ceguera' 
        
        sheet['C54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C54'].font = Font(name = 'Arial', size= 7)
        sheet['C54'] ='Sordomudez' 
        
        sheet['C55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C55'].font = Font(name = 'Arial', size= 7)
        sheet['C55'] ='Enfermedad Cerebro vascular' 
        
        sheet['C56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C56'].font = Font(name = 'Arial', size= 7)
        sheet['C56'] ='Trastornos específicos del desarrollo del habla y lenguaje' 
        
        sheet['C57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C57'].font = Font(name = 'Arial', size= 7)
        sheet['C57'] ='Disartria' 
        
        sheet['C58'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C58'].font = Font(name = 'Arial', size= 7)
        sheet['C58'] ='Disfagia' 
        
        ########################################################
        ########## DISCAPACIDAD MENTAL #########################
        ########################################################
        sheet['B61'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B61'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B61'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Mental (5005152)' 
                
        sheet['B62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B62'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B62'].fill = fill
        sheet['B62'].border = border
        sheet['B62'] = 'Atenciones'
        
        sheet['D62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D62'].fill = fill
        sheet['D62'].border = border
        sheet['D62'] = 'Total'
        
        sheet['E62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E62'].fill = fill
        sheet['E62'].border = border
        sheet['E62'] = 'Niños         (1d - 11a)'
        
        sheet['F62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F62'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F62'].fill = fill
        sheet['F62'].border = border
        sheet['F62'] = 'Adolescentes (12a - 17a)'
        
        sheet['G62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G62'].fill = fill
        sheet['G62'].border = border
        sheet['G62'] = 'Jóvenes (18a - 29a)'
        
        sheet['H62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H62'].fill = fill
        sheet['H62'].border = border
        sheet['H62'] = 'Adultos (30a - 59a)'
        
        sheet['I62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I62'].fill = fill
        sheet['I62'].border = border
        sheet['I62'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=63, max_row=66, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B63'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B63'] ='TRASTORNOS DE APRENDIZAJE' 
        
        sheet['B64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B64'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B64'] ='RETRASO MENTAL LEVE, MODERADO, SEVERO' 
        
        sheet['B65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B65'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B65'] ='TRASTORNOS DEL ESPECTRO AUTISTA' 
        
        sheet['B66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B66'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B66'] ='OTROS TRASTORNOS DE SALUD MENTAL' 
        
        sheet['B67'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B67'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B67'] ='SUB TOTAL' 
        
        ##########
        
        sheet['C63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C63'].font = Font(name = 'Arial', size= 7)
        sheet['C63'] ='Trastornos del aprendizaje' 
        
        sheet['C64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C64'].font = Font(name = 'Arial', size= 7)
        sheet['C64'] ='Retardo Mental: Leve, moderado, severo' 
        
        sheet['C65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C65'].font = Font(name = 'Arial', size= 7)
        sheet['C65'] ='Trastornos del espectro autista' 
        
        sheet['C66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C66'].font = Font(name = 'Arial', size= 7)
        sheet['C66'] ='Otras alteraciones de salud mental' 
                
        ##################################################
        ########## CERTIFICACION #########################
        ##################################################
        sheet['B69'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B69'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B69'] ='PERSONAS CON DISCAPACIDAD CERTIFICADAS EN ESTABLECIMIENTOS DE SALUD (3000689)' 
                
        sheet['B70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B70'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B70'].fill = fill
        sheet['B70'].border = border
        sheet['B70'] = 'Atenciones'
        
        sheet['D70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D70'].fill = fill
        sheet['D70'].border = border
        sheet['D70'] = 'Total'
        
        sheet['E70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E70'].fill = fill
        sheet['E70'].border = border
        sheet['E70'] = 'Niños         (1d - 11a)'
        
        sheet['F70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F70'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F70'].fill = fill
        sheet['F70'].border = border
        sheet['F70'] = 'Adolescentes (12a - 17a)'
        
        sheet['G70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G70'].fill = fill
        sheet['G70'].border = border
        sheet['G70'] = 'Jóvenes (18a - 29a)'
        
        sheet['H70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H70'].fill = fill
        sheet['H70'].border = border
        sheet['H70'] = 'Adultos (30a - 59a)'
        
        sheet['I70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I70'].fill = fill
        sheet['I70'].border = border
        sheet['I70'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=71, max_row=74, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B71'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B71'] ='Certificación de Discapacidad (0515204)' 
        
        sheet['B74'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B74'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B74'] ='Certificación de Incapacidad (0515205)' 
        
        sheet['B75'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B75'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B75'] ='SUB TOTAL' 
        
        sheet['C71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C71'].font = Font(name = 'Arial', size= 7)
        sheet['C71'] ='Evaluación' 
        
        sheet['C72'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C72'].font = Font(name = 'Arial', size= 7)
        sheet['C72'] ='Calificación' 
        
        sheet['C73'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C73'].font = Font(name = 'Arial', size= 7)
        sheet['C73'] ='Certificación' 

        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B77'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B77'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B77'] ='PERSONAS CON DISCAPACIDAD RECIBEN SERVICIOS DE REHABILITACIÓN BASADA EN LA COMUNIDAD (3000690)' 
        
        sheet['B78'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B78'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B78'] ='CAPACITACIÓN A AGENTES COMUNITARIOS EN REHABILITACIÓN BASADA EN LA COMUNIDAD (5005155)' 
        
        sheet['B82'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B82'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B82'] ='Capacitación a Agentes Comunitarios  (APP138)' 
        
        sheet['D80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D80'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D80'].fill = fill
        sheet['D80'].border = border
        sheet['D80'] = 'Taller'
        
        sheet['F80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F80'].fill = fill
        sheet['F80'].border = border
        sheet['F80'] = 'Sesion Educativa'
        
        sheet['H80'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H80'].fill = fill
        sheet['H80'].border = border
        sheet['H80'] = 'Sesion Demostrativa'
        
        sheet['D81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D81'].fill = fill
        sheet['D81'].border = border
        sheet['D81'] = 'N°'
        
        sheet['E81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E81'].fill = fill
        sheet['E81'].border = border
        sheet['E81'] = 'Capacitados'
        
        sheet['F81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F81'].fill = fill
        sheet['F81'].border = border
        sheet['F81'] = 'N°'
        
        sheet['G81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G81'].fill = fill
        sheet['G81'].border = border
        sheet['G81'] = 'Capacitados'
        
        sheet['H81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H81'].fill = fill
        sheet['H81'].border = border
        sheet['H81'] = 'N° '
        
        sheet['I81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I81'].fill = fill
        sheet['I81'].border = border
        sheet['I81'] = 'Capacitados'
        
        
        #borde plomo
        for row in sheet.iter_rows(min_row=82, max_row=82, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        ################################################
        ########## VISITAS RBC #########################
        ################################################
        sheet['B84'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B84'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B84'] ='Vistas a alas familias Rehabilitacion Basada en la Comunidad' 
                
        sheet['B85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B85'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B85'].fill = fill
        sheet['B85'].border = border
        sheet['B85'] = 'Visitas'
        
        sheet['D85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D85'].fill = fill
        sheet['D85'].border = border
        sheet['D85'] = 'Total'
        
        sheet['E85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E85'].fill = fill
        sheet['E85'].border = border
        sheet['E85'] = 'Niños         (1d - 11a)'
        
        sheet['F85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F85'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F85'].fill = fill
        sheet['F85'].border = border
        sheet['F85'] = 'Adolescentes (12a - 17a)'
        
        sheet['G85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G85'].fill = fill
        sheet['G85'].border = border
        sheet['G85'] = 'Jóvenes (18a - 29a)'
        
        sheet['H85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H85'].fill = fill
        sheet['H85'].border = border
        sheet['H85'] = 'Adultos (30a - 59a)'
        
        sheet['I85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I85'].fill = fill
        sheet['I85'].border = border
        sheet['I85'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=86, max_row=90, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = borde_plomo
        
        sheet['B86'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B86'].font = Font(name = 'Arial', size= 8)
        sheet['B86'] ='1º Visita' 
        
        sheet['B87'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B87'].font = Font(name = 'Arial', size= 8)
        sheet['B87'] ='2º Visita' 
        
        sheet['B88'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B88'].font = Font(name = 'Arial', size= 8)
        sheet['B88'] ='3º Visita' 
        
        sheet['B89'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B89'].font = Font(name = 'Arial', size= 8)
        sheet['B89'] ='4º a Visita (trazador)' 
        
        sheet['B90'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B90'].font = Font(name = 'Arial', size= 8)
        sheet['B90'] ='5º a + Visitas' 
        
        sheet['B91'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B91'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B91'] ='SUB TOTAL' 
        
        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B93'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B93'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B93'] ='Capacitación a Actores Sociales para la aplicación de la estrategia de Rehabilitación Basada en la Comunidad' 
                
        sheet['B94'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B94'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B94'] ='Actividades con Gobiernos Locales:' 
        
        sheet['B97'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B97'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B97'] ='Actividad con Comité Multisectorial (APP96)' 
        
        sheet['D95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D95'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D95'].fill = fill
        sheet['D95'].border = border
        sheet['D95'] = 'Taller'
        
        sheet['F95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F95'].fill = fill
        sheet['F95'].border = border
        sheet['F95'] = 'Sesion Educativa'
        
        sheet['H95'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H95'].fill = fill
        sheet['H95'].border = border
        sheet['H95'] = 'Sesion Demostrativa'
        
        sheet['D96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D96'].fill = fill
        sheet['D96'].border = border
        sheet['D96'] = 'N°'
        
        sheet['E96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E96'].fill = fill
        sheet['E96'].border = border
        sheet['E96'] = 'Capacitados'
        
        sheet['F96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F96'].fill = fill
        sheet['F96'].border = border
        sheet['F96'] = 'N°'
        
        sheet['G96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G96'].fill = fill
        sheet['G96'].border = border
        sheet['G96'] = 'Capacitados'
        
        sheet['H96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H96'].fill = fill
        sheet['H96'].border = border
        sheet['H96'] = 'N° '
        
        sheet['I96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I96'].fill = fill
        sheet['I96'].border = border
        sheet['I96'] = 'Capacitados'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=97, max_row=97, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        #############################################################################
        #############################################################################                
        # cambina celdas
        sheet.merge_cells('C6:D6')
        sheet.merge_cells('C7:E7')
        
        sheet.merge_cells('B18:B19')
        sheet.merge_cells('B20:B22')
        sheet.merge_cells('B27:B33')
        sheet.merge_cells('B37:B43')
        sheet.merge_cells('B44:B45')
        
        # sensorial
        sheet.merge_cells('B57:B58')
        
        sheet.merge_cells('B15:C15')
        sheet.merge_cells('B51:C51')
        
        # mental
        sheet.merge_cells('B62:C62')
        
        #certificado
        sheet.merge_cells('B70:C70')
        
        sheet.merge_cells('B71:B73')
        
        #RBC
        sheet.merge_cells('B85:C85')
        
        #capa
        sheet.merge_cells('D80:E80')
        sheet.merge_cells('F80:G80')
        sheet.merge_cells('H80:I80')

        sheet.merge_cells('D95:E95')
        sheet.merge_cells('F95:G95')
        sheet.merge_cells('H95:I95')
        
        #capacitacion
        sheet.merge_cells('B82:C82')
        sheet.merge_cells('B97:C97')
        
        #visita
        sheet.merge_cells('B86:C86')
        sheet.merge_cells('B87:C87')
        sheet.merge_cells('B88:C88')
        sheet.merge_cells('B89:C89')
        sheet.merge_cells('B90:C90')
        
        # Definir ubicaciones específicas para cada columna y su suma total
        columnas_ubicaciones = {
            'PROVINCIA': 'D10',
            'DIS_1': 'E16', 
            'DIS_2': 'F16',
            'DIS_3': 'G16',
            'DIS_4': 'H16',
            'DIS_5': 'I16',
            'DIS_6': 'E17',
            'DIS_7': 'F17',
            'DIS_8': 'G17',
            'DIS_9': 'H17',
            'DIS_10': 'I17',
            'DIS_11': 'E18',
            'DIS_12': 'F18',
            'DIS_13': 'G18',
            'DIS_14': 'H18',
            'DIS_15': 'I18',
            'DIS_16': 'E19',
            'DIS_17': 'F19',
            'DIS_18': 'G19',
            'DIS_19': 'H19',
            'DIS_20': 'I19',
            'DIS_21': 'E20',
            'DIS_22': 'F20',
            'DIS_23': 'G20',
            'DIS_24': 'H20',
            'DIS_25': 'I20',
            'DIS_26': 'E21',
            'DIS_27': 'F21',
            'DIS_28': 'G21',
            'DIS_29': 'H21',
            'DIS_30': 'I21',
            'DIS_31': 'E22',
            'DIS_32': 'F22',
            'DIS_33': 'G22',
            'DIS_34': 'H22',
            'DIS_35': 'I22',
            'DIS_36': 'E23',
            'DIS_37': 'F23',
            'DIS_38': 'G23',
            'DIS_39': 'H23',
            'DIS_40': 'I23',
            'DIS_41': 'E24',
            'DIS_42': 'F24',
            'DIS_43': 'G24',
            'DIS_44': 'H24',
            'DIS_45': 'I24',
            'DIS_46': 'E25',
            'DIS_47': 'F25',
            'DIS_48': 'G25',
            'DIS_49': 'H25',
            'DIS_50': 'I25',
            'DIS_51': 'E26',
            'DIS_52': 'F26',
            'DIS_53': 'G26',
            'DIS_54': 'H26',
            'DIS_55': 'I26',
            'DIS_56': 'E27',
            'DIS_57': 'F27',
            'DIS_58': 'G27',
            'DIS_59': 'H27',
            'DIS_60': 'I27',
            'DIS_61': 'E28',
            'DIS_62': 'F28',
            'DIS_63': 'G28',
            'DIS_64': 'H28',
            'DIS_65': 'I28',
            'DIS_66': 'E29',
            'DIS_67': 'F29',
            'DIS_68': 'G29',
            'DIS_69': 'H29',
            'DIS_70': 'I29',
            'DIS_71': 'E30',
            'DIS_72': 'F30',
            'DIS_73': 'G30',
            'DIS_74': 'H30',
            'DIS_75': 'I30',
            'DIS_76': 'E31',
            'DIS_77': 'F31',
            'DIS_78': 'G31',
            'DIS_79': 'H31',
            'DIS_80': 'I31',
            'DIS_81': 'E32',
            'DIS_82': 'F32',
            'DIS_83': 'G32',
            'DIS_84': 'H32',
            'DIS_85': 'I32',
            'DIS_86': 'E33',
            'DIS_87': 'F33',
            'DIS_88': 'G33',
            'DIS_89': 'H33',
            'DIS_90': 'I33',
            'DIS_91': 'E34',
            'DIS_92': 'F34',
            'DIS_93': 'G34',
            'DIS_94': 'H34',
            'DIS_95': 'I34',
            'DIS_96': 'E35',
            'DIS_97': 'F35',
            'DIS_98': 'G35',
            'DIS_99': 'H35',
            'DIS_100': 'I35',
            'DIS_101': 'E36',
            'DIS_102': 'F36',
            'DIS_103': 'G36',
            'DIS_104': 'H36',
            'DIS_105': 'I36',
            'DIS_106': 'E37',
            'DIS_107': 'F37',
            'DIS_108': 'G37',
            'DIS_109': 'H37',
            'DIS_110': 'I37',
            'DIS_111': 'E38',
            'DIS_112': 'F38',
            'DIS_113': 'G38',
            'DIS_114': 'H38',
            'DIS_115': 'I38',
            'DIS_116': 'E39',
            'DIS_117': 'F39',
            'DIS_118': 'G39',
            'DIS_119': 'H39',
            'DIS_120': 'I39',
            'DIS_121': 'E40',
            'DIS_122': 'F40',
            'DIS_123': 'G40',
            'DIS_124': 'H40',
            'DIS_125': 'I40',
            'DIS_126': 'E41',
            'DIS_127': 'F41',
            'DIS_128': 'G41',
            'DIS_129': 'H41',
            'DIS_130': 'I41', 
            'DIS_131': 'E42',
            'DIS_132': 'F42',
            'DIS_133': 'G42',
            'DIS_134': 'H42',
            'DIS_135': 'I42', 
            'DIS_136': 'E43',
            'DIS_137': 'F43',
            'DIS_138': 'G43',
            'DIS_139': 'H43',
            'DIS_140': 'I43', 
            'DIS_141': 'E44',
            'DIS_142': 'F44',
            'DIS_143': 'G44',
            'DIS_144': 'H44',
            'DIS_145': 'I44', 
            'DIS_146': 'E45',
            'DIS_147': 'F45',
            'DIS_148': 'G45',
            'DIS_149': 'H45',
            'DIS_150': 'I45', 
            'DIS_151': 'E46',
            'DIS_152': 'F46',
            'DIS_153': 'G46',
            'DIS_154': 'H46',
            'DIS_155': 'I46', 
            'DIS_156': 'E47',
            'DIS_157': 'F47',
            'DIS_158': 'G47',
            'DIS_159': 'H47',
            'DIS_160': 'I47',            
        }
        
        col_ubi_sensorial = {    
            'PROVINCIA': 'D10',
            'DIS_161': 'E52',
            'DIS_162': 'F52',
            'DIS_163': 'G52',
            'DIS_164': 'H52',
            'DIS_165': 'I52',
            'DIS_166': 'E53',
            'DIS_167': 'F53',
            'DIS_168': 'G53',
            'DIS_169': 'H53',
            'DIS_170': 'I53',
            'DIS_171': 'E54',
            'DIS_172': 'F54',
            'DIS_173': 'G54',
            'DIS_174': 'H54',
            'DIS_175': 'I54',
            'DIS_176': 'E55',
            'DIS_177': 'F55',
            'DIS_178': 'G55',
            'DIS_179': 'H55',
            'DIS_180': 'I55',
            'DIS_181': 'E56',
            'DIS_182': 'F56',
            'DIS_183': 'G56',
            'DIS_184': 'H56',
            'DIS_185': 'I56',
            'DIS_186': 'E57',
            'DIS_187': 'F57',
            'DIS_188': 'G57',
            'DIS_189': 'H57',
            'DIS_190': 'I57',
            'DIS_191': 'E58',
            'DIS_192': 'F58',
            'DIS_193': 'G58',
            'DIS_194': 'H58',
            'DIS_195': 'I58',
        }
        
        col_ubi_mental = {    
            'PROVINCIA': 'D10',
            'DIS_196': 'E63',
            'DIS_197': 'F63',
            'DIS_198': 'G63',
            'DIS_199': 'H63',
            'DIS_200': 'I63',
            'DIS_201': 'E64',
            'DIS_202': 'F64',
            'DIS_203': 'G64',
            'DIS_204': 'H64',
            'DIS_205': 'I64',
            'DIS_206': 'E65',
            'DIS_207': 'F65',
            'DIS_208': 'G65',
            'DIS_209': 'H65',
            'DIS_210': 'I65',
            'DIS_211': 'E66',
            'DIS_212': 'F66',
            'DIS_213': 'G66',
            'DIS_214': 'H66',
            'DIS_215': 'I66',
        }
        
        col_ubi_certificado = {    
            'PROVINCIA': 'D10',
            'DIS_216': 'E71',
            'DIS_217': 'F71',
            'DIS_218': 'G71',
            'DIS_219': 'H71',
            'DIS_220': 'I71',
            'DIS_221': 'E72',
            'DIS_222': 'F72',
            'DIS_223': 'G72',
            'DIS_224': 'H72',
            'DIS_225': 'I72',
            'DIS_226': 'E73',
            'DIS_227': 'F73',
            'DIS_228': 'G73',
            'DIS_229': 'H73',
            'DIS_230': 'I73',
            'DIS_231': 'E74',
            'DIS_232': 'F74',
            'DIS_233': 'G74',
            'DIS_234': 'H74',
            'DIS_235': 'I74',
        }
        
        col_ubi_capacitacion = {    
            'PROVINCIA': 'D10',
            'DIS_273': 'D12',
            'DIS_274': 'E12',
        }
        
        col_ubi_agente = {    
            'PROVINCIA': 'D10',
            'DIS_236': 'D82',
            'DIS_237': 'E82',
            'DIS_238': 'F82',
            'DIS_239': 'G82',
            'DIS_240': 'H82',
            'DIS_241': 'I82',
        }      
        
        col_ubi_rbc = {    
            'PROVINCIA': 'D10',
            'DIS_242': 'E86',
            'DIS_243': 'F86',
            'DIS_244': 'G86',
            'DIS_245': 'H86',
            'DIS_246': 'I86',
            'DIS_247': 'E87',
            'DIS_248': 'F87',
            'DIS_249': 'G87',
            'DIS_250': 'H87',
            'DIS_251': 'I87',
            'DIS_252': 'E88',
            'DIS_253': 'F88',
            'DIS_254': 'G88',
            'DIS_255': 'H88',
            'DIS_256': 'I88',
            'DIS_257': 'E89',
            'DIS_258': 'F89',
            'DIS_259': 'G89',
            'DIS_260': 'H89',
            'DIS_261': 'I89',
            'DIS_262': 'E90',
            'DIS_263': 'F90',
            'DIS_264': 'G90',
            'DIS_265': 'H90',
            'DIS_266': 'I90'
        }
        
        col_ubi_comite = {    
            'PROVINCIA': 'D10',
            'DIS_267': 'D97',
            'DIS_268': 'E97',
            'DIS_269': 'F97',
            'DIS_270': 'G97',
            'DIS_271': 'H97',
            'DIS_272': 'I97',
        }
        
        # Inicializar diccionario para almacenar sumas por columna
        column_sums = {
            'DIS_1': 0,
            'DIS_2': 0,
            'DIS_3': 0,
            'DIS_4': 0,
            'DIS_5': 0,
            'DIS_6': 0,
            'DIS_7': 0,
            'DIS_8': 0,
            'DIS_9': 0,
            'DIS_10': 0,
            'DIS_11': 0,
            'DIS_12': 0,
            'DIS_13': 0,
            'DIS_14': 0,
            'DIS_15': 0,
            'DIS_16': 0,
            'DIS_17': 0,
            'DIS_18': 0,
            'DIS_19': 0,
            'DIS_20': 0,
            'DIS_21': 0,
            'DIS_22': 0,
            'DIS_23': 0,
            'DIS_24': 0,
            'DIS_25': 0,
            'DIS_26': 0,
            'DIS_27': 0,
            'DIS_28': 0,
            'DIS_29': 0,
            'DIS_30': 0,
            'DIS_31': 0,
            'DIS_32': 0,
            'DIS_33': 0,
            'DIS_34': 0,
            'DIS_35': 0,
            'DIS_36': 0,
            'DIS_37': 0,
            'DIS_38': 0,
            'DIS_39': 0,
            'DIS_40': 0,
            'DIS_41': 0,
            'DIS_42': 0,
            'DIS_43': 0,
            'DIS_44': 0,
            'DIS_45': 0,
            'DIS_46': 0,
            'DIS_47': 0,
            'DIS_48': 0,
            'DIS_49': 0,
            'DIS_50': 0,
            'DIS_51': 0,
            'DIS_52': 0,
            'DIS_53': 0,
            'DIS_54': 0,
            'DIS_55': 0,
            'DIS_56': 0,
            'DIS_57': 0,
            'DIS_58': 0,
            'DIS_59': 0,
            'DIS_60': 0,
            'DIS_61': 0,
            'DIS_62': 0,
            'DIS_63': 0,
            'DIS_64': 0,
            'DIS_65': 0,
            'DIS_66': 0,
            'DIS_67': 0,
            'DIS_68': 0,
            'DIS_69': 0,
            'DIS_70': 0,
            'DIS_71': 0,
            'DIS_72': 0,
            'DIS_73': 0,
            'DIS_74': 0,
            'DIS_75': 0,
            'DIS_76': 0,
            'DIS_77': 0,
            'DIS_78': 0,
            'DIS_79': 0,
            'DIS_80': 0,
            'DIS_81': 0,
            'DIS_82': 0,
            'DIS_83': 0,
            'DIS_84': 0,
            'DIS_85': 0,
            'DIS_86': 0,
            'DIS_87': 0,
            'DIS_88': 0,
            'DIS_89': 0,
            'DIS_90': 0,
            'DIS_91': 0,
            'DIS_92': 0,
            'DIS_93': 0,
            'DIS_94': 0,
            'DIS_95': 0,
            'DIS_96': 0,
            'DIS_97': 0,
            'DIS_98': 0,
            'DIS_99': 0,
            'DIS_100': 0,
            'DIS_101': 0,
            'DIS_102': 0,
            'DIS_103': 0,
            'DIS_104': 0,
            'DIS_105': 0,
            'DIS_106': 0,
            'DIS_107': 0,
            'DIS_108': 0,
            'DIS_109': 0,
            'DIS_110': 0,
            'DIS_111': 0,
            'DIS_112': 0,
            'DIS_113': 0,
            'DIS_114': 0,
            'DIS_115': 0,
            'DIS_116': 0,
            'DIS_117': 0,
            'DIS_118': 0,
            'DIS_119': 0,
            'DIS_120': 0,
            'DIS_121': 0,
            'DIS_122': 0,
            'DIS_123': 0,
            'DIS_124': 0,
            'DIS_125': 0,
            'DIS_126': 0,
            'DIS_127': 0,
            'DIS_128': 0,
            'DIS_129': 0,
            'DIS_130': 0, 
            'DIS_131': 0,
            'DIS_132': 0,
            'DIS_133': 0,
            'DIS_134': 0,
            'DIS_135': 0, 
            'DIS_136': 0,
            'DIS_137': 0,
            'DIS_138': 0,
            'DIS_139': 0,
            'DIS_140': 0, 
            'DIS_141': 0,
            'DIS_142': 0,
            'DIS_143': 0,
            'DIS_144': 0,
            'DIS_145': 0, 
            'DIS_146': 0,
            'DIS_147': 0,
            'DIS_148': 0,
            'DIS_149': 0,
            'DIS_150': 0, 
            'DIS_151': 0,
            'DIS_152': 0,
            'DIS_153': 0,
            'DIS_154': 0,
            'DIS_155': 0, 
            'DIS_156': 0,
            'DIS_157': 0,
            'DIS_158': 0,
            'DIS_159': 0,
            'DIS_160': 0,    
        }
        
        col_sum_sensorial = {       
            'DIS_161': 0,
            'DIS_162': 0,
            'DIS_163': 0,
            'DIS_164': 0,
            'DIS_165': 0,
            'DIS_166': 0,
            'DIS_167': 0,
            'DIS_168': 0,
            'DIS_169': 0,
            'DIS_170': 0,
            'DIS_171': 0,
            'DIS_172': 0,
            'DIS_173': 0,
            'DIS_174': 0,
            'DIS_175': 0,
            'DIS_176': 0,
            'DIS_177': 0,
            'DIS_178': 0,
            'DIS_179': 0,
            'DIS_180': 0,
            'DIS_181': 0,
            'DIS_182': 0,
            'DIS_183': 0,
            'DIS_184': 0,
            'DIS_185': 0,
            'DIS_186': 0,
            'DIS_187': 0,
            'DIS_188': 0,
            'DIS_189': 0,
            'DIS_190': 0,
            'DIS_191': 0,
            'DIS_192': 0,
            'DIS_193': 0,
            'DIS_194': 0,
            'DIS_195': 0,
        } 

        col_sum_mental = {    
            'DIS_196': 0,
            'DIS_197': 0,
            'DIS_198': 0,
            'DIS_199': 0,
            'DIS_200': 0,
            'DIS_201': 0,
            'DIS_202': 0,
            'DIS_203': 0,
            'DIS_204': 0,
            'DIS_205': 0,
            'DIS_206': 0,
            'DIS_207': 0,
            'DIS_208': 0,
            'DIS_209': 0,
            'DIS_210': 0,
            'DIS_211': 0,
            'DIS_212': 0,
            'DIS_213': 0,
            'DIS_214': 0,
            'DIS_215': 0,
        }
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_certificado = {       
            'DIS_216': 0,
            'DIS_217': 0,
            'DIS_218': 0,
            'DIS_219': 0,
            'DIS_220': 0,
            'DIS_221': 0,
            'DIS_222': 0,
            'DIS_223': 0,
            'DIS_224': 0,
            'DIS_225': 0,
            'DIS_226': 0,
            'DIS_227': 0,
            'DIS_228': 0,
            'DIS_229': 0,
            'DIS_230': 0,
            'DIS_231': 0,
            'DIS_232': 0,
            'DIS_233': 0,
            'DIS_234': 0,
            'DIS_235': 0,
        }  
        
        col_sum_capacitacion = {    
            'DIS_273': 0,
            'DIS_274': 0,
        }
        
        col_sum_agente = {    
            'DIS_236': 0,
            'DIS_237': 0,
            'DIS_238': 0,
            'DIS_239': 0,
            'DIS_240': 0,
            'DIS_241': 0,
        }      
        
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_rbc = {       
            'DIS_242': 0,
            'DIS_243': 0,
            'DIS_244': 0,
            'DIS_245': 0,
            'DIS_246': 0,
            'DIS_247': 0,
            'DIS_248': 0,
            'DIS_249': 0,
            'DIS_250': 0,
            'DIS_251': 0,
            'DIS_252': 0,
            'DIS_253': 0,
            'DIS_254': 0,
            'DIS_255': 0,
            'DIS_256': 0,
            'DIS_257': 0,
            'DIS_258': 0,
            'DIS_259': 0,
            'DIS_260': 0,
            'DIS_261': 0,
            'DIS_262': 0,
            'DIS_263': 0,
            'DIS_264': 0,
            'DIS_265': 0,
            'DIS_266': 0,
        } 
        
        col_sum_comite = {    
            'DIS_267': 0,
            'DIS_268': 0,
            'DIS_269': 0,
            'DIS_270': 0,
            'DIS_271': 0,
            'DIS_272': 0,
        }
                    
        ############################
        ###  DISCAPACIDAD FISICA ###
        ############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_red:
            for col_name in column_sums:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(columnas_ubicaciones.keys()).index(col_name) + 1
                    column_sums[col_name] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila: {row}")                        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_name, total_cell in columnas_ubicaciones.items():
            if col_name in column_sums:
                # Obtener la celda correspondiente según la ubicación
                cell = sheet[total_cell]
                # Asignar el valor de la suma a la celda
                cell.value = column_sums[col_name]
                # Aplicar formato a la celda
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        
        # Sumar los valores del diccionario      
        total_sum_cat_1 =  sum([column_sums['DIS_1'], column_sums['DIS_2'], column_sums['DIS_3'],column_sums['DIS_4'],column_sums['DIS_5']])
        total_sum_cat_2 =  sum([column_sums['DIS_6'], column_sums['DIS_7'], column_sums['DIS_8'],column_sums['DIS_9'],column_sums['DIS_10']])
        total_sum_cat_3 =  sum([column_sums['DIS_11'], column_sums['DIS_12'], column_sums['DIS_13'],column_sums['DIS_14'],column_sums['DIS_15']])
        total_sum_cat_4 =  sum([column_sums['DIS_16'], column_sums['DIS_17'], column_sums['DIS_18'],column_sums['DIS_19'],column_sums['DIS_20']])
        total_sum_cat_5 =  sum([column_sums['DIS_21'], column_sums['DIS_22'], column_sums['DIS_23'],column_sums['DIS_24'],column_sums['DIS_25']])
        total_sum_cat_6 =  sum([column_sums['DIS_26'], column_sums['DIS_27'], column_sums['DIS_28'],column_sums['DIS_29'],column_sums['DIS_30']])
        total_sum_cat_7 =  sum([column_sums['DIS_31'], column_sums['DIS_32'], column_sums['DIS_33'],column_sums['DIS_34'],column_sums['DIS_35']])
        total_sum_cat_8 =  sum([column_sums['DIS_36'], column_sums['DIS_37'], column_sums['DIS_38'],column_sums['DIS_39'],column_sums['DIS_40']])
        total_sum_cat_9 =  sum([column_sums['DIS_41'], column_sums['DIS_42'], column_sums['DIS_43'],column_sums['DIS_44'],column_sums['DIS_45']])
        total_sum_cat_10 =  sum([column_sums['DIS_46'], column_sums['DIS_47'], column_sums['DIS_48'],column_sums['DIS_49'],column_sums['DIS_50']])
        total_sum_cat_11 =  sum([column_sums['DIS_51'], column_sums['DIS_52'], column_sums['DIS_53'],column_sums['DIS_54'],column_sums['DIS_55']])
        total_sum_cat_12 =  sum([column_sums['DIS_56'], column_sums['DIS_57'], column_sums['DIS_58'],column_sums['DIS_59'],column_sums['DIS_60']])
        total_sum_cat_13 =  sum([column_sums['DIS_61'], column_sums['DIS_62'], column_sums['DIS_63'],column_sums['DIS_64'],column_sums['DIS_65']])
        total_sum_cat_14 =  sum([column_sums['DIS_66'], column_sums['DIS_67'], column_sums['DIS_68'],column_sums['DIS_69'],column_sums['DIS_70']])
        total_sum_cat_15 =  sum([column_sums['DIS_71'], column_sums['DIS_72'], column_sums['DIS_73'],column_sums['DIS_74'],column_sums['DIS_75']])
        total_sum_cat_16 =  sum([column_sums['DIS_76'], column_sums['DIS_77'], column_sums['DIS_78'],column_sums['DIS_79'],column_sums['DIS_80']])   
        total_sum_cat_17 =  sum([column_sums['DIS_81'], column_sums['DIS_82'], column_sums['DIS_83'],column_sums['DIS_84'],column_sums['DIS_85']])
        total_sum_cat_18 =  sum([column_sums['DIS_86'], column_sums['DIS_87'], column_sums['DIS_88'],column_sums['DIS_89'],column_sums['DIS_90']])
        total_sum_cat_19 =  sum([column_sums['DIS_91'], column_sums['DIS_92'], column_sums['DIS_93'],column_sums['DIS_94'],column_sums['DIS_95']])
        total_sum_cat_20 =  sum([column_sums['DIS_96'], column_sums['DIS_97'], column_sums['DIS_98'],column_sums['DIS_99'],column_sums['DIS_100']])
        total_sum_cat_21 =  sum([column_sums['DIS_101'], column_sums['DIS_102'], column_sums['DIS_103'],column_sums['DIS_104'],column_sums['DIS_105']])
        total_sum_cat_22 =  sum([column_sums['DIS_106'], column_sums['DIS_107'], column_sums['DIS_108'],column_sums['DIS_109'],column_sums['DIS_110']])
        total_sum_cat_23 =  sum([column_sums['DIS_111'], column_sums['DIS_112'], column_sums['DIS_113'],column_sums['DIS_114'],column_sums['DIS_115']])
        total_sum_cat_24 =  sum([column_sums['DIS_116'], column_sums['DIS_117'], column_sums['DIS_118'],column_sums['DIS_119'],column_sums['DIS_120']])
        total_sum_cat_25 =  sum([column_sums['DIS_121'], column_sums['DIS_122'], column_sums['DIS_123'],column_sums['DIS_124'],column_sums['DIS_125']])
        total_sum_cat_26 =  sum([column_sums['DIS_126'], column_sums['DIS_127'], column_sums['DIS_128'],column_sums['DIS_129'],column_sums['DIS_130']])
        total_sum_cat_27 =  sum([column_sums['DIS_131'], column_sums['DIS_132'], column_sums['DIS_133'],column_sums['DIS_134'],column_sums['DIS_135']])
        total_sum_cat_28 =  sum([column_sums['DIS_136'], column_sums['DIS_137'], column_sums['DIS_138'],column_sums['DIS_139'],column_sums['DIS_140']])
        total_sum_cat_29 =  sum([column_sums['DIS_141'], column_sums['DIS_142'], column_sums['DIS_143'],column_sums['DIS_144'],column_sums['DIS_145']])
        total_sum_cat_30 =  sum([column_sums['DIS_146'], column_sums['DIS_147'], column_sums['DIS_148'],column_sums['DIS_149'],column_sums['DIS_150']])
        total_sum_cat_31 =  sum([column_sums['DIS_151'], column_sums['DIS_152'], column_sums['DIS_153'],column_sums['DIS_154'],column_sums['DIS_155']])
        total_sum_cat_32 =  sum([column_sums['DIS_156'], column_sums['DIS_157'], column_sums['DIS_158'],column_sums['DIS_159'],column_sums['DIS_160']])

        sheet['D16'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D16'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D16'] = total_sum_cat_1     
        
        sheet['D17'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D17'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D17'] = total_sum_cat_2 
        
        sheet['D18'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D18'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D18'] = total_sum_cat_3    
        
        sheet['D19'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D19'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D19'] = total_sum_cat_4    
        
        sheet['D20'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D20'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D20'] = total_sum_cat_5    
        
        sheet['D21'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D21'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D21'] = total_sum_cat_6    
        
        sheet['D22'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D22'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D22'] = total_sum_cat_7    
        
        sheet['D23'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D23'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D23'] = total_sum_cat_8    
        
        sheet['D24'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D24'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D24'] = total_sum_cat_9    
        
        sheet['D25'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D25'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D25'] = total_sum_cat_10 
        
        sheet['D26'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D26'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D26'] = total_sum_cat_11
                
        sheet['D27'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D27'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D27'] = total_sum_cat_12    
        
        sheet['D28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D28'] = total_sum_cat_13   
        
        sheet['D29'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D29'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D29'] = total_sum_cat_14   
        
        sheet['D30'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D30'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D30'] = total_sum_cat_15   
        
        sheet['D31'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D31'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D31'] = total_sum_cat_16   
        
        sheet['D32'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D32'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D32'] = total_sum_cat_17         
        
        sheet['D33'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D33'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D33'] = total_sum_cat_18   
        
        sheet['D34'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D34'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D34'] = total_sum_cat_19   
        
        sheet['D35'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D35'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D35'] = total_sum_cat_20   
        
        sheet['D36'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D36'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D36'] = total_sum_cat_21   
        
        sheet['D37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D37'] = total_sum_cat_22   
        
        sheet['D38'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D38'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D38'] = total_sum_cat_23   
        
        sheet['D39'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D39'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D39'] = total_sum_cat_24   
        
        sheet['D40'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D40'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D40'] = total_sum_cat_25  
        
        sheet['D41'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D41'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D41'] = total_sum_cat_26 
        
        sheet['D42'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D42'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D42'] = total_sum_cat_27   
        
        sheet['D43'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D43'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D43'] = total_sum_cat_28   
        
        sheet['D44'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D44'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D44'] = total_sum_cat_29  
        
        sheet['D45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D45'] = total_sum_cat_30  
        
        sheet['D46'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D46'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D46'] = total_sum_cat_31
        
        sheet['D47'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D47'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D47'] = total_sum_cat_32
        
        # Sumar los valores del VERTICAL      
        total_sum_cat_vertical_1 =  sum([column_sums['DIS_1'],column_sums['DIS_6'], column_sums['DIS_11'],column_sums['DIS_16'],column_sums['DIS_21'],column_sums['DIS_26'],column_sums['DIS_31'],column_sums['DIS_36'],column_sums['DIS_41'],column_sums['DIS_46'],column_sums['DIS_51'],column_sums['DIS_56'],column_sums['DIS_61'],column_sums['DIS_66'],column_sums['DIS_71'],column_sums['DIS_76'],column_sums['DIS_81'],column_sums['DIS_86'],column_sums['DIS_91'],column_sums['DIS_96'],column_sums['DIS_101'],column_sums['DIS_106'] ,column_sums['DIS_111'],column_sums['DIS_116'],column_sums['DIS_121'],column_sums['DIS_126'],column_sums['DIS_131'],column_sums['DIS_136'],column_sums['DIS_141'],column_sums['DIS_146'],column_sums['DIS_151'],column_sums['DIS_156']])
        total_sum_cat_vertical_2 =  sum([column_sums['DIS_2'],column_sums['DIS_7'], column_sums['DIS_12'],column_sums['DIS_17'],column_sums['DIS_22'],column_sums['DIS_27'],column_sums['DIS_32'],column_sums['DIS_37'],column_sums['DIS_42'],column_sums['DIS_47'],column_sums['DIS_52'],column_sums['DIS_57'],column_sums['DIS_62'],column_sums['DIS_67'],column_sums['DIS_72'],column_sums['DIS_77'],column_sums['DIS_82'],column_sums['DIS_87'],column_sums['DIS_92'],column_sums['DIS_97'],column_sums['DIS_102'],column_sums['DIS_107'] ,column_sums['DIS_112'],column_sums['DIS_117'],column_sums['DIS_122'],column_sums['DIS_127'],column_sums['DIS_132'],column_sums['DIS_137'],column_sums['DIS_142'],column_sums['DIS_147'],column_sums['DIS_152'],column_sums['DIS_157']])
        total_sum_cat_vertical_3 =  sum([column_sums['DIS_3'],column_sums['DIS_8'], column_sums['DIS_13'],column_sums['DIS_18'],column_sums['DIS_23'],column_sums['DIS_28'],column_sums['DIS_33'],column_sums['DIS_38'],column_sums['DIS_43'],column_sums['DIS_48'],column_sums['DIS_53'],column_sums['DIS_58'],column_sums['DIS_63'],column_sums['DIS_68'],column_sums['DIS_73'],column_sums['DIS_78'],column_sums['DIS_83'],column_sums['DIS_88'],column_sums['DIS_93'],column_sums['DIS_98'],column_sums['DIS_103'],column_sums['DIS_108'] ,column_sums['DIS_113'],column_sums['DIS_118'],column_sums['DIS_123'],column_sums['DIS_128'],column_sums['DIS_133'],column_sums['DIS_138'],column_sums['DIS_143'],column_sums['DIS_148'],column_sums['DIS_153'],column_sums['DIS_158']])
        total_sum_cat_vertical_4 =  sum([column_sums['DIS_4'],column_sums['DIS_9'], column_sums['DIS_14'],column_sums['DIS_19'],column_sums['DIS_24'],column_sums['DIS_29'],column_sums['DIS_34'],column_sums['DIS_39'],column_sums['DIS_44'],column_sums['DIS_49'],column_sums['DIS_54'],column_sums['DIS_59'],column_sums['DIS_64'],column_sums['DIS_69'],column_sums['DIS_74'],column_sums['DIS_79'],column_sums['DIS_84'],column_sums['DIS_89'],column_sums['DIS_94'],column_sums['DIS_99'],column_sums['DIS_104'],column_sums['DIS_109'] ,column_sums['DIS_114'],column_sums['DIS_119'],column_sums['DIS_124'],column_sums['DIS_129'],column_sums['DIS_134'],column_sums['DIS_139'],column_sums['DIS_144'],column_sums['DIS_149'],column_sums['DIS_154'],column_sums['DIS_159']])
        total_sum_cat_vertical_5 =  sum([column_sums['DIS_5'],column_sums['DIS_10'],column_sums['DIS_15'],column_sums['DIS_20'],column_sums['DIS_25'],column_sums['DIS_30'],column_sums['DIS_35'],column_sums['DIS_40'],column_sums['DIS_45'],column_sums['DIS_50'],column_sums['DIS_55'],column_sums['DIS_60'],column_sums['DIS_65'],column_sums['DIS_70'],column_sums['DIS_75'],column_sums['DIS_80'],column_sums['DIS_85'],column_sums['DIS_90'],column_sums['DIS_95'],column_sums['DIS_100'],column_sums['DIS_105'],column_sums['DIS_110'],column_sums['DIS_115'],column_sums['DIS_120'],column_sums['DIS_125'],column_sums['DIS_130'],column_sums['DIS_135'],column_sums['DIS_140'],column_sums['DIS_145'],column_sums['DIS_150'],column_sums['DIS_155'],column_sums['DIS_160']])

        sheet['E48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E48'] = total_sum_cat_vertical_1     
        
        sheet['F48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F48'] = total_sum_cat_vertical_2 
        
        sheet['G48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G48'] = total_sum_cat_vertical_3    
        
        sheet['H48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H48'] = total_sum_cat_vertical_4    
        
        sheet['I48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I48'] = total_sum_cat_vertical_5    
        ##########################################################################
        
        ###############################
        ###  DISCAPACIDAD SENSORIAL ###
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_sensorial_red:
            for col_sensorial in col_sum_sensorial:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_sensorial.keys()).index(col_sensorial) + 1
                    col_sum_sensorial[col_sensorial] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_sensorial, total_cell_sensorial in col_ubi_sensorial.items():
            if col_sensorial in col_sum_sensorial:
                # Obtener la celda correspondiente según la ubicación
                cell_sensorial = sheet[total_cell_sensorial]
                # Asignar el valor de la suma a la celda
                cell_sensorial.value = col_sum_sensorial[col_sensorial]
                # Aplicar formato a la celda
                cell_sensorial.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_sensorial.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_sensorial.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_sensorial['DIS_161'], col_sum_sensorial['DIS_162'], col_sum_sensorial['DIS_163'], col_sum_sensorial['DIS_164'], col_sum_sensorial['DIS_165']])
        t_sum_cat_2 =  sum([col_sum_sensorial['DIS_166'], col_sum_sensorial['DIS_167'], col_sum_sensorial['DIS_168'], col_sum_sensorial['DIS_169'], col_sum_sensorial['DIS_170']])
        t_sum_cat_3 =  sum([col_sum_sensorial['DIS_171'], col_sum_sensorial['DIS_172'], col_sum_sensorial['DIS_173'], col_sum_sensorial['DIS_174'], col_sum_sensorial['DIS_175']])
        t_sum_cat_4 =  sum([col_sum_sensorial['DIS_176'], col_sum_sensorial['DIS_177'], col_sum_sensorial['DIS_178'], col_sum_sensorial['DIS_179'], col_sum_sensorial['DIS_180']])
        t_sum_cat_5 =  sum([col_sum_sensorial['DIS_181'], col_sum_sensorial['DIS_182'], col_sum_sensorial['DIS_183'], col_sum_sensorial['DIS_184'], col_sum_sensorial['DIS_185']])
        t_sum_cat_6 =  sum([col_sum_sensorial['DIS_186'], col_sum_sensorial['DIS_187'], col_sum_sensorial['DIS_188'], col_sum_sensorial['DIS_189'], col_sum_sensorial['DIS_190']])
        t_sum_cat_7 =  sum([col_sum_sensorial['DIS_191'], col_sum_sensorial['DIS_192'], col_sum_sensorial['DIS_193'], col_sum_sensorial['DIS_194'], col_sum_sensorial['DIS_195']])
        
        sheet['D52'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D52'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D52'] = t_sum_cat_1     
        
        sheet['D53'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D53'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D53'] = t_sum_cat_2 
        
        sheet['D54'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D54'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D54'] = t_sum_cat_3    
        
        sheet['D55'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D55'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D55'] = t_sum_cat_4    
        
        sheet['D56'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D56'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D56'] = t_sum_cat_5    
        
        sheet['D57'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D57'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D57'] = t_sum_cat_6    
        
        sheet['D58'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D58'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D58'] = t_sum_cat_7    
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_sensorial['DIS_161'],col_sum_sensorial['DIS_166'],col_sum_sensorial['DIS_171'],col_sum_sensorial['DIS_176'],col_sum_sensorial['DIS_181'],col_sum_sensorial['DIS_186'],col_sum_sensorial['DIS_191']])
        t_sum_cat_vertical_2 =  sum([col_sum_sensorial['DIS_162'],col_sum_sensorial['DIS_167'],col_sum_sensorial['DIS_172'],col_sum_sensorial['DIS_177'],col_sum_sensorial['DIS_182'],col_sum_sensorial['DIS_187'],col_sum_sensorial['DIS_192']])
        t_sum_cat_vertical_3 =  sum([col_sum_sensorial['DIS_163'],col_sum_sensorial['DIS_168'],col_sum_sensorial['DIS_173'],col_sum_sensorial['DIS_178'],col_sum_sensorial['DIS_183'],col_sum_sensorial['DIS_188'],col_sum_sensorial['DIS_193']])
        t_sum_cat_vertical_4 =  sum([col_sum_sensorial['DIS_164'],col_sum_sensorial['DIS_169'],col_sum_sensorial['DIS_174'],col_sum_sensorial['DIS_179'],col_sum_sensorial['DIS_184'],col_sum_sensorial['DIS_189'],col_sum_sensorial['DIS_194']])
        t_sum_cat_vertical_5 =  sum([col_sum_sensorial['DIS_165'],col_sum_sensorial['DIS_170'],col_sum_sensorial['DIS_175'],col_sum_sensorial['DIS_180'],col_sum_sensorial['DIS_185'],col_sum_sensorial['DIS_190'],col_sum_sensorial['DIS_195']])
        
        sheet['E59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E59'] = t_sum_cat_vertical_1     
        
        sheet['F59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F59'] = t_sum_cat_vertical_2 
        
        sheet['G59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G59'] = t_sum_cat_vertical_3    
        
        sheet['H59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H59'] = t_sum_cat_vertical_4    
        
        sheet['I59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I59'] = t_sum_cat_vertical_5    
        ##########################################################################
                
        ###############################
        ###  DISCAPACIDAD MENTAL ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_mental_red:
            for col_mental in col_sum_mental:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_mental.keys()).index(col_mental) + 1
                    col_sum_mental[col_mental] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_mental, total_cell_mental in col_ubi_mental.items():
            if col_mental in col_sum_mental:
                # Obtener la celda correspondiente según la ubicación
                cell_mental = sheet[total_cell_mental]
                # Asignar el valor de la suma a la celda
                cell_mental.value = col_sum_mental[col_mental]
                # Aplicar formato a la celda
                cell_mental.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_mental.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_mental.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_mental['DIS_196'], col_sum_mental['DIS_197'], col_sum_mental['DIS_198'], col_sum_mental['DIS_199'], col_sum_mental['DIS_200']])
        t_sum_cat_2 =  sum([col_sum_mental['DIS_201'], col_sum_mental['DIS_202'], col_sum_mental['DIS_203'], col_sum_mental['DIS_204'], col_sum_mental['DIS_205']])
        t_sum_cat_3 =  sum([col_sum_mental['DIS_206'], col_sum_mental['DIS_207'], col_sum_mental['DIS_208'], col_sum_mental['DIS_209'], col_sum_mental['DIS_210']])
        t_sum_cat_4 =  sum([col_sum_mental['DIS_211'], col_sum_mental['DIS_212'], col_sum_mental['DIS_213'], col_sum_mental['DIS_214'], col_sum_mental['DIS_215']])
        
        sheet['D63'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D63'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D63'] = t_sum_cat_1     
        
        sheet['D64'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D64'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D64'] = t_sum_cat_2 
        
        sheet['D65'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D65'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D65'] = t_sum_cat_3    
        
        sheet['D66'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D66'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D66'] = t_sum_cat_4    

        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_mental['DIS_196'],col_sum_mental['DIS_201'],col_sum_mental['DIS_206'],col_sum_mental['DIS_211']])
        t_sum_cat_vertical_2 =  sum([col_sum_mental['DIS_197'],col_sum_mental['DIS_202'],col_sum_mental['DIS_207'],col_sum_mental['DIS_212']])
        t_sum_cat_vertical_3 =  sum([col_sum_mental['DIS_198'],col_sum_mental['DIS_203'],col_sum_mental['DIS_208'],col_sum_mental['DIS_213']])
        t_sum_cat_vertical_4 =  sum([col_sum_mental['DIS_199'],col_sum_mental['DIS_204'],col_sum_mental['DIS_209'],col_sum_mental['DIS_214']])
        t_sum_cat_vertical_5 =  sum([col_sum_mental['DIS_200'],col_sum_mental['DIS_205'],col_sum_mental['DIS_210'],col_sum_mental['DIS_215']])
        
        sheet['E67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E67'] = t_sum_cat_vertical_1     
        
        sheet['F67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F67'] = t_sum_cat_vertical_2 
        
        sheet['G67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G67'] = t_sum_cat_vertical_3    
        
        sheet['H67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H67'] = t_sum_cat_vertical_4    
        
        sheet['I67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I67'] = t_sum_cat_vertical_5    
        ##########################################################################
        
        #################################
        ###  DISCAPACIDAD CERTIFICADO ###
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_certificado_red:
            for col_certificado in col_sum_certificado:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_certificado.keys()).index(col_certificado) + 1
                    col_sum_certificado[col_certificado] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_certificado, total_cell_certificado in col_ubi_certificado.items():
            if col_certificado in col_sum_certificado:
                # Obtener la celda correspondiente según la ubicación
                cell_certificado = sheet[total_cell_certificado]
                # Asignar el valor de la suma a la celda
                cell_certificado.value = col_sum_certificado[col_certificado]
                # Aplicar formato a la celda
                cell_certificado.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_certificado.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_certificado.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        # Sumar los valores del diccionario      
        t_sum_cat_cert_1 =  sum([col_sum_certificado['DIS_216'], col_sum_certificado['DIS_217'], col_sum_certificado['DIS_218'], col_sum_certificado['DIS_219'], col_sum_certificado['DIS_220']])
        t_sum_cat_cert_2 =  sum([col_sum_certificado['DIS_221'], col_sum_certificado['DIS_222'], col_sum_certificado['DIS_223'], col_sum_certificado['DIS_224'], col_sum_certificado['DIS_225']])
        t_sum_cat_cert_3 =  sum([col_sum_certificado['DIS_226'], col_sum_certificado['DIS_227'], col_sum_certificado['DIS_228'], col_sum_certificado['DIS_229'], col_sum_certificado['DIS_230']])
        t_sum_cat_cert_4 =  sum([col_sum_certificado['DIS_231'], col_sum_certificado['DIS_232'], col_sum_certificado['DIS_233'], col_sum_certificado['DIS_234'], col_sum_certificado['DIS_235']])

        sheet['D71'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D71'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D71'] = t_sum_cat_cert_1     
        
        sheet['D72'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D72'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D72'] = t_sum_cat_cert_2 
        
        sheet['D73'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D73'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D73'] = t_sum_cat_cert_3 
        
        sheet['D74'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D74'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D74'] = t_sum_cat_cert_4 
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vert_1 =  sum([col_sum_certificado['DIS_216'],col_sum_certificado['DIS_221'],col_sum_certificado['DIS_226'],col_sum_certificado['DIS_231']])
        t_sum_cat_vert_2 =  sum([col_sum_certificado['DIS_217'],col_sum_certificado['DIS_222'],col_sum_certificado['DIS_227'],col_sum_certificado['DIS_232']])
        t_sum_cat_vert_3 =  sum([col_sum_certificado['DIS_218'],col_sum_certificado['DIS_223'],col_sum_certificado['DIS_228'],col_sum_certificado['DIS_233']])
        t_sum_cat_vert_4 =  sum([col_sum_certificado['DIS_219'],col_sum_certificado['DIS_224'],col_sum_certificado['DIS_229'],col_sum_certificado['DIS_234']])
        t_sum_cat_vert_5 =  sum([col_sum_certificado['DIS_220'],col_sum_certificado['DIS_225'],col_sum_certificado['DIS_230'],col_sum_certificado['DIS_235']])
        
        sheet['E75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E75'] = t_sum_cat_vert_1     
        
        sheet['F75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F75'] = t_sum_cat_vert_2 
        
        sheet['G75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G75'] = t_sum_cat_vert_3    
        
        sheet['H75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H75'] = t_sum_cat_vert_4    
        
        sheet['I75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I75'] = t_sum_cat_vert_5    
        
        #################################
        ###  DISCAPACIDAD RBC ###########
        #################################       
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_rbc_red:
            for col_rbc in col_sum_rbc:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_rbc.keys()).index(col_rbc) + 1
                    col_sum_rbc[col_rbc] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
                    
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_rbc, total_cell_rbc in col_ubi_rbc.items():
            if col_rbc in col_sum_rbc:
                # Obtener la celda correspondiente según la ubicación
                cell_rbc = sheet[total_cell_rbc]
                # Asignar el valor de la suma a la celda
                cell_rbc.value = col_sum_rbc[col_rbc]
                # Aplicar formato a la celda
                cell_rbc.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_rbc.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_rbc.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        ## Sumar los valores del diccionario      
        t_sum_cat_rbc_1 =  sum([col_sum_rbc['DIS_242'], col_sum_rbc['DIS_243'], col_sum_rbc['DIS_244'], col_sum_rbc['DIS_245'], col_sum_rbc['DIS_246']])
        t_sum_cat_rbc_2 =  sum([col_sum_rbc['DIS_247'], col_sum_rbc['DIS_248'], col_sum_rbc['DIS_249'], col_sum_rbc['DIS_250'], col_sum_rbc['DIS_251']])
        t_sum_cat_rbc_3 =  sum([col_sum_rbc['DIS_252'], col_sum_rbc['DIS_253'], col_sum_rbc['DIS_254'], col_sum_rbc['DIS_255'], col_sum_rbc['DIS_256']])
        t_sum_cat_rbc_4 =  sum([col_sum_rbc['DIS_257'], col_sum_rbc['DIS_258'], col_sum_rbc['DIS_259'], col_sum_rbc['DIS_260'], col_sum_rbc['DIS_261']])
        t_sum_cat_rbc_5 =  sum([col_sum_rbc['DIS_262'], col_sum_rbc['DIS_263'], col_sum_rbc['DIS_264'], col_sum_rbc['DIS_265'], col_sum_rbc['DIS_266']])

        sheet['D86'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D86'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D86'] = t_sum_cat_rbc_1     
        
        sheet['D87'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D87'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D87'] = t_sum_cat_rbc_2 
        
        sheet['D88'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D88'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D88'] = t_sum_cat_rbc_3     
        
        sheet['D89'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D89'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D89'] = t_sum_cat_rbc_4 
        
        sheet['D90'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D90'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D90'] = t_sum_cat_rbc_5 
        
        # Sumar los valores del VERTICAL      
        t_sum_vert_rbc_1 =  sum([col_sum_rbc['DIS_242'],col_sum_rbc['DIS_247'],col_sum_rbc['DIS_252'],col_sum_rbc['DIS_257'],col_sum_rbc['DIS_262']])
        t_sum_vert_rbc_2 =  sum([col_sum_rbc['DIS_243'],col_sum_rbc['DIS_248'],col_sum_rbc['DIS_253'],col_sum_rbc['DIS_258'],col_sum_rbc['DIS_263']])
        t_sum_vert_rbc_3 =  sum([col_sum_rbc['DIS_244'],col_sum_rbc['DIS_249'],col_sum_rbc['DIS_254'],col_sum_rbc['DIS_259'],col_sum_rbc['DIS_264']])
        t_sum_vert_rbc_4 =  sum([col_sum_rbc['DIS_245'],col_sum_rbc['DIS_250'],col_sum_rbc['DIS_255'],col_sum_rbc['DIS_260'],col_sum_rbc['DIS_265']])
        t_sum_vert_rbc_5 =  sum([col_sum_rbc['DIS_246'],col_sum_rbc['DIS_251'],col_sum_rbc['DIS_256'],col_sum_rbc['DIS_261'],col_sum_rbc['DIS_266']])
        
        sheet['E91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E91'] = t_sum_vert_rbc_1
        
        sheet['F91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F91'] = t_sum_vert_rbc_2 
        
        sheet['G91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G91'] = t_sum_vert_rbc_3    
        
        sheet['H91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H91'] = t_sum_vert_rbc_4    
        
        sheet['I91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I91'] = t_sum_vert_rbc_5   
        
        #################################
        ###  CAPACITACION PERSONAL ######
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_capacitacion_red:
            for col_capacitacion in col_sum_capacitacion:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_capacitacion.keys()).index(col_capacitacion) + 1
                    col_sum_capacitacion[col_capacitacion] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_capacitacion, total_cell_capacitacion in col_ubi_capacitacion.items():
            if col_capacitacion in col_sum_capacitacion:
                # Obtener la celda correspondiente según la ubicación
                cell_capacitacion = sheet[total_cell_capacitacion]
                # Asignar el valor de la suma a la celda
                cell_capacitacion.value = col_sum_capacitacion[col_capacitacion]
                # Aplicar formato a la celda
                cell_capacitacion.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_capacitacion.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_capacitacion.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_capacitacion['DIS_273']])
        t_sum_cat_2 = sum([col_sum_capacitacion['DIS_274']])
        
        sheet['D12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D12'] = t_sum_cat_1     
        
        sheet['E12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E12'] = t_sum_cat_2 
        
        ###############################
        ###  CAPACITACION AGENTE ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_agente_red:
            for col_agente in col_sum_agente:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_agente.keys()).index(col_agente) + 1
                    col_sum_agente[col_agente] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_agente, total_cell_agente in col_ubi_agente.items():
            if col_agente in col_sum_agente:
                # Obtener la celda correspondiente según la ubicación
                cell_agente = sheet[total_cell_agente]
                # Asignar el valor de la suma a la celda
                cell_agente.value = col_sum_agente[col_agente]
                # Aplicar formato a la celda
                cell_agente.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_agente.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_agente.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_agente['DIS_236']])
        t_sum_cat_2 = sum([col_sum_agente['DIS_237']])
        t_sum_cat_3 = sum([col_sum_agente['DIS_238']])
        t_sum_cat_4 = sum([col_sum_agente['DIS_239']])
        t_sum_cat_5 = sum([col_sum_agente['DIS_240']])
        t_sum_cat_6 = sum([col_sum_agente['DIS_241']])
        
        sheet['D82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D82'] = t_sum_cat_1     
        
        sheet['E82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E82'] = t_sum_cat_2 
        
        sheet['F82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F82'] = t_sum_cat_3
        
        sheet['G82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G82'] = t_sum_cat_4 
        
        sheet['H82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H82'] = t_sum_cat_5
        
        sheet['I82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I82'] = t_sum_cat_6 
        
        ############################
        ###  CAPACITACION COMITE ###
        #############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_comite_red:
            for col_comite in col_sum_comite:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_comite.keys()).index(col_comite) + 1
                    col_sum_comite[col_comite] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_comite, total_cell_comite in col_ubi_comite.items():
            if col_comite in col_sum_comite:
                # Obtener la celda correspondiente según la ubicación
                cell_comite = sheet[total_cell_comite]
                # Asignar el valor de la suma a la celda
                cell_comite.value = col_sum_comite[col_comite]
                # Aplicar formato a la celda
                cell_comite.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_comite.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_comite.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_comite['DIS_267']])
        t_sum_cat_2 = sum([col_sum_comite['DIS_268']])
        t_sum_cat_3 = sum([col_sum_comite['DIS_269']])
        t_sum_cat_4 = sum([col_sum_comite['DIS_270']])
        t_sum_cat_5 = sum([col_sum_comite['DIS_271']])
        t_sum_cat_6 = sum([col_sum_comite['DIS_272']])
        
        sheet['D97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D97'] = t_sum_cat_1     
        
        sheet['E97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E97'] = t_sum_cat_2 
        
        sheet['F97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F97'] = t_sum_cat_3
        
        sheet['G97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G97'] = t_sum_cat_4 
        
        sheet['H97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H97'] = t_sum_cat_5
        
        sheet['I97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I97'] = t_sum_cat_6 
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_operacional_redes.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        workbook.save(response)

        return response


################################################
# REPORTE POR MICRO-REDES
################################################
def get_microredes(request, microredes_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }

    return render(request, 'discapacidad/microredes.html', context)

def p_microredes(request):
    redes_param = request.GET.get('redes')
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'redes_param': redes_param,
        'microredes': microredes
    }
    return render(request, 'discapacidad/partials/p_microredes.html', context)

def rpt_operacional_fisico_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,                
                    SUM(dis_1) AS dis_1,
                    SUM(dis_2) AS dis_2,
                    SUM(dis_3) AS dis_3,
                    SUM(dis_4) AS dis_4,
                    SUM(dis_5) AS dis_5,
                    SUM(dis_6) AS dis_6,
                    SUM(dis_7) AS dis_7,
                    SUM(dis_8) AS dis_8,
                    SUM(dis_9) AS dis_9,
                    SUM(dis_10) AS dis_10,
                    SUM(dis_11) AS dis_11,
                    SUM(dis_12) AS dis_12,
                    SUM(dis_13) AS dis_13,
                    SUM(dis_14) AS dis_14,
                    SUM(dis_15) AS dis_15,
                    SUM(dis_16) AS dis_16,
                    SUM(dis_17) AS dis_17,
                    SUM(dis_18) AS dis_18,
                    SUM(dis_19) AS dis_19,
                    SUM(dis_20) AS dis_20,
                    SUM(dis_21) AS dis_21,
                    SUM(dis_22) AS dis_22,
                    SUM(dis_23) AS dis_23,
                    SUM(dis_24) AS dis_24,
                    SUM(dis_25) AS dis_25,
                    SUM(dis_26) AS dis_26,
                    SUM(dis_27) AS dis_27,
                    SUM(dis_28) AS dis_28,
                    SUM(dis_29) AS dis_29,
                    SUM(dis_30) AS dis_30,
                    SUM(dis_31) AS dis_31,
                    SUM(dis_32) AS dis_32,
                    SUM(dis_33) AS dis_33,
                    SUM(dis_34) AS dis_34,
                    SUM(dis_35) AS dis_35,
                    SUM(dis_36) AS dis_36,
                    SUM(dis_37) AS dis_37,
                    SUM(dis_38) AS dis_38,
                    SUM(dis_39) AS dis_39,
                    SUM(dis_40) AS dis_40,
                    SUM(dis_41) AS dis_41,
                    SUM(dis_42) AS dis_42,
                    SUM(dis_43) AS dis_43,
                    SUM(dis_44) AS dis_44,
                    SUM(dis_45) AS dis_45,
                    SUM(dis_46) AS dis_46,
                    SUM(dis_47) AS dis_47,
                    SUM(dis_48) AS dis_48,
                    SUM(dis_49) AS dis_49,
                    SUM(dis_50) AS dis_50,
                    SUM(dis_51) AS dis_51,
                    SUM(dis_52) AS dis_52,
                    SUM(dis_53) AS dis_53,
                    SUM(dis_54) AS dis_54,
                    SUM(dis_55) AS dis_55,
                    SUM(dis_56) AS dis_56,
                    SUM(dis_57) AS dis_57,
                    SUM(dis_58) AS dis_58,
                    SUM(dis_59) AS dis_59,
                    SUM(dis_60) AS dis_60,
                    SUM(dis_61) AS dis_61,
                    SUM(dis_62) AS dis_62,
                    SUM(dis_63) AS dis_63,
                    SUM(dis_64) AS dis_64,
                    SUM(dis_65) AS dis_65,
                    SUM(dis_66) AS dis_66,
                    SUM(dis_67) AS dis_67,
                    SUM(dis_68) AS dis_68,
                    SUM(dis_69) AS dis_69,
                    SUM(dis_70) AS dis_70,
                    SUM(dis_71) AS dis_71,
                    SUM(dis_72) AS dis_72,
                    SUM(dis_73) AS dis_73,
                    SUM(dis_74) AS dis_74,
                    SUM(dis_75) AS dis_75,
                    SUM(dis_76) AS dis_76,
                    SUM(dis_77) AS dis_77,
                    SUM(dis_78) AS dis_78,
                    SUM(dis_79) AS dis_79,
                    SUM(dis_80) AS dis_80,
                    SUM(dis_81) AS dis_81,
                    SUM(dis_82) AS dis_82,
                    SUM(dis_83) AS dis_83,
                    SUM(dis_84) AS dis_84,
                    SUM(dis_85) AS dis_85,
                    SUM(dis_86) AS dis_86,
                    SUM(dis_87) AS dis_87,
                    SUM(dis_88) AS dis_88,
                    SUM(dis_89) AS dis_89,
                    SUM(dis_90) AS dis_90,
                    SUM(dis_91) AS dis_91,
                    SUM(dis_92) AS dis_92,
                    SUM(dis_93) AS dis_93,
                    SUM(dis_94) AS dis_94,
                    SUM(dis_95) AS dis_95,
                    SUM(dis_96) AS dis_96,
                    SUM(dis_97) AS dis_97,
                    SUM(dis_98) AS dis_98,
                    SUM(dis_99) AS dis_99,
                    SUM(dis_100) AS dis_100,
                    SUM(dis_101) AS dis_101,
                    SUM(dis_102) AS dis_102,
                    SUM(dis_103) AS dis_103,
                    SUM(dis_104) AS dis_104,
                    SUM(dis_105) AS dis_105,
                    SUM(dis_106) AS dis_106,
                    SUM(dis_107) AS dis_107,
                    SUM(dis_108) AS dis_108,
                    SUM(dis_109) AS dis_109,
                    SUM(dis_110) AS dis_110,
                    SUM(dis_111) AS dis_111,
                    SUM(dis_112) AS dis_112,
                    SUM(dis_113) AS dis_113,
                    SUM(dis_114) AS dis_114,
                    SUM(dis_115) AS dis_115,
                    SUM(dis_116) AS dis_116,
                    SUM(dis_117) AS dis_117,
                    SUM(dis_118) AS dis_118,
                    SUM(dis_119) AS dis_119,
                    SUM(dis_120) AS dis_120,
                    SUM(dis_121) AS dis_121,
                    SUM(dis_122) AS dis_122,
                    SUM(dis_123) AS dis_123,
                    SUM(dis_124) AS dis_124,
                    SUM(dis_125) AS dis_125,
                    SUM(dis_126) AS dis_126,
                    SUM(dis_127) AS dis_127,
                    SUM(dis_128) AS dis_128,
                    SUM(dis_129) AS dis_129,
                    SUM(dis_130) AS dis_130, 
                    SUM(dis_131) AS dis_131,
                    SUM(dis_132) AS dis_132,
                    SUM(dis_133) AS dis_133,
                    SUM(dis_134) AS dis_134,
                    SUM(dis_135) AS dis_135,
                    SUM(dis_136) AS dis_136,
                    SUM(dis_137) AS dis_137,
                    SUM(dis_138) AS dis_138,
                    SUM(dis_139) AS dis_139,
                    SUM(dis_140) AS dis_140, 
                    SUM(dis_141) AS dis_141,
                    SUM(dis_142) AS dis_142,
                    SUM(dis_143) AS dis_143,
                    SUM(dis_144) AS dis_144,
                    SUM(dis_145) AS dis_145,
                    SUM(dis_146) AS dis_146,
                    SUM(dis_147) AS dis_147,
                    SUM(dis_148) AS dis_148,
                    SUM(dis_149) AS dis_149,
                    SUM(dis_150) AS dis_150,
                    SUM(dis_151) AS dis_151,
                    SUM(dis_152) AS dis_152,
                    SUM(dis_153) AS dis_153,
                    SUM(dis_154) AS dis_154,
                    SUM(dis_155) AS dis_155,
                    SUM(dis_156) AS dis_156,
                    SUM(dis_157) AS dis_157,
                    SUM(dis_158) AS dis_158,
                    SUM(dis_159) AS dis_159,
                    SUM(dis_160) AS dis_160 
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_1,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_2,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_3,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_4,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_5,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_6,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_7,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_8,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_9,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_10,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_11,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_12,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_13,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_14,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_15,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_16,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_17,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_18,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_19,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_20,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_21,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_22,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_23,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_24,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_25,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_26,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_27,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_28,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_29,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_30,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_31,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_32,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_33,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_34,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_35,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_36,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_37,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_38,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_39,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_40,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_41,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_42,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_43,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_44,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_45,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_46,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_47,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_48,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_49,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_50,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_51,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_52,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_53,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_54,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_55,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_56,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_57,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_58,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_59,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_60,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_61,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_62,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_63,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_64,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_65,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_66,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_67,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_68,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_69,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_70,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_71,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_72,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_73,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_74,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_75,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_76,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_77,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_78,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_79,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_80,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_81,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_82,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_83,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_84,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_85,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_86,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_87,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_88,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_89,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_90,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_91,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_92,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_93,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_94,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_95,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_96,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_97,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_98,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_99,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_100,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_101,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_102,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_103,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_104,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_105,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_106,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_107,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_108,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_109,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_110,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_111,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_112,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_113,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_114,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_115,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_116,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_117,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_118,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_119,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_120,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_121,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_122,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_123,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_124,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_125,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_126,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_127,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_128,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_129,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_130,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_131,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_132,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_133,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_134,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_135,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_136,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_137,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_138,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_139,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_140,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_141,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_142,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_143,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_144,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_145,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_146,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_147,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_148,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_149,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_150,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_151,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_152,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_153,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_154,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_155,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_156,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_157,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_158,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_159,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_160
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_microred = cursor.fetchall()
    return resultado_microred

def rpt_operacional_sensorial_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,   
                    SUM(dis_161) AS dis_161,
                    SUM(dis_162) AS dis_162,
                    SUM(dis_163) AS dis_163,
                    SUM(dis_164) AS dis_164,
                    SUM(dis_165) AS dis_165,
                    SUM(dis_166) AS dis_166,
                    SUM(dis_167) AS dis_167,
                    SUM(dis_168) AS dis_168,
                    SUM(dis_169) AS dis_169,
                    SUM(dis_170) AS dis_170,
                    SUM(dis_171) AS dis_171,
                    SUM(dis_172) AS dis_172,
                    SUM(dis_173) AS dis_173,
                    SUM(dis_174) AS dis_174,
                    SUM(dis_175) AS dis_175,
                    SUM(dis_176) AS dis_176,
                    SUM(dis_177) AS dis_177,
                    SUM(dis_178) AS dis_178,
                    SUM(dis_179) AS dis_179,
                    SUM(dis_180) AS dis_180,
                    SUM(dis_181) AS dis_181,
                    SUM(dis_182) AS dis_182,
                    SUM(dis_183) AS dis_183,
                    SUM(dis_184) AS dis_184,
                    SUM(dis_185) AS dis_185,
                    SUM(dis_186) AS dis_186,
                    SUM(dis_187) AS dis_187,
                    SUM(dis_188) AS dis_188,
                    SUM(dis_189) AS dis_189,
                    SUM(dis_190) AS dis_190,
                    SUM(dis_191) AS dis_191,
                    SUM(dis_192) AS dis_192,
                    SUM(dis_193) AS dis_193,
                    SUM(dis_194) AS dis_194,
                    SUM(dis_195) AS dis_195
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_161,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_162,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_163,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_164,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_165,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_166,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_167,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_168,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_169,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_170,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_171,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_172,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_173,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_174,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_175,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_176,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_177,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_178,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_179,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_180,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_181,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_182,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_183,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_184,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_185,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_186,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_187,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_188,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_189,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_190,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_191,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_192,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_193,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_194,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_195
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s  
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_sensorial_microred = cursor.fetchall()
    
    return resultado_sensorial_microred

def rpt_operacional_certificado_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,   
                    SUM(dis_216) AS dis_216,
                    SUM(dis_217) AS dis_217,
                    SUM(dis_218) AS dis_218,
                    SUM(dis_219) AS dis_219,
                    SUM(dis_220) AS dis_220,
                    SUM(dis_221) AS dis_221,
                    SUM(dis_222) AS dis_222,
                    SUM(dis_223) AS dis_223,
                    SUM(dis_224) AS dis_224,
                    SUM(dis_225) AS dis_225,
                    SUM(dis_226) AS dis_226,
                    SUM(dis_227) AS dis_227,
                    SUM(dis_228) AS dis_228,
                    SUM(dis_229) AS dis_229,
                    SUM(dis_230) AS dis_230,
                    SUM(dis_231) AS dis_231,
                    SUM(dis_232) AS dis_232,
                    SUM(dis_233) AS dis_233,
                    SUM(dis_234) AS dis_234,
                    SUM(dis_235) AS dis_235
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_216,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_217,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_218,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_219,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_220,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_221,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_222,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_223,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_224,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_225,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_226,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_227,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_228,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_229,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_230,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_231,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_232,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_233,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_234,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_235
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed,  TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_certificado_microred = cursor.fetchall()  
    return resultado_certificado_microred

def rpt_operacional_rbc_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,   
                    SUM(dis_242) AS dis_242,
                    SUM(dis_243) AS dis_243,
                    SUM(dis_244) AS dis_244,
                    SUM(dis_245) AS dis_245,
                    SUM(dis_246) AS dis_246,
                    SUM(dis_247) AS dis_247,
                    SUM(dis_248) AS dis_248,
                    SUM(dis_249) AS dis_249,
                    SUM(dis_250) AS dis_250,
                    SUM(dis_251) AS dis_251,
                    SUM(dis_252) AS dis_252,
                    SUM(dis_253) AS dis_253,
                    SUM(dis_254) AS dis_254,
                    SUM(dis_255) AS dis_255,
                    SUM(dis_256) AS dis_256,
                    SUM(dis_257) AS dis_257,
                    SUM(dis_258) AS dis_258,
                    SUM(dis_259) AS dis_259,
                    SUM(dis_260) AS dis_260, 
                    SUM(dis_261) AS dis_261, 
                    SUM(dis_262) AS dis_262, 
                    SUM(dis_263) AS dis_263, 
                    SUM(dis_264) AS dis_264, 
                    SUM(dis_265) AS dis_265, 
                    SUM(dis_266) AS dis_266
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_242,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_243,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_244,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_245,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_246,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_247,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_248,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_249,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_250,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_251,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_252,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_253,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_254,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_255,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_256,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_257,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_258,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_259,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_260,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_261,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_262,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_263,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_264,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_265,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_266
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_rbc_microred = cursor.fetchall()

    return resultado_rbc_microred

def rpt_operacional_mental_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,   
                    SUM(dis_196) AS dis_196,
                    SUM(dis_197) AS dis_197,
                    SUM(dis_198) AS dis_198,
                    SUM(dis_199) AS dis_199,
                    SUM(dis_200) AS dis_200,
                    SUM(dis_201) AS dis_201,
                    SUM(dis_202) AS dis_202,
                    SUM(dis_203) AS dis_203,
                    SUM(dis_204) AS dis_204,
                    SUM(dis_205) AS dis_205,
                    SUM(dis_206) AS dis_206,
                    SUM(dis_207) AS dis_207,
                    SUM(dis_208) AS dis_208,
                    SUM(dis_209) AS dis_209,
                    SUM(dis_210) AS dis_210, 
                    SUM(dis_211) AS dis_211, 
                    SUM(dis_212) AS dis_212, 
                    SUM(dis_213) AS dis_213, 
                    SUM(dis_214) AS dis_214, 
                    SUM(dis_215) AS dis_215
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_196,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_197,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_198,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_199,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_200,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_201,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_202,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_203,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_204,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_205,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_206,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_207,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_208,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_209,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_210,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_211,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_212,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_213,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_214,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_215
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_mental_microred = cursor.fetchall()
    return resultado_mental_microred

def rpt_operacional_capacitacion_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,   
                    SUM(dis_273) AS dis_273,
                    SUM(dis_274) AS dis_274
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        COUNT(Categoria) AS dis_273,
                        SUM(gedad) AS dis_274
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_capacitacion_microred = cursor.fetchall()

    return resultado_capacitacion_microred

def rpt_operacional_agente_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,   
                    SUM(dis_236) AS dis_236,
                    SUM(dis_237) AS dis_237,
                    SUM(dis_238) AS dis_238,
                    SUM(dis_239) AS dis_239,
                    SUM(dis_240) AS dis_240,
                    SUM(dis_241) AS dis_241
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 THEN 1 ELSE 0 END) 	   AS dis_236,
                        SUM(CASE WHEN Categoria = 1 THEN gedad ELSE 0 END) AS dis_237,
                        SUM(CASE WHEN Categoria = 2 THEN 1 ELSE 0 END)     AS dis_238,
                        SUM(CASE WHEN Categoria = 2 THEN gedad ELSE 0 END) AS dis_239,
                        SUM(CASE WHEN Categoria = 3 THEN 1 ELSE 0 END)     AS dis_240,
                        SUM(CASE WHEN Categoria = 3 THEN gedad ELSE 0 END) AS dis_241
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_agente_microred = cursor.fetchall()

    return resultado_agente_microred

def rpt_operacional_comite_microred(codigo_red,codigo_microred,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    codigo_red,
                    red,  
                    codigo_microred,  
                    microred,   
                    SUM(dis_267) AS dis_267,
                    SUM(dis_268) AS dis_268,
                    SUM(dis_269) AS dis_269,
                    SUM(dis_270) AS dis_270,
                    SUM(dis_271) AS dis_271,
                    SUM(dis_272) AS dis_272
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red AS codigo_red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Red AS red,
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Microred AS codigo_microred,
                        MAESTRO_HIS_ESTABLECIMIENTO.MicroRed AS microred,
                        renaes,
                        SUM(CASE WHEN Actividad = 1 THEN 1 ELSE 0 END) 		AS dis_267,
                        SUM(CASE WHEN Actividad = 1 THEN Partic ELSE 0 END) AS dis_268,
                        SUM(CASE WHEN Actividad = 2 THEN 1 ELSE 0 END)      AS dis_269,
                        SUM(CASE WHEN Actividad = 2 THEN Partic ELSE 0 END) AS dis_270,
                        SUM(CASE WHEN Actividad = 3 THEN 1 ELSE 0 END)      AS dis_271,
                        SUM(CASE WHEN Actividad = 3 THEN Partic ELSE 0 END) AS dis_272
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_red = %s AND MAESTRO_HIS_ESTABLECIMIENTO.codigo_microred = %s  
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Red, MAESTRO_HIS_ESTABLECIMIENTO.Red, MAESTRO_HIS_ESTABLECIMIENTO.Codigo_MicroRed, MAESTRO_HIS_ESTABLECIMIENTO.MicroRed, TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_red, red, codigo_microred, microred
                """, [str(codigo_red)[:2], str(codigo_microred)[:2], str(fecha_inicio) + '01', str(fecha_fin) + '31'])

        resultado_comite_microred = cursor.fetchall()

    return resultado_comite_microred

class RptOperacinalMicroRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        red = request.GET.get('redes')
        microred = request.GET.get('microredes')

        # Creación de la consulta
        resultado_microred = rpt_operacional_fisico_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_sensorial_microred = rpt_operacional_sensorial_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_certificado_microred = rpt_operacional_certificado_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_rbc_microred = rpt_operacional_rbc_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_mental_microred = rpt_operacional_mental_microred(red, microred,fecha_inicio, fecha_fin)
        resultado_capacitacion_microred = rpt_operacional_capacitacion_microred(red, microred,fecha_inicio, fecha_fin)
        resultado_agente_microred = rpt_operacional_agente_microred(red, microred, fecha_inicio, fecha_fin)
        resultado_comite_microred = rpt_operacional_comite_microred(red, microred, fecha_inicio, fecha_fin)
        
        microred_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Codigo_Red=red,Codigo_MicroRed=microred
        ).values_list('MicroRed', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
        
        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambia el alto de la columna
        sheet.row_dimensions[1].height = 14
        sheet.row_dimensions[2].height = 14
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[15].height = 25
        # cambia el ancho de la columna
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 28
        sheet.column_dimensions['C'].width = 28
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 9
        sheet.column_dimensions['F'].width = 9
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 9
        sheet.column_dimensions['L'].width = 9
        # linea de division
        sheet.freeze_panes = 'AL8'
        
        # Configuración del fondo y el borde
        fill = PatternFill(patternType='solid', fgColor='00B0F0')
        border = Border(left=Side(style='thin', color='00B0F0'),
                        right=Side(style='thin', color='00B0F0'),
                        top=Side(style='thin', color='00B0F0'),
                        bottom=Side(style='thin', color='00B0F0'))

        borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                        right=Side(style='thin', color='A9A9A9'), # Plomo
                        top=Side(style='thin', color='A9A9A9'), # Plomo
                        bottom=Side(style='thin', color='A9A9A9')) # Plomo

        # crea titulo del reporte
        sheet['B1'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B1'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
        
        sheet['B2'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B2'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
        
        sheet['B4'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B4'].font = Font(name = 'Arial', size= 12, bold = True)
        sheet['B4'] = 'REPORTE DE ACTIVIDADES DEL COMPONENTE DE DISCAPACIDAD'
        
        sheet['B6'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B6'] ='DIRESA / GERESA / DISA'
        
        sheet['C6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C6'].font = Font(name = 'Arial', size= 7)
        sheet['C6'] ='JUNIN'

        sheet['B7'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B7'] ='PROV/ DIST/ RED/ MR/ ESTABLEC'
        
        sheet['C7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C7'].font = Font(name = 'Arial', size= 7)
        sheet['C7'] = microred_codigo[0]
        
        sheet['E6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['E6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['E6'] ='PERIODO'
        
        sheet['F6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F6'] ='MES INICIO'
        
        sheet['F7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F7'] ='MES FIN'
        
        sheet['G6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G6'].font = Font(name = 'Arial', size= 8)
        sheet['G6'] = fecha_inicio_codigo[0]
        
        sheet['G7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G7'].font = Font(name = 'Arial', size= 8)
        sheet['G7'] = fecha_fin_codigo[0]
        
        sheet['B9'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B9'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B9'] ='PERSONAS CON DISCAPACIDAD RECIBEN ATENCION DE REHABILITACION EN ESTABLECIMIENTOS DE SALUD (3000688)'
        
        sheet['B10'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B10'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B10'] ='Capacitación en medicina de rehabilitación integral (5004449)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=12, max_row=12, min_col=3, max_col=5):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['C12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C12'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['C12'] ='Capacitación  (C0009)' 
        
        sheet['D11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D11'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D11'].fill = fill
        sheet['D11'].border = border
        sheet['D11'] = 'N°'
                
        sheet['E11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E11'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['E11'].fill = fill
        sheet['E11'].border = border
        sheet['E11'] = 'Capacitados'
        #######################################################
        ########## DISCAPACIDAD FISICA ########################
        #######################################################
        sheet['B14'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B14'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B14'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Física (5005150)' 
                
        sheet['B15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B15'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B15'].fill = fill
        sheet['B15'].border = border
        sheet['B15'] = 'Atenciones'
        
        sheet['D15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D15'].fill = fill
        sheet['D15'].border = border
        sheet['D15'] = 'Total'
        
        sheet['E15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E15'].fill = fill
        sheet['E15'].border = border
        sheet['E15'] = 'Niños         (1d - 11a)'
        
        sheet['F15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F15'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F15'].fill = fill
        sheet['F15'].border = border
        sheet['F15'] = 'Adolescentes (12a - 17a)'
        
        sheet['G15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G15'].fill = fill
        sheet['G15'].border = border
        sheet['G15'] = 'Jóvenes (18a - 29a)'
        
        sheet['H15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H15'].fill = fill
        sheet['H15'].border = border
        sheet['H15'] = 'Adultos (30a - 59a)'
        
        sheet['I15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I15'].fill = fill
        sheet['I15'].border = border
        sheet['I15'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=16, max_row=47, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B16'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B16'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B16'] ='LESIONES MEDULARES' 
                
        sheet['B17'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B17'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B17'] ='ENFERMEDAD DE PARKINSON Y SIMILARES' 
        
        sheet['B18'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B18'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B18'] ='REHABILITACIÓN EN PACIENTES AMPUTADOS' 
                
        sheet['B20'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B20'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B20'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA NEUROLÓGICA' 
        
        sheet['B23'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B23'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B23'] ='TRASTORNOS DEL DESARROLLO DE LA FUNCIÓN MOTRIZ' 
        
        sheet['B24'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B24'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B24'] ='ATENCIÓN DE REHABILITACIÓN DE ENFERMEDAD ARTICULAR DEGENERATIVA' 
        
        sheet['B25'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B25'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B25'] ='ENCEFALOPATÍA INFANTIL' 
                
        sheet['B26'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B26'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B26'] ='SÍNDROME DOWN' 
        
        sheet['B27'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B27'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B27'] ='REHABILITACIÓN EN PATOLOGÍA DE LA COLUMNA VERTEBRAL Y OTROS TRASTORNOS POSTURALES' 
        
        sheet['B34'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B34'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B34'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD CARDIOVASCULAR' 
        
        sheet['B35'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B35'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B35'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD RESPIRATORIA' 
        
        sheet['B36'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B36'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B36'] ='ATENCIÓN DE REHABILITACIÓN EN ALTERACIONES DEL PISO PÉLVICO' 
        
        sheet['B37'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B37'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B37'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA TRAUMATOLÓGICA Y REUMATOLÓGICA' 
        
        sheet['B44'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B44'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B44'] ='ATENCIÓN DE REHABILITACIÓN ONCOLÓGICA' 
        
        sheet['B46'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B46'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B46'] ='ATENCIÓN DE REHABILITACIÓN EN DOLOR' 
        
        sheet['B47'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B47'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B47'] ='ATENCIÓN DE REHABILITACIÓN EN PACIENTES QUEMADOS' 
        ####     
        sheet['C16'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C16'].font = Font(name = 'Arial', size= 7)
        sheet['C16'] ='Lesiones medulares' 
    
        sheet['C17'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C17'].font = Font(name = 'Arial', size= 7)
        sheet['C17'] ='Enfermedad de Parkinson y similares' 
        
        sheet['C18'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C18'].font = Font(name = 'Arial', size= 7)
        sheet['C18'] ='Amputados de miembros superiores' 
        
        sheet['C19'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C19'].font = Font(name = 'Arial', size= 7)
        sheet['C19'] ='Amputados de miembros inferiores' 
        
        sheet['C20'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C20'].font = Font(name = 'Arial', size= 7)
        sheet['C20'] ='Enfermedades cerebrovasculares'
        
        sheet['C21'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C21'].font = Font(name = 'Arial', size= 7)
        sheet['C21'] ='Enfermedades musculares y de la unión mioneural'
        
        sheet['C22'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C22'].font = Font(name = 'Arial', size= 7)
        sheet['C22'] ='Lesiones de nervios periféricos'
        
        sheet['C23'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C23'].font = Font(name = 'Arial', size= 7)
        sheet['C23'] ='Trastornos del desarrollo de la funcion motriz'
        
        sheet['C24'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C24'].font = Font(name = 'Arial', size= 7)
        sheet['C24'] ='Enfermedad articular degenerativa'
        
        sheet['C25'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C25'].font = Font(name = 'Arial', size= 7)
        sheet['C25'] ='Encefalopatía infantil y otras lesiones'
        
        sheet['C26'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C26'].font = Font(name = 'Arial', size= 7)
        sheet['C26'] ='Sindrome de Down'
        
        sheet['C27'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C27'].font = Font(name = 'Arial', size= 7)
        sheet['C27'] ='Cifosis y lordosis'
        
        sheet['C28'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C28'].font = Font(name = 'Arial', size= 7)
        sheet['C28'] ='Espondilo artropatías'
        
        sheet['C29'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C29'].font = Font(name = 'Arial', size= 7)
        sheet['C29'] ='Otros trastornos de los discos intervertebrales'
        
        sheet['C30'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C30'].font = Font(name = 'Arial', size= 7)
        sheet['C30'] ='Cervicalgia, dorsalgia, lumbago'
        
        sheet['C31'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C31'].font = Font(name = 'Arial', size= 7)
        sheet['C31'] ='Otras dorsopatías deformantes'
        
        sheet['C32'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C32'].font = Font(name = 'Arial', size= 7)
        sheet['C32'] ='Otros trastornos articulares'
        
        sheet['C33'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C33'].font = Font(name = 'Arial', size= 7)
        sheet['C33'] ='Defectos en la longitud de extremidades'
        
        sheet['C34'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C34'].font = Font(name = 'Arial', size= 7)
        sheet['C34'] ='Enfermedad cardiovascular'
        
        sheet['C35'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C35'].font = Font(name = 'Arial', size= 7)
        sheet['C35'] ='Enfermedad respiratoria'
        
        sheet['C36'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C36'].font = Font(name = 'Arial', size= 7)
        sheet['C36'] ='Vejiga neurogénica y dolor'
        
        sheet['C37'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C37'].font = Font(name = 'Arial', size= 7)
        sheet['C37'] ='Incontinencia'
        
        sheet['C38'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C38'].font = Font(name = 'Arial', size= 7)
        sheet['C38'] ='Prolapso'
        
        sheet['C39'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C39'].font = Font(name = 'Arial', size= 7)
        sheet['C39'] ='Traumatismos'
        
        sheet['C40'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C40'].font = Font(name = 'Arial', size= 7)
        sheet['C40'] ='Enfermedades del tejido conectivo'
        
        sheet['C41'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C41'].font = Font(name = 'Arial', size= 7)
        sheet['C41'] ='Patología articular excluida columna'
        
        sheet['C42'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C42'].font = Font(name = 'Arial', size= 7)
        sheet['C42'] ='Lesiones infecciosas'
        
        sheet['C43'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C43'].font = Font(name = 'Arial', size= 7)
        sheet['C43'] ='Lesión biomecánica'
        
        sheet['C44'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C44'].font = Font(name = 'Arial', size= 7)
        sheet['C44'] ='Linfedema'
        
        sheet['C45'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C45'].font = Font(name = 'Arial', size= 7)
        sheet['C45'] ='Sarcopenia'
        
        sheet['C46'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C46'].font = Font(name = 'Arial', size= 7)
        sheet['C46'] ='Dolor'
        
        sheet['C47'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C47'].font = Font(name = 'Arial', size= 7)
        sheet['C47'] ='Quemaduras, corrosiones y congelaciones'
        
        ##########################################################    
        ########## DISCAPACIDAD SENSORIAL ########################
        ##########################################################
        sheet['B50'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B50'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B50'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Sensorial (5005151)' 
                
        sheet['B51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B51'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B51'].fill = fill
        sheet['B51'].border = border
        sheet['B51'] = 'Atenciones'
        
        sheet['D51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D51'].fill = fill
        sheet['D51'].border = border
        sheet['D51'] = 'Total'
        
        sheet['E51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E51'].fill = fill
        sheet['E51'].border = border
        sheet['E51'] = 'Niños         (1d - 11a)'
        
        sheet['F51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F51'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F51'].fill = fill
        sheet['F51'].border = border
        sheet['F51'] = 'Adolescentes (12a - 17a)'
        
        sheet['G51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G51'].fill = fill
        sheet['G51'].border = border
        sheet['G51'] = 'Jóvenes (18a - 29a)'
        
        sheet['H51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H51'].fill = fill
        sheet['H51'].border = border
        sheet['H51'] = 'Adultos (30a - 59a)'
        
        sheet['I51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I51'].fill = fill
        sheet['I51'].border = border
        sheet['I51'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=52, max_row=58, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B52'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B52'] ='HIPOACUSIA Y/O SORDERA' 
        
        sheet['B53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B53'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B53'] ='BAJA VISION Y/O CEGUERA' 
        
        sheet['B54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B54'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B54'] ='SORDOMUDEZ' 
        
        sheet['B55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B55'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B55'] ='ENFERMEDAD CEREBRO VASCULAR' 
        
        sheet['B56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B56'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B56'] ='TRASTORNOS ESPECIFICOS DEL DESARROLLO DEL HABLA Y LENGUAJE' 
        
        sheet['B57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B57'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B57'] ='DISARTRIA Y DISFAGIA' 
        
        sheet['B59'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B59'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B59'] ='SUB TOTAL' 
        
        ########               
        sheet['C52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C52'].font = Font(name = 'Arial', size= 7)
        sheet['C52'] ='Hipoacusia y sordera' 
        
        sheet['C53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C53'].font = Font(name = 'Arial', size= 7)
        sheet['C53'] ='Baja visión y ceguera' 
        
        sheet['C54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C54'].font = Font(name = 'Arial', size= 7)
        sheet['C54'] ='Sordomudez' 
        
        sheet['C55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C55'].font = Font(name = 'Arial', size= 7)
        sheet['C55'] ='Enfermedad Cerebro vascular' 
        
        sheet['C56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C56'].font = Font(name = 'Arial', size= 7)
        sheet['C56'] ='Trastornos específicos del desarrollo del habla y lenguaje' 
        
        sheet['C57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C57'].font = Font(name = 'Arial', size= 7)
        sheet['C57'] ='Disartria' 
        
        sheet['C58'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C58'].font = Font(name = 'Arial', size= 7)
        sheet['C58'] ='Disfagia' 
        
        ########################################################
        ########## DISCAPACIDAD MENTAL #########################
        ########################################################
        sheet['B61'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B61'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B61'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Mental (5005152)' 
                
        sheet['B62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B62'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B62'].fill = fill
        sheet['B62'].border = border
        sheet['B62'] = 'Atenciones'
        
        sheet['D62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D62'].fill = fill
        sheet['D62'].border = border
        sheet['D62'] = 'Total'
        
        sheet['E62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E62'].fill = fill
        sheet['E62'].border = border
        sheet['E62'] = 'Niños         (1d - 11a)'
        
        sheet['F62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F62'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F62'].fill = fill
        sheet['F62'].border = border
        sheet['F62'] = 'Adolescentes (12a - 17a)'
        
        sheet['G62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G62'].fill = fill
        sheet['G62'].border = border
        sheet['G62'] = 'Jóvenes (18a - 29a)'
        
        sheet['H62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H62'].fill = fill
        sheet['H62'].border = border
        sheet['H62'] = 'Adultos (30a - 59a)'
        
        sheet['I62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I62'].fill = fill
        sheet['I62'].border = border
        sheet['I62'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=63, max_row=66, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B63'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B63'] ='TRASTORNOS DE APRENDIZAJE' 
        
        sheet['B64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B64'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B64'] ='RETRASO MENTAL LEVE, MODERADO, SEVERO' 
        
        sheet['B65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B65'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B65'] ='TRASTORNOS DEL ESPECTRO AUTISTA' 
        
        sheet['B66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B66'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B66'] ='OTROS TRASTORNOS DE SALUD MENTAL' 
        
        sheet['B67'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B67'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B67'] ='SUB TOTAL' 
        
        ##########
        
        sheet['C63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C63'].font = Font(name = 'Arial', size= 7)
        sheet['C63'] ='Trastornos del aprendizaje' 
        
        sheet['C64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C64'].font = Font(name = 'Arial', size= 7)
        sheet['C64'] ='Retardo Mental: Leve, moderado, severo' 
        
        sheet['C65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C65'].font = Font(name = 'Arial', size= 7)
        sheet['C65'] ='Trastornos del espectro autista' 
        
        sheet['C66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C66'].font = Font(name = 'Arial', size= 7)
        sheet['C66'] ='Otras alteraciones de salud mental' 
                
        ##################################################
        ########## CERTIFICACION #########################
        ##################################################
        sheet['B69'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B69'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B69'] ='PERSONAS CON DISCAPACIDAD CERTIFICADAS EN ESTABLECIMIENTOS DE SALUD (3000689)' 
                
        sheet['B70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B70'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B70'].fill = fill
        sheet['B70'].border = border
        sheet['B70'] = 'Atenciones'
        
        sheet['D70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D70'].fill = fill
        sheet['D70'].border = border
        sheet['D70'] = 'Total'
        
        sheet['E70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E70'].fill = fill
        sheet['E70'].border = border
        sheet['E70'] = 'Niños         (1d - 11a)'
        
        sheet['F70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F70'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F70'].fill = fill
        sheet['F70'].border = border
        sheet['F70'] = 'Adolescentes (12a - 17a)'
        
        sheet['G70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G70'].fill = fill
        sheet['G70'].border = border
        sheet['G70'] = 'Jóvenes (18a - 29a)'
        
        sheet['H70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H70'].fill = fill
        sheet['H70'].border = border
        sheet['H70'] = 'Adultos (30a - 59a)'
        
        sheet['I70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I70'].fill = fill
        sheet['I70'].border = border
        sheet['I70'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=71, max_row=74, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B71'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B71'] ='Certificación de Discapacidad (0515204)' 
        
        sheet['B74'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B74'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B74'] ='Certificación de Incapacidad (0515205)' 
        
        sheet['B75'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B75'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B75'] ='SUB TOTAL' 
        
        sheet['C71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C71'].font = Font(name = 'Arial', size= 7)
        sheet['C71'] ='Evaluación' 
        
        sheet['C72'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C72'].font = Font(name = 'Arial', size= 7)
        sheet['C72'] ='Calificación' 
        
        sheet['C73'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C73'].font = Font(name = 'Arial', size= 7)
        sheet['C73'] ='Certificación' 

        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B77'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B77'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B77'] ='PERSONAS CON DISCAPACIDAD RECIBEN SERVICIOS DE REHABILITACIÓN BASADA EN LA COMUNIDAD (3000690)' 
        
        sheet['B78'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B78'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B78'] ='CAPACITACIÓN A AGENTES COMUNITARIOS EN REHABILITACIÓN BASADA EN LA COMUNIDAD (5005155)' 
        
        sheet['B82'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B82'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B82'] ='Capacitación a Agentes Comunitarios  (APP138)' 
        
        sheet['D80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D80'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D80'].fill = fill
        sheet['D80'].border = border
        sheet['D80'] = 'Taller'
        
        sheet['F80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F80'].fill = fill
        sheet['F80'].border = border
        sheet['F80'] = 'Sesion Educativa'
        
        sheet['H80'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H80'].fill = fill
        sheet['H80'].border = border
        sheet['H80'] = 'Sesion Demostrativa'
        
        sheet['D81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D81'].fill = fill
        sheet['D81'].border = border
        sheet['D81'] = 'N°'
        
        sheet['E81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E81'].fill = fill
        sheet['E81'].border = border
        sheet['E81'] = 'Capacitados'
        
        sheet['F81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F81'].fill = fill
        sheet['F81'].border = border
        sheet['F81'] = 'N°'
        
        sheet['G81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G81'].fill = fill
        sheet['G81'].border = border
        sheet['G81'] = 'Capacitados'
        
        sheet['H81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H81'].fill = fill
        sheet['H81'].border = border
        sheet['H81'] = 'N° '
        
        sheet['I81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I81'].fill = fill
        sheet['I81'].border = border
        sheet['I81'] = 'Capacitados'
        
        
        #borde plomo
        for row in sheet.iter_rows(min_row=82, max_row=82, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        ################################################
        ########## VISITAS RBC #########################
        ################################################
        sheet['B84'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B84'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B84'] ='Vistas a alas familias Rehabilitacion Basada en la Comunidad' 
                
        sheet['B85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B85'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B85'].fill = fill
        sheet['B85'].border = border
        sheet['B85'] = 'Visitas'
        
        sheet['D85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D85'].fill = fill
        sheet['D85'].border = border
        sheet['D85'] = 'Total'
        
        sheet['E85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E85'].fill = fill
        sheet['E85'].border = border
        sheet['E85'] = 'Niños         (1d - 11a)'
        
        sheet['F85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F85'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F85'].fill = fill
        sheet['F85'].border = border
        sheet['F85'] = 'Adolescentes (12a - 17a)'
        
        sheet['G85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G85'].fill = fill
        sheet['G85'].border = border
        sheet['G85'] = 'Jóvenes (18a - 29a)'
        
        sheet['H85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H85'].fill = fill
        sheet['H85'].border = border
        sheet['H85'] = 'Adultos (30a - 59a)'
        
        sheet['I85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I85'].fill = fill
        sheet['I85'].border = border
        sheet['I85'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=86, max_row=90, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = borde_plomo
        
        sheet['B86'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B86'].font = Font(name = 'Arial', size= 8)
        sheet['B86'] ='1º Visita' 
        
        sheet['B87'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B87'].font = Font(name = 'Arial', size= 8)
        sheet['B87'] ='2º Visita' 
        
        sheet['B88'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B88'].font = Font(name = 'Arial', size= 8)
        sheet['B88'] ='3º Visita' 
        
        sheet['B89'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B89'].font = Font(name = 'Arial', size= 8)
        sheet['B89'] ='4º a Visita (trazador)' 
        
        sheet['B90'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B90'].font = Font(name = 'Arial', size= 8)
        sheet['B90'] ='5º a + Visitas' 
        
        sheet['B91'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B91'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B91'] ='SUB TOTAL' 
        
        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B93'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B93'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B93'] ='Capacitación a Actores Sociales para la aplicación de la estrategia de Rehabilitación Basada en la Comunidad' 
                
        sheet['B94'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B94'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B94'] ='Actividades con Gobiernos Locales:' 
        
        sheet['B97'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B97'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B97'] ='Actividad con Comité Multisectorial (APP96)' 
        
        sheet['D95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D95'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D95'].fill = fill
        sheet['D95'].border = border
        sheet['D95'] = 'Taller'
        
        sheet['F95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F95'].fill = fill
        sheet['F95'].border = border
        sheet['F95'] = 'Sesion Educativa'
        
        sheet['H95'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H95'].fill = fill
        sheet['H95'].border = border
        sheet['H95'] = 'Sesion Demostrativa'
        
        sheet['D96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D96'].fill = fill
        sheet['D96'].border = border
        sheet['D96'] = 'N°'
        
        sheet['E96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E96'].fill = fill
        sheet['E96'].border = border
        sheet['E96'] = 'Capacitados'
        
        sheet['F96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F96'].fill = fill
        sheet['F96'].border = border
        sheet['F96'] = 'N°'
        
        sheet['G96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G96'].fill = fill
        sheet['G96'].border = border
        sheet['G96'] = 'Capacitados'
        
        sheet['H96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H96'].fill = fill
        sheet['H96'].border = border
        sheet['H96'] = 'N° '
        
        sheet['I96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I96'].fill = fill
        sheet['I96'].border = border
        sheet['I96'] = 'Capacitados'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=97, max_row=97, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        #############################################################################
        #############################################################################                
        # cambina celdas
        sheet.merge_cells('C6:D6')
        sheet.merge_cells('C7:E7')
        
        sheet.merge_cells('B18:B19')
        sheet.merge_cells('B20:B22')
        sheet.merge_cells('B27:B33')
        sheet.merge_cells('B37:B43')
        sheet.merge_cells('B44:B45')
        
        # sensorial
        sheet.merge_cells('B57:B58')
        
        sheet.merge_cells('B15:C15')
        sheet.merge_cells('B51:C51')
        
        # mental
        sheet.merge_cells('B62:C62')
        
        #certificado
        sheet.merge_cells('B70:C70')
        
        sheet.merge_cells('B71:B73')
        
        #RBC
        sheet.merge_cells('B85:C85')
        
        #capa
        sheet.merge_cells('D80:E80')
        sheet.merge_cells('F80:G80')
        sheet.merge_cells('H80:I80')

        sheet.merge_cells('D95:E95')
        sheet.merge_cells('F95:G95')
        sheet.merge_cells('H95:I95')
        
        #capacitacion
        sheet.merge_cells('B82:C82')
        sheet.merge_cells('B97:C97')
        
        #visita
        sheet.merge_cells('B86:C86')
        sheet.merge_cells('B87:C87')
        sheet.merge_cells('B88:C88')
        sheet.merge_cells('B89:C89')
        sheet.merge_cells('B90:C90')
        
        # Definir ubicaciones específicas para cada columna y su suma total
        columnas_ubicaciones = {
            'PROVINCIA': 'D10',
            'DIS_1': 'E16', 
            'DIS_2': 'F16',
            'DIS_3': 'G16',
            'DIS_4': 'H16',
            'DIS_5': 'I16',
            'DIS_6': 'E17',
            'DIS_7': 'F17',
            'DIS_8': 'G17',
            'DIS_9': 'H17',
            'DIS_10': 'I17',
            'DIS_11': 'E18',
            'DIS_12': 'F18',
            'DIS_13': 'G18',
            'DIS_14': 'H18',
            'DIS_15': 'I18',
            'DIS_16': 'E19',
            'DIS_17': 'F19',
            'DIS_18': 'G19',
            'DIS_19': 'H19',
            'DIS_20': 'I19',
            'DIS_21': 'E20',
            'DIS_22': 'F20',
            'DIS_23': 'G20',
            'DIS_24': 'H20',
            'DIS_25': 'I20',
            'DIS_26': 'E21',
            'DIS_27': 'F21',
            'DIS_28': 'G21',
            'DIS_29': 'H21',
            'DIS_30': 'I21',
            'DIS_31': 'E22',
            'DIS_32': 'F22',
            'DIS_33': 'G22',
            'DIS_34': 'H22',
            'DIS_35': 'I22',
            'DIS_36': 'E23',
            'DIS_37': 'F23',
            'DIS_38': 'G23',
            'DIS_39': 'H23',
            'DIS_40': 'I23',
            'DIS_41': 'E24',
            'DIS_42': 'F24',
            'DIS_43': 'G24',
            'DIS_44': 'H24',
            'DIS_45': 'I24',
            'DIS_46': 'E25',
            'DIS_47': 'F25',
            'DIS_48': 'G25',
            'DIS_49': 'H25',
            'DIS_50': 'I25',
            'DIS_51': 'E26',
            'DIS_52': 'F26',
            'DIS_53': 'G26',
            'DIS_54': 'H26',
            'DIS_55': 'I26',
            'DIS_56': 'E27',
            'DIS_57': 'F27',
            'DIS_58': 'G27',
            'DIS_59': 'H27',
            'DIS_60': 'I27',
            'DIS_61': 'E28',
            'DIS_62': 'F28',
            'DIS_63': 'G28',
            'DIS_64': 'H28',
            'DIS_65': 'I28',
            'DIS_66': 'E29',
            'DIS_67': 'F29',
            'DIS_68': 'G29',
            'DIS_69': 'H29',
            'DIS_70': 'I29',
            'DIS_71': 'E30',
            'DIS_72': 'F30',
            'DIS_73': 'G30',
            'DIS_74': 'H30',
            'DIS_75': 'I30',
            'DIS_76': 'E31',
            'DIS_77': 'F31',
            'DIS_78': 'G31',
            'DIS_79': 'H31',
            'DIS_80': 'I31',
            'DIS_81': 'E32',
            'DIS_82': 'F32',
            'DIS_83': 'G32',
            'DIS_84': 'H32',
            'DIS_85': 'I32',
            'DIS_86': 'E33',
            'DIS_87': 'F33',
            'DIS_88': 'G33',
            'DIS_89': 'H33',
            'DIS_90': 'I33',
            'DIS_91': 'E34',
            'DIS_92': 'F34',
            'DIS_93': 'G34',
            'DIS_94': 'H34',
            'DIS_95': 'I34',
            'DIS_96': 'E35',
            'DIS_97': 'F35',
            'DIS_98': 'G35',
            'DIS_99': 'H35',
            'DIS_100': 'I35',
            'DIS_101': 'E36',
            'DIS_102': 'F36',
            'DIS_103': 'G36',
            'DIS_104': 'H36',
            'DIS_105': 'I36',
            'DIS_106': 'E37',
            'DIS_107': 'F37',
            'DIS_108': 'G37',
            'DIS_109': 'H37',
            'DIS_110': 'I37',
            'DIS_111': 'E38',
            'DIS_112': 'F38',
            'DIS_113': 'G38',
            'DIS_114': 'H38',
            'DIS_115': 'I38',
            'DIS_116': 'E39',
            'DIS_117': 'F39',
            'DIS_118': 'G39',
            'DIS_119': 'H39',
            'DIS_120': 'I39',
            'DIS_121': 'E40',
            'DIS_122': 'F40',
            'DIS_123': 'G40',
            'DIS_124': 'H40',
            'DIS_125': 'I40',
            'DIS_126': 'E41',
            'DIS_127': 'F41',
            'DIS_128': 'G41',
            'DIS_129': 'H41',
            'DIS_130': 'I41', 
            'DIS_131': 'E42',
            'DIS_132': 'F42',
            'DIS_133': 'G42',
            'DIS_134': 'H42',
            'DIS_135': 'I42', 
            'DIS_136': 'E43',
            'DIS_137': 'F43',
            'DIS_138': 'G43',
            'DIS_139': 'H43',
            'DIS_140': 'I43', 
            'DIS_141': 'E44',
            'DIS_142': 'F44',
            'DIS_143': 'G44',
            'DIS_144': 'H44',
            'DIS_145': 'I44', 
            'DIS_146': 'E45',
            'DIS_147': 'F45',
            'DIS_148': 'G45',
            'DIS_149': 'H45',
            'DIS_150': 'I45', 
            'DIS_151': 'E46',
            'DIS_152': 'F46',
            'DIS_153': 'G46',
            'DIS_154': 'H46',
            'DIS_155': 'I46', 
            'DIS_156': 'E47',
            'DIS_157': 'F47',
            'DIS_158': 'G47',
            'DIS_159': 'H47',
            'DIS_160': 'I47',            
        }
        
        col_ubi_sensorial = {    
            'PROVINCIA': 'D10',
            'DIS_161': 'E52',
            'DIS_162': 'F52',
            'DIS_163': 'G52',
            'DIS_164': 'H52',
            'DIS_165': 'I52',
            'DIS_166': 'E53',
            'DIS_167': 'F53',
            'DIS_168': 'G53',
            'DIS_169': 'H53',
            'DIS_170': 'I53',
            'DIS_171': 'E54',
            'DIS_172': 'F54',
            'DIS_173': 'G54',
            'DIS_174': 'H54',
            'DIS_175': 'I54',
            'DIS_176': 'E55',
            'DIS_177': 'F55',
            'DIS_178': 'G55',
            'DIS_179': 'H55',
            'DIS_180': 'I55',
            'DIS_181': 'E56',
            'DIS_182': 'F56',
            'DIS_183': 'G56',
            'DIS_184': 'H56',
            'DIS_185': 'I56',
            'DIS_186': 'E57',
            'DIS_187': 'F57',
            'DIS_188': 'G57',
            'DIS_189': 'H57',
            'DIS_190': 'I57',
            'DIS_191': 'E58',
            'DIS_192': 'F58',
            'DIS_193': 'G58',
            'DIS_194': 'H58',
            'DIS_195': 'I58',
        }
        
        col_ubi_mental = {    
            'PROVINCIA': 'D10',
            'DIS_196': 'E63',
            'DIS_197': 'F63',
            'DIS_198': 'G63',
            'DIS_199': 'H63',
            'DIS_200': 'I63',
            'DIS_201': 'E64',
            'DIS_202': 'F64',
            'DIS_203': 'G64',
            'DIS_204': 'H64',
            'DIS_205': 'I64',
            'DIS_206': 'E65',
            'DIS_207': 'F65',
            'DIS_208': 'G65',
            'DIS_209': 'H65',
            'DIS_210': 'I65',
            'DIS_211': 'E66',
            'DIS_212': 'F66',
            'DIS_213': 'G66',
            'DIS_214': 'H66',
            'DIS_215': 'I66',
        }
        
        col_ubi_certificado = {    
            'PROVINCIA': 'D10',
            'DIS_216': 'E71',
            'DIS_217': 'F71',
            'DIS_218': 'G71',
            'DIS_219': 'H71',
            'DIS_220': 'I71',
            'DIS_221': 'E72',
            'DIS_222': 'F72',
            'DIS_223': 'G72',
            'DIS_224': 'H72',
            'DIS_225': 'I72',
            'DIS_226': 'E73',
            'DIS_227': 'F73',
            'DIS_228': 'G73',
            'DIS_229': 'H73',
            'DIS_230': 'I73',
            'DIS_231': 'E74',
            'DIS_232': 'F74',
            'DIS_233': 'G74',
            'DIS_234': 'H74',
            'DIS_235': 'I74',
        }
        
        col_ubi_capacitacion = {    
            'PROVINCIA': 'D10',
            'DIS_273': 'D12',
            'DIS_274': 'E12',
        }
        
        col_ubi_agente = {    
            'PROVINCIA': 'D10',
            'DIS_236': 'D82',
            'DIS_237': 'E82',
            'DIS_238': 'F82',
            'DIS_239': 'G82',
            'DIS_240': 'H82',
            'DIS_241': 'I82',
        }      
        
        col_ubi_rbc = {    
            'PROVINCIA': 'D10',
            'DIS_242': 'E86',
            'DIS_243': 'F86',
            'DIS_244': 'G86',
            'DIS_245': 'H86',
            'DIS_246': 'I86',
            'DIS_247': 'E87',
            'DIS_248': 'F87',
            'DIS_249': 'G87',
            'DIS_250': 'H87',
            'DIS_251': 'I87',
            'DIS_252': 'E88',
            'DIS_253': 'F88',
            'DIS_254': 'G88',
            'DIS_255': 'H88',
            'DIS_256': 'I88',
            'DIS_257': 'E89',
            'DIS_258': 'F89',
            'DIS_259': 'G89',
            'DIS_260': 'H89',
            'DIS_261': 'I89',
            'DIS_262': 'E90',
            'DIS_263': 'F90',
            'DIS_264': 'G90',
            'DIS_265': 'H90',
            'DIS_266': 'I90'
        }
        
        col_ubi_comite = {    
            'PROVINCIA': 'D10',
            'DIS_267': 'D97',
            'DIS_268': 'E97',
            'DIS_269': 'F97',
            'DIS_270': 'G97',
            'DIS_271': 'H97',
            'DIS_272': 'I97',
        }
        
        # Inicializar diccionario para almacenar sumas por columna
        column_sums = {
            'DIS_1': 0,
            'DIS_2': 0,
            'DIS_3': 0,
            'DIS_4': 0,
            'DIS_5': 0,
            'DIS_6': 0,
            'DIS_7': 0,
            'DIS_8': 0,
            'DIS_9': 0,
            'DIS_10': 0,
            'DIS_11': 0,
            'DIS_12': 0,
            'DIS_13': 0,
            'DIS_14': 0,
            'DIS_15': 0,
            'DIS_16': 0,
            'DIS_17': 0,
            'DIS_18': 0,
            'DIS_19': 0,
            'DIS_20': 0,
            'DIS_21': 0,
            'DIS_22': 0,
            'DIS_23': 0,
            'DIS_24': 0,
            'DIS_25': 0,
            'DIS_26': 0,
            'DIS_27': 0,
            'DIS_28': 0,
            'DIS_29': 0,
            'DIS_30': 0,
            'DIS_31': 0,
            'DIS_32': 0,
            'DIS_33': 0,
            'DIS_34': 0,
            'DIS_35': 0,
            'DIS_36': 0,
            'DIS_37': 0,
            'DIS_38': 0,
            'DIS_39': 0,
            'DIS_40': 0,
            'DIS_41': 0,
            'DIS_42': 0,
            'DIS_43': 0,
            'DIS_44': 0,
            'DIS_45': 0,
            'DIS_46': 0,
            'DIS_47': 0,
            'DIS_48': 0,
            'DIS_49': 0,
            'DIS_50': 0,
            'DIS_51': 0,
            'DIS_52': 0,
            'DIS_53': 0,
            'DIS_54': 0,
            'DIS_55': 0,
            'DIS_56': 0,
            'DIS_57': 0,
            'DIS_58': 0,
            'DIS_59': 0,
            'DIS_60': 0,
            'DIS_61': 0,
            'DIS_62': 0,
            'DIS_63': 0,
            'DIS_64': 0,
            'DIS_65': 0,
            'DIS_66': 0,
            'DIS_67': 0,
            'DIS_68': 0,
            'DIS_69': 0,
            'DIS_70': 0,
            'DIS_71': 0,
            'DIS_72': 0,
            'DIS_73': 0,
            'DIS_74': 0,
            'DIS_75': 0,
            'DIS_76': 0,
            'DIS_77': 0,
            'DIS_78': 0,
            'DIS_79': 0,
            'DIS_80': 0,
            'DIS_81': 0,
            'DIS_82': 0,
            'DIS_83': 0,
            'DIS_84': 0,
            'DIS_85': 0,
            'DIS_86': 0,
            'DIS_87': 0,
            'DIS_88': 0,
            'DIS_89': 0,
            'DIS_90': 0,
            'DIS_91': 0,
            'DIS_92': 0,
            'DIS_93': 0,
            'DIS_94': 0,
            'DIS_95': 0,
            'DIS_96': 0,
            'DIS_97': 0,
            'DIS_98': 0,
            'DIS_99': 0,
            'DIS_100': 0,
            'DIS_101': 0,
            'DIS_102': 0,
            'DIS_103': 0,
            'DIS_104': 0,
            'DIS_105': 0,
            'DIS_106': 0,
            'DIS_107': 0,
            'DIS_108': 0,
            'DIS_109': 0,
            'DIS_110': 0,
            'DIS_111': 0,
            'DIS_112': 0,
            'DIS_113': 0,
            'DIS_114': 0,
            'DIS_115': 0,
            'DIS_116': 0,
            'DIS_117': 0,
            'DIS_118': 0,
            'DIS_119': 0,
            'DIS_120': 0,
            'DIS_121': 0,
            'DIS_122': 0,
            'DIS_123': 0,
            'DIS_124': 0,
            'DIS_125': 0,
            'DIS_126': 0,
            'DIS_127': 0,
            'DIS_128': 0,
            'DIS_129': 0,
            'DIS_130': 0, 
            'DIS_131': 0,
            'DIS_132': 0,
            'DIS_133': 0,
            'DIS_134': 0,
            'DIS_135': 0, 
            'DIS_136': 0,
            'DIS_137': 0,
            'DIS_138': 0,
            'DIS_139': 0,
            'DIS_140': 0, 
            'DIS_141': 0,
            'DIS_142': 0,
            'DIS_143': 0,
            'DIS_144': 0,
            'DIS_145': 0, 
            'DIS_146': 0,
            'DIS_147': 0,
            'DIS_148': 0,
            'DIS_149': 0,
            'DIS_150': 0, 
            'DIS_151': 0,
            'DIS_152': 0,
            'DIS_153': 0,
            'DIS_154': 0,
            'DIS_155': 0, 
            'DIS_156': 0,
            'DIS_157': 0,
            'DIS_158': 0,
            'DIS_159': 0,
            'DIS_160': 0,    
        }
        
        col_sum_sensorial = {       
            'DIS_161': 0,
            'DIS_162': 0,
            'DIS_163': 0,
            'DIS_164': 0,
            'DIS_165': 0,
            'DIS_166': 0,
            'DIS_167': 0,
            'DIS_168': 0,
            'DIS_169': 0,
            'DIS_170': 0,
            'DIS_171': 0,
            'DIS_172': 0,
            'DIS_173': 0,
            'DIS_174': 0,
            'DIS_175': 0,
            'DIS_176': 0,
            'DIS_177': 0,
            'DIS_178': 0,
            'DIS_179': 0,
            'DIS_180': 0,
            'DIS_181': 0,
            'DIS_182': 0,
            'DIS_183': 0,
            'DIS_184': 0,
            'DIS_185': 0,
            'DIS_186': 0,
            'DIS_187': 0,
            'DIS_188': 0,
            'DIS_189': 0,
            'DIS_190': 0,
            'DIS_191': 0,
            'DIS_192': 0,
            'DIS_193': 0,
            'DIS_194': 0,
            'DIS_195': 0,
        } 

        col_sum_mental = {    
            'DIS_196': 0,
            'DIS_197': 0,
            'DIS_198': 0,
            'DIS_199': 0,
            'DIS_200': 0,
            'DIS_201': 0,
            'DIS_202': 0,
            'DIS_203': 0,
            'DIS_204': 0,
            'DIS_205': 0,
            'DIS_206': 0,
            'DIS_207': 0,
            'DIS_208': 0,
            'DIS_209': 0,
            'DIS_210': 0,
            'DIS_211': 0,
            'DIS_212': 0,
            'DIS_213': 0,
            'DIS_214': 0,
            'DIS_215': 0,
        }
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_certificado = {       
            'DIS_216': 0,
            'DIS_217': 0,
            'DIS_218': 0,
            'DIS_219': 0,
            'DIS_220': 0,
            'DIS_221': 0,
            'DIS_222': 0,
            'DIS_223': 0,
            'DIS_224': 0,
            'DIS_225': 0,
            'DIS_226': 0,
            'DIS_227': 0,
            'DIS_228': 0,
            'DIS_229': 0,
            'DIS_230': 0,
            'DIS_231': 0,
            'DIS_232': 0,
            'DIS_233': 0,
            'DIS_234': 0,
            'DIS_235': 0,
        }  
        
        col_sum_capacitacion = {    
            'DIS_273': 0,
            'DIS_274': 0,
        }
        
        col_sum_agente = {    
            'DIS_236': 0,
            'DIS_237': 0,
            'DIS_238': 0,
            'DIS_239': 0,
            'DIS_240': 0,
            'DIS_241': 0,
        }      
        
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_rbc = {       
            'DIS_242': 0,
            'DIS_243': 0,
            'DIS_244': 0,
            'DIS_245': 0,
            'DIS_246': 0,
            'DIS_247': 0,
            'DIS_248': 0,
            'DIS_249': 0,
            'DIS_250': 0,
            'DIS_251': 0,
            'DIS_252': 0,
            'DIS_253': 0,
            'DIS_254': 0,
            'DIS_255': 0,
            'DIS_256': 0,
            'DIS_257': 0,
            'DIS_258': 0,
            'DIS_259': 0,
            'DIS_260': 0,
            'DIS_261': 0,
            'DIS_262': 0,
            'DIS_263': 0,
            'DIS_264': 0,
            'DIS_265': 0,
            'DIS_266': 0,
        } 
        
        col_sum_comite = {    
            'DIS_267': 0,
            'DIS_268': 0,
            'DIS_269': 0,
            'DIS_270': 0,
            'DIS_271': 0,
            'DIS_272': 0,
        }
                    
        ############################
        ###  DISCAPACIDAD FISICA ###
        ############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_microred:
            for col_name in column_sums:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(columnas_ubicaciones.keys()).index(col_name) + 3
                    column_sums[col_name] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila: {row}")                        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_name, total_cell in columnas_ubicaciones.items():
            if col_name in column_sums:
                # Obtener la celda correspondiente según la ubicación
                cell = sheet[total_cell]
                # Asignar el valor de la suma a la celda
                cell.value = column_sums[col_name]
                # Aplicar formato a la celda
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        
        # Sumar los valores del diccionario      
        total_sum_cat_1 =  sum([column_sums['DIS_1'], column_sums['DIS_2'], column_sums['DIS_3'],column_sums['DIS_4'],column_sums['DIS_5']])
        total_sum_cat_2 =  sum([column_sums['DIS_6'], column_sums['DIS_7'], column_sums['DIS_8'],column_sums['DIS_9'],column_sums['DIS_10']])
        total_sum_cat_3 =  sum([column_sums['DIS_11'], column_sums['DIS_12'], column_sums['DIS_13'],column_sums['DIS_14'],column_sums['DIS_15']])
        total_sum_cat_4 =  sum([column_sums['DIS_16'], column_sums['DIS_17'], column_sums['DIS_18'],column_sums['DIS_19'],column_sums['DIS_20']])
        total_sum_cat_5 =  sum([column_sums['DIS_21'], column_sums['DIS_22'], column_sums['DIS_23'],column_sums['DIS_24'],column_sums['DIS_25']])
        total_sum_cat_6 =  sum([column_sums['DIS_26'], column_sums['DIS_27'], column_sums['DIS_28'],column_sums['DIS_29'],column_sums['DIS_30']])
        total_sum_cat_7 =  sum([column_sums['DIS_31'], column_sums['DIS_32'], column_sums['DIS_33'],column_sums['DIS_34'],column_sums['DIS_35']])
        total_sum_cat_8 =  sum([column_sums['DIS_36'], column_sums['DIS_37'], column_sums['DIS_38'],column_sums['DIS_39'],column_sums['DIS_40']])
        total_sum_cat_9 =  sum([column_sums['DIS_41'], column_sums['DIS_42'], column_sums['DIS_43'],column_sums['DIS_44'],column_sums['DIS_45']])
        total_sum_cat_10 =  sum([column_sums['DIS_46'], column_sums['DIS_47'], column_sums['DIS_48'],column_sums['DIS_49'],column_sums['DIS_50']])
        total_sum_cat_11 =  sum([column_sums['DIS_51'], column_sums['DIS_52'], column_sums['DIS_53'],column_sums['DIS_54'],column_sums['DIS_55']])
        total_sum_cat_12 =  sum([column_sums['DIS_56'], column_sums['DIS_57'], column_sums['DIS_58'],column_sums['DIS_59'],column_sums['DIS_60']])
        total_sum_cat_13 =  sum([column_sums['DIS_61'], column_sums['DIS_62'], column_sums['DIS_63'],column_sums['DIS_64'],column_sums['DIS_65']])
        total_sum_cat_14 =  sum([column_sums['DIS_66'], column_sums['DIS_67'], column_sums['DIS_68'],column_sums['DIS_69'],column_sums['DIS_70']])
        total_sum_cat_15 =  sum([column_sums['DIS_71'], column_sums['DIS_72'], column_sums['DIS_73'],column_sums['DIS_74'],column_sums['DIS_75']])
        total_sum_cat_16 =  sum([column_sums['DIS_76'], column_sums['DIS_77'], column_sums['DIS_78'],column_sums['DIS_79'],column_sums['DIS_80']])   
        total_sum_cat_17 =  sum([column_sums['DIS_81'], column_sums['DIS_82'], column_sums['DIS_83'],column_sums['DIS_84'],column_sums['DIS_85']])
        total_sum_cat_18 =  sum([column_sums['DIS_86'], column_sums['DIS_87'], column_sums['DIS_88'],column_sums['DIS_89'],column_sums['DIS_90']])
        total_sum_cat_19 =  sum([column_sums['DIS_91'], column_sums['DIS_92'], column_sums['DIS_93'],column_sums['DIS_94'],column_sums['DIS_95']])
        total_sum_cat_20 =  sum([column_sums['DIS_96'], column_sums['DIS_97'], column_sums['DIS_98'],column_sums['DIS_99'],column_sums['DIS_100']])
        total_sum_cat_21 =  sum([column_sums['DIS_101'], column_sums['DIS_102'], column_sums['DIS_103'],column_sums['DIS_104'],column_sums['DIS_105']])
        total_sum_cat_22 =  sum([column_sums['DIS_106'], column_sums['DIS_107'], column_sums['DIS_108'],column_sums['DIS_109'],column_sums['DIS_110']])
        total_sum_cat_23 =  sum([column_sums['DIS_111'], column_sums['DIS_112'], column_sums['DIS_113'],column_sums['DIS_114'],column_sums['DIS_115']])
        total_sum_cat_24 =  sum([column_sums['DIS_116'], column_sums['DIS_117'], column_sums['DIS_118'],column_sums['DIS_119'],column_sums['DIS_120']])
        total_sum_cat_25 =  sum([column_sums['DIS_121'], column_sums['DIS_122'], column_sums['DIS_123'],column_sums['DIS_124'],column_sums['DIS_125']])
        total_sum_cat_26 =  sum([column_sums['DIS_126'], column_sums['DIS_127'], column_sums['DIS_128'],column_sums['DIS_129'],column_sums['DIS_130']])
        total_sum_cat_27 =  sum([column_sums['DIS_131'], column_sums['DIS_132'], column_sums['DIS_133'],column_sums['DIS_134'],column_sums['DIS_135']])
        total_sum_cat_28 =  sum([column_sums['DIS_136'], column_sums['DIS_137'], column_sums['DIS_138'],column_sums['DIS_139'],column_sums['DIS_140']])
        total_sum_cat_29 =  sum([column_sums['DIS_141'], column_sums['DIS_142'], column_sums['DIS_143'],column_sums['DIS_144'],column_sums['DIS_145']])
        total_sum_cat_30 =  sum([column_sums['DIS_146'], column_sums['DIS_147'], column_sums['DIS_148'],column_sums['DIS_149'],column_sums['DIS_150']])
        total_sum_cat_31 =  sum([column_sums['DIS_151'], column_sums['DIS_152'], column_sums['DIS_153'],column_sums['DIS_154'],column_sums['DIS_155']])
        total_sum_cat_32 =  sum([column_sums['DIS_156'], column_sums['DIS_157'], column_sums['DIS_158'],column_sums['DIS_159'],column_sums['DIS_160']])

        sheet['D16'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D16'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D16'] = total_sum_cat_1     
        
        sheet['D17'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D17'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D17'] = total_sum_cat_2 
        
        sheet['D18'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D18'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D18'] = total_sum_cat_3    
        
        sheet['D19'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D19'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D19'] = total_sum_cat_4    
        
        sheet['D20'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D20'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D20'] = total_sum_cat_5    
        
        sheet['D21'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D21'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D21'] = total_sum_cat_6    
        
        sheet['D22'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D22'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D22'] = total_sum_cat_7    
        
        sheet['D23'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D23'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D23'] = total_sum_cat_8    
        
        sheet['D24'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D24'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D24'] = total_sum_cat_9    
        
        sheet['D25'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D25'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D25'] = total_sum_cat_10 
        
        sheet['D26'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D26'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D26'] = total_sum_cat_11
                
        sheet['D27'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D27'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D27'] = total_sum_cat_12    
        
        sheet['D28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D28'] = total_sum_cat_13   
        
        sheet['D29'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D29'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D29'] = total_sum_cat_14   
        
        sheet['D30'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D30'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D30'] = total_sum_cat_15   
        
        sheet['D31'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D31'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D31'] = total_sum_cat_16   
        
        sheet['D32'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D32'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D32'] = total_sum_cat_17         
        
        sheet['D33'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D33'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D33'] = total_sum_cat_18   
        
        sheet['D34'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D34'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D34'] = total_sum_cat_19   
        
        sheet['D35'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D35'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D35'] = total_sum_cat_20   
        
        sheet['D36'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D36'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D36'] = total_sum_cat_21   
        
        sheet['D37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D37'] = total_sum_cat_22   
        
        sheet['D38'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D38'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D38'] = total_sum_cat_23   
        
        sheet['D39'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D39'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D39'] = total_sum_cat_24   
        
        sheet['D40'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D40'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D40'] = total_sum_cat_25  
        
        sheet['D41'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D41'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D41'] = total_sum_cat_26 
        
        sheet['D42'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D42'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D42'] = total_sum_cat_27   
        
        sheet['D43'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D43'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D43'] = total_sum_cat_28   
        
        sheet['D44'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D44'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D44'] = total_sum_cat_29  
        
        sheet['D45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D45'] = total_sum_cat_30  
        
        sheet['D46'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D46'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D46'] = total_sum_cat_31
        
        sheet['D47'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D47'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D47'] = total_sum_cat_32
        
        # Sumar los valores del VERTICAL      
        total_sum_cat_vertical_1 =  sum([column_sums['DIS_1'],column_sums['DIS_6'], column_sums['DIS_11'],column_sums['DIS_16'],column_sums['DIS_21'],column_sums['DIS_26'],column_sums['DIS_31'],column_sums['DIS_36'],column_sums['DIS_41'],column_sums['DIS_46'],column_sums['DIS_51'],column_sums['DIS_56'],column_sums['DIS_61'],column_sums['DIS_66'],column_sums['DIS_71'],column_sums['DIS_76'],column_sums['DIS_81'],column_sums['DIS_86'],column_sums['DIS_91'],column_sums['DIS_96'],column_sums['DIS_101'],column_sums['DIS_106'] ,column_sums['DIS_111'],column_sums['DIS_116'],column_sums['DIS_121'],column_sums['DIS_126'],column_sums['DIS_131'],column_sums['DIS_136'],column_sums['DIS_141'],column_sums['DIS_146'],column_sums['DIS_151'],column_sums['DIS_156']])
        total_sum_cat_vertical_2 =  sum([column_sums['DIS_2'],column_sums['DIS_7'], column_sums['DIS_12'],column_sums['DIS_17'],column_sums['DIS_22'],column_sums['DIS_27'],column_sums['DIS_32'],column_sums['DIS_37'],column_sums['DIS_42'],column_sums['DIS_47'],column_sums['DIS_52'],column_sums['DIS_57'],column_sums['DIS_62'],column_sums['DIS_67'],column_sums['DIS_72'],column_sums['DIS_77'],column_sums['DIS_82'],column_sums['DIS_87'],column_sums['DIS_92'],column_sums['DIS_97'],column_sums['DIS_102'],column_sums['DIS_107'] ,column_sums['DIS_112'],column_sums['DIS_117'],column_sums['DIS_122'],column_sums['DIS_127'],column_sums['DIS_132'],column_sums['DIS_137'],column_sums['DIS_142'],column_sums['DIS_147'],column_sums['DIS_152'],column_sums['DIS_157']])
        total_sum_cat_vertical_3 =  sum([column_sums['DIS_3'],column_sums['DIS_8'], column_sums['DIS_13'],column_sums['DIS_18'],column_sums['DIS_23'],column_sums['DIS_28'],column_sums['DIS_33'],column_sums['DIS_38'],column_sums['DIS_43'],column_sums['DIS_48'],column_sums['DIS_53'],column_sums['DIS_58'],column_sums['DIS_63'],column_sums['DIS_68'],column_sums['DIS_73'],column_sums['DIS_78'],column_sums['DIS_83'],column_sums['DIS_88'],column_sums['DIS_93'],column_sums['DIS_98'],column_sums['DIS_103'],column_sums['DIS_108'] ,column_sums['DIS_113'],column_sums['DIS_118'],column_sums['DIS_123'],column_sums['DIS_128'],column_sums['DIS_133'],column_sums['DIS_138'],column_sums['DIS_143'],column_sums['DIS_148'],column_sums['DIS_153'],column_sums['DIS_158']])
        total_sum_cat_vertical_4 =  sum([column_sums['DIS_4'],column_sums['DIS_9'], column_sums['DIS_14'],column_sums['DIS_19'],column_sums['DIS_24'],column_sums['DIS_29'],column_sums['DIS_34'],column_sums['DIS_39'],column_sums['DIS_44'],column_sums['DIS_49'],column_sums['DIS_54'],column_sums['DIS_59'],column_sums['DIS_64'],column_sums['DIS_69'],column_sums['DIS_74'],column_sums['DIS_79'],column_sums['DIS_84'],column_sums['DIS_89'],column_sums['DIS_94'],column_sums['DIS_99'],column_sums['DIS_104'],column_sums['DIS_109'] ,column_sums['DIS_114'],column_sums['DIS_119'],column_sums['DIS_124'],column_sums['DIS_129'],column_sums['DIS_134'],column_sums['DIS_139'],column_sums['DIS_144'],column_sums['DIS_149'],column_sums['DIS_154'],column_sums['DIS_159']])
        total_sum_cat_vertical_5 =  sum([column_sums['DIS_5'],column_sums['DIS_10'],column_sums['DIS_15'],column_sums['DIS_20'],column_sums['DIS_25'],column_sums['DIS_30'],column_sums['DIS_35'],column_sums['DIS_40'],column_sums['DIS_45'],column_sums['DIS_50'],column_sums['DIS_55'],column_sums['DIS_60'],column_sums['DIS_65'],column_sums['DIS_70'],column_sums['DIS_75'],column_sums['DIS_80'],column_sums['DIS_85'],column_sums['DIS_90'],column_sums['DIS_95'],column_sums['DIS_100'],column_sums['DIS_105'],column_sums['DIS_110'],column_sums['DIS_115'],column_sums['DIS_120'],column_sums['DIS_125'],column_sums['DIS_130'],column_sums['DIS_135'],column_sums['DIS_140'],column_sums['DIS_145'],column_sums['DIS_150'],column_sums['DIS_155'],column_sums['DIS_160']])

        sheet['E48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E48'] = total_sum_cat_vertical_1     
        
        sheet['F48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F48'] = total_sum_cat_vertical_2 
        
        sheet['G48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G48'] = total_sum_cat_vertical_3    
        
        sheet['H48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H48'] = total_sum_cat_vertical_4    
        
        sheet['I48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I48'] = total_sum_cat_vertical_5    
        ##########################################################################
        
        ###############################
        ###  DISCAPACIDAD SENSORIAL ###
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_sensorial_microred:
            for col_sensorial in col_sum_sensorial:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_sensorial.keys()).index(col_sensorial) + 3
                    col_sum_sensorial[col_sensorial] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_sensorial, total_cell_sensorial in col_ubi_sensorial.items():
            if col_sensorial in col_sum_sensorial:
                # Obtener la celda correspondiente según la ubicación
                cell_sensorial = sheet[total_cell_sensorial]
                # Asignar el valor de la suma a la celda
                cell_sensorial.value = col_sum_sensorial[col_sensorial]
                # Aplicar formato a la celda
                cell_sensorial.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_sensorial.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_sensorial.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_sensorial['DIS_161'], col_sum_sensorial['DIS_162'], col_sum_sensorial['DIS_163'], col_sum_sensorial['DIS_164'], col_sum_sensorial['DIS_165']])
        t_sum_cat_2 =  sum([col_sum_sensorial['DIS_166'], col_sum_sensorial['DIS_167'], col_sum_sensorial['DIS_168'], col_sum_sensorial['DIS_169'], col_sum_sensorial['DIS_170']])
        t_sum_cat_3 =  sum([col_sum_sensorial['DIS_171'], col_sum_sensorial['DIS_172'], col_sum_sensorial['DIS_173'], col_sum_sensorial['DIS_174'], col_sum_sensorial['DIS_175']])
        t_sum_cat_4 =  sum([col_sum_sensorial['DIS_176'], col_sum_sensorial['DIS_177'], col_sum_sensorial['DIS_178'], col_sum_sensorial['DIS_179'], col_sum_sensorial['DIS_180']])
        t_sum_cat_5 =  sum([col_sum_sensorial['DIS_181'], col_sum_sensorial['DIS_182'], col_sum_sensorial['DIS_183'], col_sum_sensorial['DIS_184'], col_sum_sensorial['DIS_185']])
        t_sum_cat_6 =  sum([col_sum_sensorial['DIS_186'], col_sum_sensorial['DIS_187'], col_sum_sensorial['DIS_188'], col_sum_sensorial['DIS_189'], col_sum_sensorial['DIS_190']])
        t_sum_cat_7 =  sum([col_sum_sensorial['DIS_191'], col_sum_sensorial['DIS_192'], col_sum_sensorial['DIS_193'], col_sum_sensorial['DIS_194'], col_sum_sensorial['DIS_195']])
        
        sheet['D52'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D52'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D52'] = t_sum_cat_1     
        
        sheet['D53'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D53'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D53'] = t_sum_cat_2 
        
        sheet['D54'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D54'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D54'] = t_sum_cat_3    
        
        sheet['D55'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D55'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D55'] = t_sum_cat_4    
        
        sheet['D56'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D56'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D56'] = t_sum_cat_5    
        
        sheet['D57'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D57'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D57'] = t_sum_cat_6    
        
        sheet['D58'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D58'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D58'] = t_sum_cat_7    
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_sensorial['DIS_161'],col_sum_sensorial['DIS_166'],col_sum_sensorial['DIS_171'],col_sum_sensorial['DIS_176'],col_sum_sensorial['DIS_181'],col_sum_sensorial['DIS_186'],col_sum_sensorial['DIS_191']])
        t_sum_cat_vertical_2 =  sum([col_sum_sensorial['DIS_162'],col_sum_sensorial['DIS_167'],col_sum_sensorial['DIS_172'],col_sum_sensorial['DIS_177'],col_sum_sensorial['DIS_182'],col_sum_sensorial['DIS_187'],col_sum_sensorial['DIS_192']])
        t_sum_cat_vertical_3 =  sum([col_sum_sensorial['DIS_163'],col_sum_sensorial['DIS_168'],col_sum_sensorial['DIS_173'],col_sum_sensorial['DIS_178'],col_sum_sensorial['DIS_183'],col_sum_sensorial['DIS_188'],col_sum_sensorial['DIS_193']])
        t_sum_cat_vertical_4 =  sum([col_sum_sensorial['DIS_164'],col_sum_sensorial['DIS_169'],col_sum_sensorial['DIS_174'],col_sum_sensorial['DIS_179'],col_sum_sensorial['DIS_184'],col_sum_sensorial['DIS_189'],col_sum_sensorial['DIS_194']])
        t_sum_cat_vertical_5 =  sum([col_sum_sensorial['DIS_165'],col_sum_sensorial['DIS_170'],col_sum_sensorial['DIS_175'],col_sum_sensorial['DIS_180'],col_sum_sensorial['DIS_185'],col_sum_sensorial['DIS_190'],col_sum_sensorial['DIS_195']])
        
        sheet['E59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E59'] = t_sum_cat_vertical_1     
        
        sheet['F59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F59'] = t_sum_cat_vertical_2 
        
        sheet['G59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G59'] = t_sum_cat_vertical_3    
        
        sheet['H59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H59'] = t_sum_cat_vertical_4    
        
        sheet['I59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I59'] = t_sum_cat_vertical_5    
        ##########################################################################
                
        ###############################
        ###  DISCAPACIDAD MENTAL ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_mental_microred:
            for col_mental in col_sum_mental:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_mental.keys()).index(col_mental) + 3
                    col_sum_mental[col_mental] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_mental, total_cell_mental in col_ubi_mental.items():
            if col_mental in col_sum_mental:
                # Obtener la celda correspondiente según la ubicación
                cell_mental = sheet[total_cell_mental]
                # Asignar el valor de la suma a la celda
                cell_mental.value = col_sum_mental[col_mental]
                # Aplicar formato a la celda
                cell_mental.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_mental.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_mental.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_mental['DIS_196'], col_sum_mental['DIS_197'], col_sum_mental['DIS_198'], col_sum_mental['DIS_199'], col_sum_mental['DIS_200']])
        t_sum_cat_2 =  sum([col_sum_mental['DIS_201'], col_sum_mental['DIS_202'], col_sum_mental['DIS_203'], col_sum_mental['DIS_204'], col_sum_mental['DIS_205']])
        t_sum_cat_3 =  sum([col_sum_mental['DIS_206'], col_sum_mental['DIS_207'], col_sum_mental['DIS_208'], col_sum_mental['DIS_209'], col_sum_mental['DIS_210']])
        t_sum_cat_4 =  sum([col_sum_mental['DIS_211'], col_sum_mental['DIS_212'], col_sum_mental['DIS_213'], col_sum_mental['DIS_214'], col_sum_mental['DIS_215']])
        
        sheet['D63'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D63'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D63'] = t_sum_cat_1     
        
        sheet['D64'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D64'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D64'] = t_sum_cat_2 
        
        sheet['D65'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D65'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D65'] = t_sum_cat_3    
        
        sheet['D66'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D66'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D66'] = t_sum_cat_4    

        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_mental['DIS_196'],col_sum_mental['DIS_201'],col_sum_mental['DIS_206'],col_sum_mental['DIS_211']])
        t_sum_cat_vertical_2 =  sum([col_sum_mental['DIS_197'],col_sum_mental['DIS_202'],col_sum_mental['DIS_207'],col_sum_mental['DIS_212']])
        t_sum_cat_vertical_3 =  sum([col_sum_mental['DIS_198'],col_sum_mental['DIS_203'],col_sum_mental['DIS_208'],col_sum_mental['DIS_213']])
        t_sum_cat_vertical_4 =  sum([col_sum_mental['DIS_199'],col_sum_mental['DIS_204'],col_sum_mental['DIS_209'],col_sum_mental['DIS_214']])
        t_sum_cat_vertical_5 =  sum([col_sum_mental['DIS_200'],col_sum_mental['DIS_205'],col_sum_mental['DIS_210'],col_sum_mental['DIS_215']])
        
        sheet['E67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E67'] = t_sum_cat_vertical_1     
        
        sheet['F67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F67'] = t_sum_cat_vertical_2 
        
        sheet['G67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G67'] = t_sum_cat_vertical_3    
        
        sheet['H67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H67'] = t_sum_cat_vertical_4    
        
        sheet['I67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I67'] = t_sum_cat_vertical_5    
        ##########################################################################
        
        #################################
        ###  DISCAPACIDAD CERTIFICADO ###
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_certificado_microred:
            for col_certificado in col_sum_certificado:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_certificado.keys()).index(col_certificado) + 3
                    col_sum_certificado[col_certificado] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_certificado, total_cell_certificado in col_ubi_certificado.items():
            if col_certificado in col_sum_certificado:
                # Obtener la celda correspondiente según la ubicación
                cell_certificado = sheet[total_cell_certificado]
                # Asignar el valor de la suma a la celda
                cell_certificado.value = col_sum_certificado[col_certificado]
                # Aplicar formato a la celda
                cell_certificado.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_certificado.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_certificado.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        # Sumar los valores del diccionario      
        t_sum_cat_cert_1 =  sum([col_sum_certificado['DIS_216'], col_sum_certificado['DIS_217'], col_sum_certificado['DIS_218'], col_sum_certificado['DIS_219'], col_sum_certificado['DIS_220']])
        t_sum_cat_cert_2 =  sum([col_sum_certificado['DIS_221'], col_sum_certificado['DIS_222'], col_sum_certificado['DIS_223'], col_sum_certificado['DIS_224'], col_sum_certificado['DIS_225']])
        t_sum_cat_cert_3 =  sum([col_sum_certificado['DIS_226'], col_sum_certificado['DIS_227'], col_sum_certificado['DIS_228'], col_sum_certificado['DIS_229'], col_sum_certificado['DIS_230']])
        t_sum_cat_cert_4 =  sum([col_sum_certificado['DIS_231'], col_sum_certificado['DIS_232'], col_sum_certificado['DIS_233'], col_sum_certificado['DIS_234'], col_sum_certificado['DIS_235']])

        sheet['D71'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D71'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D71'] = t_sum_cat_cert_1     
        
        sheet['D72'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D72'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D72'] = t_sum_cat_cert_2 
        
        sheet['D73'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D73'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D73'] = t_sum_cat_cert_3 
        
        sheet['D74'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D74'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D74'] = t_sum_cat_cert_4 
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vert_1 =  sum([col_sum_certificado['DIS_216'],col_sum_certificado['DIS_221'],col_sum_certificado['DIS_226'],col_sum_certificado['DIS_231']])
        t_sum_cat_vert_2 =  sum([col_sum_certificado['DIS_217'],col_sum_certificado['DIS_222'],col_sum_certificado['DIS_227'],col_sum_certificado['DIS_232']])
        t_sum_cat_vert_3 =  sum([col_sum_certificado['DIS_218'],col_sum_certificado['DIS_223'],col_sum_certificado['DIS_228'],col_sum_certificado['DIS_233']])
        t_sum_cat_vert_4 =  sum([col_sum_certificado['DIS_219'],col_sum_certificado['DIS_224'],col_sum_certificado['DIS_229'],col_sum_certificado['DIS_234']])
        t_sum_cat_vert_5 =  sum([col_sum_certificado['DIS_220'],col_sum_certificado['DIS_225'],col_sum_certificado['DIS_230'],col_sum_certificado['DIS_235']])
        
        sheet['E75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E75'] = t_sum_cat_vert_1     
        
        sheet['F75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F75'] = t_sum_cat_vert_2 
        
        sheet['G75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G75'] = t_sum_cat_vert_3    
        
        sheet['H75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H75'] = t_sum_cat_vert_4    
        
        sheet['I75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I75'] = t_sum_cat_vert_5    
        
        #################################
        ###  DISCAPACIDAD RBC ###########
        #################################       
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_rbc_microred:
            for col_rbc in col_sum_rbc:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_rbc.keys()).index(col_rbc) + 3
                    col_sum_rbc[col_rbc] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
                    
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_rbc, total_cell_rbc in col_ubi_rbc.items():
            if col_rbc in col_sum_rbc:
                # Obtener la celda correspondiente según la ubicación
                cell_rbc = sheet[total_cell_rbc]
                # Asignar el valor de la suma a la celda
                cell_rbc.value = col_sum_rbc[col_rbc]
                # Aplicar formato a la celda
                cell_rbc.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_rbc.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_rbc.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        ## Sumar los valores del diccionario      
        t_sum_cat_rbc_1 =  sum([col_sum_rbc['DIS_242'], col_sum_rbc['DIS_243'], col_sum_rbc['DIS_244'], col_sum_rbc['DIS_245'], col_sum_rbc['DIS_246']])
        t_sum_cat_rbc_2 =  sum([col_sum_rbc['DIS_247'], col_sum_rbc['DIS_248'], col_sum_rbc['DIS_249'], col_sum_rbc['DIS_250'], col_sum_rbc['DIS_251']])
        t_sum_cat_rbc_3 =  sum([col_sum_rbc['DIS_252'], col_sum_rbc['DIS_253'], col_sum_rbc['DIS_254'], col_sum_rbc['DIS_255'], col_sum_rbc['DIS_256']])
        t_sum_cat_rbc_4 =  sum([col_sum_rbc['DIS_257'], col_sum_rbc['DIS_258'], col_sum_rbc['DIS_259'], col_sum_rbc['DIS_260'], col_sum_rbc['DIS_261']])
        t_sum_cat_rbc_5 =  sum([col_sum_rbc['DIS_262'], col_sum_rbc['DIS_263'], col_sum_rbc['DIS_264'], col_sum_rbc['DIS_265'], col_sum_rbc['DIS_266']])

        sheet['D86'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D86'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D86'] = t_sum_cat_rbc_1     
        
        sheet['D87'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D87'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D87'] = t_sum_cat_rbc_2 
        
        sheet['D88'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D88'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D88'] = t_sum_cat_rbc_3     
        
        sheet['D89'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D89'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D89'] = t_sum_cat_rbc_4 
        
        sheet['D90'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D90'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D90'] = t_sum_cat_rbc_5 
        
        # Sumar los valores del VERTICAL      
        t_sum_vert_rbc_1 =  sum([col_sum_rbc['DIS_242'],col_sum_rbc['DIS_247'],col_sum_rbc['DIS_252'],col_sum_rbc['DIS_257'],col_sum_rbc['DIS_262']])
        t_sum_vert_rbc_2 =  sum([col_sum_rbc['DIS_243'],col_sum_rbc['DIS_248'],col_sum_rbc['DIS_253'],col_sum_rbc['DIS_258'],col_sum_rbc['DIS_263']])
        t_sum_vert_rbc_3 =  sum([col_sum_rbc['DIS_244'],col_sum_rbc['DIS_249'],col_sum_rbc['DIS_254'],col_sum_rbc['DIS_259'],col_sum_rbc['DIS_264']])
        t_sum_vert_rbc_4 =  sum([col_sum_rbc['DIS_245'],col_sum_rbc['DIS_250'],col_sum_rbc['DIS_255'],col_sum_rbc['DIS_260'],col_sum_rbc['DIS_265']])
        t_sum_vert_rbc_5 =  sum([col_sum_rbc['DIS_246'],col_sum_rbc['DIS_251'],col_sum_rbc['DIS_256'],col_sum_rbc['DIS_261'],col_sum_rbc['DIS_266']])
        
        sheet['E91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E91'] = t_sum_vert_rbc_1
        
        sheet['F91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F91'] = t_sum_vert_rbc_2 
        
        sheet['G91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G91'] = t_sum_vert_rbc_3    
        
        sheet['H91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H91'] = t_sum_vert_rbc_4    
        
        sheet['I91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I91'] = t_sum_vert_rbc_5   
        
        #################################
        ###  CAPACITACION PERSONAL ######
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_capacitacion_microred:
            for col_capacitacion in col_sum_capacitacion:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_capacitacion.keys()).index(col_capacitacion) + 3
                    col_sum_capacitacion[col_capacitacion] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_capacitacion, total_cell_capacitacion in col_ubi_capacitacion.items():
            if col_capacitacion in col_sum_capacitacion:
                # Obtener la celda correspondiente según la ubicación
                cell_capacitacion = sheet[total_cell_capacitacion]
                # Asignar el valor de la suma a la celda
                cell_capacitacion.value = col_sum_capacitacion[col_capacitacion]
                # Aplicar formato a la celda
                cell_capacitacion.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_capacitacion.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_capacitacion.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_capacitacion['DIS_273']])
        t_sum_cat_2 = sum([col_sum_capacitacion['DIS_274']])
        
        sheet['D12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D12'] = t_sum_cat_1     
        
        sheet['E12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E12'] = t_sum_cat_2 
        
        ###############################
        ###  CAPACITACION AGENTE ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_agente_microred:
            for col_agente in col_sum_agente:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_agente.keys()).index(col_agente) + 3
                    col_sum_agente[col_agente] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_agente, total_cell_agente in col_ubi_agente.items():
            if col_agente in col_sum_agente:
                # Obtener la celda correspondiente según la ubicación
                cell_agente = sheet[total_cell_agente]
                # Asignar el valor de la suma a la celda
                cell_agente.value = col_sum_agente[col_agente]
                # Aplicar formato a la celda
                cell_agente.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_agente.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_agente.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_agente['DIS_236']])
        t_sum_cat_2 = sum([col_sum_agente['DIS_237']])
        t_sum_cat_3 = sum([col_sum_agente['DIS_238']])
        t_sum_cat_4 = sum([col_sum_agente['DIS_239']])
        t_sum_cat_5 = sum([col_sum_agente['DIS_240']])
        t_sum_cat_6 = sum([col_sum_agente['DIS_241']])
        
        sheet['D82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D82'] = t_sum_cat_1     
        
        sheet['E82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E82'] = t_sum_cat_2 
        
        sheet['F82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F82'] = t_sum_cat_3
        
        sheet['G82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G82'] = t_sum_cat_4 
        
        sheet['H82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H82'] = t_sum_cat_5
        
        sheet['I82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I82'] = t_sum_cat_6 
        
        ############################
        ###  CAPACITACION COMITE ###
        #############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_comite_microred:
            for col_comite in col_sum_comite:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_comite.keys()).index(col_comite) + 3
                    col_sum_comite[col_comite] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_comite, total_cell_comite in col_ubi_comite.items():
            if col_comite in col_sum_comite:
                # Obtener la celda correspondiente según la ubicación
                cell_comite = sheet[total_cell_comite]
                # Asignar el valor de la suma a la celda
                cell_comite.value = col_sum_comite[col_comite]
                # Aplicar formato a la celda
                cell_comite.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_comite.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_comite.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_comite['DIS_267']])
        t_sum_cat_2 = sum([col_sum_comite['DIS_268']])
        t_sum_cat_3 = sum([col_sum_comite['DIS_269']])
        t_sum_cat_4 = sum([col_sum_comite['DIS_270']])
        t_sum_cat_5 = sum([col_sum_comite['DIS_271']])
        t_sum_cat_6 = sum([col_sum_comite['DIS_272']])
        
        sheet['D97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D97'] = t_sum_cat_1     
        
        sheet['E97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E97'] = t_sum_cat_2 
        
        sheet['F97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F97'] = t_sum_cat_3
        
        sheet['G97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G97'] = t_sum_cat_4 
        
        sheet['H97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H97'] = t_sum_cat_5
        
        sheet['I97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I97'] = t_sum_cat_6 
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_operacional_microredes.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        workbook.save(response)

        return response

################################################
# REPORTE POR ESTABLECIMIENTOS
################################################
def get_establecimientos(request,establecimiento_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request,'discapacidad/establecimientos.html', context)

def p_microredes_establec(request):
    redes_param = request.GET.get('redes') 
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    return render(request, 'discapacidad/partials/p_microredes_establec.html', context)

def p_establecimientos(request):
    microredes = request.GET.get('p_microredes_establec')    
    codigo_red = request.GET.get('redes')
    establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_MicroRed=microredes,Codigo_Red=codigo_red,Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_Unico','Nombre_Establecimiento').distinct()
    context= {
        'establec': establec
    }
    return render(request, 'discapacidad/partials/p_establecimientos.html', context)

def rpt_operacional_fisico_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,               
                    SUM(dis_1) AS dis_1,
                    SUM(dis_2) AS dis_2,
                    SUM(dis_3) AS dis_3,
                    SUM(dis_4) AS dis_4,
                    SUM(dis_5) AS dis_5,
                    SUM(dis_6) AS dis_6,
                    SUM(dis_7) AS dis_7,
                    SUM(dis_8) AS dis_8,
                    SUM(dis_9) AS dis_9,
                    SUM(dis_10) AS dis_10,
                    SUM(dis_11) AS dis_11,
                    SUM(dis_12) AS dis_12,
                    SUM(dis_13) AS dis_13,
                    SUM(dis_14) AS dis_14,
                    SUM(dis_15) AS dis_15,
                    SUM(dis_16) AS dis_16,
                    SUM(dis_17) AS dis_17,
                    SUM(dis_18) AS dis_18,
                    SUM(dis_19) AS dis_19,
                    SUM(dis_20) AS dis_20,
                    SUM(dis_21) AS dis_21,
                    SUM(dis_22) AS dis_22,
                    SUM(dis_23) AS dis_23,
                    SUM(dis_24) AS dis_24,
                    SUM(dis_25) AS dis_25,
                    SUM(dis_26) AS dis_26,
                    SUM(dis_27) AS dis_27,
                    SUM(dis_28) AS dis_28,
                    SUM(dis_29) AS dis_29,
                    SUM(dis_30) AS dis_30,
                    SUM(dis_31) AS dis_31,
                    SUM(dis_32) AS dis_32,
                    SUM(dis_33) AS dis_33,
                    SUM(dis_34) AS dis_34,
                    SUM(dis_35) AS dis_35,
                    SUM(dis_36) AS dis_36,
                    SUM(dis_37) AS dis_37,
                    SUM(dis_38) AS dis_38,
                    SUM(dis_39) AS dis_39,
                    SUM(dis_40) AS dis_40,
                    SUM(dis_41) AS dis_41,
                    SUM(dis_42) AS dis_42,
                    SUM(dis_43) AS dis_43,
                    SUM(dis_44) AS dis_44,
                    SUM(dis_45) AS dis_45,
                    SUM(dis_46) AS dis_46,
                    SUM(dis_47) AS dis_47,
                    SUM(dis_48) AS dis_48,
                    SUM(dis_49) AS dis_49,
                    SUM(dis_50) AS dis_50,
                    SUM(dis_51) AS dis_51,
                    SUM(dis_52) AS dis_52,
                    SUM(dis_53) AS dis_53,
                    SUM(dis_54) AS dis_54,
                    SUM(dis_55) AS dis_55,
                    SUM(dis_56) AS dis_56,
                    SUM(dis_57) AS dis_57,
                    SUM(dis_58) AS dis_58,
                    SUM(dis_59) AS dis_59,
                    SUM(dis_60) AS dis_60,
                    SUM(dis_61) AS dis_61,
                    SUM(dis_62) AS dis_62,
                    SUM(dis_63) AS dis_63,
                    SUM(dis_64) AS dis_64,
                    SUM(dis_65) AS dis_65,
                    SUM(dis_66) AS dis_66,
                    SUM(dis_67) AS dis_67,
                    SUM(dis_68) AS dis_68,
                    SUM(dis_69) AS dis_69,
                    SUM(dis_70) AS dis_70,
                    SUM(dis_71) AS dis_71,
                    SUM(dis_72) AS dis_72,
                    SUM(dis_73) AS dis_73,
                    SUM(dis_74) AS dis_74,
                    SUM(dis_75) AS dis_75,
                    SUM(dis_76) AS dis_76,
                    SUM(dis_77) AS dis_77,
                    SUM(dis_78) AS dis_78,
                    SUM(dis_79) AS dis_79,
                    SUM(dis_80) AS dis_80,
                    SUM(dis_81) AS dis_81,
                    SUM(dis_82) AS dis_82,
                    SUM(dis_83) AS dis_83,
                    SUM(dis_84) AS dis_84,
                    SUM(dis_85) AS dis_85,
                    SUM(dis_86) AS dis_86,
                    SUM(dis_87) AS dis_87,
                    SUM(dis_88) AS dis_88,
                    SUM(dis_89) AS dis_89,
                    SUM(dis_90) AS dis_90,
                    SUM(dis_91) AS dis_91,
                    SUM(dis_92) AS dis_92,
                    SUM(dis_93) AS dis_93,
                    SUM(dis_94) AS dis_94,
                    SUM(dis_95) AS dis_95,
                    SUM(dis_96) AS dis_96,
                    SUM(dis_97) AS dis_97,
                    SUM(dis_98) AS dis_98,
                    SUM(dis_99) AS dis_99,
                    SUM(dis_100) AS dis_100,
                    SUM(dis_101) AS dis_101,
                    SUM(dis_102) AS dis_102,
                    SUM(dis_103) AS dis_103,
                    SUM(dis_104) AS dis_104,
                    SUM(dis_105) AS dis_105,
                    SUM(dis_106) AS dis_106,
                    SUM(dis_107) AS dis_107,
                    SUM(dis_108) AS dis_108,
                    SUM(dis_109) AS dis_109,
                    SUM(dis_110) AS dis_110,
                    SUM(dis_111) AS dis_111,
                    SUM(dis_112) AS dis_112,
                    SUM(dis_113) AS dis_113,
                    SUM(dis_114) AS dis_114,
                    SUM(dis_115) AS dis_115,
                    SUM(dis_116) AS dis_116,
                    SUM(dis_117) AS dis_117,
                    SUM(dis_118) AS dis_118,
                    SUM(dis_119) AS dis_119,
                    SUM(dis_120) AS dis_120,
                    SUM(dis_121) AS dis_121,
                    SUM(dis_122) AS dis_122,
                    SUM(dis_123) AS dis_123,
                    SUM(dis_124) AS dis_124,
                    SUM(dis_125) AS dis_125,
                    SUM(dis_126) AS dis_126,
                    SUM(dis_127) AS dis_127,
                    SUM(dis_128) AS dis_128,
                    SUM(dis_129) AS dis_129,
                    SUM(dis_130) AS dis_130, 
                    SUM(dis_131) AS dis_131,
                    SUM(dis_132) AS dis_132,
                    SUM(dis_133) AS dis_133,
                    SUM(dis_134) AS dis_134,
                    SUM(dis_135) AS dis_135,
                    SUM(dis_136) AS dis_136,
                    SUM(dis_137) AS dis_137,
                    SUM(dis_138) AS dis_138,
                    SUM(dis_139) AS dis_139,
                    SUM(dis_140) AS dis_140, 
                    SUM(dis_141) AS dis_141,
                    SUM(dis_142) AS dis_142,
                    SUM(dis_143) AS dis_143,
                    SUM(dis_144) AS dis_144,
                    SUM(dis_145) AS dis_145,
                    SUM(dis_146) AS dis_146,
                    SUM(dis_147) AS dis_147,
                    SUM(dis_148) AS dis_148,
                    SUM(dis_149) AS dis_149,
                    SUM(dis_150) AS dis_150,
                    SUM(dis_151) AS dis_151,
                    SUM(dis_152) AS dis_152,
                    SUM(dis_153) AS dis_153,
                    SUM(dis_154) AS dis_154,
                    SUM(dis_155) AS dis_155,
                    SUM(dis_156) AS dis_156,
                    SUM(dis_157) AS dis_157,
                    SUM(dis_158) AS dis_158,
                    SUM(dis_159) AS dis_159,
                    SUM(dis_160) AS dis_160 
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_1,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_2,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_3,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_4,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_5,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_6,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_7,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_8,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_9,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_10,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_11,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_12,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_13,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_14,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_15,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_16,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_17,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_18,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_19,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_20,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_21,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_22,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_23,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_24,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_25,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_26,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_27,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_28,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_29,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_30,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_31,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_32,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_33,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_34,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_35,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_36,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_37,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_38,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_39,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_40,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_41,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_42,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_43,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_44,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_45,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_46,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_47,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_48,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_49,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_50,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_51,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_52,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_53,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_54,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_55,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_56,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_57,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_58,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_59,
                        SUM(CASE WHEN Categoria = 12 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_60,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_61,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_62,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_63,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_64,
                        SUM(CASE WHEN Categoria = 13 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_65,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_66,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_67,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_68,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_69,
                        SUM(CASE WHEN Categoria = 14 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_70,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_71,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_72,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_73,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_74,
                        SUM(CASE WHEN Categoria = 15 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_75,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_76,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_77,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_78,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_79,
                        SUM(CASE WHEN Categoria = 16 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_80,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_81,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_82,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_83,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_84,
                        SUM(CASE WHEN Categoria = 17 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_85,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_86,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_87,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_88,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_89,
                        SUM(CASE WHEN Categoria = 18 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_90,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_91,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_92,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_93,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_94,
                        SUM(CASE WHEN Categoria = 19 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_95,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_96,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_97,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_98,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_99,
                        SUM(CASE WHEN Categoria = 20 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_100,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_101,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_102,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_103,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_104,
                        SUM(CASE WHEN Categoria = 21 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_105,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_106,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_107,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_108,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_109,
                        SUM(CASE WHEN Categoria = 22 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_110,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_111,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_112,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_113,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_114,
                        SUM(CASE WHEN Categoria = 23 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_115,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_116,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_117,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_118,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_119,
                        SUM(CASE WHEN Categoria = 24 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_120,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_121,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_122,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_123,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_124,
                        SUM(CASE WHEN Categoria = 25 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_125,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_126,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_127,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_128,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_129,
                        SUM(CASE WHEN Categoria = 26 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_130,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_131,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_132,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_133,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_134,
                        SUM(CASE WHEN Categoria = 27 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_135,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_136,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_137,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_138,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_139,
                        SUM(CASE WHEN Categoria = 28 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_140,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_141,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_142,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_143,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_144,
                        SUM(CASE WHEN Categoria = 29 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_145,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_146,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_147,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_148,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_149,
                        SUM(CASE WHEN Categoria = 30 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_150,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_151,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_152,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_153,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_154,
                        SUM(CASE WHEN Categoria = 31 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_155,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_156,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_157,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_158,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_159,
                        SUM(CASE WHEN Categoria = 32 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_160
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_02_FISICA_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])       
        resultado_establec = cursor.fetchall()
    return resultado_establec

def rpt_operacional_sensorial_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,   
                    SUM(dis_161) AS dis_161,
                    SUM(dis_162) AS dis_162,
                    SUM(dis_163) AS dis_163,
                    SUM(dis_164) AS dis_164,
                    SUM(dis_165) AS dis_165,
                    SUM(dis_166) AS dis_166,
                    SUM(dis_167) AS dis_167,
                    SUM(dis_168) AS dis_168,
                    SUM(dis_169) AS dis_169,
                    SUM(dis_170) AS dis_170,
                    SUM(dis_171) AS dis_171,
                    SUM(dis_172) AS dis_172,
                    SUM(dis_173) AS dis_173,
                    SUM(dis_174) AS dis_174,
                    SUM(dis_175) AS dis_175,
                    SUM(dis_176) AS dis_176,
                    SUM(dis_177) AS dis_177,
                    SUM(dis_178) AS dis_178,
                    SUM(dis_179) AS dis_179,
                    SUM(dis_180) AS dis_180,
                    SUM(dis_181) AS dis_181,
                    SUM(dis_182) AS dis_182,
                    SUM(dis_183) AS dis_183,
                    SUM(dis_184) AS dis_184,
                    SUM(dis_185) AS dis_185,
                    SUM(dis_186) AS dis_186,
                    SUM(dis_187) AS dis_187,
                    SUM(dis_188) AS dis_188,
                    SUM(dis_189) AS dis_189,
                    SUM(dis_190) AS dis_190,
                    SUM(dis_191) AS dis_191,
                    SUM(dis_192) AS dis_192,
                    SUM(dis_193) AS dis_193,
                    SUM(dis_194) AS dis_194,
                    SUM(dis_195) AS dis_195
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_161,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_162,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_163,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_164,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_165,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_166,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_167,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_168,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_169,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_170,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_171,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_172,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_173,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_174,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_175,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_176,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_177,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_178,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_179,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_180,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_181,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_182,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_183,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_184,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_185,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_186,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_187,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_188,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_189,
                        SUM(CASE WHEN Categoria = 6 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_190,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_191,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_192,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_193,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_194,
                        SUM(CASE WHEN Categoria = 7 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_195
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                    AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_sensorial_establec = cursor.fetchall()    
    return resultado_sensorial_establec

def rpt_operacional_certificado_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,   
                    SUM(dis_216) AS dis_216,
                    SUM(dis_217) AS dis_217,
                    SUM(dis_218) AS dis_218,
                    SUM(dis_219) AS dis_219,
                    SUM(dis_220) AS dis_220,
                    SUM(dis_221) AS dis_221,
                    SUM(dis_222) AS dis_222,
                    SUM(dis_223) AS dis_223,
                    SUM(dis_224) AS dis_224,
                    SUM(dis_225) AS dis_225,
                    SUM(dis_226) AS dis_226,
                    SUM(dis_227) AS dis_227,
                    SUM(dis_228) AS dis_228,
                    SUM(dis_229) AS dis_229,
                    SUM(dis_230) AS dis_230,
                    SUM(dis_231) AS dis_231,
                    SUM(dis_232) AS dis_232,
                    SUM(dis_233) AS dis_233,
                    SUM(dis_234) AS dis_234,
                    SUM(dis_235) AS dis_235
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_216,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_217,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_218,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_219,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_220,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_221,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_222,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_223,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_224,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_225,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_226,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_227,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_228,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_229,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_230,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_231,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_232,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_233,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_234,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_235
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico,  TRAMA_BASE_DISCAPACIDAD_RPT_04_CERTIFICADO_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_certificado_establec = cursor.fetchall()
    return resultado_certificado_establec

def rpt_operacional_rbc_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,   
                    SUM(dis_242) AS dis_242,
                    SUM(dis_243) AS dis_243,
                    SUM(dis_244) AS dis_244,
                    SUM(dis_245) AS dis_245,
                    SUM(dis_246) AS dis_246,
                    SUM(dis_247) AS dis_247,
                    SUM(dis_248) AS dis_248,
                    SUM(dis_249) AS dis_249,
                    SUM(dis_250) AS dis_250,
                    SUM(dis_251) AS dis_251,
                    SUM(dis_252) AS dis_252,
                    SUM(dis_253) AS dis_253,
                    SUM(dis_254) AS dis_254,
                    SUM(dis_255) AS dis_255,
                    SUM(dis_256) AS dis_256,
                    SUM(dis_257) AS dis_257,
                    SUM(dis_258) AS dis_258,
                    SUM(dis_259) AS dis_259,
                    SUM(dis_260) AS dis_260, 
                    SUM(dis_261) AS dis_261, 
                    SUM(dis_262) AS dis_262, 
                    SUM(dis_263) AS dis_263, 
                    SUM(dis_264) AS dis_264, 
                    SUM(dis_265) AS dis_265, 
                    SUM(dis_266) AS dis_266 
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_242,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_243,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_244,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_245,
                        SUM(CASE WHEN Categoria = 1 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_246,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_247,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_248,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_249,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_250,
                        SUM(CASE WHEN Categoria = 2 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_251,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_252,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_253,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_254,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_255,
                        SUM(CASE WHEN Categoria = 3 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_256,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_257,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_258,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_259,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_260,
                        SUM(CASE WHEN Categoria = 4 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_261,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_262,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_263,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_264,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_265,
                        SUM(CASE WHEN Categoria = 5 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_266
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_05_RBC_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_rbc_establec = cursor.fetchall()
    return resultado_rbc_establec

def rpt_operacional_mental_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,   
                    SUM(dis_196) AS dis_196,
                    SUM(dis_197) AS dis_197,
                    SUM(dis_198) AS dis_198,
                    SUM(dis_199) AS dis_199,
                    SUM(dis_200) AS dis_200,
                    SUM(dis_201) AS dis_201,
                    SUM(dis_202) AS dis_202,
                    SUM(dis_203) AS dis_203,
                    SUM(dis_204) AS dis_204,
                    SUM(dis_205) AS dis_205,
                    SUM(dis_206) AS dis_206,
                    SUM(dis_207) AS dis_207,
                    SUM(dis_208) AS dis_208,
                    SUM(dis_209) AS dis_209,
                    SUM(dis_210) AS dis_210, 
                    SUM(dis_211) AS dis_211, 
                    SUM(dis_212) AS dis_212, 
                    SUM(dis_213) AS dis_213, 
                    SUM(dis_214) AS dis_214, 
                    SUM(dis_215) AS dis_215
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_196,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_197,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_198,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_199,
                        SUM(CASE WHEN Categoria = 8 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_200,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_201,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_202,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_203,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_204,
                        SUM(CASE WHEN Categoria = 9 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_205,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_206,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_207,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_208,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_209,
                        SUM(CASE WHEN Categoria = 10 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_210,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 1 THEN 1 ELSE 0 END) AS dis_211,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 2 THEN 1 ELSE 0 END) AS dis_212,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 3 THEN 1 ELSE 0 END) AS dis_213,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 4 THEN 1 ELSE 0 END) AS dis_214,
                        SUM(CASE WHEN Categoria = 11 AND gedad = 5 THEN 1 ELSE 0 END) AS dis_215
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_03_SENSORIAL_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_mental_establec = cursor.fetchall()
    return resultado_mental_establec

def rpt_operacional_capacitacion_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,   
                    SUM(dis_273) AS dis_273,
                    SUM(dis_274) AS dis_274
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        COUNT(Categoria) AS dis_273,
                        SUM(gedad) AS dis_274
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_01_CAPACITACION_MEDICINA_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_capacitacion_establec = cursor.fetchall()
    return resultado_capacitacion_establec

def rpt_operacional_agente_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,   
                    SUM(dis_236) AS dis_236,
                    SUM(dis_237) AS dis_237,
                    SUM(dis_238) AS dis_238,
                    SUM(dis_239) AS dis_239,
                    SUM(dis_240) AS dis_240,
                    SUM(dis_241) AS dis_241
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        SUM(CASE WHEN Categoria = 1 THEN 1 ELSE 0 END) 	   AS dis_236,
                        SUM(CASE WHEN Categoria = 1 THEN gedad ELSE 0 END) AS dis_237,
                        SUM(CASE WHEN Categoria = 2 THEN 1 ELSE 0 END)     AS dis_238,
                        SUM(CASE WHEN Categoria = 2 THEN gedad ELSE 0 END) AS dis_239,
                        SUM(CASE WHEN Categoria = 3 THEN 1 ELSE 0 END)     AS dis_240,
                        SUM(CASE WHEN Categoria = 3 THEN gedad ELSE 0 END) AS dis_241
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_06_CAPACITACION_AGENTE_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_agente_establec = cursor.fetchall()
    return resultado_agente_establec

def rpt_operacional_comite_establec(establec,fecha_inicio, fecha_fin):
    with connection.cursor() as cursor:
        # Crear una tabla temporal
        # Insertar los datos agrupados y las sumas en la tabla temporal
        cursor.execute("""
                SELECT
                    renaes,   
                    SUM(dis_267) AS dis_267,
                    SUM(dis_268) AS dis_268,
                    SUM(dis_269) AS dis_269,
                    SUM(dis_270) AS dis_270,
                    SUM(dis_271) AS dis_271,
                    SUM(dis_272) AS dis_272
                FROM (
                    SELECT
                        MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico AS codigo_unico,
                        renaes,
                        SUM(CASE WHEN Actividad = 1 THEN 1 ELSE 0 END) 		AS dis_267,
                        SUM(CASE WHEN Actividad = 1 THEN Partic ELSE 0 END) AS dis_268,
                        SUM(CASE WHEN Actividad = 2 THEN 1 ELSE 0 END)      AS dis_269,
                        SUM(CASE WHEN Actividad = 2 THEN Partic ELSE 0 END) AS dis_270,
                        SUM(CASE WHEN Actividad = 3 THEN 1 ELSE 0 END)      AS dis_271,
                        SUM(CASE WHEN Actividad = 3 THEN Partic ELSE 0 END) AS dis_272
                    FROM TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL
                    LEFT JOIN MAESTRO_HIS_ESTABLECIMIENTO ON TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes = MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico
                    WHERE MAESTRO_HIS_ESTABLECIMIENTO.codigo_unico = %s   
                        AND TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.periodo BETWEEN CAST(%s AS INT) AND CAST(%s AS INT)
                    GROUP BY MAESTRO_HIS_ESTABLECIMIENTO.Codigo_Unico, TRAMA_BASE_DISCAPACIDAD_RPT_08_PREV_TALLER_EESS_NOMINAL.renaes
                ) subquery
                GROUP BY codigo_unico, renaes
                """, [str(establec), str(fecha_inicio) + '01', str(fecha_fin) + '31'])
        resultado_comite_establec = cursor.fetchall()
    return resultado_comite_establec

class RptOperacinalEstablec(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')
        establec = request.GET.get('p_establecimiento')

        # Creación de la consulta
        resultado_establec = rpt_operacional_fisico_establec(establec,fecha_inicio, fecha_fin)
        resultado_sensorial_establec = rpt_operacional_sensorial_establec(establec, fecha_inicio, fecha_fin)
        resultado_certificado_establec = rpt_operacional_certificado_establec(establec, fecha_inicio, fecha_fin)
        resultado_rbc_establec = rpt_operacional_rbc_establec(establec, fecha_inicio, fecha_fin)
        resultado_mental_establec = rpt_operacional_mental_establec(establec,fecha_inicio, fecha_fin)
        resultado_capacitacion_establec = rpt_operacional_capacitacion_establec(establec,fecha_inicio, fecha_fin)
        resultado_agente_establec = rpt_operacional_agente_establec(establec, fecha_inicio, fecha_fin)
        resultado_comite_establec = rpt_operacional_comite_establec(establec, fecha_inicio, fecha_fin)
        
        establec_codigo = list(MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
            Codigo_Unico=establec
        ).values_list('Nombre_Establecimiento', flat=True).distinct())
        
        fecha_inicio_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_inicio
        ).values_list('Mes', flat=True).distinct())
        
        fecha_fin_codigo = list(DimPeriodo.objects.filter(
            Periodo__startswith=fecha_fin
        ).values_list('Mes', flat=True).distinct())
        
        # Crear un nuevo libro de Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # cambia el alto de la columna
        sheet.row_dimensions[1].height = 14
        sheet.row_dimensions[2].height = 14
        sheet.row_dimensions[4].height = 25
        sheet.row_dimensions[15].height = 25
        # cambia el ancho de la columna
        sheet.column_dimensions['A'].width = 2
        sheet.column_dimensions['B'].width = 28
        sheet.column_dimensions['C'].width = 28
        sheet.column_dimensions['D'].width = 9
        sheet.column_dimensions['E'].width = 9
        sheet.column_dimensions['F'].width = 9
        sheet.column_dimensions['G'].width = 9
        sheet.column_dimensions['H'].width = 9
        sheet.column_dimensions['I'].width = 9
        sheet.column_dimensions['J'].width = 9
        sheet.column_dimensions['K'].width = 9
        sheet.column_dimensions['L'].width = 9
        # linea de division
        sheet.freeze_panes = 'AL8'
        
        # Configuración del fondo y el borde
        fill = PatternFill(patternType='solid', fgColor='00B0F0')
        border = Border(left=Side(style='thin', color='00B0F0'),
                        right=Side(style='thin', color='00B0F0'),
                        top=Side(style='thin', color='00B0F0'),
                        bottom=Side(style='thin', color='00B0F0'))

        borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                        right=Side(style='thin', color='A9A9A9'), # Plomo
                        top=Side(style='thin', color='A9A9A9'), # Plomo
                        bottom=Side(style='thin', color='A9A9A9')) # Plomo

        # crea titulo del reporte
        sheet['B1'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B1'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
        
        sheet['B2'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B2'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
        
        sheet['B4'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B4'].font = Font(name = 'Arial', size= 12, bold = True)
        sheet['B4'] = 'REPORTE DE ACTIVIDADES DEL COMPONENTE DE DISCAPACIDAD'
        
        sheet['B6'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B6'] ='DIRESA / GERESA / DISA'
        
        sheet['C6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C6'].font = Font(name = 'Arial', size= 7)
        sheet['C6'] ='JUNIN'

        sheet['B7'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B7'] ='PROV/ DIST/ RED/ MR/ ESTABLEC'
        
        sheet['C7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C7'].font = Font(name = 'Arial', size= 7)
        sheet['C7'] = establec_codigo[0]
        
        sheet['E6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['E6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['E6'] ='PERIODO'
        
        sheet['F6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F6'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F6'] ='MES INICIO'
        
        sheet['F7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['F7'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['F7'] ='MES FIN'
        
        sheet['G6'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G6'].font = Font(name = 'Arial', size= 8)
        sheet['G6'] = fecha_inicio_codigo[0]
        
        sheet['G7'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['G7'].font = Font(name = 'Arial', size= 8)
        sheet['G7'] = fecha_fin_codigo[0]
        
        sheet['B9'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B9'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B9'] ='PERSONAS CON DISCAPACIDAD RECIBEN ATENCION DE REHABILITACION EN ESTABLECIMIENTOS DE SALUD (3000688)'
        
        sheet['B10'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B10'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['B10'] ='Capacitación en medicina de rehabilitación integral (5004449)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=12, max_row=12, min_col=3, max_col=5):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['C12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['C12'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['C12'] ='Capacitación  (C0009)' 
        
        sheet['D11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D11'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D11'].fill = fill
        sheet['D11'].border = border
        sheet['D11'] = 'N°'
                
        sheet['E11'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E11'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['E11'].fill = fill
        sheet['E11'].border = border
        sheet['E11'] = 'Capacitados'
        #######################################################
        ########## DISCAPACIDAD FISICA ########################
        #######################################################
        sheet['B14'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B14'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B14'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Física (5005150)' 
                
        sheet['B15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B15'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B15'].fill = fill
        sheet['B15'].border = border
        sheet['B15'] = 'Atenciones'
        
        sheet['D15'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D15'].fill = fill
        sheet['D15'].border = border
        sheet['D15'] = 'Total'
        
        sheet['E15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E15'].fill = fill
        sheet['E15'].border = border
        sheet['E15'] = 'Niños         (1d - 11a)'
        
        sheet['F15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F15'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F15'].fill = fill
        sheet['F15'].border = border
        sheet['F15'] = 'Adolescentes (12a - 17a)'
        
        sheet['G15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G15'].fill = fill
        sheet['G15'].border = border
        sheet['G15'] = 'Jóvenes (18a - 29a)'
        
        sheet['H15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H15'].fill = fill
        sheet['H15'].border = border
        sheet['H15'] = 'Adultos (30a - 59a)'
        
        sheet['I15'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I15'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I15'].fill = fill
        sheet['I15'].border = border
        sheet['I15'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=16, max_row=47, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B16'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B16'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B16'] ='LESIONES MEDULARES' 
                
        sheet['B17'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B17'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B17'] ='ENFERMEDAD DE PARKINSON Y SIMILARES' 
        
        sheet['B18'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B18'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B18'] ='REHABILITACIÓN EN PACIENTES AMPUTADOS' 
                
        sheet['B20'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B20'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B20'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA NEUROLÓGICA' 
        
        sheet['B23'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B23'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B23'] ='TRASTORNOS DEL DESARROLLO DE LA FUNCIÓN MOTRIZ' 
        
        sheet['B24'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B24'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B24'] ='ATENCIÓN DE REHABILITACIÓN DE ENFERMEDAD ARTICULAR DEGENERATIVA' 
        
        sheet['B25'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B25'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B25'] ='ENCEFALOPATÍA INFANTIL' 
                
        sheet['B26'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B26'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B26'] ='SÍNDROME DOWN' 
        
        sheet['B27'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B27'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B27'] ='REHABILITACIÓN EN PATOLOGÍA DE LA COLUMNA VERTEBRAL Y OTROS TRASTORNOS POSTURALES' 
        
        sheet['B34'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B34'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B34'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD CARDIOVASCULAR' 
        
        sheet['B35'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B35'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B35'] ='ATENCIÓN DE REHABILITACIÓN EN ENFERMEDAD RESPIRATORIA' 
        
        sheet['B36'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B36'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B36'] ='ATENCIÓN DE REHABILITACIÓN EN ALTERACIONES DEL PISO PÉLVICO' 
        
        sheet['B37'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B37'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B37'] ='ATENCIÓN DE REHABILITACIÓN EN PATOLOGÍA TRAUMATOLÓGICA Y REUMATOLÓGICA' 
        
        sheet['B44'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B44'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B44'] ='ATENCIÓN DE REHABILITACIÓN ONCOLÓGICA' 
        
        sheet['B46'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B46'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B46'] ='ATENCIÓN DE REHABILITACIÓN EN DOLOR' 
        
        sheet['B47'].alignment = Alignment(horizontal= "left", vertical="center",wrap_text=True)
        sheet['B47'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B47'] ='ATENCIÓN DE REHABILITACIÓN EN PACIENTES QUEMADOS' 
        ####     
        sheet['C16'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C16'].font = Font(name = 'Arial', size= 7)
        sheet['C16'] ='Lesiones medulares' 
    
        sheet['C17'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C17'].font = Font(name = 'Arial', size= 7)
        sheet['C17'] ='Enfermedad de Parkinson y similares' 
        
        sheet['C18'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C18'].font = Font(name = 'Arial', size= 7)
        sheet['C18'] ='Amputados de miembros superiores' 
        
        sheet['C19'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C19'].font = Font(name = 'Arial', size= 7)
        sheet['C19'] ='Amputados de miembros inferiores' 
        
        sheet['C20'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C20'].font = Font(name = 'Arial', size= 7)
        sheet['C20'] ='Enfermedades cerebrovasculares'
        
        sheet['C21'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C21'].font = Font(name = 'Arial', size= 7)
        sheet['C21'] ='Enfermedades musculares y de la unión mioneural'
        
        sheet['C22'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C22'].font = Font(name = 'Arial', size= 7)
        sheet['C22'] ='Lesiones de nervios periféricos'
        
        sheet['C23'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C23'].font = Font(name = 'Arial', size= 7)
        sheet['C23'] ='Trastornos del desarrollo de la funcion motriz'
        
        sheet['C24'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C24'].font = Font(name = 'Arial', size= 7)
        sheet['C24'] ='Enfermedad articular degenerativa'
        
        sheet['C25'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C25'].font = Font(name = 'Arial', size= 7)
        sheet['C25'] ='Encefalopatía infantil y otras lesiones'
        
        sheet['C26'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C26'].font = Font(name = 'Arial', size= 7)
        sheet['C26'] ='Sindrome de Down'
        
        sheet['C27'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C27'].font = Font(name = 'Arial', size= 7)
        sheet['C27'] ='Cifosis y lordosis'
        
        sheet['C28'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C28'].font = Font(name = 'Arial', size= 7)
        sheet['C28'] ='Espondilo artropatías'
        
        sheet['C29'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C29'].font = Font(name = 'Arial', size= 7)
        sheet['C29'] ='Otros trastornos de los discos intervertebrales'
        
        sheet['C30'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C30'].font = Font(name = 'Arial', size= 7)
        sheet['C30'] ='Cervicalgia, dorsalgia, lumbago'
        
        sheet['C31'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C31'].font = Font(name = 'Arial', size= 7)
        sheet['C31'] ='Otras dorsopatías deformantes'
        
        sheet['C32'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C32'].font = Font(name = 'Arial', size= 7)
        sheet['C32'] ='Otros trastornos articulares'
        
        sheet['C33'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C33'].font = Font(name = 'Arial', size= 7)
        sheet['C33'] ='Defectos en la longitud de extremidades'
        
        sheet['C34'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C34'].font = Font(name = 'Arial', size= 7)
        sheet['C34'] ='Enfermedad cardiovascular'
        
        sheet['C35'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C35'].font = Font(name = 'Arial', size= 7)
        sheet['C35'] ='Enfermedad respiratoria'
        
        sheet['C36'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C36'].font = Font(name = 'Arial', size= 7)
        sheet['C36'] ='Vejiga neurogénica y dolor'
        
        sheet['C37'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C37'].font = Font(name = 'Arial', size= 7)
        sheet['C37'] ='Incontinencia'
        
        sheet['C38'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C38'].font = Font(name = 'Arial', size= 7)
        sheet['C38'] ='Prolapso'
        
        sheet['C39'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C39'].font = Font(name = 'Arial', size= 7)
        sheet['C39'] ='Traumatismos'
        
        sheet['C40'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C40'].font = Font(name = 'Arial', size= 7)
        sheet['C40'] ='Enfermedades del tejido conectivo'
        
        sheet['C41'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C41'].font = Font(name = 'Arial', size= 7)
        sheet['C41'] ='Patología articular excluida columna'
        
        sheet['C42'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C42'].font = Font(name = 'Arial', size= 7)
        sheet['C42'] ='Lesiones infecciosas'
        
        sheet['C43'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C43'].font = Font(name = 'Arial', size= 7)
        sheet['C43'] ='Lesión biomecánica'
        
        sheet['C44'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C44'].font = Font(name = 'Arial', size= 7)
        sheet['C44'] ='Linfedema'
        
        sheet['C45'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C45'].font = Font(name = 'Arial', size= 7)
        sheet['C45'] ='Sarcopenia'
        
        sheet['C46'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C46'].font = Font(name = 'Arial', size= 7)
        sheet['C46'] ='Dolor'
        
        sheet['C47'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C47'].font = Font(name = 'Arial', size= 7)
        sheet['C47'] ='Quemaduras, corrosiones y congelaciones'
        
        ##########################################################    
        ########## DISCAPACIDAD SENSORIAL ########################
        ##########################################################
        sheet['B50'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B50'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B50'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Sensorial (5005151)' 
                
        sheet['B51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B51'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B51'].fill = fill
        sheet['B51'].border = border
        sheet['B51'] = 'Atenciones'
        
        sheet['D51'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D51'].fill = fill
        sheet['D51'].border = border
        sheet['D51'] = 'Total'
        
        sheet['E51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E51'].fill = fill
        sheet['E51'].border = border
        sheet['E51'] = 'Niños         (1d - 11a)'
        
        sheet['F51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F51'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F51'].fill = fill
        sheet['F51'].border = border
        sheet['F51'] = 'Adolescentes (12a - 17a)'
        
        sheet['G51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G51'].fill = fill
        sheet['G51'].border = border
        sheet['G51'] = 'Jóvenes (18a - 29a)'
        
        sheet['H51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H51'].fill = fill
        sheet['H51'].border = border
        sheet['H51'] = 'Adultos (30a - 59a)'
        
        sheet['I51'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I51'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I51'].fill = fill
        sheet['I51'].border = border
        sheet['I51'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=52, max_row=58, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B52'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B52'] ='HIPOACUSIA Y/O SORDERA' 
        
        sheet['B53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B53'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B53'] ='BAJA VISION Y/O CEGUERA' 
        
        sheet['B54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B54'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B54'] ='SORDOMUDEZ' 
        
        sheet['B55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B55'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B55'] ='ENFERMEDAD CEREBRO VASCULAR' 
        
        sheet['B56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B56'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B56'] ='TRASTORNOS ESPECIFICOS DEL DESARROLLO DEL HABLA Y LENGUAJE' 
        
        sheet['B57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B57'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B57'] ='DISARTRIA Y DISFAGIA' 
        
        sheet['B59'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B59'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B59'] ='SUB TOTAL' 
        
        ########               
        sheet['C52'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C52'].font = Font(name = 'Arial', size= 7)
        sheet['C52'] ='Hipoacusia y sordera' 
        
        sheet['C53'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C53'].font = Font(name = 'Arial', size= 7)
        sheet['C53'] ='Baja visión y ceguera' 
        
        sheet['C54'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C54'].font = Font(name = 'Arial', size= 7)
        sheet['C54'] ='Sordomudez' 
        
        sheet['C55'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C55'].font = Font(name = 'Arial', size= 7)
        sheet['C55'] ='Enfermedad Cerebro vascular' 
        
        sheet['C56'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C56'].font = Font(name = 'Arial', size= 7)
        sheet['C56'] ='Trastornos específicos del desarrollo del habla y lenguaje' 
        
        sheet['C57'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C57'].font = Font(name = 'Arial', size= 7)
        sheet['C57'] ='Disartria' 
        
        sheet['C58'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C58'].font = Font(name = 'Arial', size= 7)
        sheet['C58'] ='Disfagia' 
        
        ########################################################
        ########## DISCAPACIDAD MENTAL #########################
        ########################################################
        sheet['B61'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B61'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B61'] ='Atención de Rehabilitación en Personas con Discapacidad de Tipo Mental (5005152)' 
                
        sheet['B62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B62'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B62'].fill = fill
        sheet['B62'].border = border
        sheet['B62'] = 'Atenciones'
        
        sheet['D62'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D62'].fill = fill
        sheet['D62'].border = border
        sheet['D62'] = 'Total'
        
        sheet['E62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E62'].fill = fill
        sheet['E62'].border = border
        sheet['E62'] = 'Niños         (1d - 11a)'
        
        sheet['F62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F62'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F62'].fill = fill
        sheet['F62'].border = border
        sheet['F62'] = 'Adolescentes (12a - 17a)'
        
        sheet['G62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G62'].fill = fill
        sheet['G62'].border = border
        sheet['G62'] = 'Jóvenes (18a - 29a)'
        
        sheet['H62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H62'].fill = fill
        sheet['H62'].border = border
        sheet['H62'] = 'Adultos (30a - 59a)'
        
        sheet['I62'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I62'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I62'].fill = fill
        sheet['I62'].border = border
        sheet['I62'] = 'A Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=63, max_row=66, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B63'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B63'] ='TRASTORNOS DE APRENDIZAJE' 
        
        sheet['B64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B64'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B64'] ='RETRASO MENTAL LEVE, MODERADO, SEVERO' 
        
        sheet['B65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B65'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B65'] ='TRASTORNOS DEL ESPECTRO AUTISTA' 
        
        sheet['B66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B66'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B66'] ='OTROS TRASTORNOS DE SALUD MENTAL' 
        
        sheet['B67'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['B67'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B67'] ='SUB TOTAL' 
        
        ##########
        
        sheet['C63'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C63'].font = Font(name = 'Arial', size= 7)
        sheet['C63'] ='Trastornos del aprendizaje' 
        
        sheet['C64'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C64'].font = Font(name = 'Arial', size= 7)
        sheet['C64'] ='Retardo Mental: Leve, moderado, severo' 
        
        sheet['C65'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C65'].font = Font(name = 'Arial', size= 7)
        sheet['C65'] ='Trastornos del espectro autista' 
        
        sheet['C66'].alignment = Alignment(horizontal= "left", vertical="center", wrap_text=True)
        sheet['C66'].font = Font(name = 'Arial', size= 7)
        sheet['C66'] ='Otras alteraciones de salud mental' 
                
        ##################################################
        ########## CERTIFICACION #########################
        ##################################################
        sheet['B69'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B69'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B69'] ='PERSONAS CON DISCAPACIDAD CERTIFICADAS EN ESTABLECIMIENTOS DE SALUD (3000689)' 
                
        sheet['B70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B70'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B70'].fill = fill
        sheet['B70'].border = border
        sheet['B70'] = 'Atenciones'
        
        sheet['D70'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D70'].fill = fill
        sheet['D70'].border = border
        sheet['D70'] = 'Total'
        
        sheet['E70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E70'].fill = fill
        sheet['E70'].border = border
        sheet['E70'] = 'Niños         (1d - 11a)'
        
        sheet['F70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F70'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F70'].fill = fill
        sheet['F70'].border = border
        sheet['F70'] = 'Adolescentes (12a - 17a)'
        
        sheet['G70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G70'].fill = fill
        sheet['G70'].border = border
        sheet['G70'] = 'Jóvenes (18a - 29a)'
        
        sheet['H70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H70'].fill = fill
        sheet['H70'].border = border
        sheet['H70'] = 'Adultos (30a - 59a)'
        
        sheet['I70'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I70'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I70'].fill = fill
        sheet['I70'].border = border
        sheet['I70'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=71, max_row=74, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        sheet['B71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B71'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B71'] ='Certificación de Discapacidad (0515204)' 
        
        sheet['B74'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B74'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B74'] ='Certificación de Incapacidad (0515205)' 
        
        sheet['B75'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B75'].font = Font(name = 'Arial', size= 7, bold = True)
        sheet['B75'] ='SUB TOTAL' 
        
        sheet['C71'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C71'].font = Font(name = 'Arial', size= 7)
        sheet['C71'] ='Evaluación' 
        
        sheet['C72'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C72'].font = Font(name = 'Arial', size= 7)
        sheet['C72'] ='Calificación' 
        
        sheet['C73'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['C73'].font = Font(name = 'Arial', size= 7)
        sheet['C73'] ='Certificación' 

        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B77'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B77'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B77'] ='PERSONAS CON DISCAPACIDAD RECIBEN SERVICIOS DE REHABILITACIÓN BASADA EN LA COMUNIDAD (3000690)' 
        
        sheet['B78'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B78'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B78'] ='CAPACITACIÓN A AGENTES COMUNITARIOS EN REHABILITACIÓN BASADA EN LA COMUNIDAD (5005155)' 
        
        sheet['B82'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B82'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B82'] ='Capacitación a Agentes Comunitarios  (APP138)' 
        
        sheet['D80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D80'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D80'].fill = fill
        sheet['D80'].border = border
        sheet['D80'] = 'Taller'
        
        sheet['F80'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F80'].fill = fill
        sheet['F80'].border = border
        sheet['F80'] = 'Sesion Educativa'
        
        sheet['H80'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H80'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H80'].fill = fill
        sheet['H80'].border = border
        sheet['H80'] = 'Sesion Demostrativa'
        
        sheet['D81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D81'].fill = fill
        sheet['D81'].border = border
        sheet['D81'] = 'N°'
        
        sheet['E81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E81'].fill = fill
        sheet['E81'].border = border
        sheet['E81'] = 'Capacitados'
        
        sheet['F81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F81'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F81'].fill = fill
        sheet['F81'].border = border
        sheet['F81'] = 'N°'
        
        sheet['G81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G81'].fill = fill
        sheet['G81'].border = border
        sheet['G81'] = 'Capacitados'
        
        sheet['H81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H81'].fill = fill
        sheet['H81'].border = border
        sheet['H81'] = 'N° '
        
        sheet['I81'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I81'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I81'].fill = fill
        sheet['I81'].border = border
        sheet['I81'] = 'Capacitados'
        
        
        #borde plomo
        for row in sheet.iter_rows(min_row=82, max_row=82, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        
        ################################################
        ########## VISITAS RBC #########################
        ################################################
        sheet['B84'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B84'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B84'] ='Vistas a alas familias Rehabilitacion Basada en la Comunidad' 
                
        sheet['B85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['B85'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['B85'].fill = fill
        sheet['B85'].border = border
        sheet['B85'] = 'Visitas'
        
        sheet['D85'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['D85'].fill = fill
        sheet['D85'].border = border
        sheet['D85'] = 'Total'
        
        sheet['E85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E85'].fill = fill
        sheet['E85'].border = border
        sheet['E85'] = 'Niños         (1d - 11a)'
        
        sheet['F85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F85'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F85'].fill = fill
        sheet['F85'].border = border
        sheet['F85'] = 'Adolescentes (12a - 17a)'
        
        sheet['G85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G85'].fill = fill
        sheet['G85'].border = border
        sheet['G85'] = 'Jóvenes (18a - 29a)'
        
        sheet['H85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H85'].fill = fill
        sheet['H85'].border = border
        sheet['H85'] = 'Adultos (30a - 59a)'
        
        sheet['I85'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I85'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I85'].fill = fill
        sheet['I85'].border = border
        sheet['I85'] = 'A. Mayores (60a +)'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=86, max_row=90, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = borde_plomo
        
        sheet['B86'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B86'].font = Font(name = 'Arial', size= 8)
        sheet['B86'] ='1º Visita' 
        
        sheet['B87'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B87'].font = Font(name = 'Arial', size= 8)
        sheet['B87'] ='2º Visita' 
        
        sheet['B88'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B88'].font = Font(name = 'Arial', size= 8)
        sheet['B88'] ='3º Visita' 
        
        sheet['B89'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B89'].font = Font(name = 'Arial', size= 8)
        sheet['B89'] ='4º a Visita (trazador)' 
        
        sheet['B90'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B90'].font = Font(name = 'Arial', size= 8)
        sheet['B90'] ='5º a + Visitas' 
        
        sheet['B91'].alignment = Alignment(horizontal= "right", vertical="center")
        sheet['B91'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B91'] ='SUB TOTAL' 
        
        #########################################################
        ########## CAPACITACION AGENTES COMUNITARIOS ############
        #########################################################
        sheet['B93'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B93'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B93'] ='Capacitación a Actores Sociales para la aplicación de la estrategia de Rehabilitación Basada en la Comunidad' 
                
        sheet['B94'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B94'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B94'] ='Actividades con Gobiernos Locales:' 
        
        sheet['B97'].alignment = Alignment(horizontal= "left", vertical="center")
        sheet['B97'].font = Font(name = 'Arial', size= 8, bold = True)
        sheet['B97'] ='Actividad con Comité Multisectorial (APP96)' 
        
        sheet['D95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D95'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
        sheet['D95'].fill = fill
        sheet['D95'].border = border
        sheet['D95'] = 'Taller'
        
        sheet['F95'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['F95'].fill = fill
        sheet['F95'].border = border
        sheet['F95'] = 'Sesion Educativa'
        
        sheet['H95'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H95'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H95'].fill = fill
        sheet['H95'].border = border
        sheet['H95'] = 'Sesion Demostrativa'
        
        sheet['D96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['D96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['D96'].fill = fill
        sheet['D96'].border = border
        sheet['D96'] = 'N°'
        
        sheet['E96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['E96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['E96'].fill = fill
        sheet['E96'].border = border
        sheet['E96'] = 'Capacitados'
        
        sheet['F96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['F96'].font = Font(name = 'Arial', size= 7, bold = True,color='FFFFFF')
        sheet['F96'].fill = fill
        sheet['F96'].border = border
        sheet['F96'] = 'N°'
        
        sheet['G96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['G96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['G96'].fill = fill
        sheet['G96'].border = border
        sheet['G96'] = 'Capacitados'
        
        sheet['H96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['H96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['H96'].fill = fill
        sheet['H96'].border = border
        sheet['H96'] = 'N° '
        
        sheet['I96'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
        sheet['I96'].font = Font(name = 'Arial', size= 7, bold = True, color='FFFFFF')
        sheet['I96'].fill = fill
        sheet['I96'].border = border
        sheet['I96'] = 'Capacitados'
        
        #borde plomo
        for row in sheet.iter_rows(min_row=97, max_row=97, min_col=2, max_col=9):
            for cell in row:
                # Aplicar estilos de alineación a cada celda
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = borde_plomo
        #############################################################################
        #############################################################################                
        # cambina celdas
        sheet.merge_cells('C6:D6')
        sheet.merge_cells('C7:E7')
        
        sheet.merge_cells('B18:B19')
        sheet.merge_cells('B20:B22')
        sheet.merge_cells('B27:B33')
        sheet.merge_cells('B37:B43')
        sheet.merge_cells('B44:B45')
        
        # sensorial
        sheet.merge_cells('B57:B58')
        
        sheet.merge_cells('B15:C15')
        sheet.merge_cells('B51:C51')
        
        # mental
        sheet.merge_cells('B62:C62')
        
        #certificado
        sheet.merge_cells('B70:C70')
        
        sheet.merge_cells('B71:B73')
        
        #RBC
        sheet.merge_cells('B85:C85')
        
        #capa
        sheet.merge_cells('D80:E80')
        sheet.merge_cells('F80:G80')
        sheet.merge_cells('H80:I80')

        sheet.merge_cells('D95:E95')
        sheet.merge_cells('F95:G95')
        sheet.merge_cells('H95:I95')
        
        #capacitacion
        sheet.merge_cells('B82:C82')
        sheet.merge_cells('B97:C97')
        
        #visita
        sheet.merge_cells('B86:C86')
        sheet.merge_cells('B87:C87')
        sheet.merge_cells('B88:C88')
        sheet.merge_cells('B89:C89')
        sheet.merge_cells('B90:C90')
        
        # Definir ubicaciones específicas para cada columna y su suma total
        columnas_ubicaciones = {
            'PROVINCIA': 'D10',
            'DIS_1': 'E16', 
            'DIS_2': 'F16',
            'DIS_3': 'G16',
            'DIS_4': 'H16',
            'DIS_5': 'I16',
            'DIS_6': 'E17',
            'DIS_7': 'F17',
            'DIS_8': 'G17',
            'DIS_9': 'H17',
            'DIS_10': 'I17',
            'DIS_11': 'E18',
            'DIS_12': 'F18',
            'DIS_13': 'G18',
            'DIS_14': 'H18',
            'DIS_15': 'I18',
            'DIS_16': 'E19',
            'DIS_17': 'F19',
            'DIS_18': 'G19',
            'DIS_19': 'H19',
            'DIS_20': 'I19',
            'DIS_21': 'E20',
            'DIS_22': 'F20',
            'DIS_23': 'G20',
            'DIS_24': 'H20',
            'DIS_25': 'I20',
            'DIS_26': 'E21',
            'DIS_27': 'F21',
            'DIS_28': 'G21',
            'DIS_29': 'H21',
            'DIS_30': 'I21',
            'DIS_31': 'E22',
            'DIS_32': 'F22',
            'DIS_33': 'G22',
            'DIS_34': 'H22',
            'DIS_35': 'I22',
            'DIS_36': 'E23',
            'DIS_37': 'F23',
            'DIS_38': 'G23',
            'DIS_39': 'H23',
            'DIS_40': 'I23',
            'DIS_41': 'E24',
            'DIS_42': 'F24',
            'DIS_43': 'G24',
            'DIS_44': 'H24',
            'DIS_45': 'I24',
            'DIS_46': 'E25',
            'DIS_47': 'F25',
            'DIS_48': 'G25',
            'DIS_49': 'H25',
            'DIS_50': 'I25',
            'DIS_51': 'E26',
            'DIS_52': 'F26',
            'DIS_53': 'G26',
            'DIS_54': 'H26',
            'DIS_55': 'I26',
            'DIS_56': 'E27',
            'DIS_57': 'F27',
            'DIS_58': 'G27',
            'DIS_59': 'H27',
            'DIS_60': 'I27',
            'DIS_61': 'E28',
            'DIS_62': 'F28',
            'DIS_63': 'G28',
            'DIS_64': 'H28',
            'DIS_65': 'I28',
            'DIS_66': 'E29',
            'DIS_67': 'F29',
            'DIS_68': 'G29',
            'DIS_69': 'H29',
            'DIS_70': 'I29',
            'DIS_71': 'E30',
            'DIS_72': 'F30',
            'DIS_73': 'G30',
            'DIS_74': 'H30',
            'DIS_75': 'I30',
            'DIS_76': 'E31',
            'DIS_77': 'F31',
            'DIS_78': 'G31',
            'DIS_79': 'H31',
            'DIS_80': 'I31',
            'DIS_81': 'E32',
            'DIS_82': 'F32',
            'DIS_83': 'G32',
            'DIS_84': 'H32',
            'DIS_85': 'I32',
            'DIS_86': 'E33',
            'DIS_87': 'F33',
            'DIS_88': 'G33',
            'DIS_89': 'H33',
            'DIS_90': 'I33',
            'DIS_91': 'E34',
            'DIS_92': 'F34',
            'DIS_93': 'G34',
            'DIS_94': 'H34',
            'DIS_95': 'I34',
            'DIS_96': 'E35',
            'DIS_97': 'F35',
            'DIS_98': 'G35',
            'DIS_99': 'H35',
            'DIS_100': 'I35',
            'DIS_101': 'E36',
            'DIS_102': 'F36',
            'DIS_103': 'G36',
            'DIS_104': 'H36',
            'DIS_105': 'I36',
            'DIS_106': 'E37',
            'DIS_107': 'F37',
            'DIS_108': 'G37',
            'DIS_109': 'H37',
            'DIS_110': 'I37',
            'DIS_111': 'E38',
            'DIS_112': 'F38',
            'DIS_113': 'G38',
            'DIS_114': 'H38',
            'DIS_115': 'I38',
            'DIS_116': 'E39',
            'DIS_117': 'F39',
            'DIS_118': 'G39',
            'DIS_119': 'H39',
            'DIS_120': 'I39',
            'DIS_121': 'E40',
            'DIS_122': 'F40',
            'DIS_123': 'G40',
            'DIS_124': 'H40',
            'DIS_125': 'I40',
            'DIS_126': 'E41',
            'DIS_127': 'F41',
            'DIS_128': 'G41',
            'DIS_129': 'H41',
            'DIS_130': 'I41', 
            'DIS_131': 'E42',
            'DIS_132': 'F42',
            'DIS_133': 'G42',
            'DIS_134': 'H42',
            'DIS_135': 'I42', 
            'DIS_136': 'E43',
            'DIS_137': 'F43',
            'DIS_138': 'G43',
            'DIS_139': 'H43',
            'DIS_140': 'I43', 
            'DIS_141': 'E44',
            'DIS_142': 'F44',
            'DIS_143': 'G44',
            'DIS_144': 'H44',
            'DIS_145': 'I44', 
            'DIS_146': 'E45',
            'DIS_147': 'F45',
            'DIS_148': 'G45',
            'DIS_149': 'H45',
            'DIS_150': 'I45', 
            'DIS_151': 'E46',
            'DIS_152': 'F46',
            'DIS_153': 'G46',
            'DIS_154': 'H46',
            'DIS_155': 'I46', 
            'DIS_156': 'E47',
            'DIS_157': 'F47',
            'DIS_158': 'G47',
            'DIS_159': 'H47',
            'DIS_160': 'I47',            
        }
        
        col_ubi_sensorial = {    
            'PROVINCIA': 'D10',
            'DIS_161': 'E52',
            'DIS_162': 'F52',
            'DIS_163': 'G52',
            'DIS_164': 'H52',
            'DIS_165': 'I52',
            'DIS_166': 'E53',
            'DIS_167': 'F53',
            'DIS_168': 'G53',
            'DIS_169': 'H53',
            'DIS_170': 'I53',
            'DIS_171': 'E54',
            'DIS_172': 'F54',
            'DIS_173': 'G54',
            'DIS_174': 'H54',
            'DIS_175': 'I54',
            'DIS_176': 'E55',
            'DIS_177': 'F55',
            'DIS_178': 'G55',
            'DIS_179': 'H55',
            'DIS_180': 'I55',
            'DIS_181': 'E56',
            'DIS_182': 'F56',
            'DIS_183': 'G56',
            'DIS_184': 'H56',
            'DIS_185': 'I56',
            'DIS_186': 'E57',
            'DIS_187': 'F57',
            'DIS_188': 'G57',
            'DIS_189': 'H57',
            'DIS_190': 'I57',
            'DIS_191': 'E58',
            'DIS_192': 'F58',
            'DIS_193': 'G58',
            'DIS_194': 'H58',
            'DIS_195': 'I58',
        }
        
        col_ubi_mental = {    
            'PROVINCIA': 'D10',
            'DIS_196': 'E63',
            'DIS_197': 'F63',
            'DIS_198': 'G63',
            'DIS_199': 'H63',
            'DIS_200': 'I63',
            'DIS_201': 'E64',
            'DIS_202': 'F64',
            'DIS_203': 'G64',
            'DIS_204': 'H64',
            'DIS_205': 'I64',
            'DIS_206': 'E65',
            'DIS_207': 'F65',
            'DIS_208': 'G65',
            'DIS_209': 'H65',
            'DIS_210': 'I65',
            'DIS_211': 'E66',
            'DIS_212': 'F66',
            'DIS_213': 'G66',
            'DIS_214': 'H66',
            'DIS_215': 'I66',
        }
        
        col_ubi_certificado = {    
            'PROVINCIA': 'D10',
            'DIS_216': 'E71',
            'DIS_217': 'F71',
            'DIS_218': 'G71',
            'DIS_219': 'H71',
            'DIS_220': 'I71',
            'DIS_221': 'E72',
            'DIS_222': 'F72',
            'DIS_223': 'G72',
            'DIS_224': 'H72',
            'DIS_225': 'I72',
            'DIS_226': 'E73',
            'DIS_227': 'F73',
            'DIS_228': 'G73',
            'DIS_229': 'H73',
            'DIS_230': 'I73',
            'DIS_231': 'E74',
            'DIS_232': 'F74',
            'DIS_233': 'G74',
            'DIS_234': 'H74',
            'DIS_235': 'I74',
        }
        
        col_ubi_capacitacion = {    
            'PROVINCIA': 'D10',
            'DIS_273': 'D12',
            'DIS_274': 'E12',
        }
        
        col_ubi_agente = {    
            'PROVINCIA': 'D10',
            'DIS_236': 'D82',
            'DIS_237': 'E82',
            'DIS_238': 'F82',
            'DIS_239': 'G82',
            'DIS_240': 'H82',
            'DIS_241': 'I82',
        }      
        
        col_ubi_rbc = {    
            'PROVINCIA': 'D10',
            'DIS_242': 'E86',
            'DIS_243': 'F86',
            'DIS_244': 'G86',
            'DIS_245': 'H86',
            'DIS_246': 'I86',
            'DIS_247': 'E87',
            'DIS_248': 'F87',
            'DIS_249': 'G87',
            'DIS_250': 'H87',
            'DIS_251': 'I87',
            'DIS_252': 'E88',
            'DIS_253': 'F88',
            'DIS_254': 'G88',
            'DIS_255': 'H88',
            'DIS_256': 'I88',
            'DIS_257': 'E89',
            'DIS_258': 'F89',
            'DIS_259': 'G89',
            'DIS_260': 'H89',
            'DIS_261': 'I89',
            'DIS_262': 'E90',
            'DIS_263': 'F90',
            'DIS_264': 'G90',
            'DIS_265': 'H90',
            'DIS_266': 'I90'
        }
        
        col_ubi_comite = {    
            'PROVINCIA': 'D10',
            'DIS_267': 'D97',
            'DIS_268': 'E97',
            'DIS_269': 'F97',
            'DIS_270': 'G97',
            'DIS_271': 'H97',
            'DIS_272': 'I97',
        }
        
        # Inicializar diccionario para almacenar sumas por columna
        column_sums = {
            'DIS_1': 0,
            'DIS_2': 0,
            'DIS_3': 0,
            'DIS_4': 0,
            'DIS_5': 0,
            'DIS_6': 0,
            'DIS_7': 0,
            'DIS_8': 0,
            'DIS_9': 0,
            'DIS_10': 0,
            'DIS_11': 0,
            'DIS_12': 0,
            'DIS_13': 0,
            'DIS_14': 0,
            'DIS_15': 0,
            'DIS_16': 0,
            'DIS_17': 0,
            'DIS_18': 0,
            'DIS_19': 0,
            'DIS_20': 0,
            'DIS_21': 0,
            'DIS_22': 0,
            'DIS_23': 0,
            'DIS_24': 0,
            'DIS_25': 0,
            'DIS_26': 0,
            'DIS_27': 0,
            'DIS_28': 0,
            'DIS_29': 0,
            'DIS_30': 0,
            'DIS_31': 0,
            'DIS_32': 0,
            'DIS_33': 0,
            'DIS_34': 0,
            'DIS_35': 0,
            'DIS_36': 0,
            'DIS_37': 0,
            'DIS_38': 0,
            'DIS_39': 0,
            'DIS_40': 0,
            'DIS_41': 0,
            'DIS_42': 0,
            'DIS_43': 0,
            'DIS_44': 0,
            'DIS_45': 0,
            'DIS_46': 0,
            'DIS_47': 0,
            'DIS_48': 0,
            'DIS_49': 0,
            'DIS_50': 0,
            'DIS_51': 0,
            'DIS_52': 0,
            'DIS_53': 0,
            'DIS_54': 0,
            'DIS_55': 0,
            'DIS_56': 0,
            'DIS_57': 0,
            'DIS_58': 0,
            'DIS_59': 0,
            'DIS_60': 0,
            'DIS_61': 0,
            'DIS_62': 0,
            'DIS_63': 0,
            'DIS_64': 0,
            'DIS_65': 0,
            'DIS_66': 0,
            'DIS_67': 0,
            'DIS_68': 0,
            'DIS_69': 0,
            'DIS_70': 0,
            'DIS_71': 0,
            'DIS_72': 0,
            'DIS_73': 0,
            'DIS_74': 0,
            'DIS_75': 0,
            'DIS_76': 0,
            'DIS_77': 0,
            'DIS_78': 0,
            'DIS_79': 0,
            'DIS_80': 0,
            'DIS_81': 0,
            'DIS_82': 0,
            'DIS_83': 0,
            'DIS_84': 0,
            'DIS_85': 0,
            'DIS_86': 0,
            'DIS_87': 0,
            'DIS_88': 0,
            'DIS_89': 0,
            'DIS_90': 0,
            'DIS_91': 0,
            'DIS_92': 0,
            'DIS_93': 0,
            'DIS_94': 0,
            'DIS_95': 0,
            'DIS_96': 0,
            'DIS_97': 0,
            'DIS_98': 0,
            'DIS_99': 0,
            'DIS_100': 0,
            'DIS_101': 0,
            'DIS_102': 0,
            'DIS_103': 0,
            'DIS_104': 0,
            'DIS_105': 0,
            'DIS_106': 0,
            'DIS_107': 0,
            'DIS_108': 0,
            'DIS_109': 0,
            'DIS_110': 0,
            'DIS_111': 0,
            'DIS_112': 0,
            'DIS_113': 0,
            'DIS_114': 0,
            'DIS_115': 0,
            'DIS_116': 0,
            'DIS_117': 0,
            'DIS_118': 0,
            'DIS_119': 0,
            'DIS_120': 0,
            'DIS_121': 0,
            'DIS_122': 0,
            'DIS_123': 0,
            'DIS_124': 0,
            'DIS_125': 0,
            'DIS_126': 0,
            'DIS_127': 0,
            'DIS_128': 0,
            'DIS_129': 0,
            'DIS_130': 0, 
            'DIS_131': 0,
            'DIS_132': 0,
            'DIS_133': 0,
            'DIS_134': 0,
            'DIS_135': 0, 
            'DIS_136': 0,
            'DIS_137': 0,
            'DIS_138': 0,
            'DIS_139': 0,
            'DIS_140': 0, 
            'DIS_141': 0,
            'DIS_142': 0,
            'DIS_143': 0,
            'DIS_144': 0,
            'DIS_145': 0, 
            'DIS_146': 0,
            'DIS_147': 0,
            'DIS_148': 0,
            'DIS_149': 0,
            'DIS_150': 0, 
            'DIS_151': 0,
            'DIS_152': 0,
            'DIS_153': 0,
            'DIS_154': 0,
            'DIS_155': 0, 
            'DIS_156': 0,
            'DIS_157': 0,
            'DIS_158': 0,
            'DIS_159': 0,
            'DIS_160': 0,    
        }
        
        col_sum_sensorial = {       
            'DIS_161': 0,
            'DIS_162': 0,
            'DIS_163': 0,
            'DIS_164': 0,
            'DIS_165': 0,
            'DIS_166': 0,
            'DIS_167': 0,
            'DIS_168': 0,
            'DIS_169': 0,
            'DIS_170': 0,
            'DIS_171': 0,
            'DIS_172': 0,
            'DIS_173': 0,
            'DIS_174': 0,
            'DIS_175': 0,
            'DIS_176': 0,
            'DIS_177': 0,
            'DIS_178': 0,
            'DIS_179': 0,
            'DIS_180': 0,
            'DIS_181': 0,
            'DIS_182': 0,
            'DIS_183': 0,
            'DIS_184': 0,
            'DIS_185': 0,
            'DIS_186': 0,
            'DIS_187': 0,
            'DIS_188': 0,
            'DIS_189': 0,
            'DIS_190': 0,
            'DIS_191': 0,
            'DIS_192': 0,
            'DIS_193': 0,
            'DIS_194': 0,
            'DIS_195': 0,
        } 

        col_sum_mental = {    
            'DIS_196': 0,
            'DIS_197': 0,
            'DIS_198': 0,
            'DIS_199': 0,
            'DIS_200': 0,
            'DIS_201': 0,
            'DIS_202': 0,
            'DIS_203': 0,
            'DIS_204': 0,
            'DIS_205': 0,
            'DIS_206': 0,
            'DIS_207': 0,
            'DIS_208': 0,
            'DIS_209': 0,
            'DIS_210': 0,
            'DIS_211': 0,
            'DIS_212': 0,
            'DIS_213': 0,
            'DIS_214': 0,
            'DIS_215': 0,
        }
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_certificado = {       
            'DIS_216': 0,
            'DIS_217': 0,
            'DIS_218': 0,
            'DIS_219': 0,
            'DIS_220': 0,
            'DIS_221': 0,
            'DIS_222': 0,
            'DIS_223': 0,
            'DIS_224': 0,
            'DIS_225': 0,
            'DIS_226': 0,
            'DIS_227': 0,
            'DIS_228': 0,
            'DIS_229': 0,
            'DIS_230': 0,
            'DIS_231': 0,
            'DIS_232': 0,
            'DIS_233': 0,
            'DIS_234': 0,
            'DIS_235': 0,
        }  
        
        col_sum_capacitacion = {    
            'DIS_273': 0,
            'DIS_274': 0,
        }
        
        col_sum_agente = {    
            'DIS_236': 0,
            'DIS_237': 0,
            'DIS_238': 0,
            'DIS_239': 0,
            'DIS_240': 0,
            'DIS_241': 0,
        }      
        
        # Inicializar diccionario para almacenar sumas por columna
        col_sum_rbc = {       
            'DIS_242': 0,
            'DIS_243': 0,
            'DIS_244': 0,
            'DIS_245': 0,
            'DIS_246': 0,
            'DIS_247': 0,
            'DIS_248': 0,
            'DIS_249': 0,
            'DIS_250': 0,
            'DIS_251': 0,
            'DIS_252': 0,
            'DIS_253': 0,
            'DIS_254': 0,
            'DIS_255': 0,
            'DIS_256': 0,
            'DIS_257': 0,
            'DIS_258': 0,
            'DIS_259': 0,
            'DIS_260': 0,
            'DIS_261': 0,
            'DIS_262': 0,
            'DIS_263': 0,
            'DIS_264': 0,
            'DIS_265': 0,
            'DIS_266': 0,
        } 
        
        col_sum_comite = {    
            'DIS_267': 0,
            'DIS_268': 0,
            'DIS_269': 0,
            'DIS_270': 0,
            'DIS_271': 0,
            'DIS_272': 0,
        }
        
        ############################
        ###  DISCAPACIDAD FISICA ###
        ############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_establec:
            for col_name in column_sums:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(columnas_ubicaciones.keys()).index(col_name)
                    column_sums[col_name] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila dis_fisica: {row}")                        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_name, total_cell in columnas_ubicaciones.items():
            if col_name in column_sums:
                # Obtener la celda correspondiente según la ubicación
                cell = sheet[total_cell]
                # Asignar el valor de la suma a la celda
                cell.value = column_sums[col_name]
                # Aplicar formato a la celda
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        
        # Sumar los valores del diccionario      
        total_sum_cat_1 =  sum([column_sums['DIS_1'], column_sums['DIS_2'], column_sums['DIS_3'],column_sums['DIS_4'],column_sums['DIS_5']])
        total_sum_cat_2 =  sum([column_sums['DIS_6'], column_sums['DIS_7'], column_sums['DIS_8'],column_sums['DIS_9'],column_sums['DIS_10']])
        total_sum_cat_3 =  sum([column_sums['DIS_11'], column_sums['DIS_12'], column_sums['DIS_13'],column_sums['DIS_14'],column_sums['DIS_15']])
        total_sum_cat_4 =  sum([column_sums['DIS_16'], column_sums['DIS_17'], column_sums['DIS_18'],column_sums['DIS_19'],column_sums['DIS_20']])
        total_sum_cat_5 =  sum([column_sums['DIS_21'], column_sums['DIS_22'], column_sums['DIS_23'],column_sums['DIS_24'],column_sums['DIS_25']])
        total_sum_cat_6 =  sum([column_sums['DIS_26'], column_sums['DIS_27'], column_sums['DIS_28'],column_sums['DIS_29'],column_sums['DIS_30']])
        total_sum_cat_7 =  sum([column_sums['DIS_31'], column_sums['DIS_32'], column_sums['DIS_33'],column_sums['DIS_34'],column_sums['DIS_35']])
        total_sum_cat_8 =  sum([column_sums['DIS_36'], column_sums['DIS_37'], column_sums['DIS_38'],column_sums['DIS_39'],column_sums['DIS_40']])
        total_sum_cat_9 =  sum([column_sums['DIS_41'], column_sums['DIS_42'], column_sums['DIS_43'],column_sums['DIS_44'],column_sums['DIS_45']])
        total_sum_cat_10 =  sum([column_sums['DIS_46'], column_sums['DIS_47'], column_sums['DIS_48'],column_sums['DIS_49'],column_sums['DIS_50']])
        total_sum_cat_11 =  sum([column_sums['DIS_51'], column_sums['DIS_52'], column_sums['DIS_53'],column_sums['DIS_54'],column_sums['DIS_55']])
        total_sum_cat_12 =  sum([column_sums['DIS_56'], column_sums['DIS_57'], column_sums['DIS_58'],column_sums['DIS_59'],column_sums['DIS_60']])
        total_sum_cat_13 =  sum([column_sums['DIS_61'], column_sums['DIS_62'], column_sums['DIS_63'],column_sums['DIS_64'],column_sums['DIS_65']])
        total_sum_cat_14 =  sum([column_sums['DIS_66'], column_sums['DIS_67'], column_sums['DIS_68'],column_sums['DIS_69'],column_sums['DIS_70']])
        total_sum_cat_15 =  sum([column_sums['DIS_71'], column_sums['DIS_72'], column_sums['DIS_73'],column_sums['DIS_74'],column_sums['DIS_75']])
        total_sum_cat_16 =  sum([column_sums['DIS_76'], column_sums['DIS_77'], column_sums['DIS_78'],column_sums['DIS_79'],column_sums['DIS_80']])   
        total_sum_cat_17 =  sum([column_sums['DIS_81'], column_sums['DIS_82'], column_sums['DIS_83'],column_sums['DIS_84'],column_sums['DIS_85']])
        total_sum_cat_18 =  sum([column_sums['DIS_86'], column_sums['DIS_87'], column_sums['DIS_88'],column_sums['DIS_89'],column_sums['DIS_90']])
        total_sum_cat_19 =  sum([column_sums['DIS_91'], column_sums['DIS_92'], column_sums['DIS_93'],column_sums['DIS_94'],column_sums['DIS_95']])
        total_sum_cat_20 =  sum([column_sums['DIS_96'], column_sums['DIS_97'], column_sums['DIS_98'],column_sums['DIS_99'],column_sums['DIS_100']])
        total_sum_cat_21 =  sum([column_sums['DIS_101'], column_sums['DIS_102'], column_sums['DIS_103'],column_sums['DIS_104'],column_sums['DIS_105']])
        total_sum_cat_22 =  sum([column_sums['DIS_106'], column_sums['DIS_107'], column_sums['DIS_108'],column_sums['DIS_109'],column_sums['DIS_110']])
        total_sum_cat_23 =  sum([column_sums['DIS_111'], column_sums['DIS_112'], column_sums['DIS_113'],column_sums['DIS_114'],column_sums['DIS_115']])
        total_sum_cat_24 =  sum([column_sums['DIS_116'], column_sums['DIS_117'], column_sums['DIS_118'],column_sums['DIS_119'],column_sums['DIS_120']])
        total_sum_cat_25 =  sum([column_sums['DIS_121'], column_sums['DIS_122'], column_sums['DIS_123'],column_sums['DIS_124'],column_sums['DIS_125']])
        total_sum_cat_26 =  sum([column_sums['DIS_126'], column_sums['DIS_127'], column_sums['DIS_128'],column_sums['DIS_129'],column_sums['DIS_130']])
        total_sum_cat_27 =  sum([column_sums['DIS_131'], column_sums['DIS_132'], column_sums['DIS_133'],column_sums['DIS_134'],column_sums['DIS_135']])
        total_sum_cat_28 =  sum([column_sums['DIS_136'], column_sums['DIS_137'], column_sums['DIS_138'],column_sums['DIS_139'],column_sums['DIS_140']])
        total_sum_cat_29 =  sum([column_sums['DIS_141'], column_sums['DIS_142'], column_sums['DIS_143'],column_sums['DIS_144'],column_sums['DIS_145']])
        total_sum_cat_30 =  sum([column_sums['DIS_146'], column_sums['DIS_147'], column_sums['DIS_148'],column_sums['DIS_149'],column_sums['DIS_150']])
        total_sum_cat_31 =  sum([column_sums['DIS_151'], column_sums['DIS_152'], column_sums['DIS_153'],column_sums['DIS_154'],column_sums['DIS_155']])
        total_sum_cat_32 =  sum([column_sums['DIS_156'], column_sums['DIS_157'], column_sums['DIS_158'],column_sums['DIS_159'],column_sums['DIS_160']])

        sheet['D16'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D16'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D16'] = total_sum_cat_1     
        
        sheet['D17'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D17'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D17'] = total_sum_cat_2 
        
        sheet['D18'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D18'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D18'] = total_sum_cat_3    
        
        sheet['D19'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D19'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D19'] = total_sum_cat_4    
        
        sheet['D20'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D20'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D20'] = total_sum_cat_5    
        
        sheet['D21'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D21'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D21'] = total_sum_cat_6    
        
        sheet['D22'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D22'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D22'] = total_sum_cat_7    
        
        sheet['D23'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D23'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D23'] = total_sum_cat_8    
        
        sheet['D24'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D24'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D24'] = total_sum_cat_9    
        
        sheet['D25'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D25'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D25'] = total_sum_cat_10 
        
        sheet['D26'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D26'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D26'] = total_sum_cat_11
                
        sheet['D27'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D27'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D27'] = total_sum_cat_12    
        
        sheet['D28'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D28'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D28'] = total_sum_cat_13   
        
        sheet['D29'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D29'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D29'] = total_sum_cat_14   
        
        sheet['D30'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D30'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D30'] = total_sum_cat_15   
        
        sheet['D31'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D31'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D31'] = total_sum_cat_16   
        
        sheet['D32'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D32'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D32'] = total_sum_cat_17         
        
        sheet['D33'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D33'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D33'] = total_sum_cat_18   
        
        sheet['D34'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D34'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D34'] = total_sum_cat_19   
        
        sheet['D35'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D35'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D35'] = total_sum_cat_20   
        
        sheet['D36'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D36'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D36'] = total_sum_cat_21   
        
        sheet['D37'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D37'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D37'] = total_sum_cat_22   
        
        sheet['D38'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D38'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D38'] = total_sum_cat_23   
        
        sheet['D39'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D39'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D39'] = total_sum_cat_24   
        
        sheet['D40'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D40'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D40'] = total_sum_cat_25  
        
        sheet['D41'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D41'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D41'] = total_sum_cat_26 
        
        sheet['D42'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D42'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D42'] = total_sum_cat_27   
        
        sheet['D43'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D43'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D43'] = total_sum_cat_28   
        
        sheet['D44'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D44'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D44'] = total_sum_cat_29  
        
        sheet['D45'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D45'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D45'] = total_sum_cat_30  
        
        sheet['D46'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D46'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D46'] = total_sum_cat_31
        
        sheet['D47'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D47'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D47'] = total_sum_cat_32
        
        # Sumar los valores del VERTICAL      
        total_sum_cat_vertical_1 =  sum([column_sums['DIS_1'],column_sums['DIS_6'], column_sums['DIS_11'],column_sums['DIS_16'],column_sums['DIS_21'],column_sums['DIS_26'],column_sums['DIS_31'],column_sums['DIS_36'],column_sums['DIS_41'],column_sums['DIS_46'],column_sums['DIS_51'],column_sums['DIS_56'],column_sums['DIS_61'],column_sums['DIS_66'],column_sums['DIS_71'],column_sums['DIS_76'],column_sums['DIS_81'],column_sums['DIS_86'],column_sums['DIS_91'],column_sums['DIS_96'],column_sums['DIS_101'],column_sums['DIS_106'] ,column_sums['DIS_111'],column_sums['DIS_116'],column_sums['DIS_121'],column_sums['DIS_126'],column_sums['DIS_131'],column_sums['DIS_136'],column_sums['DIS_141'],column_sums['DIS_146'],column_sums['DIS_151'],column_sums['DIS_156']])
        total_sum_cat_vertical_2 =  sum([column_sums['DIS_2'],column_sums['DIS_7'], column_sums['DIS_12'],column_sums['DIS_17'],column_sums['DIS_22'],column_sums['DIS_27'],column_sums['DIS_32'],column_sums['DIS_37'],column_sums['DIS_42'],column_sums['DIS_47'],column_sums['DIS_52'],column_sums['DIS_57'],column_sums['DIS_62'],column_sums['DIS_67'],column_sums['DIS_72'],column_sums['DIS_77'],column_sums['DIS_82'],column_sums['DIS_87'],column_sums['DIS_92'],column_sums['DIS_97'],column_sums['DIS_102'],column_sums['DIS_107'] ,column_sums['DIS_112'],column_sums['DIS_117'],column_sums['DIS_122'],column_sums['DIS_127'],column_sums['DIS_132'],column_sums['DIS_137'],column_sums['DIS_142'],column_sums['DIS_147'],column_sums['DIS_152'],column_sums['DIS_157']])
        total_sum_cat_vertical_3 =  sum([column_sums['DIS_3'],column_sums['DIS_8'], column_sums['DIS_13'],column_sums['DIS_18'],column_sums['DIS_23'],column_sums['DIS_28'],column_sums['DIS_33'],column_sums['DIS_38'],column_sums['DIS_43'],column_sums['DIS_48'],column_sums['DIS_53'],column_sums['DIS_58'],column_sums['DIS_63'],column_sums['DIS_68'],column_sums['DIS_73'],column_sums['DIS_78'],column_sums['DIS_83'],column_sums['DIS_88'],column_sums['DIS_93'],column_sums['DIS_98'],column_sums['DIS_103'],column_sums['DIS_108'] ,column_sums['DIS_113'],column_sums['DIS_118'],column_sums['DIS_123'],column_sums['DIS_128'],column_sums['DIS_133'],column_sums['DIS_138'],column_sums['DIS_143'],column_sums['DIS_148'],column_sums['DIS_153'],column_sums['DIS_158']])
        total_sum_cat_vertical_4 =  sum([column_sums['DIS_4'],column_sums['DIS_9'], column_sums['DIS_14'],column_sums['DIS_19'],column_sums['DIS_24'],column_sums['DIS_29'],column_sums['DIS_34'],column_sums['DIS_39'],column_sums['DIS_44'],column_sums['DIS_49'],column_sums['DIS_54'],column_sums['DIS_59'],column_sums['DIS_64'],column_sums['DIS_69'],column_sums['DIS_74'],column_sums['DIS_79'],column_sums['DIS_84'],column_sums['DIS_89'],column_sums['DIS_94'],column_sums['DIS_99'],column_sums['DIS_104'],column_sums['DIS_109'] ,column_sums['DIS_114'],column_sums['DIS_119'],column_sums['DIS_124'],column_sums['DIS_129'],column_sums['DIS_134'],column_sums['DIS_139'],column_sums['DIS_144'],column_sums['DIS_149'],column_sums['DIS_154'],column_sums['DIS_159']])
        total_sum_cat_vertical_5 =  sum([column_sums['DIS_5'],column_sums['DIS_10'],column_sums['DIS_15'],column_sums['DIS_20'],column_sums['DIS_25'],column_sums['DIS_30'],column_sums['DIS_35'],column_sums['DIS_40'],column_sums['DIS_45'],column_sums['DIS_50'],column_sums['DIS_55'],column_sums['DIS_60'],column_sums['DIS_65'],column_sums['DIS_70'],column_sums['DIS_75'],column_sums['DIS_80'],column_sums['DIS_85'],column_sums['DIS_90'],column_sums['DIS_95'],column_sums['DIS_100'],column_sums['DIS_105'],column_sums['DIS_110'],column_sums['DIS_115'],column_sums['DIS_120'],column_sums['DIS_125'],column_sums['DIS_130'],column_sums['DIS_135'],column_sums['DIS_140'],column_sums['DIS_145'],column_sums['DIS_150'],column_sums['DIS_155'],column_sums['DIS_160']])

        sheet['E48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E48'] = total_sum_cat_vertical_1     
        
        sheet['F48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F48'] = total_sum_cat_vertical_2 
        
        sheet['G48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G48'] = total_sum_cat_vertical_3    
        
        sheet['H48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H48'] = total_sum_cat_vertical_4    
        
        sheet['I48'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I48'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I48'] = total_sum_cat_vertical_5    
        ##########################################################################
        
        ###############################
        ###  DISCAPACIDAD SENSORIAL ###
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_sensorial_establec:
            for col_sensorial in col_sum_sensorial:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_sensorial.keys()).index(col_sensorial)
                    col_sum_sensorial[col_sensorial] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila sensorial: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_sensorial, total_cell_sensorial in col_ubi_sensorial.items():
            if col_sensorial in col_sum_sensorial:
                # Obtener la celda correspondiente según la ubicación
                cell_sensorial = sheet[total_cell_sensorial]
                # Asignar el valor de la suma a la celda
                cell_sensorial.value = col_sum_sensorial[col_sensorial]
                # Aplicar formato a la celda
                cell_sensorial.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_sensorial.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_sensorial.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_sensorial['DIS_161'], col_sum_sensorial['DIS_162'], col_sum_sensorial['DIS_163'], col_sum_sensorial['DIS_164'], col_sum_sensorial['DIS_165']])
        t_sum_cat_2 =  sum([col_sum_sensorial['DIS_166'], col_sum_sensorial['DIS_167'], col_sum_sensorial['DIS_168'], col_sum_sensorial['DIS_169'], col_sum_sensorial['DIS_170']])
        t_sum_cat_3 =  sum([col_sum_sensorial['DIS_171'], col_sum_sensorial['DIS_172'], col_sum_sensorial['DIS_173'], col_sum_sensorial['DIS_174'], col_sum_sensorial['DIS_175']])
        t_sum_cat_4 =  sum([col_sum_sensorial['DIS_176'], col_sum_sensorial['DIS_177'], col_sum_sensorial['DIS_178'], col_sum_sensorial['DIS_179'], col_sum_sensorial['DIS_180']])
        t_sum_cat_5 =  sum([col_sum_sensorial['DIS_181'], col_sum_sensorial['DIS_182'], col_sum_sensorial['DIS_183'], col_sum_sensorial['DIS_184'], col_sum_sensorial['DIS_185']])
        t_sum_cat_6 =  sum([col_sum_sensorial['DIS_186'], col_sum_sensorial['DIS_187'], col_sum_sensorial['DIS_188'], col_sum_sensorial['DIS_189'], col_sum_sensorial['DIS_190']])
        t_sum_cat_7 =  sum([col_sum_sensorial['DIS_191'], col_sum_sensorial['DIS_192'], col_sum_sensorial['DIS_193'], col_sum_sensorial['DIS_194'], col_sum_sensorial['DIS_195']])
        
        sheet['D52'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D52'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D52'] = t_sum_cat_1     
        
        sheet['D53'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D53'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D53'] = t_sum_cat_2 
        
        sheet['D54'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D54'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D54'] = t_sum_cat_3    
        
        sheet['D55'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D55'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D55'] = t_sum_cat_4    
        
        sheet['D56'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D56'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D56'] = t_sum_cat_5    
        
        sheet['D57'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D57'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D57'] = t_sum_cat_6    
        
        sheet['D58'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D58'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D58'] = t_sum_cat_7    
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_sensorial['DIS_161'],col_sum_sensorial['DIS_166'],col_sum_sensorial['DIS_171'],col_sum_sensorial['DIS_176'],col_sum_sensorial['DIS_181'],col_sum_sensorial['DIS_186'],col_sum_sensorial['DIS_191']])
        t_sum_cat_vertical_2 =  sum([col_sum_sensorial['DIS_162'],col_sum_sensorial['DIS_167'],col_sum_sensorial['DIS_172'],col_sum_sensorial['DIS_177'],col_sum_sensorial['DIS_182'],col_sum_sensorial['DIS_187'],col_sum_sensorial['DIS_192']])
        t_sum_cat_vertical_3 =  sum([col_sum_sensorial['DIS_163'],col_sum_sensorial['DIS_168'],col_sum_sensorial['DIS_173'],col_sum_sensorial['DIS_178'],col_sum_sensorial['DIS_183'],col_sum_sensorial['DIS_188'],col_sum_sensorial['DIS_193']])
        t_sum_cat_vertical_4 =  sum([col_sum_sensorial['DIS_164'],col_sum_sensorial['DIS_169'],col_sum_sensorial['DIS_174'],col_sum_sensorial['DIS_179'],col_sum_sensorial['DIS_184'],col_sum_sensorial['DIS_189'],col_sum_sensorial['DIS_194']])
        t_sum_cat_vertical_5 =  sum([col_sum_sensorial['DIS_165'],col_sum_sensorial['DIS_170'],col_sum_sensorial['DIS_175'],col_sum_sensorial['DIS_180'],col_sum_sensorial['DIS_185'],col_sum_sensorial['DIS_190'],col_sum_sensorial['DIS_195']])
        
        sheet['E59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E59'] = t_sum_cat_vertical_1     
        
        sheet['F59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F59'] = t_sum_cat_vertical_2 
        
        sheet['G59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G59'] = t_sum_cat_vertical_3    
        
        sheet['H59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H59'] = t_sum_cat_vertical_4    
        
        sheet['I59'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I59'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I59'] = t_sum_cat_vertical_5    
        ##########################################################################
                
        ###############################
        ###  DISCAPACIDAD MENTAL ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_mental_establec:
            for col_mental in col_sum_mental:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_mental.keys()).index(col_mental)
                    col_sum_mental[col_mental] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila mental: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_mental, total_cell_mental in col_ubi_mental.items():
            if col_mental in col_sum_mental:
                # Obtener la celda correspondiente según la ubicación
                cell_mental = sheet[total_cell_mental]
                # Asignar el valor de la suma a la celda
                cell_mental.value = col_sum_mental[col_mental]
                # Aplicar formato a la celda
                cell_mental.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_mental.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_mental.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 =  sum([col_sum_mental['DIS_196'], col_sum_mental['DIS_197'], col_sum_mental['DIS_198'], col_sum_mental['DIS_199'], col_sum_mental['DIS_200']])
        t_sum_cat_2 =  sum([col_sum_mental['DIS_201'], col_sum_mental['DIS_202'], col_sum_mental['DIS_203'], col_sum_mental['DIS_204'], col_sum_mental['DIS_205']])
        t_sum_cat_3 =  sum([col_sum_mental['DIS_206'], col_sum_mental['DIS_207'], col_sum_mental['DIS_208'], col_sum_mental['DIS_209'], col_sum_mental['DIS_210']])
        t_sum_cat_4 =  sum([col_sum_mental['DIS_211'], col_sum_mental['DIS_212'], col_sum_mental['DIS_213'], col_sum_mental['DIS_214'], col_sum_mental['DIS_215']])
        
        sheet['D63'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D63'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D63'] = t_sum_cat_1     
        
        sheet['D64'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D64'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D64'] = t_sum_cat_2 
        
        sheet['D65'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D65'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D65'] = t_sum_cat_3    
        
        sheet['D66'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D66'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D66'] = t_sum_cat_4    

        # Sumar los valores del VERTICAL      
        t_sum_cat_vertical_1 =  sum([col_sum_mental['DIS_196'],col_sum_mental['DIS_201'],col_sum_mental['DIS_206'],col_sum_mental['DIS_211']])
        t_sum_cat_vertical_2 =  sum([col_sum_mental['DIS_197'],col_sum_mental['DIS_202'],col_sum_mental['DIS_207'],col_sum_mental['DIS_212']])
        t_sum_cat_vertical_3 =  sum([col_sum_mental['DIS_198'],col_sum_mental['DIS_203'],col_sum_mental['DIS_208'],col_sum_mental['DIS_213']])
        t_sum_cat_vertical_4 =  sum([col_sum_mental['DIS_199'],col_sum_mental['DIS_204'],col_sum_mental['DIS_209'],col_sum_mental['DIS_214']])
        t_sum_cat_vertical_5 =  sum([col_sum_mental['DIS_200'],col_sum_mental['DIS_205'],col_sum_mental['DIS_210'],col_sum_mental['DIS_215']])
        
        sheet['E67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E67'] = t_sum_cat_vertical_1     
        
        sheet['F67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F67'] = t_sum_cat_vertical_2 
        
        sheet['G67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G67'] = t_sum_cat_vertical_3    
        
        sheet['H67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H67'] = t_sum_cat_vertical_4    
        
        sheet['I67'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I67'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I67'] = t_sum_cat_vertical_5    
        ##########################################################################
        
        #################################
        ###  DISCAPACIDAD CERTIFICADO ###
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_certificado_establec:
            for col_certificado in col_sum_certificado:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_certificado.keys()).index(col_certificado)
                    col_sum_certificado[col_certificado] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila certificado: {row}")
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_certificado, total_cell_certificado in col_ubi_certificado.items():
            if col_certificado in col_sum_certificado:
                # Obtener la celda correspondiente según la ubicación
                cell_certificado = sheet[total_cell_certificado]
                # Asignar el valor de la suma a la celda
                cell_certificado.value = col_sum_certificado[col_certificado]
                # Aplicar formato a la celda
                cell_certificado.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_certificado.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_certificado.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        # Sumar los valores del diccionario      
        t_sum_cat_cert_1 =  sum([col_sum_certificado['DIS_216'], col_sum_certificado['DIS_217'], col_sum_certificado['DIS_218'], col_sum_certificado['DIS_219'], col_sum_certificado['DIS_220']])
        t_sum_cat_cert_2 =  sum([col_sum_certificado['DIS_221'], col_sum_certificado['DIS_222'], col_sum_certificado['DIS_223'], col_sum_certificado['DIS_224'], col_sum_certificado['DIS_225']])
        t_sum_cat_cert_3 =  sum([col_sum_certificado['DIS_226'], col_sum_certificado['DIS_227'], col_sum_certificado['DIS_228'], col_sum_certificado['DIS_229'], col_sum_certificado['DIS_230']])
        t_sum_cat_cert_4 =  sum([col_sum_certificado['DIS_231'], col_sum_certificado['DIS_232'], col_sum_certificado['DIS_233'], col_sum_certificado['DIS_234'], col_sum_certificado['DIS_235']])

        sheet['D71'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D71'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D71'] = t_sum_cat_cert_1     
        
        sheet['D72'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D72'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D72'] = t_sum_cat_cert_2 
        
        sheet['D73'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D73'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D73'] = t_sum_cat_cert_3 
        
        sheet['D74'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D74'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D74'] = t_sum_cat_cert_4 
        
        # Sumar los valores del VERTICAL      
        t_sum_cat_vert_1 =  sum([col_sum_certificado['DIS_216'],col_sum_certificado['DIS_221'],col_sum_certificado['DIS_226'],col_sum_certificado['DIS_231']])
        t_sum_cat_vert_2 =  sum([col_sum_certificado['DIS_217'],col_sum_certificado['DIS_222'],col_sum_certificado['DIS_227'],col_sum_certificado['DIS_232']])
        t_sum_cat_vert_3 =  sum([col_sum_certificado['DIS_218'],col_sum_certificado['DIS_223'],col_sum_certificado['DIS_228'],col_sum_certificado['DIS_233']])
        t_sum_cat_vert_4 =  sum([col_sum_certificado['DIS_219'],col_sum_certificado['DIS_224'],col_sum_certificado['DIS_229'],col_sum_certificado['DIS_234']])
        t_sum_cat_vert_5 =  sum([col_sum_certificado['DIS_220'],col_sum_certificado['DIS_225'],col_sum_certificado['DIS_230'],col_sum_certificado['DIS_235']])
        
        sheet['E75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E75'] = t_sum_cat_vert_1     
        
        sheet['F75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F75'] = t_sum_cat_vert_2 
        
        sheet['G75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G75'] = t_sum_cat_vert_3    
        
        sheet['H75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H75'] = t_sum_cat_vert_4    
        
        sheet['I75'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I75'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I75'] = t_sum_cat_vert_5    
        
        #################################
        ###  DISCAPACIDAD RBC ###########
        #################################       
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_rbc_establec:
            for col_rbc in col_sum_rbc:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_rbc.keys()).index(col_rbc)
                    col_sum_rbc[col_rbc] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila rbc: {row}")
                    
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_rbc, total_cell_rbc in col_ubi_rbc.items():
            if col_rbc in col_sum_rbc:
                # Obtener la celda correspondiente según la ubicación
                cell_rbc = sheet[total_cell_rbc]
                # Asignar el valor de la suma a la celda
                cell_rbc.value = col_sum_rbc[col_rbc]
                # Aplicar formato a la celda
                cell_rbc.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_rbc.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_rbc.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
                
        ## Sumar los valores del diccionario      
        t_sum_cat_rbc_1 =  sum([col_sum_rbc['DIS_242'], col_sum_rbc['DIS_243'], col_sum_rbc['DIS_244'], col_sum_rbc['DIS_245'], col_sum_rbc['DIS_246']])
        t_sum_cat_rbc_2 =  sum([col_sum_rbc['DIS_247'], col_sum_rbc['DIS_248'], col_sum_rbc['DIS_249'], col_sum_rbc['DIS_250'], col_sum_rbc['DIS_251']])
        t_sum_cat_rbc_3 =  sum([col_sum_rbc['DIS_252'], col_sum_rbc['DIS_253'], col_sum_rbc['DIS_254'], col_sum_rbc['DIS_255'], col_sum_rbc['DIS_256']])
        t_sum_cat_rbc_4 =  sum([col_sum_rbc['DIS_257'], col_sum_rbc['DIS_258'], col_sum_rbc['DIS_259'], col_sum_rbc['DIS_260'], col_sum_rbc['DIS_261']])
        t_sum_cat_rbc_5 =  sum([col_sum_rbc['DIS_262'], col_sum_rbc['DIS_263'], col_sum_rbc['DIS_264'], col_sum_rbc['DIS_265'], col_sum_rbc['DIS_266']])

        sheet['D86'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D86'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D86'] = t_sum_cat_rbc_1     
        
        sheet['D87'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D87'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D87'] = t_sum_cat_rbc_2 
        
        sheet['D88'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D88'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D88'] = t_sum_cat_rbc_3     
        
        sheet['D89'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D89'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D89'] = t_sum_cat_rbc_4 
        
        sheet['D90'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D90'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D90'] = t_sum_cat_rbc_5 
        
        # Sumar los valores del VERTICAL      
        t_sum_vert_rbc_1 =  sum([col_sum_rbc['DIS_242'],col_sum_rbc['DIS_247'],col_sum_rbc['DIS_252'],col_sum_rbc['DIS_257'],col_sum_rbc['DIS_262']])
        t_sum_vert_rbc_2 =  sum([col_sum_rbc['DIS_243'],col_sum_rbc['DIS_248'],col_sum_rbc['DIS_253'],col_sum_rbc['DIS_258'],col_sum_rbc['DIS_263']])
        t_sum_vert_rbc_3 =  sum([col_sum_rbc['DIS_244'],col_sum_rbc['DIS_249'],col_sum_rbc['DIS_254'],col_sum_rbc['DIS_259'],col_sum_rbc['DIS_264']])
        t_sum_vert_rbc_4 =  sum([col_sum_rbc['DIS_245'],col_sum_rbc['DIS_250'],col_sum_rbc['DIS_255'],col_sum_rbc['DIS_260'],col_sum_rbc['DIS_265']])
        t_sum_vert_rbc_5 =  sum([col_sum_rbc['DIS_246'],col_sum_rbc['DIS_251'],col_sum_rbc['DIS_256'],col_sum_rbc['DIS_261'],col_sum_rbc['DIS_266']])
        
        sheet['E91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E91'] = t_sum_vert_rbc_1
        
        sheet['F91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F91'] = t_sum_vert_rbc_2 
        
        sheet['G91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G91'] = t_sum_vert_rbc_3    
        
        sheet['H91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H91'] = t_sum_vert_rbc_4    
        
        sheet['I91'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I91'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I91'] = t_sum_vert_rbc_5   
        
        #################################
        ###  CAPACITACION PERSONAL ######
        #################################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_capacitacion_establec:
            for col_capacitacion in col_sum_capacitacion:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_capacitacion.keys()).index(col_capacitacion)
                    col_sum_capacitacion[col_capacitacion] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila capacitacion: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_capacitacion, total_cell_capacitacion in col_ubi_capacitacion.items():
            if col_capacitacion in col_sum_capacitacion:
                # Obtener la celda correspondiente según la ubicación
                cell_capacitacion = sheet[total_cell_capacitacion]
                # Asignar el valor de la suma a la celda
                cell_capacitacion.value = col_sum_capacitacion[col_capacitacion]
                # Aplicar formato a la celda
                cell_capacitacion.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_capacitacion.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_capacitacion.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_capacitacion['DIS_273']])
        t_sum_cat_2 = sum([col_sum_capacitacion['DIS_274']])
        
        sheet['D12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D12'] = t_sum_cat_1     
        
        sheet['E12'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E12'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E12'] = t_sum_cat_2 
        
        ###############################
        ###  CAPACITACION AGENTE ######
        ###############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_agente_establec:
            for col_agente in col_sum_agente:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_agente.keys()).index(col_agente)
                    col_sum_agente[col_agente] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila agente: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_agente, total_cell_agente in col_ubi_agente.items():
            if col_agente in col_sum_agente:
                # Obtener la celda correspondiente según la ubicación
                cell_agente = sheet[total_cell_agente]
                # Asignar el valor de la suma a la celda
                cell_agente.value = col_sum_agente[col_agente]
                # Aplicar formato a la celda
                cell_agente.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_agente.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_agente.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_agente['DIS_236']])
        t_sum_cat_2 = sum([col_sum_agente['DIS_237']])
        t_sum_cat_3 = sum([col_sum_agente['DIS_238']])
        t_sum_cat_4 = sum([col_sum_agente['DIS_239']])
        t_sum_cat_5 = sum([col_sum_agente['DIS_240']])
        t_sum_cat_6 = sum([col_sum_agente['DIS_241']])
        
        sheet['D82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D82'] = t_sum_cat_1     
        
        sheet['E82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E82'] = t_sum_cat_2 
        
        sheet['F82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F82'] = t_sum_cat_3
        
        sheet['G82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G82'] = t_sum_cat_4 
        
        sheet['H82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H82'] = t_sum_cat_5
        
        sheet['I82'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I82'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I82'] = t_sum_cat_6 
        
        ############################
        ###  CAPACITACION COMITE ###
        #############################
        # Procesar los datos y calcular las sumas por columna
        for row in resultado_comite_establec:
            for col_comite in col_sum_comite:
                try:
                    # Obtener el índice de la columna según el nombre (DIS_1 -> 1, DIS_2 -> 2, etc.)
                    col_index = list(col_ubi_comite.keys()).index(col_comite)
                    col_sum_comite[col_comite] += int(row[col_index])
                except IndexError:
                    print(f"Error al procesar la fila comite: {row}")
        
        # Escribir las sumas totales por columna en la hoja de cálculo
        for col_comite, total_cell_comite in col_ubi_comite.items():
            if col_comite in col_sum_comite:
                # Obtener la celda correspondiente según la ubicación
                cell_comite = sheet[total_cell_comite]
                # Asignar el valor de la suma a la celda
                cell_comite.value = col_sum_comite[col_comite]
                # Aplicar formato a la celda
                cell_comite.alignment = Alignment(horizontal="center", vertical="center")  # Alinear al centro
                cell_comite.font = Font(name='Arial', size=9)  # Establecer fuente, tamaño y negrita
                cell_comite.number_format = '0'  # Formato de número para mostrar como entero sin decimales       
        # Sumar los valores del diccionario      
        t_sum_cat_1 = sum([col_sum_comite['DIS_267']])
        t_sum_cat_2 = sum([col_sum_comite['DIS_268']])
        t_sum_cat_3 = sum([col_sum_comite['DIS_269']])
        t_sum_cat_4 = sum([col_sum_comite['DIS_270']])
        t_sum_cat_5 = sum([col_sum_comite['DIS_271']])
        t_sum_cat_6 = sum([col_sum_comite['DIS_272']])
        
        sheet['D97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['D97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['D97'] = t_sum_cat_1     
        
        sheet['E97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['E97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['E97'] = t_sum_cat_2 
        
        sheet['F97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['F97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['F97'] = t_sum_cat_3
        
        sheet['G97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['G97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['G97'] = t_sum_cat_4 
        
        sheet['H97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['H97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['H97'] = t_sum_cat_5
        
        sheet['I97'].alignment = Alignment(horizontal= "center", vertical="center")
        sheet['I97'].font = Font(name = 'Arial', size= 9, bold = True)
        sheet['I97'] = t_sum_cat_6 
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_operacional_establec.xlsx"

        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        workbook.save(response)

        return response

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.shortcuts import render, redirect

# DIRECTORIO MUNICIPIO 
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView
from .padron_models import Directorio_municipio, Directorio_salud
from .forms import Directorio_MunicipioForm , Directorio_SaludForm

# TABLERO SELLO 
from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo
from django.db.models.functions import Substr
import logging

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from .utils import generar_operacional

from django.db.models.functions import Substr

from datetime import datetime
import locale

logger = logging.getLogger(__name__)



class DirectorioMunicipioCreateView(CreateView):
    model = Directorio_municipio
    form_class = Directorio_MunicipioForm
    template_name = 'municipio/directorio_form.html'
    success_url = reverse_lazy('municipio-list')

    def get_initial(self):
        # Inicializar datos como se hacía en la función original
        #empleados = Empleado.objects.get(user=self.request.user)

        initial_data = {
            'estado_auditoria': '0',
        }
        return initial_data


class DirectorioMunicipioListView(ListView):
    model = Directorio_municipio
    template_name = 'municipio/directorio_list.html'
    context_object_name = 'municipios'

    def get_queryset(self):
        return Directorio_municipio.objects.filter(user=self.request.user)


def directorio_municipalidad_detail(request, municipio_directorio_id):
    if request.method == 'GET':
        directorio_municipalidad = get_object_or_404(Directorio_municipio, pk=municipio_directorio_id)
        form = Directorio_MunicipioForm(instance=directorio_municipalidad)
        context = {
            'directorio_municipalidad': directorio_municipalidad,
            'form': form
        }
        return render(request, 'municipio/directorio_detail.html', context)
    else:
        try:
            directorio_municipalidad = get_object_or_404(
                Directorio_municipio, pk=municipio_directorio_id)
            form = Directorio_MunicipioForm(request.POST, request.FILES, instance=directorio_municipalidad)
            form.save()
            return redirect('municipio-list')
        except ValueError:
            return render(request, 'municipio/directorio_detail.html', {'directorio_municipalidad': directorio_municipalidad, 'form': form, 'error': 'Error actualizar'})


class DirectorioMunicipioListViewPublic(ListView):
    model = Directorio_municipio
    template_name = 'municipio/directorio_public.html'
    context_object_name = 'municipios'

    def get_queryset(self):
        return Directorio_municipio.objects.filter(estado_auditoria='1')


### SALUD 
class DirectorioSaludCreateView(CreateView):
    model = Directorio_salud
    form_class = Directorio_SaludForm
    template_name = 'salud/directorio_form.html'
    success_url = reverse_lazy('salud-list')

    def get_initial(self):
        # Inicializar datos como se hacía en la función original
        #empleados = Empleado.objects.get(user=self.request.user)

        initial_data = {
            'estado_auditoria': '0',
        }
        return initial_data


class DirectorioSaludListView(ListView):
    model = Directorio_salud
    template_name = 'salud/directorio_list.html'
    context_object_name = 'salud'

    def get_queryset(self):
        return Directorio_salud.objects.filter(user=self.request.user)


def directorio_salud_detail(request, salud_directorio_id):
    if request.method == 'GET':
        directorio_salud = get_object_or_404(Directorio_salud, pk=salud_directorio_id)
        form = Directorio_SaludForm(instance=directorio_salud)
        context = {
            'directorio_salud': directorio_salud,
            'form': form
        }
        return render(request, 'salud/directorio_detail.html', context)
    else:
        try:
            directorio_salud = get_object_or_404(
                Directorio_salud, pk=salud_directorio_id)
            form = Directorio_SaludForm(request.POST, request.FILES, instance=directorio_salud)
            form.save()
            return redirect('salud-list')
        except ValueError:
            return render(request, 'salud/directorio_detail.html', {'directorio_salud': directorio_salud, 'form': form, 'error': 'Error actualizar'})


class DirectorioSaludListViewPublic(ListView):
    model = Directorio_salud
    template_name = 'salud/directorio_public.html'
    context_object_name = 'salud'

    def get_queryset(self):
        return Directorio_salud.objects.filter(estado_auditoria='1')


#####  DE SELLO MUNICIPAL INDICADOR 1
def obtener_distritos(provincia):
    distritos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Provincia=provincia).values('Distrito').distinct().order_by('Distrito')
    return list(distritos)


def obtener_avance(provincia, distrito):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.obtener_avance(%s, %s)",
            [provincia, distrito]
        )
        return cursor.fetchall()


def obtener_ranking(mes):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.obtener_ranking(%s)", [mes]
        )
        return cursor.fetchall()


def index_sello(request):
    mes_seleccionado = request.GET.get('mes', 'SETIEMBRE')
    provincia_seleccionada = request.GET.get('provincia')
    distrito_seleccionado = request.GET.get('distrito')

    provincias = MAESTRO_HIS_ESTABLECIMIENTO.objects.values_list('Provincia', flat=True).distinct().order_by('Provincia')

    # Si la solicitud es AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Verificar si se solicitan distritos
            if 'get_distritos' in request.GET:
                distritos = obtener_distritos(provincia_seleccionada)
                return JsonResponse(distritos, safe=False)

            # Obtener datos de avance y ranking
            resultados_avance = obtener_avance(provincia_seleccionada, distrito_seleccionado)
            resultados_ranking = obtener_ranking(mes_seleccionado)

            # Procesar los resultados
            data = {
                'fechas': [row[2] for row in resultados_avance],
                'num': [float(row[3]) for row in resultados_avance],
                'den': [float(row[4]) for row in resultados_avance],
                'avance': [float(row[5]) for row in resultados_avance],
                
                'provincia': [row[0] for row in resultados_ranking],
                'distrito': [row[1] for row in resultados_ranking],
                'num_r': [float(row[2]) for row in resultados_ranking],
                'den_r': [float(row[3]) for row in resultados_ranking],
                'avance_r': [float(row[4]) for row in resultados_ranking],
            }

            return JsonResponse(data)

        except Exception as e:
            logger.error(f"Error al obtener datos: {str(e)}")
            return JsonResponse({'error': str(e)}, status=500)

    # Si no es una solicitud AJAX, renderiza la página principal
    return render(request, 'sello/index_sello.html', {
        'provincias': provincias,
        'mes_seleccionado': mes_seleccionado,
    })


#--- PROVINCIAS -------------------------------------------------------------
def sello_get_provincias(request,provincias_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    context = {
                'provincias': provincias,
            }
    
    return render(request, 'sello/provincias.html', context)


def sello_get_distritos(request, distritos_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'sello/distritos.html', context)


def sello_p_distritos(request):
    provincia_param = request.GET.get('provincia')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 'sello/partials/p_distritos.html', context)


def obtener_seguimiento_distrito(provincia, distrito):
    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM public.fn_seguimiento_sello(%s, %s)",
            [provincia, distrito]
        )
        return cursor.fetchall()


class RptOperacinalDist(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        provincia = request.GET.get('provincia')
        distritos = request.GET.get('distritos')
        
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_distrito(provincia, distritos)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_seguimiento_sello_distrito.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

def fill_worksheet(ws, results): 
    # cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 3
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 3
    ws.row_dimensions[7].height = 3
    ws.row_dimensions[8].height = 25
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 33
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 4
    ws.column_dimensions['K'].width = 4
    ws.column_dimensions['L'].width = 4
    ws.column_dimensions['M'].width = 8
    ws.column_dimensions['N'].width = 30
    ws.column_dimensions['O'].width = 30
    ws.column_dimensions['P'].width = 30
    ws.column_dimensions['Q'].width = 9
    ws.column_dimensions['R'].width = 8
    ws.column_dimensions['S'].width = 12
    ws.column_dimensions['T'].width = 15
    ws.column_dimensions['U'].width = 10
    ws.column_dimensions['V'].width = 8
    ws.column_dimensions['W'].width = 8
    ws.column_dimensions['X'].width = 9
    ws.column_dimensions['Y'].width = 9
    ws.column_dimensions['Z'].width = 8
    ws.column_dimensions['AA'].width = 10
    # linea de division
    ws.freeze_panes = 'M9'
    # Configuración del fondo y el borde
    fill = PatternFill(patternType='solid', fgColor='00B0F0')
    # Definir el color anaranjado usando PatternFill
    orange_fill = PatternFill(patternType='solid', fgColor='FFA500')
    # Definir los estilos
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    

    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL INDICADOR 1 - SELLO MUNICIPAL'
    
    ws['B6'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B6'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B6'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'
        
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['B8'].fill = fill
    ws['B8'].border = border
    ws['B8'] = 'COD PAD'
    
    ws['C8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['C8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['C8'].fill = fill
    ws['C8'].border = border
    ws['C8'] = 'CNV'
    
    ws['D8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['D8'].fill = fill
    ws['D8'].border = border
    ws['D8'] = 'CUI'      
    
    ws['E8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['E8'].fill = fill
    ws['E8'].border = border
    ws['E8'] = 'DNI' 
    
    ws['F8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['F8'].fill = fill
    ws['F8'].border = border
    ws['F8'] = 'NOMBRE COMPLETO DE NIÑO/A'     
    
    ws['G8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['G8'].fill = gray_fill
    ws['G8'].border = border
    ws['G8'] = 'VAL DNI'    
    
    ws['H8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['H8'].fill = fill
    ws['H8'].border = border
    ws['H8'] = 'SEXO'    
    
    ws['I8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['I8'].fill = fill
    ws['I8'].border = border
    ws['I8'] = 'FECHA NAC'    
    
    ws['J8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['J8'].fill = fill
    ws['J8'].border = border
    ws['J8'] = 'ED A'  
    
    ws['K8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['K8'].fill = fill
    ws['K8'].border = border
    ws['K8'] = 'ED M'  
    
    ws['L8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['L8'].fill = fill
    ws['L8'].border = border
    ws['L8'] = 'EDA D'  
    
    ws['M8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['M8'].fill = fill
    ws['M8'].border = border
    ws['M8'] = 'AREA'  
    
    ws['N8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['N8'].fill = fill
    ws['N8'].border = border
    ws['N8'] = 'EJE VIAL'  
    
    ws['O8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['O8'].fill = fill
    ws['O8'].border = border
    ws['O8'] = 'DIRECCION'  
    
    ws['P8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['P8'].fill = fill
    ws['P8'].border = border
    ws['P8'] = 'REFERENCIA'  
    
    ws['Q8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Q8'].fill = gray_fill
    ws['Q8'].border = border
    ws['Q8'] = 'VAL DIR'    
    
    ws['R8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['R8'].fill = fill
    ws['R8'].border = border
    ws['R8'] = 'UBIGUEO' 
    
    ws['S8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['S8'].fill = fill
    ws['S8'].border = border
    ws['S8'] = 'PROVINCIA' 
    
    ws['T8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['T8'].fill = fill
    ws['T8'].border = border
    ws['T8'] = 'DISTRITO' 
    
    ws['U8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['U8'].fill = fill
    ws['U8'].border = border
    ws['U8'] = 'VISITADO' 
    
    ws['V8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['V8'].fill = fill
    ws['V8'].border = border
    ws['V8'] = 'ENCONTRADO' 
    
    ws['W8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['W8'].fill = gray_fill
    ws['W8'].border = border
    ws['W8'] = 'VAL V/E'   
    
    ws['X8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['X8'].fill = fill
    ws['X8'].border = border
    ws['X8'] = 'DNI MADRE' 
    
    ws['Y8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['Y8'].fill = fill
    ws['Y8'].border = border
    ws['Y8'] = 'CELULAR' 
    
    ws['Z8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z8'].font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    ws['Z8'].fill = fill
    ws['Z8'].border = border
    ws['Z8'] = 'MES EVAL'  
    
    ws['AA8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA8'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AA8'].fill = orange_fill
    ws['AA8'].border = border
    ws['AA8'] = 'IND'         
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    # Escribir datos
    for row, record in enumerate(results, start=9):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [6, 14, 15, 16]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar color en la columna 27
            if col == 27:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=8,  bold = True, color='FFFFFF')  # Letra blanca
                    elif value_upper == "CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')  # Fondo verde
                        cell.font = Font(name='Arial', size=8,  bold = True, color='FFFFFF')  # Letra blanca
                    else:
                        cell.font = Font(name='Arial', size=8)
                else:
                    cell.font = Font(name='Arial', size=8)
            
            # Aplicar color de letra en las columnas 7 y 17
            elif col in [7, 17]:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.font = Font(name='Arial', size=8, color="FF0000")  # Letra roja
                    elif value_upper == "CUMPLE":
                        cell.font = Font(name='Arial', size=8, color="00B050")  # Letra verde
                    else:
                        cell.font = Font(name='Arial', size=8)
                else:
                    cell.font = Font(name='Arial', size=8)
            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)  # Fuente normal para otras columnas

            cell.border = border
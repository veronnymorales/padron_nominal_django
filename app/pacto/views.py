import logging
from django.shortcuts import render
from django.http import JsonResponse
from .queries import (obtener_distritos, obtener_ranking_pacto, obtener_ranking_distritos, obtener_seguimiento_pacto, obtener_seguimiento_pacto_distrito)

from base.models import MAESTRO_HIS_ESTABLECIMIENTO, Actualizacion

from django.db.models.functions import Substr

from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo, Actualizacion
from django.db.models.functions import Substr

# REPORTE EXCEL
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models.functions import Substr

from datetime import datetime
import locale

from django.db.models import IntegerField  # Importar IntegerField
from django.db.models.functions import Cast, Substr  # Importar Cast y Substr
# linea de border 
from openpyxl.utils import column_index_from_string

# Reporte excel
from datetime import datetime
import getpass  # Para obtener el nombre del usuario
from django.contrib.auth.models import User  # O tu modelo de usuario personalizado
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

User = get_user_model()

from django.db.models import IntegerField             # Importar IntegerField
from django.db.models.functions import Cast, Substr     # Importar Cast y Substr

logger = logging.getLogger(__name__)


def index_pacto_nino(request):
    actualizacion = Actualizacion.objects.all()

    # Provincias para el primer <select>
    provincias = (MAESTRO_HIS_ESTABLECIMIENTO.objects
                    .values_list('Provincia', flat=True)
                    .distinct()
                    .order_by('Provincia'))
    
    # Obtener parámetros
    departamento_selecionado = request.GET.get('departamento')
    provincia_seleccionada = request.GET.get('provincia')
    distrito_seleccionado = request.GET.get('distrito')

    # -- Manejo de distritos via HTMX (retorna template parcial) --
    if 'get_distritos' in request.GET:
        if provincia_seleccionada:
            distritos = obtener_distritos(provincia_seleccionada)
        else:
            distritos = []
        
        return render(request, "pn_situacion_actual/partials/_distritos_options.html", {
            "distritos": distritos
        })

    # -- Si es una solicitud AJAX, devolvemos JsonResponse con la data --
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Intentamos procesar la data sin mostrar errores
        try:
            # Llamada redundante a get_distritos (por si la plantilla antigua lo usara)
            if 'get_distritos' in request.GET:
                distritos = obtener_distritos(provincia_seleccionada)
                return JsonResponse(distritos, safe=False)
              
            # OBTENER RANKING ZONA
            resultados_ranking_pacto = obtener_ranking_pacto()
            print("resultados_ranking_pacto:", resultados_ranking_pacto)
            
                        
            # OBTENER RANKING DISTRITO
            resultados_ranking_distrito = obtener_ranking_distritos()
            #print("resultados_ranking_distrito:", resultados_ranking_distrito)
            
            # Estructura de datos inicial
            data = {
                # DATA RANKING ZONA
                'zona': [],
                'den': [],
                'num': [],
                'avance': [],
                'estado': [],              
                
                # DATA RANKING DISTRITO
                'zona_ranking': [],
                'provincia_ranking': [],
                'distrito_ranking': [],
                'den_ranking': [],
                'num_ranking': [],
                'avance_ranking': [],
                'estado_ranking': [],
                
            }

            # ----------------------------------------------------------------------------
            # 1. Procesar RANKING ZONA
            # ----------------------------------------------------------------------------
            for row in resultados_ranking_pacto:
                # En lugar de lanzar error, checamos si la tupla es la longitud esperada:
                if len(row) == 5:
                    data['zona'].append(row[0])
                    # Cambia null (None) a 0
                    data['den'].append(str(row[1]) if row[1] is not None else '0')
                    data['num'].append(str(row[2]) if row[2] is not None else '0')
                    data['avance'].append(str(row[3]) if row[3] is not None else '0')
                    data['estado'].append(str(row[4]) if row[4] is not None else 'RIESGO')
                else:
                    logger.warning(f"Fila con estructura inválida: {row}")
            
            # ----------------------------------------------------------------------------
            # 2. Procesar RANKING DISTRITOS
            # ----------------------------------------------------------------------------
            for row in resultados_ranking_distrito:
                # En lugar de lanzar error, checamos si la tupla es la longitud esperada:
                if len(row) == 7:
                    data['zona_ranking'].append(row[0])
                    data['provincia_ranking'].append(row[1])
                    data['distrito_ranking'].append(row[2])
                    # Cambia null (None) a 0
                    data['den_ranking'].append(str(row[3]) if row[3] is not None else '0')
                    data['num_ranking'].append(str(row[4]) if row[4] is not None else '0')
                    data['avance_ranking'].append(str(row[5]) if row[5] is not None else '0')
                    data['estado_ranking'].append(str(row[6]) if row[6] is not None else 'RIESGO')
                else:
                    logger.warning(f"Fila con estructura inválida: {row}")
            
            return JsonResponse(data)
        
        except:
            # Si ocurre alguna excepción global, la silenciamos (no mostramos nada)
            return JsonResponse({}, status=200)
        

    # -- Si no es AJAX, render normal de la plantilla --
    return render(request, 'pacto/index_pacto_nino.html', {
        'actualizacion': actualizacion,
        'provincias': provincias,
    })


#------------------------
## SEGUIMIENTO NOMINAL
#------------------------

## SEGUIMIENTO POR PROVINCIAS
def get_provincias_pacto(request, provincia_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    context = {
                'provincias': provincias,
            }
    
    return render(request, 'compromiso_nino/provincias.html', context)

## SEGUIMIENTO POR DISTRITOS
def get_distritos_pacto(request, distrito_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(periodo_filtrado=Substr('Periodo', 1, 6))
                .values('Mes','periodo_filtrado')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'compromiso_nino/distritos.html', context)

def p_distrito_pacto(request):
    provincia_param = request.GET.get('provincia')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 'compromiso_nino/partials/p_distritos.html', context)


## REPORTE DE EXCEL
class RptPactoProvincia(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_departamento = 'JUNIN'
        p_provincia = request.GET.get('provincia')
        p_edades =  request.GET.get('edades','')
        p_cumple = request.GET.get('cumple', '') 

        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_pacto(p_departamento, p_provincia, p_edades, p_cumple)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_compromiso(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_compromiso_provincia.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response
    
class RptPactoDistrito(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_departamento = 'JUNIN'
        p_provincia = request.GET.get('provincia')
        p_distrito = request.GET.get('distrito')
        p_edades =  request.GET.get('edades','')
        p_cumple = request.GET.get('cumple', '') 

        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_pacto_distrito(p_departamento, p_provincia, p_distrito, p_edades, p_cumple)
        

        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet_compromiso(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_pacto_distrito.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response
    

def fill_worksheet_compromiso(ws, results): 
    # cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 12
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 25
    
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 9
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 9
    ws.column_dimensions['H'].width = 17
    ws.column_dimensions['I'].width = 9
    ws.column_dimensions['J'].width = 35
    ws.column_dimensions['K'].width = 6
    ws.column_dimensions['L'].width = 9
    ws.column_dimensions['M'].width = 9
    ws.column_dimensions['N'].width = 4
    ws.column_dimensions['O'].width = 4
    ws.column_dimensions['P'].width = 4
    ws.column_dimensions['Q'].width = 22
    ws.column_dimensions['R'].width = 25
    ws.column_dimensions['S'].width = 35
    ws.column_dimensions['T'].width = 25
    ws.column_dimensions['U'].width = 7
    ws.column_dimensions['V'].width = 9
    ws.column_dimensions['W'].width = 15
    ws.column_dimensions['X'].width = 15
    ws.column_dimensions['Y'].width = 9
    ws.column_dimensions['Z'].width = 25    
    ws.column_dimensions['AA'].width = 9
    ws.column_dimensions['AB'].width = 9       
    ws.column_dimensions['AC'].width = 9
    ws.column_dimensions['AD'].width = 9    
    ws.column_dimensions['AE'].width = 9
    ws.column_dimensions['AF'].width = 9   
    ws.column_dimensions['AG'].width = 9    
    ws.column_dimensions['AH'].width = 9   
    ws.column_dimensions['AI'].width = 25
    ws.column_dimensions['AJ'].width = 9    
    ws.column_dimensions['AK'].width = 25
    ws.column_dimensions['AL'].width = 18    
    ws.column_dimensions['AM'].width = 9
    ws.column_dimensions['AN'].width = 9    
    ws.column_dimensions['AO'].width = 25
    ws.column_dimensions['AP'].width = 9    
    ws.column_dimensions['AQ'].width = 9
    ws.column_dimensions['AR'].width = 9    
    ws.column_dimensions['AS'].width = 25
    ws.column_dimensions['AT'].width = 9   
    ws.column_dimensions['AU'].width = 9
    ws.column_dimensions['AV'].width = 18
    ws.column_dimensions['AW'].width = 9
    ws.column_dimensions['AX'].width = 9
    ws.column_dimensions['AY'].width = 13
    ws.column_dimensions['AZ'].width = 9
    ws.column_dimensions['BA'].width = 13
    ws.column_dimensions['BB'].width = 10
    ws.column_dimensions['BC'].width = 15
    ws.column_dimensions['BD'].width = 9
    ws.column_dimensions['BE'].width = 11
    ws.column_dimensions['BF'].width = 5
    ws.column_dimensions['BG'].width = 9
    ws.column_dimensions['BH'].width = 6
    ws.column_dimensions['BI'].width = 6
    ws.column_dimensions['BJ'].width = 25
    ws.column_dimensions['BK'].width = 7
    ws.column_dimensions['BL'].width = 5 
    ws.column_dimensions['BM'].width = 9
    ws.column_dimensions['BN'].width = 5 
    ws.column_dimensions['BO'].width = 18
    ws.column_dimensions['BP'].width = 5
    ws.column_dimensions['BQ'].width = 18
    ws.column_dimensions['BR'].width = 9
    ws.column_dimensions['BS'].width = 5
    ws.column_dimensions['BT'].width = 18
    ws.column_dimensions['BU'].width = 18
    ws.column_dimensions['BV'].width = 18
    ws.column_dimensions['BW'].width = 5
    
    # linea de division
    ws.freeze_panes = 'H7'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
        # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    # Definir el estilo de relleno celeste
    celeste_fill = PatternFill(start_color='FF87CEEB', end_color='FF87CEEB', fill_type='solid')
    # Morado más claro
    morado_claro_fill = PatternFill(start_color='FFE9D8FF', end_color='FFE9D8FF', fill_type='solid')
    # Plomo más claro
    plomo_claro_fill = PatternFill(start_color='FFEDEDED', end_color='FFEDEDED', fill_type='solid')
    # Azul más claro
    azul_claro_fill = PatternFill(start_color='FFD8EFFA', end_color='FFD8EFFA', fill_type='solid')
    # Naranja más claro
    naranja_claro_fill = PatternFill(start_color='FFFFEBD8', end_color='FFFFEBD8', fill_type='solid')
    # Verde más claro
    verde_claro_fill = PatternFill(start_color='FFBDF7BD', end_color='FFBDF7BD', fill_type='solid')
    # Guinda (bordó / burdeos)
    guinda_claro_fill = PatternFill(start_color='FFE8A8A6', end_color='FFE8A8A6', fill_type='solid')

        
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    border_negro = Border(left=Side(style='thin', color='000000'), # negro
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')) 
    
    # Merge cells 
    # numerador y denominador
    ws.merge_cells('B5:Q5') 
    ws.merge_cells('R5:AB5')
    ws.merge_cells('AC5:AG5')
    ws.merge_cells('AH5:AO5')
    ws.merge_cells('AQ5:AV5')
    ws.merge_cells('AW5:BD5')
    ws.merge_cells('BF5:BW5')

    # Combina cela
    ws['B5'] = 'DATOS DEL MENOR - VARIABLES DE IDENTIFICACION'
    ws['R5'] = 'DIRECCION COMPLETA DEL MENOR'
    ws['AC5'] = 'VISITAS DOMICILARIAS'
    ws['AH5'] = 'VARIBALES CON INFORMACION DE SALUD'
    ws['AQ5'] = 'DATOS DE LA MADRE'
    ws['AW5'] = 'AUDITORIA DE LOS REGISTROS'
    ws['BF5'] = 'INFORMACION HIS MINSA - ULTIMA ATENCION REGIONAL'

    ### numerador y denominador 
    
    ws['B5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['B5'].fill = fill
    ws['B5'].border = border_negro
    
    ws['R5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['R5'].fill = gray_fill
    ws['R5'].border = border_negro
    
    ### intervalo 
    ws['AC5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AC5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['AC5'].fill = green_fill
    ws['AC5'].border = border_negro
    
    ws['AH5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['AH5'].fill = blue_fill
    ws['AH5'].border = border_negro

    ws['AQ5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AQ5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['AQ5'].fill = yellow_fill
    ws['AQ5'].border = border_negro
    
    ws['AW5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AW5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['AW5'].fill = green_fill_2
    ws['AW5'].border = border_negro
    
    ws['BF5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BF5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['BF5'].fill = celeste_fill
    ws['BF5'].border = border_negro
    
    ### BORDE DE CELDAS CONBINADAS
    
    # NUM y DEN
    inicio_columna = 'B'
    fin_columna = 'BW'
    fila = 5
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # NUM y DEN
    inicio_columna = 'B'
    fin_columna = 'BW'
    fila = 6
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
            
    ##### imprimer fecha y hora del reporte
    fecha_hora_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nombre_usuario = getpass.getuser()

    # Obtener el usuario actualmente autenticado
    try:
        user = User.objects.get(is_active=True)
    except User.DoesNotExist:
        user = None
    except User.MultipleObjectsReturned:
        # Manejar el caso donde hay múltiples usuarios activos
        user = User.objects.filter(is_active=True).first()  # Por ejemplo, obtener el primero
    # Asignar fecha y hora a la celda A1
    ws['V1'].value = 'Fecha y Hora:'
    ws['W1'].value = fecha_hora_actual

    # Asignar nombre de usuario a la celda A2
    ws['V2'].value = 'Usuario:'
    ws['W2'].value = nombre_usuario
    
    # Formatear las etiquetas en negrita
    etiqueta_font = Font(name='Arial', size=8)
    ws['V1'].font = etiqueta_font
    ws['W1'].font = etiqueta_font
    ws['V2'].font = etiqueta_font
    ws['W2'].font = etiqueta_font

    # Alinear el texto
    ws['V1'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W1'].alignment = Alignment(horizontal="left", vertical="center")
    ws['V2'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W2'].alignment = Alignment(horizontal="left", vertical="center")
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL PADRON NOMINAL DE LA REGION JUNIN'
    
    ws['B3'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B3'].font = Font(name = 'Arial', size= 7, color='0000CC')
    ws['B3'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'
        
    ws['AP5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP5'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AP5'].fill = celeste_fill
    
    ws['BE5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BE5'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BE5'].fill = orange_fill

    
    
    
    
    ws['B6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['B6'].fill = fill
    ws['B6'].border = border_negro
    ws['B6'] = 'ID'
    
    ws['C6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['C6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['C6'].fill = fill
    ws['C6'].border = border_negro 
    ws['C6'] = 'COD PADRON'
    
    ws['D6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['D6'].fill = fill
    ws['D6'].border = border
    ws['D6'] = 'TIP DOC'      
    
    ws['E6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['E6'].fill = fill
    ws['E6'].border = border
    ws['E6'] = 'CNV' 
    
    ws['F6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['F6'].fill = fill
    ws['F6'].border = border
    ws['F6'] = 'CUI'     
    
    ws['G6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['G6'].fill = fill
    ws['G6'].border = border
    ws['G6'] = 'DNI'    
    
    ws['H6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['H6'].fill = fill
    ws['H6'].border = border
    ws['H6'] = 'ESTADO TRAMITE'    
    
    ws['I6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['I6'].fill = fill
    ws['I6'].border = border
    ws['I6'] = 'IND DNI'    
    
    ws['J6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['J6'].fill = fill
    ws['J6'].border = border
    ws['J6'] = 'NOMBRE DE NIÑO/A'  
    
    ws['K6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['K6'].fill = fill
    ws['K6'].border = border
    ws['K6'] = 'SEXO'  
    
    ws['L6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['L6'].fill = fill
    ws['L6'].border = border
    ws['L6'] = 'SEGURO'  
    
    ws['M6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['M6'].fill = fill
    ws['M6'].border = border
    ws['M6'] = 'FECHA NAC'  
    
    ws['N6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['N6'].fill = fill
    ws['N6'].border = border
    ws['N6'] = 'ED A'  
    
    ws['O6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['O6'].fill = fill
    ws['O6'].border = border
    ws['O6'] = 'ED M' 
    
    ws['P6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['P6'].fill = fill
    ws['P6'].border = border
    ws['P6'] = 'ED D'  
    
    ws['Q6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Q6'].fill = fill
    ws['Q6'].border = border
    ws['Q6'] = 'EDAD ACTUAL' 
    
    ws['R6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['R6'].fill = gray_fill
    ws['R6'].border = border
    ws['R6'] = 'EJE VIAL'  
    
    ws['S6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['S6'].fill = gray_fill
    ws['S6'].border = border
    ws['S6'] = 'DIRECCION' 
    
    ws['T6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['T6'].fill = gray_fill
    ws['T6'].border = border
    ws['T6'] = 'REFERENCIA'  
    
    ws['U6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['U6'].fill = gray_fill
    ws['U6'].border = border
    ws['U6'] = 'UBIGUEO' 
    
    ws['V6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['V6'].fill = gray_fill
    ws['V6'].border = border
    ws['V6'] = 'DEP'  
    
    ws['W6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['W6'].fill = gray_fill
    ws['W6'].border = border
    ws['W6'] = 'PROVINCIA' 
        
    ws['X6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['X6'].fill = gray_fill
    ws['X6'].border = border
    ws['X6'] = 'DISTRITO' 

    ws['Y6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Y6'].fill = gray_fill
    ws['Y6'].border = border
    ws['Y6'] = 'COD CP'  
    
    ws['Z6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Z6'].fill = gray_fill
    ws['Z6'].border = border
    ws['Z6'] = 'CENTRO POBLADO' 

    ws['AA6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AA6'].fill = gray_fill
    ws['AA6'].border = border
    ws['AA6'] = 'AREA'  
    
    ws['AB6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AB6'].fill = gray_fill
    ws['AB6'].border = border
    ws['AB6'] = 'IND DIR' 
    
    ws['AC6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AC6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AC6'].fill = green_fill
    ws['AC6'].border = border
    ws['AC6'] = 'MENOR VISITADO'  
    
    ws['AD6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AD6'].fill = green_fill
    ws['AD6'].border = border
    ws['AD6'] = 'MENOR ENCON' 
    
    ws['AE6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AE6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AE6'].fill = green_fill
    ws['AE6'].border = border
    ws['AE6'] = 'FECHA VISITA'  
    
    ws['AF6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AF6'].fill = green_fill
    ws['AF6'].border = border
    ws['AF6'] = 'IND VIS' 
    
    ws['AG6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AG6'].fill = green_fill
    ws['AG6'].border = border
    ws['AG6'] = 'TRANSITO'  
    
    ws['AH6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AH6'].fill = blue_fill
    ws['AH6'].border = border
    ws['AH6'] = 'COD NAC' 
    
    ws['AI6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AI6'].fill = blue_fill
    ws['AI6'].border = border
    ws['AI6'] = 'ESTABLECIMIENTO DE NACIMIENTO'  
    
    ws['AJ6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AJ6'].fill = blue_fill
    ws['AJ6'].border = border
    ws['AJ6'] = 'COD EESS' 
    
    ws['AK6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AK6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AK6'].fill = blue_fill
    ws['AK6'].border = border
    ws['AK6'] = 'ESTABLECIMIENTO DE ATENCION' 
    
    ws['AL6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AL6'].fill = blue_fill
    ws['AL6'].border = border
    ws['AL6'] = 'FRECUENCIA ATENCION'  
    
    ws['AM6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AM6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AM6'].fill = blue_fill
    ws['AM6'].border = border
    ws['AM6'] = 'IND SALUD' 
    
    ws['AN6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AN6'].fill = blue_fill
    ws['AN6'].border = border
    ws['AN6'] = 'COD ADS'  
    
    ws['AO6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AO6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AO6'].fill = blue_fill
    ws['AO6'].border = border
    ws['AO6'] = 'ESTABLECIMIENTO DE ADSCRIPCION (SIS)' 
    
    ws['AP6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AP6'].fill = celeste_fill
    ws['AP6'].border = border_negro
    ws['AP6'] = 'PROG. SOCIAL'  
    
    ws['AQ6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AQ6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AQ6'].fill = yellow_fill
    ws['AQ6'].border = border
    ws['AQ6'] = 'TIPO DOC' 
    
    ws['AR6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AR6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AR6'].fill = yellow_fill
    ws['AR6'].border = border
    ws['AR6'] = 'NUM DOC'  
    
    ws['AS6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AS6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AS6'].fill = yellow_fill
    ws['AS6'].border = border
    ws['AS6'] = 'NOMBRE DE LA MADRE' 
    
    ws['AT6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AT6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AT6'].fill = yellow_fill
    ws['AT6'].border = border
    ws['AT6'] = 'CELULAR'  
    
    ws['AU6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AU6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AU6'].fill = yellow_fill
    ws['AU6'].border = border
    ws['AU6'] = 'IND MADRE' 
    
    ws['AV6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AV6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AV6'].fill = yellow_fill
    ws['AV6'].border = border
    ws['AV6'] = 'CORREO ELECTRONICO' 
    
    ws['AW6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AW6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AW6'].fill = green_fill_2
    ws['AW6'].border = border
    ws['AW6'] = 'ESTADO REGISTRO' 
    
    ws['AX6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AX6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AX6'].fill = green_fill_2
    ws['AX6'].border = border
    ws['AX6'] = 'FECHA REGISTRO' 
    
    ws['AY6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AY6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AY6'].fill = green_fill_2
    ws['AY6'].border = border
    ws['AY6'] = 'USUARIO'  
    
    ws['AZ6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AZ6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AZ6'].fill = green_fill_2
    ws['AZ6'].border = border
    ws['AZ6'] = 'FECHA MODIF'       
    
    ws['BA6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BA6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BA6'].fill = green_fill_2
    ws['BA6'].border = border
    ws['BA6'] = 'USUARIO MODIF' 
    
    ws['BB6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BB6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BB6'].fill = green_fill_2
    ws['BB6'].border = border
    ws['BB6'] = 'ENTIDAD'  
    
    ws['BC6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BC6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BC6'].fill = green_fill_2
    ws['BC6'].border = border
    ws['BC6'] = 'TIPO REGISTRO'  
    
    ws['BD6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BD6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BD6'].fill = green_fill_2
    ws['BD6'].border = border
    ws['BD6'] = 'FECHA CORTE'  
    
    ws['BE6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BE6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BE6'].fill = orange_fill
    ws['BE6'].border = border
    ws['BE6'] = 'INDICADOR' 
    
    ws['BF6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BF6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BF6'].fill = celeste_fill
    ws['BF6'].border = border
    ws['BF6'] = 'DEN'  
    
    ws['BG6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BG6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BG6'].fill = celeste_fill
    ws['BG6'].border = border
    ws['BG6'] = 'FECHA ULT ATE'       
    
    ws['BH6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BH6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BH6'].fill = celeste_fill
    ws['BH6'].border = border
    ws['BH6'] = 'RENAES' 
    
    ws['BI6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BI6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BI6'].fill = celeste_fill
    ws['BI6'].border = border
    ws['BI6'] = 'ID EST'  
    
    ws['BJ6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BJ6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BJ6'].fill = celeste_fill
    ws['BJ6'].border = border
    ws['BJ6'] = 'NOMBRE ESTABLECIMIENTO DE ATENCION'  
    
    ws['BK6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BK6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BK6'].fill = celeste_fill
    ws['BK6'].border = border
    ws['BK6'] = 'UBIG ESTA'  
    
    ws['BL6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BL6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BL6'].fill = celeste_fill
    ws['BL6'].border = border
    ws['BL6'] = 'COD DISA'  
    
    ws['BM6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BM6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BM6'].fill = celeste_fill
    ws['BM6'].border = border
    ws['BM6'] = 'DISA'  
    
    ws['BN6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BN6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BN6'].fill = celeste_fill
    ws['BN6'].border = border
    ws['BN6'] = 'COD RED'       
    
    ws['BO6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BO6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BO6'].fill = celeste_fill
    ws['BO6'].border = border
    ws['BO6'] = 'RED DE SALUD' 
    
    ws['BP6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BP6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BP6'].fill = celeste_fill
    ws['BP6'].border = border
    ws['BP6'] = 'COD MICRO'  
    
    ws['BQ6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BQ6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BQ6'].fill = celeste_fill
    ws['BQ6'].border = border
    ws['BQ6'] = 'NOMBRE DE MICRORED'  
    
    ws['BR6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BR6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BR6'].fill = celeste_fill
    ws['BR6'].border = border
    ws['BR6'] = 'COD UNICO' 
    
    ws['BS6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BS6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BS6'].fill = celeste_fill
    ws['BS6'].border = border
    ws['BS6'] = 'COD SEC'   
    
    ws['BT6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BT6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BT6'].fill = celeste_fill
    ws['BT6'].border = border
    ws['BT6'] = 'SECTOR'  
    
    ws['BU6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BU6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BU6'].fill = celeste_fill
    ws['BU6'].border = border
    ws['BU6'] = 'PROVINCIA'  
    
    ws['BV6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BV6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BV6'].fill = celeste_fill
    ws['BV6'].border = border
    ws['BV6'] = 'DISTRITO'  
    
    ws['BW6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BW6'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BW6'].fill = celeste_fill
    ws['BW6'].border = border
    ws['BW6'] = 'CAT EST'  
    
        
        # Define styles
    promo_fill = PatternFill(patternType='solid', fgColor='FFD966')  # Yellow fill for promo
    font_normal = Font(name='Arial', size=8)
    font_bold_white = Font(name='Arial', size=7, bold=True, color='FFFFFF')
    font_red_bold = Font(name='Arial', size=7, bold=True, color='FF0000')
    font_green_bold = Font(name='Arial', size=7, bold=True, color='00FF00')
    font_red = Font(name='Arial', size=7, color='FF0000')
    font_green = Font(name='Arial', size=7, color='00B050')
    font_check = Font(name='Arial', size=10, color='00B050')
    font_x = Font(name='Arial', size=10, color='FF0000')
    plomo_claro_font = Font(name='Arial', size=7, color='FFEDEDED', bold=False)
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True, color='FFFFFF')
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    # Definir los caracteres especiales de check y X
    check_mark = '✓'  # Unicode para check
    x_mark = '✗'  # Unicode para X
    sub_cumple = 'CUMPLE'
    sub_no_cumple = 'NO CUMPLE'
    
    
    # Escribir datos
    for row, record in enumerate(results, start=7):
        for col, value in enumerate(record, start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [10, 17, 18, 19, 20,24,26,35, 37, 41,45, 62, 67, 69, 74 ]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar color en la columna 27
            if col == 57:
                if value == 0:
                    cell.value = sub_no_cumple  # Insertar check
                    cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                    cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                elif value == 1:
                    cell.value = sub_cumple  
                    cell.fill = PatternFill(patternType='solid', fgColor='00FF00')  # Fondo verde
                    cell.font = Font(name='Arial', size=7,  bold = True, color='FFFFFF')  # Letra blanca
                else:
                    cell.font = Font(name='Arial', size=7)
            
            # Aplicar color de letra SUB INDICADORES
            elif col in [9, 32, 39,47]:
                if value == 0:
                    cell.value = sub_no_cumple  # Insertar check
                    cell.font = Font(name='Arial', size=7, color="FF0000")  # Letra roja
                elif value == 1:
                    cell.value = sub_cumple # Insertar check
                    cell.font = Font(name='Arial', size=7, color="00B050")  # Letra verde
                else:
                        cell.font = Font(name='Arial', size=7)
                        
            elif col in [28,39,47]:
                if value == '0':
                    cell.value = sub_no_cumple  # Insertar check
                    cell.font = Font(name='Arial', size=7, color="FF0000")  # Letra roja
                elif value == '1':
                    cell.value = sub_cumple # Insertar check
                    cell.font = Font(name='Arial', size=7, color="00B050")  # Letra verde
                else:
                    cell.font = Font(name='Arial', size=7)
                        
            # Aplicar color de letra SUB INDICADORES
            elif col in [9, 28, 33,39,47]:
                cell.font = Font(name='Arial', size=8, color="FF000033")

            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)  # Fuente normal para otras columnas

            # Aplicar caracteres especiales check y X
            if col in [33,58]:
                if value == 1:
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == 0:
                    cell.value = x_mark  # Insertar X
                    cell.font = Font(name='Arial', size=10, color='FF0000')  # Letra roja
                else:
                    cell.font = Font(name='Arial', size=8)  # Fuente normal si no es 1 o 0
            
            if col in [33,58]:
                if value == '1':
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == '0':
                    cell.value = x_mark  # Insertar X
                    cell.font = Font(name='Arial', size=10, color='FF0000')  # Letra roja
                else:
                    cell.font = Font(name='Arial', size=8) 
            
                        
            cell.border = border
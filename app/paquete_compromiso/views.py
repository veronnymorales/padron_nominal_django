from django.shortcuts import render

# TABLERO PAQUETE GESTANTE 
from django.db import connection
from django.http import JsonResponse
from base.models import MAESTRO_HIS_ESTABLECIMIENTO, DimPeriodo, Actualizacion
from django.db.models.functions import Substr
import logging

from .queries import (obtener_avance_paquete_compromiso, obtener_variables_paquete_compromiso, obtener_avance_regional_mensual_paquete_compromiso,
                    obtener_cobertura_por_zona, obtener_cobertura_por_provincia, obtener_cobertura_por_distrito, 
                    obtener_seguimiento_paquete_compromiso,
                    obtener_avance_regional_mensual_paquete_compromiso, obtener_avance_cobertura_paquete_compromiso, 
                    obtener_cobertura_por_red, obtener_cobertura_por_microred, obtener_cobertura_por_establecimiento, 
                    obtener_seguimiento_paquete_compromiso_red, obtener_seguimiento_paquete_compromiso_microred, obtener_seguimiento_paquete_compromiso_establecimiento)

# report excel
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import openpyxl
from openpyxl.utils import get_column_letter

from django.db.models.functions import Substr

from datetime import datetime
import locale

from django.db.models import IntegerField  # Importar IntegerField
from django.db.models.functions import Cast, Substr  # Importar Cast y Substr
# linea de border 
from openpyxl.utils import column_index_from_string

# Reporte excel
from datetime import datetime
import getpass  # Para obtener el nombre del usuario
from django.contrib.auth.models import User  # O tu modelo de usuario personalizado
from django.http import HttpResponse
from io import BytesIO
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

User = get_user_model()

from django.db.models import IntegerField             # Importar IntegerField
from django.db.models.functions import Cast, Substr     # Importar Cast y Substr

logger = logging.getLogger(__name__)


def BASE(request):
    actualizacion = Actualizacion.objects.all()
    return render(request,'paquete_compromiso/index_paquete_compromiso.html', {"actualizacion": actualizacion})
############################
## COMPONENTES Y GRAFICOS 
############################
def process_avance_por_region(resultados_avance_por_region):
    data = {    
        'r_numerador_resumen': [],
        'r_denominador_resumen': [],
        'r_avance_resumen': []
    }
    
    # Si no hay resultados, retornar valores por defecto
    if not resultados_avance_por_region:
        data['r_numerador_resumen'] = [0]
        data['r_denominador_resumen'] = [0]
        data['r_avance_resumen'] = [0.0]
        #print(f"[PROCESS] Sin datos de entrada, usando valores por defecto: {data}")
        return data
    
    # Procesar el primer (y único) registro
    row = resultados_avance_por_region[0]
    try:
        # Mapear las claves correctas del resultado
        numerador = row.get('num', 0)
        denominador = row.get('den', 0) 
        avance = row.get('avance', 0.0)

        # Asegurar que los valores no sean None
        numerador = numerador if numerador is not None else 0
        denominador = denominador if denominador is not None else 0
        avance = avance if avance is not None else 0.0

        data['r_numerador_resumen'].append(int(numerador))
        data['r_denominador_resumen'].append(int(denominador))
        data['r_avance_resumen'].append(float(avance))
        
        #print(f"[PROCESS] Datos procesados: Num={numerador}, Den={denominador}, Avance={avance}%")
        
    except (KeyError, ValueError, TypeError) as e:
        logger.error(f"Error procesando datos: {e}, Row: {row}")
        # Valores por defecto en caso de error
        data['r_numerador_resumen'] = [0]
        data['r_denominador_resumen'] = [0]
        data['r_avance_resumen'] = [0.0]
    
    #print(f"[PROCESS] Resultado final: {data}")
    return data

def process_avance_regional_mensual_paquete_compromiso(resultados_avance_regional_mensual_paquete_compromiso):
    """Procesa los resultados del graficos"""
    data = {
        'num_1': [],
        'den_1': [],
        'cob_1': [],
        'num_2': [],
        'den_2': [],
        'cob_2': [],
        'num_3': [],
        'den_3': [],
        'cob_3': [],
        'num_4': [],
        'den_4': [],
        'cob_4': [],
        'num_5': [],
        'den_5': [],
        'cob_5': [],
        'num_6': [],
        'den_6': [],
        'cob_6': [],
        'num_7': [],
        'den_7': [],
        'cob_7': [],
        'num_8': [],
        'den_8': [],
        'cob_8': [],                
        'num_9': [],
        'den_9': [],
        'cob_9': [],
        'num_10': [],
        'den_10': [],
        'cob_10': [],
        'num_11': [],
        'den_11': [],
        'cob_11': [],
        'num_12': [],
        'den_12': [],
        'cob_12': [],
    }
    for index, row in enumerate(resultados_avance_regional_mensual_paquete_compromiso):
        try:
            # Verifica que el diccionario tenga las claves necesarias
            required_keys = {'num_1','den_1','cob_1','num_2','den_2','cob_2','num_3','den_3','cob_3','num_4','den_4','cob_4','num_5','den_5','cob_5','num_6','den_6','cob_6','num_7','den_7','cob_7','num_8','den_8','cob_8','num_9','den_9','cob_9','num_10','den_10','cob_10','num_11','den_11','cob_11','num_12','den_12','cob_12'}
            
            if not required_keys.issubset(row.keys()):
                raise ValueError(f"La fila {index} no tiene las claves necesarias: {row}")
            # Extraer cada valor, convirtiendo a float
            num_1_value = float(row.get('num_1', 0.0))
            den_1_value = float(row.get('den_1', 0.0))
            cob_1_value = float(row.get('cob_1', 0.0))
            num_2_value = float(row.get('num_2', 0.0))
            den_2_value = float(row.get('den_2', 0.0))
            cob_2_value = float(row.get('cob_2', 0.0))
            num_3_value = float(row.get('num_3', 0.0))
            den_3_value = float(row.get('den_3', 0.0))
            cob_3_value = float(row.get('cob_3', 0.0))
            num_4_value = float(row.get('num_4', 0.0))
            den_4_value = float(row.get('den_4', 0.0))
            cob_4_value = float(row.get('cob_4', 0.0))
            num_5_value = float(row.get('num_5', 0.0))
            den_5_value = float(row.get('den_5', 0.0))
            cob_5_value = float(row.get('cob_5', 0.0))
            num_6_value = float(row.get('num_6', 0.0))
            den_6_value = float(row.get('den_6', 0.0))
            cob_6_value = float(row.get('cob_6', 0.0))
            num_7_value = float(row.get('num_7', 0.0))
            den_7_value = float(row.get('den_7', 0.0))
            cob_7_value = float(row.get('cob_7', 0.0))
            num_8_value = float(row.get('num_8', 0.0))
            den_8_value = float(row.get('den_8', 0.0))
            cob_8_value = float(row.get('cob_8', 0.0))
            num_9_value = float(row.get('num_9', 0.0))
            den_9_value = float(row.get('den_9', 0.0))
            cob_9_value = float(row.get('cob_9', 0.0))
            num_10_value = float(row.get('num_10', 0.0))
            den_10_value = float(row.get('den_10', 0.0))
            cob_10_value = float(row.get('cob_10', 0.0))
            num_11_value = float(row.get('num_11', 0.0))
            den_11_value = float(row.get('den_11', 0.0))
            cob_11_value = float(row.get('cob_11', 0.0))
            num_12_value = float(row.get('num_12', 0.0))
            den_12_value = float(row.get('den_12', 0.0))
            cob_12_value = float(row.get('cob_12', 0.0))
            
            data['num_1'].append(num_1_value)
            data['den_1'].append(den_1_value)
            data['cob_1'].append(cob_1_value)
            data['num_2'].append(num_2_value)
            data['den_2'].append(den_2_value)
            data['cob_2'].append(cob_2_value)
            data['num_3'].append(num_3_value)
            data['den_3'].append(den_3_value)
            data['cob_3'].append(cob_3_value)
            data['num_4'].append(num_4_value)
            data['den_4'].append(den_4_value)
            data['cob_4'].append(cob_4_value)
            data['num_5'].append(num_5_value)
            data['den_5'].append(den_5_value)
            data['cob_5'].append(cob_5_value)
            data['num_6'].append(num_6_value)
            data['den_6'].append(den_6_value)
            data['cob_6'].append(cob_6_value)
            data['num_7'].append(num_7_value)
            data['den_7'].append(den_7_value)
            data['cob_7'].append(cob_7_value)
            data['num_8'].append(num_8_value)
            data['den_8'].append(den_8_value)
            data['cob_8'].append(cob_8_value)
            data['num_9'].append(num_9_value)
            data['den_9'].append(den_9_value)
            data['cob_9'].append(cob_9_value)
            data['num_10'].append(num_10_value)
            data['den_10'].append(den_10_value)
            data['cob_10'].append(cob_10_value)
            data['num_11'].append(num_11_value)
            data['den_11'].append(den_11_value)
            data['cob_11'].append(cob_11_value)
            data['num_12'].append(num_12_value)
            data['den_12'].append(den_12_value)
            data['cob_12'].append(cob_12_value)

        except Exception as e:
            logger.error(f"Error procesando la fila {index}: {str(e)}")
    return data

def process_variables_por_region(resultados_variables_por_region):
    data = {    
        'v_den_variable': [],
        'v_num_cred': [],
        'v_avance_cred': [],
        'v_num_cred_rn': [],
        'v_avance_cred_rn': [],
        'v_num_cred_mensual': [],    
        'v_avance_cred_mensual': [],
        'v_num_vac':[], 
        'v_avance_vac':[],
        'v_num_vac_antineumococica': [],
        'v_avance_vac_antineumococica': [],
        'v_num_vac_antipolio': [],
        'v_avance_vac_antipolio': [],
        'v_num_vac_pentavalente': [],
        'v_avance_vac_pentavalente': [],
        'v_num_vac_rotavirus': [],
        'v_avance_vac_rotavirus': [],
        'v_num_esq': [],
        'v_avance_esq': [],
        'v_num_esq4M': [],
        'v_avance_esq4M': [],
        'v_num_esq6M': [],
        'v_avance_esq6M': [],
        'v_num_esq6M_trat': [],
        'v_avance_esq6M_trat': [],
        'v_num_esq6M_multi': [],
        'v_avance_esq6M_multi': [],
        'v_num_dosaje_Hb': [],
        'v_avance_num_dosaje_Hb': [],
        'v_num_DNIemision': [],
        'v_avance_DNIemision': []
    }
    
    # Si no hay resultados, retornar valores por defecto
    if not resultados_variables_por_region:
        data['v_den_variable'] = [0]
        data['v_num_cred'] = [0]
        data['v_avance_cred'] = [0.0]
        data['v_num_cred_rn'] = [0]
        data['v_avance_cred_rn'] = [0.0]
        data['v_num_cred_mensual'] = [0]
        data['v_avance_cred_mensual'] = [0.0]
        data['v_num_vac'] = [0]
        data['v_avance_vac'] = [0.0]
        data['v_num_vac_antineumococica'] = [0]
        data['v_avance_vac_antineumococica'] = [0.0]
        data['v_num_vac_antipolio'] = [0]
        data['v_avance_vac_antipolio'] = [0.0]
        data['v_num_vac_pentavalente'] = [0]
        data['v_avance_vac_pentavalente'] = [0.0]
        data['v_num_vac_rotavirus'] = [0]
        data['v_avance_vac_rotavirus'] = [0.0]
        data['v_num_esq'] = [0]
        data['v_avance_esq'] = [0.0]
        data['v_num_esq4M'] = [0]
        data['v_avance_esq4M'] = [0.0]
        data['v_num_esq6M'] = [0]
        data['v_avance_esq6M'] = [0.0]
        data['v_num_esq6M_trat'] = [0]
        data['v_avance_esq6M_trat'] = [0.0]
        data['v_num_esq6M_multi'] = [0]
        data['v_avance_esq6M_multi'] = [0.0]
        data['v_num_dosaje_Hb'] = [0]
        data['v_avance_num_dosaje_Hb'] = [0.0]
        data['v_num_DNIemision'] = [0]
        data['v_avance_DNIemision'] = [0.0]
        return data
    
    # Procesar el primer (y único) registro
    row = resultados_variables_por_region[0]
    try:
        # Mapear las claves correctas del resultado        
        den_variable = row.get('den_variable', 0)
        num_cred = row.get('num_cred', 0)
        avance_cred = row.get('avance_cred', 0.0)
        num_cred_rn = row.get('num_cred_rn', 0)
        avance_cred_rn = row.get('avance_cred_rn', 0.0)
        num_cred_mensual = row.get('num_cred_mensual', 0)
        avance_cred_mensual = row.get('avance_cred_mensual', 0.0)
        num_vac = row.get('num_vac', 0)     
        avance_vac = row.get('avance_vac', 0.0)
        num_vac_antineumococica = row.get('num_vac_antineumococica', 0)
        avance_vac_antineumococica = row.get('avance_vac_antineumococica', 0.0)
        num_vac_antipolio = row.get('num_vac_antipolio', 0)
        avance_vac_antipolio = row.get('avance_vac_antipolio', 0.0)
        num_vac_pentavalente = row.get('num_vac_pentavalente', 0)
        avance_vac_pentavalente = row.get('avance_vac_pentavalente', 0.0)
        num_vac_rotavirus = row.get('num_vac_rotavirus', 0)
        avance_vac_rotavirus = row.get('avance_vac_rotavirus', 0.0)
        num_esq = row.get('num_esq', 0)
        avance_esq = row.get('avance_esq', 0.0)
        num_esq4M = row.get('num_esq4M', 0)
        avance_esq4M = row.get('avance_esq4M', 0.0)
        num_esq6M = row.get('num_esq6M', 0)
        avance_esq6M = row.get('avance_esq6M', 0.0)
        num_esq6M_trat = row.get('num_esq6M_trat', 0)
        avance_esq6M_trat = row.get('avance_esq6M_trat', 0.0)
        num_esq6M_multi = row.get('num_esq6M_multi', 0)
        avance_esq6M_multi = row.get('avance_esq6M_multi', 0.0)
        num_dosaje_Hb = row.get('num_dosaje_Hb', 0)
        avance_num_dosaje_Hb = row.get('avance_num_dosaje_Hb', 0.0)
        num_DNIemision = row.get('num_DNIemision', 0)
        avance_DNIemision = row.get('avance_DNIemision', 0.0)


        # Asegurar que los valores no sean None        
        den_variable = den_variable if den_variable is not None else 0
        num_cred = num_cred if num_cred is not None else 0
        avance_cred = avance_cred if avance_cred is not None else 0.0
        num_cred_rn = num_cred_rn if num_cred_rn is not None else 0
        avance_cred_rn = avance_cred_rn if avance_cred_rn is not None else 0.0
        num_cred_mensual = num_cred_mensual if num_cred_mensual is not None else 0
        avance_cred_mensual = avance_cred_mensual if avance_cred_mensual is not None else 0.0
        num_vac = num_vac if num_vac is not None else 0
        avance_vac = avance_vac if avance_vac is not None else 0.0
        num_vac_antineumococica = num_vac_antineumococica if num_vac_antineumococica is not None else 0
        avance_vac_antineumococica = avance_vac_antineumococica if avance_vac_antineumococica is not None else 0.0
        num_vac_antipolio = num_vac_antipolio if num_vac_antipolio is not None else 0
        avance_vac_antipolio = avance_vac_antipolio if avance_vac_antipolio is not None else 0.0
        num_vac_pentavalente = num_vac_pentavalente if num_vac_pentavalente is not None else 0
        avance_vac_pentavalente = avance_vac_pentavalente if avance_vac_pentavalente is not None else 0.0
        num_vac_rotavirus = num_vac_rotavirus if num_vac_rotavirus is not None else 0
        avance_vac_rotavirus = avance_vac_rotavirus if avance_vac_rotavirus is not None else 0.0
        num_esq = num_esq if num_esq is not None else 0
        avance_esq = avance_esq if avance_esq is not None else 0.0
        num_esq4M = num_esq4M if num_esq4M is not None else 0
        avance_esq4M = avance_esq4M if avance_esq4M is not None else 0.0
        num_esq6M = num_esq6M if num_esq6M is not None else 0
        avance_esq6M = avance_esq6M if avance_esq6M is not None else 0.0
        num_esq6M_trat = num_esq6M_trat if num_esq6M_trat is not None else 0
        avance_esq6M_trat = avance_esq6M_trat if avance_esq6M_trat is not None else 0.0
        num_esq6M_multi = num_esq6M_multi if num_esq6M_multi is not None else 0
        avance_esq6M_multi = avance_esq6M_multi if avance_esq6M_multi is not None else 0.0
        num_dosaje_Hb = num_dosaje_Hb if num_dosaje_Hb is not None else 0
        avance_num_dosaje_Hb = avance_num_dosaje_Hb if avance_num_dosaje_Hb is not None else 0.0
        num_DNIemision = num_DNIemision if num_DNIemision is not None else 0
        avance_DNIemision = avance_DNIemision if avance_DNIemision is not None else 0.0

        data['v_den_variable'].append(int(den_variable))
        data['v_num_cred'].append(int(num_cred))
        data['v_avance_cred'].append(float(avance_cred))
        data['v_num_cred_rn'].append(int(num_cred_rn))
        data['v_avance_cred_rn'].append(float(avance_cred_rn))
        data['v_num_cred_mensual'].append(int(num_cred_mensual))
        data['v_avance_cred_mensual'].append(float(avance_cred_mensual))
        data['v_num_vac'].append(int(num_vac))
        data['v_avance_vac'].append(float(avance_vac))
        data['v_num_vac_antineumococica'].append(int(num_vac_antineumococica))
        data['v_avance_vac_antineumococica'].append(float(avance_vac_antineumococica))
        data['v_num_vac_antipolio'].append(int(num_vac_antipolio))
        data['v_avance_vac_antipolio'].append(float(avance_vac_antipolio))
        data['v_num_vac_pentavalente'].append(int(num_vac_pentavalente))
        data['v_avance_vac_pentavalente'].append(float(avance_vac_pentavalente))
        data['v_num_vac_rotavirus'].append(int(num_vac_rotavirus))
        data['v_avance_vac_rotavirus'].append(float(avance_vac_rotavirus))
        data['v_num_esq'].append(int(num_esq))
        data['v_avance_esq'].append(float(avance_esq))
        data['v_num_esq4M'].append(int(num_esq4M))
        data['v_avance_esq4M'].append(float(avance_esq4M))
        data['v_num_esq6M'].append(int(num_esq6M))
        data['v_avance_esq6M'].append(float(avance_esq6M))
        data['v_num_esq6M_trat'].append(int(num_esq6M_trat))
        data['v_avance_esq6M_trat'].append(float(avance_esq6M_trat))
        data['v_num_esq6M_multi'].append(int(num_esq6M_multi))
        data['v_avance_esq6M_multi'].append(float(avance_esq6M_multi))
        data['v_num_dosaje_Hb'].append(int(num_dosaje_Hb))
        data['v_avance_num_dosaje_Hb'].append(float(avance_num_dosaje_Hb))
        data['v_num_DNIemision'].append(int(num_DNIemision))
        data['v_avance_DNIemision'].append(float(avance_DNIemision))
        
    except (KeyError, ValueError, TypeError) as e:
        logger.error(f"Error procesando datos: {e}, Row: {row}")
        # Valores por defecto en caso de error
        data['v_den_variable'] = [0]
        data['v_num_cred'] = [0]
        data['v_avance_cred'] = [0.0]
        data['v_num_cred_rn'] = [0]
        data['v_avance_cred_rn'] = [0.0]
        data['v_num_cred_mensual'] = [0]
        data['v_avance_cred_mensual'] = [0.0]
        data['v_num_vac'] = [0]
        data['v_avance_vac'] = [0.0]
        data['v_num_vac_antineumococica'] = [0]
        data['v_avance_vac_antineumococica'] = [0.0]
        data['v_num_vac_antipolio'] = [0]
        data['v_avance_vac_antipolio'] = [0.0]
        data['v_num_vac_pentavalente'] = [0]
        data['v_avance_vac_pentavalente'] = [0.0]
        data['v_num_vac_rotavirus'] = [0]
        data['v_avance_vac_rotavirus'] = [0.0]
        data['v_num_esq'] = [0]
        data['v_avance_esq'] = [0.0]
        data['v_num_esq4M'] = [0]
        data['v_avance_esq4M'] = [0.0]
        data['v_num_esq6M'] = [0]
        data['v_avance_esq6M'] = [0.0]
        data['v_num_esq6M_trat'] = [0]
        data['v_avance_esq6M_trat'] = [0.0]
        data['v_num_esq6M_multi'] = [0]
        data['v_avance_esq6M_multi'] = [0.0]
        data['v_num_dosaje_Hb'] = [0]
        data['v_avance_num_dosaje_Hb'] = [0.0]
        data['v_num_DNIemision'] = [0]
        data['v_avance_DNIemision'] = [0.0]

    #print(f"[PROCESS] Resultado final: {data}")
    return data

def process_cobertura_por_zona(resultados_cobertura_por_zona):
    """Procesa los resultados del graficos"""
    data = {
            'z_zona': [],
            'z_den': [],
            'z_num': [],
            'z_brecha': [],
            'z_cob': [],
    }   
    for row in resultados_cobertura_por_zona:
        try:
            data['z_zona'].append(row['z_zona'])

            # Cambia null (None) a 0
            data['z_den'].append(row['z_den'] if row['z_den'] is not None else 0)
            data['z_num'].append(row['z_num'] if row['z_num'] is not None else 0)
            data['z_brecha'].append(row['z_brecha'] if row['z_brecha'] is not None else 0)
            data['z_cob'].append(row['z_cob'] if row['z_cob'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_provincia(resultados_cobertura_por_provincia):
    """Procesa los resultados del graficos"""
    data = {
            'p_provincia': [],
            'p_den': [],
            'p_num': [],
            'p_brecha': [],
            'p_cob': [],
    }   
    for row in resultados_cobertura_por_provincia:
        try:
            data['p_provincia'].append(row['p_provincia'])

            # Cambia null (None) a 0
            data['p_den'].append(row['p_den'] if row['p_den'] is not None else 0)
            data['p_num'].append(row['p_num'] if row['p_num'] is not None else 0)
            data['p_brecha'].append(row['p_brecha'] if row['p_brecha'] is not None else 0)
            data['p_cob'].append(row['p_cob'] if row['p_cob'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_distrito(resultados_cobertura_por_distrito):
    """Procesa los resultados del graficos"""
    data = {
            'd_distrito': [],
            'd_den': [],
            'd_num': [],
            'd_brecha': [],
            'd_cob': [],
    }   
    for row in resultados_cobertura_por_distrito:
        try:
            data['d_distrito'].append(row['d_distrito'])

            # Cambia null (None) a 0
            data['d_den'].append(row['d_den'] if row['d_den'] is not None else 0)
            data['d_num'].append(row['d_num'] if row['d_num'] is not None else 0)
            data['d_brecha'].append(row['d_brecha'] if row['d_brecha'] is not None else 0)
            data['d_cob'].append(row['d_cob'] if row['d_cob'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

##----------------------------
def process_avance_cobertura_paquete_compromiso(resultados_avance_cobertura_paquete_compromiso):
    """Procesa los resultados del graficos"""
    data = {
            'anio': [],
            'mes': [],
            'Ubigueo_Establecimiento':[],
            'Distrito': [],
            'Provincia': [],
            'Codigo_Red': [],
            'red': [],
            'Codigo_MicroRed': [],
            'microred': [],
            'Codigo_Unico': [],
            'Nombre_Establecimiento': [],
            'grupo_edad': [],
            'total_denominador': [],
            'total_numerador': [],
            'total_brecha': [],
            'cobertura_porcentaje': [],
    }   
    for row in resultados_avance_cobertura_paquete_compromiso:
        try:
            data['anio'].append(row['anio'])
            data['mes'].append(row['mes'])
            data['Ubigueo_Establecimiento'].append(row['Ubigueo_Establecimiento'])
            data['Distrito'].append(row['Distrito'])
            data['Provincia'].append(row['Provincia'])
            data['Codigo_Red'].append(row['Codigo_Red'])
            data['red'].append(row['red'])
            data['Codigo_MicroRed'].append(row['Codigo_MicroRed'])
            data['microred'].append(row['microred'])
            data['Codigo_Unico'].append(row['Codigo_Unico'])
            data['Nombre_Establecimiento'].append(row['Nombre_Establecimiento'])
            data['grupo_edad'].append(row['grupo_edad'])

            # Cambia null (None) a 0
            data['total_denominador'].append(row['total_denominador'] if row['total_denominador'] is not None else 0)
            data['total_numerador'].append(row['total_numerador'] if row['total_numerador'] is not None else 0)
            data['total_brecha'].append(row['total_brecha'] if row['total_brecha'] is not None else 0)
            data['cobertura_porcentaje'].append(row['cobertura_porcentaje'] if row['cobertura_porcentaje'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_edad(resultados_cobertura_por_edad):
    """Procesa los resultados del graficos"""
    data = {    
            'c_grupo_edad': [],
            'c_denominador': [],
            'c_numerador': [],
            'c_brecha': [],
            'c_cobertura': [],
    }
    for row in resultados_cobertura_por_edad:
        try:
            data['c_grupo_edad'].append(row['c_grupo_edad'])

            # Cambia null (None) a 0
            data['c_denominador'].append(row['c_denominador'] if row['c_denominador'] is not None else 0)
            data['c_numerador'].append(row['c_numerador'] if row['c_numerador'] is not None else 0)
            data['c_brecha'].append(row['c_brecha'] if row['c_brecha'] is not None else 0)
            data['c_cobertura'].append(row['c_cobertura'] if row['c_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_red(resultados_cobertura_por_red):
    """Procesa los resultados del graficos"""
    data = {
            'r_red': [],
            'r_denominador': [],
            'r_numerador': [],
            'r_brecha': [],
            'r_cobertura': [],
    }   
    for row in resultados_cobertura_por_red:
        try:
            data['r_red'].append(row['r_red'])

            # Cambia null (None) a 0
            data['r_denominador'].append(row['r_denominador'] if row['r_denominador'] is not None else 0)
            data['r_numerador'].append(row['r_numerador'] if row['r_numerador'] is not None else 0)
            data['r_brecha'].append(row['r_brecha'] if row['r_brecha'] is not None else 0)
            data['r_cobertura'].append(row['r_cobertura'] if row['r_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_microred(resultados_cobertura_por_microred):
    """Procesa los resultados del graficos"""
    data = {
            'm_microred': [],
            'm_denominador': [],
            'm_numerador': [],
            'm_brecha': [],
            'm_cobertura': [],
    }   
    for row in resultados_cobertura_por_microred:
        try:
            data['m_microred'].append(row['m_microred'])

            # Cambia null (None) a 0
            data['m_denominador'].append(row['m_denominador'] if row['m_denominador'] is not None else 0)
            data['m_numerador'].append(row['m_numerador'] if row['m_numerador'] is not None else 0)
            data['m_brecha'].append(row['m_brecha'] if row['m_brecha'] is not None else 0)
            data['m_cobertura'].append(row['m_cobertura'] if row['m_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data

def process_cobertura_por_establecimiento(resultados_cobertura_por_establecimiento):
    """Procesa los resultados del graficos"""
    data = {
            'e_establecimiento': [],
            'e_denominador': [],
            'e_numerador': [],
            'e_brecha': [],
            'e_cobertura': [],
    }   
    for row in resultados_cobertura_por_establecimiento:
        try:
            data['e_establecimiento'].append(row['e_establecimiento'])

            # Cambia null (None) a 0
            data['e_denominador'].append(row['e_denominador'] if row['e_denominador'] is not None else 0)
            data['e_numerador'].append(row['e_numerador'] if row['e_numerador'] is not None else 0)
            data['e_brecha'].append(row['e_brecha'] if row['e_brecha'] is not None else 0)
            data['e_cobertura'].append(row['e_cobertura'] if row['e_cobertura'] is not None else 0)
        except KeyError as e:
            logger.warning(f"Fila con estructura inválida (clave faltante: {e}): {row}")

    return data


#######################
## PANTALLA PRINCIPAL
#######################
def index_paquete_compromiso(request):
    actualizacion = Actualizacion.objects.all()

    # Capturamos el año que viene por GET
    anio = request.GET.get('anio', '2025')
    mes = request.GET.get('mes', '')
    provincia = request.GET.get('provincia', '')
    distrito = request.GET.get('distrito', '')

    if anio not in ['2024', '2025' , '2026']:
        anio = '2025'

    redes_h = (
        MAESTRO_HIS_ESTABLECIMIENTO
        .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL', Disa='JUNIN')
        .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
        .values('Red', 'codigo_red_filtrado')
        .distinct()
        .order_by('Red')
    )
    
    provincias_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )

    mes_seleccionado_inicio = request.GET.get('mes_inicio')
    mes_seleccionado_fin = request.GET.get('mes_fin')
    
    provincia_seleccionada = request.GET.get('provincia_h')
    distrito_seleccionado = request.GET.get('distrito_h')
    
    red_seleccionada = request.GET.get('red_h')
    microred_seleccionada = request.GET.get('p_microredes_establec_h')
    establecimiento_seleccionado = request.GET.get('p_establecimiento_h')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        try:
            # Unificar la lógica de obtención de datos
            #print(f"Parámetros recibidos - Año: {anio}, Mes inicio: {mes_seleccionado_inicio}, Mes fin: {mes_seleccionado_fin}, Provincia: {provincia_seleccionada}, Distrito: {distrito_seleccionado}")
            
            # Lógica para la sección de avance regional
            resultados_avance_por_region = obtener_avance_paquete_compromiso(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            resultados_variables_por_region = obtener_variables_paquete_compromiso(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)           
            resultados_avance_regional_mensual_paquete_compromiso = obtener_avance_regional_mensual_paquete_compromiso(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado) 
            
            resultados_cobertura_por_zona = obtener_cobertura_por_zona(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            resultados_cobertura_por_provincia = obtener_cobertura_por_provincia(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            resultados_cobertura_por_distrito = obtener_cobertura_por_distrito(anio, mes_seleccionado_inicio, mes_seleccionado_fin, provincia_seleccionada, distrito_seleccionado)
            
            # Lógica para la sección de cobertura
            resultados_avance_cobertura_paquete_compromiso = obtener_avance_cobertura_paquete_compromiso(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            
            resultados_cobertura_por_red = obtener_cobertura_por_red(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            resultados_cobertura_por_microred = obtener_cobertura_por_microred(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            resultados_cobertura_por_establecimiento = obtener_cobertura_por_establecimiento(anio, mes, red_seleccionada, microred_seleccionada, establecimiento_seleccionado, provincia, distrito)
            
            # Combinar todos los resultados en un solo diccionario
            data = {
                **process_avance_por_region(resultados_avance_por_region),
                **process_variables_por_region(resultados_variables_por_region),
                **process_avance_regional_mensual_paquete_compromiso(resultados_avance_regional_mensual_paquete_compromiso),
                **process_avance_cobertura_paquete_compromiso(resultados_avance_cobertura_paquete_compromiso),
                
                # Graficos de barras
                **process_cobertura_por_zona(resultados_cobertura_por_zona),
                **process_cobertura_por_provincia(resultados_cobertura_por_provincia),
                **process_cobertura_por_distrito(resultados_cobertura_por_distrito),
                
                **process_cobertura_por_red(resultados_cobertura_por_red),
                **process_cobertura_por_microred(resultados_cobertura_por_microred),
                **process_cobertura_por_establecimiento(resultados_cobertura_por_establecimiento)
            }

            return JsonResponse(data)

        except Exception as e:
            logger.error(f"Error al obtener datos: {str(e)}")
            return JsonResponse({'error': f"Error al obtener datos: {str(e)}"}, status=500)

    # Renderizado inicial de la página
    return render(request, 'paquete_compromiso/index_paquete_compromiso.html', {
        'mes_seleccionado_inicio': mes_seleccionado_inicio,
        'mes_seleccionado_fin': mes_seleccionado_fin,
        'actualizacion': actualizacion,
        'provincia_seleccionada': provincia_seleccionada,
        'distrito_seleccionado': distrito_seleccionado,
        'provincias_h': provincias_h,
    })

#####################################
## FILTROS HORIZONTAL
#####################################
##----------------------------------
## FILTROS HORIZONTAL POR SALUD
##----------------------------------
def get_establecimientos_paquete_compromiso_h(request,establecimiento_id):
    redes_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
                .order_by('Red')
    )
    provincias_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes_h': redes_h,
                'provincias_h': provincias_h,
                'mes_inicio':mes_inicio
    }
    return render(request,'paquete_compromiso/establecimientos_h.html', context)

def p_microredes_establec_paquete_compromiso_h(request):
    redes_param = request.GET.get('red_h') 
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Codigo_Red=redes_param, Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN').values('Codigo_MicroRed','MicroRed').distinct()
    context = {
        'microredes': microredes,
        'is_htmx': True
    }
    return render(request, 'paquete_compromiso/partials/p_microredes_establec_h.html', context)

def p_establecimientos_paquete_compromiso_h(request):
    microredes = request.GET.get('p_microredes_establec_h', '')    
    codigo_red = request.GET.get('red_h', '')
    
    # Construir el filtro dinámicamente
    filtros = {
        'Descripcion_Sector': 'GOBIERNO REGIONAL',
        'Disa': 'JUNIN'
    }
    
    if microredes:
        filtros['Codigo_MicroRed'] = microredes
    if codigo_red:
        filtros['Codigo_Red__startswith'] = codigo_red
    
    establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(**filtros).values('Codigo_Unico','Nombre_Establecimiento').distinct()
    
    # Debug: contar establecimientos encontrados
    establec_list = list(establec)
    
    context= {
        'establec': establec_list
    }
    return render(request, 'paquete_compromiso/partials/p_establecimientos_h.html', context)

##-----------------------------------
## FILTROS HORIZONTAL POR MUNICIPIO
##-----------------------------------
def p_distritos_paquete_compromiso_h(request):
    provincias_h = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    
    provincia_param = request.GET.get('provincia_h')
    
    distritos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
        Ubigueo_Establecimiento__startswith=provincia_param,
        Descripcion_Sector='GOBIERNO REGIONAL'
    ).values('Ubigueo_Establecimiento', 'Distrito').distinct().order_by('Distrito')
    context = {
        'distritos': distritos,
        'provincias_h' :  provincias_h
    }
    return render(request, 'paquete_compromiso/partials/p_distritos.html', context)

###########################################
## SEGUIMIENTO NOMINAL FILTROS
##########################################

###################---------------------------
## FILTRO AMBITO DE SALUD
##################---------------------------

## SEGUIMIENTO POR REDES
def get_redes_paquete_compromiso(request,redes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'paquete_compromiso/components/salud/redes.html', context)

## SEGUIMIENTO POR MICRO-REDES
def get_microredes_paquete_compromiso(request, microredes_id):
    redes = (
            MAESTRO_HIS_ESTABLECIMIENTO
            .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Departamento='JUNIN')
            .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
            .values('Red','codigo_red_filtrado')
            .distinct()
            .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    
    return render(request, 'paquete_compromiso/components/salud/microredes.html', context)

def p_microredes_paquete_compromiso(request):
    redes_param = request.GET.get('red', '')
    

    
    # Consulta principal
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
        Codigo_Red__startswith=redes_param, 
        Descripcion_Sector='GOBIERNO REGIONAL', 
        Disa='JUNIN'
    ).values('Codigo_MicroRed','MicroRed').distinct()
    
    microredes_list = list(microredes)
    
    context = {
        'redes_param': redes_param,
        'microredes': microredes_list
    }
    
    # Si es una petición AJAX, devolver solo las opciones
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'paquete_compromiso/partials/p_microredes_options.html', context)
    
    return render(request, 'paquete_compromiso/partials/p_microredes.html', context)

## REPORTE POR ESTABLECIMIENTO
def get_establecimientos_paquete_compromiso(request,establecimiento_id):
    redes = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL',Disa='JUNIN')
                .annotate(codigo_red_filtrado=Substr('Codigo_Red', 1, 4))
                .values('Red','codigo_red_filtrado')
                .distinct()
                .order_by('Red')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter(Anio='2024')
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'redes': redes,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request,'paquete_compromiso/components/salud/establecimientos.html', context)

def p_microredes_establec_paquete_compromiso(request):
    redes_param = request.GET.get('red') 
    

    
    # Usar startswith en lugar de igualdad exacta
    microredes = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
        Codigo_Red__startswith=redes_param, 
        Descripcion_Sector='GOBIERNO REGIONAL',
        Disa='JUNIN'
    ).values('Codigo_MicroRed','MicroRed').distinct()
    
    microredes_list = list(microredes)
    
    context = {
        'microredes': microredes_list,
        'is_htmx': True
    }
    return render(request, 'paquete_compromiso/partials/p_microredes_establec.html', context)

def p_establecimientos_paquete_compromiso(request):
    # Limpiar parámetros - obtener solo el primer valor si hay duplicados
    microredes = request.GET.get('p_microredes_establec', '').strip()
    
    # Para el código de red, obtener todos los valores y usar el primero no vacío
    red_values = request.GET.getlist('red')
    codigo_red = ''
    for red_val in red_values:
        if red_val and red_val.strip():
            codigo_red = red_val.strip()
            break
    
    # Construir el filtro dinámicamente
    filtros = {
        'Descripcion_Sector': 'GOBIERNO REGIONAL',
        'Disa': 'JUNIN'
    }
    
    # Agregar filtro de red si está disponible
    if codigo_red:
        filtros['Codigo_Red__startswith'] = codigo_red
    
    # Agregar filtro de microred si está presente
    if microredes:
        filtros['Codigo_MicroRed'] = microredes
        
        # Si hay microred pero no hay red, intentar obtener la red de la microred
        if not codigo_red:
            try:
                red_desde_microred = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(
                    Codigo_MicroRed=microredes,
                    Descripcion_Sector='GOBIERNO REGIONAL',
                    Disa='JUNIN'
                ).values('Codigo_Red').first()
                
                if red_desde_microred:
                    codigo_red_obtenido = red_desde_microred['Codigo_Red']
                    filtros['Codigo_Red'] = codigo_red_obtenido
            except Exception as e:
                pass
        
    try:
        establec = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(**filtros).values('Codigo_Unico','Nombre_Establecimiento').distinct()
        establec_list = list(establec)
    except Exception as e:
        establec_list = []
    
    context= {
        'establec': establec_list
    }
    
    # Si es una petición AJAX, devolver solo las opciones
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'paquete_compromiso/partials/p_establecimientos_options.html', context)
    
    return render(request, 'paquete_compromiso/partials/p_establecimientos.html', context)

######################---------------------------
## FILTRO AMBITO DE MUNICIPIO
######################-------------------------------

## SEGUIMIENTO POR PROVINCIA
def get_provincias_paquete_compromiso(request, provincia_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    )
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
            }
    
    return render(request, 'paquete_compromiso/components/municipio/provincias.html', context)

## SEGUIMIENTO POR DISTRITOS
def get_distritos_paquete_compromiso(request, distrito_id):
    provincias = (
                MAESTRO_HIS_ESTABLECIMIENTO
                .objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')
                .annotate(ubigueo_filtrado=Substr('Ubigueo_Establecimiento', 1, 4))
                .values('Provincia','ubigueo_filtrado')
                .distinct()
                .order_by('Provincia')
    )
    mes_inicio = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    mes_fin = (
                DimPeriodo
                .objects.filter()
                .annotate(nro_mes=Cast('NroMes', IntegerField())) 
                .values('Mes','nro_mes')
                .order_by('NroMes')
                .distinct()
    ) 
    context = {
                'provincias': provincias,
                'mes_inicio':mes_inicio,
                'mes_fin':mes_fin,
    }
    return render(request, 'paquete_compromiso/components/municipio/distritos.html', context)

def p_distrito_paquete_compromiso(request):
    provincia_param = request.GET.get('provincia', '')

    # Filtra los establecimientos por sector "GOBIERNO REGIONAL"
    establecimientos = MAESTRO_HIS_ESTABLECIMIENTO.objects.filter(Descripcion_Sector='GOBIERNO REGIONAL')

    # Filtra los establecimientos por el código de la provincia
    if provincia_param:
        establecimientos = establecimientos.filter(Ubigueo_Establecimiento__startswith=provincia_param[:4])
    # Selecciona el distrito y el código Ubigueo
    distritos = establecimientos.values('Distrito', 'Ubigueo_Establecimiento').distinct().order_by('Distrito')
    
    context = {
        'provincia': provincia_param,
        'distritos': distritos
    }
    return render(request, 'paquete_compromiso/partials/p_distritos.html', context)


########################################
## SEGUIMIENTO REPORTE EXCEL 
#######################################

## REPORTE DE EXCEL
class RptPaqueteCompromiso(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        anio = request.GET.get('anio', '2025')
        mes_inicio = request.GET.get('fecha_inicio', '')
        mes_fin = request.GET.get('fecha_fin', '')
        provincia = request.GET.get('provincia', '')
        distrito = request.GET.get('distrito', '')
        p_red = request.GET.get('red', '')
        p_microredes = request.GET.get('p_microredes', '')
        p_establecimiento = request.GET.get('p_establecimiento', '')
        p_cumple = request.GET.get('cumple', '') 

        # Creación de la consulta
        resultado_seguimiento_paquete_compromiso = obtener_seguimiento_paquete_compromiso(anio, mes_inicio, mes_fin, provincia, distrito, p_red, p_microredes, p_establecimiento, p_cumple)
        
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento_paquete_compromiso)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_paquete_compromiso.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptPnPoblacionMicroRed(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_departamento = 'JUNIN'
        p_red = request.GET.get('red', '')
        p_microred = request.GET.get('microredes', '')
        p_establec = ''
        p_edades =  request.GET.get('edades','')
        p_cumple = request.GET.get('cumple', '') 
        # Creación de la consulta
        resultado_seguimiento_microred = obtener_seguimiento_paquete_compromiso_microred(p_departamento, p_red, p_microred, p_edades, p_cumple)

        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento_microred)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_paquete_compromiso_microred.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response

class RptPnPoblacionEstablec(TemplateView):
    def get(self, request, *args, **kwargs):
        # Variables ingresadas
        p_departamento = 'JUNIN'
        p_red = request.GET.get('red','')
        p_microred = request.GET.get('p_microredes_establec','')  # Corregido
        p_establec = request.GET.get('p_establecimiento','')
        p_mes = request.GET.get('mes', '')
        p_edades = request.GET.get('edades', '')
        # Manejo seguro de fechas - usar valores por defecto si no están presentes
        p_cumple = request.GET.get('cumple', '')
        
        # Creación de la consulta
        resultado_seguimiento = obtener_seguimiento_paquete_compromiso_establecimiento(p_departamento,p_establec,p_edades,p_cumple)
                
        wb = Workbook()
        
        consultas = [
                ('Seguimiento', resultado_seguimiento)
        ]
        
        for index, (sheet_name, results) in enumerate(consultas):
            if index == 0:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        
            fill_worksheet(ws, results)
        
        ##########################################################################          
        # Establecer el nombre del archivo
        nombre_archivo = "rpt_paquete_compromiso_establecimiento.xlsx"
        # Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)

        return response


def fill_worksheet(ws, results): 
# cambia el alto de la columna
    ws.row_dimensions[1].height = 14
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 12
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 50
    ws.row_dimensions[8].height = 40
    # cambia el ancho de la columna
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 9
    ws.column_dimensions['D'].width = 9
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 9
    ws.column_dimensions['G'].width = 5
    ws.column_dimensions['H'].width = 5
    ws.column_dimensions['I'].width = 5
    ws.column_dimensions['J'].width = 6
    ws.column_dimensions['K'].width = 5
    ws.column_dimensions['L'].width = 5
    ws.column_dimensions['M'].width = 5
    ws.column_dimensions['N'].width = 5
    ws.column_dimensions['O'].width = 9
    ws.column_dimensions['P'].width = 3
    ws.column_dimensions['Q'].width = 9
    ws.column_dimensions['R'].width = 9
    ws.column_dimensions['S'].width = 9
    ws.column_dimensions['T'].width = 3
    ws.column_dimensions['U'].width = 9
    ws.column_dimensions['V'].width = 3
    ws.column_dimensions['W'].width = 9
    ws.column_dimensions['X'].width = 3
    ws.column_dimensions['Y'].width = 9
    ws.column_dimensions['Z'].width = 3
    ws.column_dimensions['AA'].width = 9
    ws.column_dimensions['AB'].width = 9
    ws.column_dimensions['AC'].width = 3
    ws.column_dimensions['AD'].width = 9
    ws.column_dimensions['AE'].width = 3
    ws.column_dimensions['AF'].width = 9
    ws.column_dimensions['AG'].width = 3
    ws.column_dimensions['AH'].width = 9
    ws.column_dimensions['AI'].width = 3
    ws.column_dimensions['AJ'].width = 9
    ws.column_dimensions['AK'].width = 3
    ws.column_dimensions['AL'].width = 9
    ws.column_dimensions['AM'].width = 3
    ws.column_dimensions['AN'].width = 9
    ws.column_dimensions['AO'].width = 3
    ws.column_dimensions['AP'].width = 9
    ws.column_dimensions['AQ'].width = 3
    ws.column_dimensions['AR'].width = 9
    ws.column_dimensions['AS'].width = 3
    ws.column_dimensions['AT'].width = 9
    ws.column_dimensions['AU'].width = 3
    ws.column_dimensions['AV'].width = 9
    ws.column_dimensions['AW'].width = 3
    ws.column_dimensions['AX'].width = 9
    ws.column_dimensions['AY'].width = 9
    ws.column_dimensions['AZ'].width = 9
    ws.column_dimensions['BA'].width = 3
    ws.column_dimensions['BB'].width = 9
    ws.column_dimensions['BC'].width = 3
    ws.column_dimensions['BD'].width = 9
    ws.column_dimensions['BE'].width = 9
    ws.column_dimensions['BF'].width = 3
    ws.column_dimensions['BG'].width = 9
    ws.column_dimensions['BH'].width = 3
    ws.column_dimensions['BI'].width = 9
    ws.column_dimensions['BJ'].width = 3
    ws.column_dimensions['BK'].width = 9
    ws.column_dimensions['BL'].width = 9
    ws.column_dimensions['BM'].width = 3
    ws.column_dimensions['BN'].width = 9
    ws.column_dimensions['BO'].width = 3
    ws.column_dimensions['BP'].width = 9
    ws.column_dimensions['BQ'].width = 3
    ws.column_dimensions['BR'].width = 9
    ws.column_dimensions['BS'].width = 9
    ws.column_dimensions['BT'].width = 3
    ws.column_dimensions['BU'].width = 9
    ws.column_dimensions['BV'].width = 3
    ws.column_dimensions['BW'].width = 9
    ws.column_dimensions['BX'].width = 9
    ws.column_dimensions['BY'].width = 9
    ws.column_dimensions['BZ'].width = 3
    ws.column_dimensions['CA'].width = 9
    ws.column_dimensions['CB'].width = 9
    ws.column_dimensions['CC'].width = 9
    ws.column_dimensions['CD'].width = 3
    ws.column_dimensions['CE'].width = 9
    ws.column_dimensions['CF'].width = 3
    ws.column_dimensions['CG'].width = 9
    ws.column_dimensions['CH'].width = 9
    ws.column_dimensions['CI'].width = 3
    ws.column_dimensions['CJ'].width = 9
    ws.column_dimensions['CK'].width = 3
    ws.column_dimensions['CL'].width = 9
    ws.column_dimensions['CM'].width = 3
    ws.column_dimensions['CN'].width = 9
    ws.column_dimensions['CO'].width = 9
    ws.column_dimensions['CP'].width = 3
    ws.column_dimensions['CQ'].width = 9
    ws.column_dimensions['CR'].width = 3
    ws.column_dimensions['CS'].width = 9
    ws.column_dimensions['CT'].width = 3
    ws.column_dimensions['CU'].width = 9
    ws.column_dimensions['CV'].width = 3
    ws.column_dimensions['CW'].width = 9
    ws.column_dimensions['CX'].width = 3
    ws.column_dimensions['CY'].width = 9
    ws.column_dimensions['CZ'].width = 3
    ws.column_dimensions['DA'].width = 9
    ws.column_dimensions['DB'].width = 9
    ws.column_dimensions['DC'].width = 3
    ws.column_dimensions['DD'].width = 9
    ws.column_dimensions['DE'].width = 10
    ws.column_dimensions['DF'].width = 6
    ws.column_dimensions['DG'].width = 6
    ws.column_dimensions['DH'].width = 10
    ws.column_dimensions['DI'].width = 10
    ws.column_dimensions['DJ'].width = 16
    ws.column_dimensions['DK'].width = 20
    ws.column_dimensions['DL'].width = 20
    ws.column_dimensions['DM'].width = 20
    ws.column_dimensions['DN'].width = 25
    ws.column_dimensions['DO'].width = 9
    ws.column_dimensions['DP'].width = 33
    
    # linea de division
    ws.freeze_panes = 'Q10'
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    # Configuración del fondo y el borde
    # Definir el color usando formato aRGB (opacidad completa 'FF' + color RGB)
    fill = PatternFill(start_color='FF60D7E0', end_color='FF60D7E0', fill_type='solid')
    # Definir el color anaranjado usando formato aRGB
    orange_fill = PatternFill(start_color='FFE0A960', end_color='FFE0A960', fill_type='solid')
    # Definir los estilos para gris
    gray_fill = PatternFill(start_color='FFD3D3D3', end_color='FFD3D3D3', fill_type='solid')
    # Definir el estilo de color verde
    green_fill = PatternFill(start_color='FF60E0B3', end_color='FF60E0B3', fill_type='solid')
    # Definir el estilo de color amarillo
    yellow_fill = PatternFill(start_color='FFE0DE60', end_color='FFE0DE60', fill_type='solid')
    # Definir el estilo de color azul
    blue_fill = PatternFill(start_color='FF60A2E0', end_color='FF60A2E0', fill_type='solid')
    # Definir el estilo de color verde 2
    green_fill_2 = PatternFill(start_color='FF60E07E', end_color='FF60E07E', fill_type='solid')   
    # Definir el estilo de relleno celeste
    celeste_fill = PatternFill(start_color='FF87CEEB', end_color='FF87CEEB', fill_type='solid')
    # Morado más claro
    morado_claro_fill = PatternFill(start_color='FFE9D8FF', end_color='FFE9D8FF', fill_type='solid')
    # Plomo más claro
    plomo_claro_fill = PatternFill(start_color='FFEDEDED', end_color='FFEDEDED', fill_type='solid')
    # Azul más claro
    azul_claro_fill = PatternFill(start_color='FFD8EFFA', end_color='FFD8EFFA', fill_type='solid')
    # Naranja más claro
    naranja_claro_fill = PatternFill(start_color='FFFFEBD8', end_color='FFFFEBD8', fill_type='solid')
    # Verde más claro
    verde_claro_fill = PatternFill(start_color='FFBDF7BD', end_color='FFBDF7BD', fill_type='solid')
    
    green_font = Font(name='Arial', size=8, color='00FF00')  # Verde
    red_font = Font(name='Arial', size=8, color='FF0000')    # Rojo
    
    border = Border(left=Side(style='thin', color='00B0F0'),
                    right=Side(style='thin', color='00B0F0'),
                    top=Side(style='thin', color='00B0F0'),
                    bottom=Side(style='thin', color='00B0F0'))
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    borde_plomo = Border(left=Side(style='thin', color='A9A9A9'), # Plomo
                    right=Side(style='thin', color='A9A9A9'), # Plomo
                    top=Side(style='thin', color='A9A9A9'), # Plomo
                    bottom=Side(style='thin', color='A9A9A9')) # Plomo
    
    border_negro = Border(left=Side(style='thin', color='000000'), # negro
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000')) 
    
    # Merge cells 
    # numerador y denominador
    ws.merge_cells('B5:P5') 
    ws.merge_cells('Q5:DG5')
    
    # BPN y Sin BPN
    ws.merge_cells('Q6:Z6') 
    ws.merge_cells('AA6:AW6')
    ws.merge_cells('AX6:BC6')
    ws.merge_cells('BD6:BJ6')
    ws.merge_cells('BK6:BQ6')
    ws.merge_cells('BR6:BV6')
    ws.merge_cells('BW6:BZ6')
    ws.merge_cells('CA6:CF6')
    ws.merge_cells('CG6:CM6')
    ws.merge_cells('CN6:CZ6')
    ws.merge_cells('DA6:DC6')
    ws.merge_cells('DD6:DG6')
    
    ws.merge_cells('B6:P6')
    ws.merge_cells('B7:C7')
    ws.merge_cells('B8:C8')
    
    # Auxiliar HORIZONTAL
    ws.merge_cells('Q7:Q8')
    ws.merge_cells('R7:R8')
    ws.merge_cells('AA7:AA8')
    ws.merge_cells('AX7:AX8')
    ws.merge_cells('AY7:AY8')
    ws.merge_cells('BD7:BD8')
    ws.merge_cells('BK7:BK8')
    ws.merge_cells('BR7:BR8')
    ws.merge_cells('BW7:BW8')
    ws.merge_cells('BX7:BX8')
    ws.merge_cells('CA7:CA8')
    ws.merge_cells('CB7:CB8')
    ws.merge_cells('CG7:CG8')
    ws.merge_cells('CN7:CN8')
    ws.merge_cells('DA7:DA8')
    ws.merge_cells('DD7:DD8')
        
    # intervalo
    ws.merge_cells('D7:H7')
    ws.merge_cells('I7:M7')
    ws.merge_cells('N7:P7')
    ws.merge_cells('S7:T7')
    ws.merge_cells('U7:V7')
    ws.merge_cells('W7:X7')
    ws.merge_cells('Y7:Z7')
    ws.merge_cells('AB7:AC7')
    ws.merge_cells('AD7:AE7')
    ws.merge_cells('AF7:AG7')
    ws.merge_cells('AH7:AI7')
    ws.merge_cells('AJ7:AK7')
    ws.merge_cells('AL7:AM7')
    ws.merge_cells('AN7:AO7')
    ws.merge_cells('AP7:AQ7')
    ws.merge_cells('AR7:AS7')
    ws.merge_cells('AT7:AU7')
    ws.merge_cells('AV7:AW7') 
    ws.merge_cells('AZ7:BA7')
    ws.merge_cells('BB7:BC7')
    ws.merge_cells('BE7:BF7')
    ws.merge_cells('BG7:BH7')
    ws.merge_cells('BI7:BJ7')
    ws.merge_cells('BL7:BM7')
    ws.merge_cells('BN7:BO7')
    ws.merge_cells('BP7:BQ7')    
    ws.merge_cells('BS7:BT7')
    ws.merge_cells('BU7:BV7')
    ws.merge_cells('BY7:BZ7')
    ws.merge_cells('CC7:CD7')
    ws.merge_cells('CE7:CF7')
    ws.merge_cells('CH7:CI7')
    ws.merge_cells('CJ7:CK7')
    ws.merge_cells('CL7:CM7')
    ws.merge_cells('CO7:CP7')
    ws.merge_cells('CQ7:CR7')
    ws.merge_cells('CS7:CT7')
    ws.merge_cells('CU7:CV7')
    ws.merge_cells('CW7:CX7')
    ws.merge_cells('CY7:CZ7')   
    ws.merge_cells('DB7:DC7')
    ws.merge_cells('DE7:DG7')
    
    
    # COD HIS
    ws.merge_cells('D8:P8')
    ws.merge_cells('S8:T8')
    ws.merge_cells('U8:V8')
    ws.merge_cells('W8:X8')
    ws.merge_cells('Y8:Z8')
    ws.merge_cells('AB8:AC8')
    ws.merge_cells('AD8:AE8')
    ws.merge_cells('AF8:AG8')
    ws.merge_cells('AH8:AI8')
    ws.merge_cells('AJ8:AK8')
    ws.merge_cells('AL8:AM8')
    ws.merge_cells('AN8:AO8')
    ws.merge_cells('AP8:AQ8')
    ws.merge_cells('AR8:AS8')
    ws.merge_cells('AT8:AU8')
    ws.merge_cells('AV8:AW8')
    
    ws.merge_cells('AZ8:BA8')
    ws.merge_cells('BB8:BC8')
    ws.merge_cells('BE8:BF8')
    ws.merge_cells('BG8:BH8')
    ws.merge_cells('BI8:BJ8')
    ws.merge_cells('BL8:BM8')
    ws.merge_cells('BN8:BO8')
    ws.merge_cells('BP8:BQ8')    
    ws.merge_cells('BS8:BT8')
    ws.merge_cells('BU8:BV8')
    ws.merge_cells('BY8:BZ8')
    ws.merge_cells('CC8:CD8')
    ws.merge_cells('CE8:CF8')
    ws.merge_cells('CH8:CI8')
    ws.merge_cells('CJ8:CK8')
    ws.merge_cells('CL8:CM8')
    ws.merge_cells('CO8:CP8')
    ws.merge_cells('CQ8:CR8')
    ws.merge_cells('CS8:CT8')
    ws.merge_cells('CU8:CV8')
    ws.merge_cells('CW8:CX8')
    ws.merge_cells('CY8:CZ8')   
    ws.merge_cells('DB8:DC8')
    ws.merge_cells('DE8:DG8')
    
    # Combina cela
    ws['B5'] = 'DENOMINADOR'
    ws['Q5'] = 'NUMERADOR'
    
    # CABECERA GRUPAL
    ws['Q6']  = 'CRED DEL RECIEN NACIDO'
    ws['AA6'] = 'CRED DEL MENOR DE 1 AÑO'
    ws['AX6'] = 'VACUNA NEUMOCOCO'
    ws['BD6'] = 'VACUNA ANTIPOLIO'
    ws['BK6'] = 'VACUNA PENTAVALENTE'
    ws['BR6'] = 'VACUNA ROTAVIRUS'
    ws['BW6'] = 'ESQUEMA DE 4 MESES'
    ws['CA6'] = 'ESQUEMA DE 6 A 11 MESES (SIN ANEMIA)'
    ws['CG6'] = 'SUPLEMENTACION DE TRATAMIENTO DE HIERRO'
    ws['CN6'] = 'SUPLEMENTACION DE TRATAMIENTO CON MULTIMICRONUTRIENTES'
    ws['DA6'] = 'DOSAJE HB'
    ws['DD6']  = 'EMISION DE DNI'
    
    # INTERVALO
    ws['Q7'] = 'NUMERADOR CRED + MENOR 1 AÑO'
    ws['R7'] = 'NUMERADOR PARCIAL'
    ws['AA7'] = 'NUMERADOR PARCIAL'
    ws['AX7'] = 'NUMERADOR VACUNA'
    ws['AY7'] = 'NUMERADOR PARCIAL'
    ws['BD7'] = 'NUMERADOR PARCIAL'
    ws['BK7'] = 'NUMERADOR PARCIAL'
    ws['BR7'] = 'NUMERADOR PARCIAL'
    ws['BW7'] = 'NUMERADOR SUPLE'
    ws['BX7'] = 'NUMERADOR 4 MESES'
    ws['CA7'] = 'NUMERADOR 6 MESES'
    ws['CB7'] = 'NUMERADOR PARCIAL'
    ws['CG7'] = 'NUMERADOR PARCIAL'
    ws['CN7'] = 'NUMERADOR PARCIAL'
    ws['DA7'] = 'NUMERADOR PARCIAL'
    ws['DD7'] = 'NUMERADOR PARCIAL'
    
    ws['D7'] = 'Niñas y niños menores de 12 meses de edad (364 días) del Padron Nominal, en el mes de medición, la determinación del corte de edad para cada periodo de medición, será el último día de cada mes'
    ws['I7'] = 'Se excluye a niños y niñas con bajo peso al nacer (menor de 2500 gramos) y/o prematuros (menor de 37 SG), registrados en CNV en línea'
    ws['N7'] = 'Consideran los distritos pertenecientes a los quintiles Q1 y Q2 de pobreza a nivel departamental'
    ws['S7'] = '1° control CRED se realiza a partir del 3er día de vida, hasta los 28 dias de edad'
    ws['U7'] = '2° CRED con intervalo mínimo de 3 dias a partir del 1° control, hasta los 28 dias de edad'
    ws['W7'] = '3° CRED con un intervalo minimo de 7 dias a partir 2° control, hasta los 28 dias de edad'
    ws['Y7'] = '4° CRED con un intervalo minimo de 7 dias a partir 3° control, hasta los 28 dias de edad'
    ws['AB7'] = '1° control es a partir de los 29 días de nacido (busqueda del dato entre 29 a 59 dias de edad)'
    ws['AD7'] = '2° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 60 a 89 días de edad)'
    ws['AF7'] = '3° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 90 a 119 días de edad)'
    ws['AH7'] = '4° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 120 a 149 días de edad)'
    ws['AJ7'] = '5° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 150 a 179 días de edad)'
    ws['AL7'] = '6° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 180 a 209 días de edad)'
    ws['AN7'] = '7° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 210 a 239 días de edad)'
    ws['AP7'] = '8° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 240 a 269 días de edad)'
    ws['AR7'] = '9° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 270 a 299 días de edad)'
    ws['AT7'] = '10° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 300 a 329 días de edad)'
    ws['AV7'] = '11° control debe realizarse con un intervalo mínimo de 28 días entre cada control (búsqueda del dato entre 330 a 364 días de edad)'
    ws['AZ7'] = '1° Dosis entre los 55 y 119 dias de edad'
    ws['BB7'] = '2° Dosis entre 28 y 70 dias despues de la 1° dosis'
    ws['BE7'] = '1° Dosis entre los 55 y 119 dias de edad'
    ws['BG7'] = '2° Dosis entre 28 y 70 dias despues de la 1° dosis, entre el rango de edad 120 a 147'
    ws['BI7'] = '3° Dosis entre 28 y 70 dias despues de la 2° dosis, entre rango de edad 148 a 217'
    ws['BL7'] = '1° Dosis entre los 55 y 119 dias de edad'
    ws['BN7'] = '2° Dosis entre 28 y 70 dias despues de la 1° dosis'
    ws['BP7'] = '3° Dosis entre 28 y 70 dias despues de la 2° dosis'   
    ws['BS7'] = '1° Dosis entre los 55 y 180 dias de edad'
    ws['BU7'] = '2° Dosis entre 28 dias despues de la 1° dosis, hasta los 240 dias de edad'
    ws['BY7'] = 'Busqueda de suplementación entre los 110 y 179 dias de edad'
    ws['CC7'] = 'Busqueda entre los 170 y 239 dias de edad'
    ws['CE7'] = 'Busqueda entre los 240 y 299 dias de edad'
    ws['CH7'] = 'Busqueda entre los 170 y 239 dias de edad'
    ws['CJ7'] = 'Busqueda entre los 240 y 299 dias de edad'
    ws['CL7'] = 'Busqueda entre los 300 y 364 dias de edad'
    ws['CO7'] = 'Esquema 6 meses entre los 170 y 209 dias de edad'
    ws['CQ7'] = 'Esquema 6 meses entre los 210 y 239 dias de edad'
    ws['CS7'] = 'Esquema 6 meses entre los 240 y 269 dias de edad'
    ws['CU7'] = 'Esquema 6 meses entre los 270 y 299 dias de edad'
    ws['CW7'] = 'Esquema 6 meses entre los 300 y 329 dias de edad'
    ws['CY7'] = 'Esquema 6 meses entre los 330 y 364 dias de edad'  
    ws['DB7'] = 'Dosaje entre los 170 a 209 dias'
    ws['DE7'] = 'La fecha de emision debe ser menor o igual al ultimo dia de evaluación,  DNI es igual o menor a 60 días de edad'
    
    # CODIGO HIS
    
    ws['S8'] = 'DX = 99381.01 ó Z001'
    ws['U8'] = 'DX = 99381.01 ó Z001'
    ws['W8'] = 'DX = 99381.01 ó Z001'
    ws['Y8'] = 'DX = 99381.01 ó Z001'
    ws['Y8'] = 'DX = 99381.01 ó Z001'
    ws['AB8'] = 'DX = 99381 ó Z001'
    ws['AD8'] = 'DX = 99381 ó Z001'
    ws['AF8'] = 'DX = 99381 ó Z001'
    ws['AH8'] = 'DX = 99381 ó Z001'
    ws['AJ8'] = 'DX = 99381 ó Z001'
    ws['AL8'] = 'DX = 99381 ó Z001'
    ws['AN8'] = 'DX = 99381 ó Z001'
    ws['AP8'] = 'DX = 99381 ó Z001'
    ws['AR8'] = 'DX = 99381 ó Z001'
    ws['AT8'] = 'DX = 99381 ó Z001'
    ws['AV8'] = 'DX = 99381 ó Z001'
    ws['AZ8'] = 'DX = 90670'
    ws['BB8'] = 'DX = 90670'
    ws['BE8'] = 'DX = 90712 ó 90713'
    ws['BG8'] = 'DX = 90712 ó 90713'
    ws['BI8'] = 'DX = 90712 ó 90713'
    ws['BL8'] = 'DX =  90723 ó 90722'
    ws['BN8'] = 'DX =  90723 ó 90722'
    ws['BP8'] = 'DX =  90723 ó 90722'   
    ws['BS8'] = 'DX = 90681'
    ws['BU8'] = 'DX = 90681'
    ws['BY8'] = 'DX = 99199.17'
    ws['CC8'] = 'DX = 99199.17'
    ws['CE8'] = 'DX = 99199.17'
    ws['CH8'] = 'DX = 99199.17 ó 99199.11 + (D500 ó D508 ó D509 ó D649 ó D539 + TD = D ó R)'
    ws['CJ8'] = 'DX = 99199.17 ó 99199.11 + (D500 ó D508 ó D509 ó D649 ó D539 + TD = D ó R)'
    ws['CL8'] = 'DX = 99199.17 ó 99199.11 + (D500 ó D508 ó D509 ó D649 ó D539 + TD = D ó R)'
    ws['CO8'] = 'DX = 99199.19'
    ws['CQ8'] = 'DX = 99199.19'
    ws['CS8'] = 'DX = 99199.19'
    ws['CU8'] = 'DX = 99199.19'
    ws['CW8'] = 'DX = 99199.19'
    ws['CY8'] = 'DX = 99199.19'  
    ws['DB8'] = 'DX = 85018 ó 85018.01'
    
    ### numerador y denominador     
    ws['B5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['B5'].fill = gray_fill
    ws['B5'].border = border_negro
    
    ws['Q5'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q5'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['Q5'].fill = naranja_claro_fill
    ws['Q5'].border = border_negro
    
    ws['B6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['B6'].font = Font(name = 'Arial', size= 10, bold = True)
    ws['B6'].fill = gray_fill
    ws['B6'].border = border_negro
    
    
    ### intervalo 
    ws['Q6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q6'].font = Font(name = 'Arial', size= 7)
    ws['Q6'].fill = morado_claro_fill
    ws['Q6'].border = border_negro
    
    ws['AA6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA6'].font = Font(name = 'Arial', size= 7)
    ws['AA6'].fill = morado_claro_fill
    ws['AA6'].border = border_negro

    ws['AX6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AX6'].font = Font(name = 'Arial', size= 7)
    ws['AX6'].fill = azul_claro_fill
    ws['AX6'].border = border_negro
    
    ws['BD6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BD6'].font = Font(name = 'Arial', size= 7)
    ws['BD6'].fill = azul_claro_fill
    ws['BD6'].border = border_negro
    
    ws['BK6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BK6'].font = Font(name = 'Arial', size= 7)
    ws['BK6'].fill = azul_claro_fill
    ws['BK6'].border = border_negro
    
    ws['BR6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BR6'].font = Font(name = 'Arial', size= 7)
    ws['BR6'].fill = azul_claro_fill
    ws['BR6'].border = border_negro
    
    ws['BW6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BW6'].font = Font(name = 'Arial', size= 7)
    ws['BW6'].fill = verde_claro_fill
    ws['BW6'].border = border_negro
    
    ws['CA6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CA6'].font = Font(name = 'Arial', size= 7)
    ws['CA6'].fill = verde_claro_fill
    ws['CA6'].border = border_negro
    
    ws['CG6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CG6'].font = Font(name = 'Arial', size= 7)
    ws['CG6'].fill = verde_claro_fill
    ws['CG6'].border = border_negro
    
    ws['CN6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CN6'].font = Font(name = 'Arial', size= 7)
    ws['CN6'].fill = verde_claro_fill
    ws['CN6'].border = border_negro
    
    ws['DA6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DA6'].font = Font(name = 'Arial', size= 7)
    ws['DA6'].fill = verde_claro_fill
    ws['DA6'].border = border_negro
    
    ws['DD6'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DD6'].font = Font(name = 'Arial', size= 7)
    ws['DD6'].fill = azul_claro_fill
    ws['DD6'].border = border_negro
    
    
    #intervalos 
    ws['Q7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q7'].font = Font(name = 'Arial', size= 7)
    ws['Q7'].fill = naranja_claro_fill
    ws['Q7'].border = border_negro
    
    ws['R7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R7'].font = Font(name = 'Arial', size= 7)
    ws['R7'].fill = plomo_claro_fill
    ws['R7'].border = border_negro
    
    ws['AA7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA7'].font = Font(name = 'Arial', size= 7)
    ws['AA7'].fill = plomo_claro_fill
    ws['AA7'].border = border_negro
    
    ws['AX7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AX7'].font = Font(name = 'Arial', size= 7)
    ws['AX7'].fill = naranja_claro_fill
    ws['AX7'].border = border_negro
    
    ws['AY7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AY7'].font = Font(name = 'Arial', size= 7)
    ws['AY7'].fill = plomo_claro_fill
    ws['AY7'].border = border_negro
    
    ws['BD7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BD7'].font = Font(name = 'Arial', size= 7)
    ws['BD7'].fill = plomo_claro_fill
    ws['BD7'].border = border_negro
    
    ws['BK7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BK7'].font = Font(name = 'Arial', size= 7)
    ws['BK7'].fill = plomo_claro_fill
    ws['BK7'].border = border_negro
    
    ws['BR7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BR7'].font = Font(name = 'Arial', size= 7)
    ws['BR7'].fill = plomo_claro_fill
    ws['BR7'].border = border_negro
    
    ws['BW7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BW7'].font = Font(name = 'Arial', size= 7)
    ws['BW7'].fill = naranja_claro_fill
    ws['BW7'].border = border_negro
    
    ws['BX7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BX7'].font = Font(name = 'Arial', size= 7)
    ws['BX7'].fill = morado_claro_fill
    ws['BX7'].border = border_negro
    
    ws['CA7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CA7'].font = Font(name = 'Arial', size= 7)
    ws['CA7'].fill = morado_claro_fill
    ws['CA7'].border = border_negro
    
    ws['CB7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CB7'].font = Font(name = 'Arial', size= 7)
    ws['CB7'].fill = plomo_claro_fill
    ws['CB7'].border = border_negro
    
    ws['CG7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CG7'].font = Font(name = 'Arial', size= 7)
    ws['CG7'].fill = plomo_claro_fill
    ws['CG7'].border = border_negro
    
    ws['CN7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CN7'].font = Font(name = 'Arial', size= 7)
    ws['CN7'].fill = plomo_claro_fill
    ws['CN7'].border = border_negro
    
    ws['DA7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DA7'].font = Font(name = 'Arial', size= 7)
    ws['DA7'].fill = naranja_claro_fill
    ws['DA7'].border = border_negro
    
    ws['DD7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DD7'].font = Font(name = 'Arial', size= 7)
    ws['DD7'].fill = naranja_claro_fill
    ws['DD7'].border = border_negro
    
    ws['D7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D7'].font = Font(name = 'Arial', size= 7)
    ws['D7'].fill = plomo_claro_fill
    ws['D7'].border = border_negro
    
    ws['I7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I7'].font = Font(name = 'Arial', size= 7)
    ws['I7'].fill = plomo_claro_fill
    ws['I7'].border = border_negro
    
    ws['N7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N7'].font = Font(name = 'Arial', size= 7)
    ws['N7'].fill = plomo_claro_fill
    ws['N7'].border = border_negro
    
    ws['S7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S7'].font = Font(name = 'Arial', size= 7)
    ws['S7'].fill = plomo_claro_fill
    ws['S7'].border = border_negro
    
    ws['U7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U7'].font = Font(name = 'Arial', size= 7)
    ws['U7'].fill = plomo_claro_fill
    ws['U7'].border = border_negro
    
    ws['W7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W7'].font = Font(name = 'Arial', size= 7)
    ws['W7'].fill = plomo_claro_fill
    ws['W7'].border = border_negro
    
    ws['Y7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y7'].font = Font(name = 'Arial', size= 7)
    ws['Y7'].fill = plomo_claro_fill
    ws['Y7'].border = border_negro
    
    ws['AB7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB7'].font = Font(name = 'Arial', size= 7)
    ws['AB7'].fill = plomo_claro_fill
    ws['AB7'].border = border_negro
    
    ws['AD7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD7'].font = Font(name = 'Arial', size= 7)
    ws['AD7'].fill = plomo_claro_fill
    ws['AD7'].border = border_negro
    
    ws['AF7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF7'].font = Font(name = 'Arial', size= 7)
    ws['AF7'].fill = plomo_claro_fill
    ws['AF7'].border = border_negro
    
    ws['AH7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH7'].font = Font(name = 'Arial', size= 7)
    ws['AH7'].fill = plomo_claro_fill
    ws['AH7'].border = border_negro
    
    ws['AJ7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ7'].font = Font(name = 'Arial', size= 7)
    ws['AJ7'].fill = plomo_claro_fill
    ws['AJ7'].border = border_negro
    
    ws['AL7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL7'].font = Font(name = 'Arial', size= 7)
    ws['AL7'].fill = plomo_claro_fill
    ws['AL7'].border = border_negro
    
    ws['AN7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN7'].font = Font(name = 'Arial', size= 7)
    ws['AN7'].fill = plomo_claro_fill
    ws['AN7'].border = border_negro
    
    ws['AP7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP7'].font = Font(name = 'Arial', size= 7)
    ws['AP7'].fill = plomo_claro_fill
    ws['AP7'].border = border_negro
    
    ws['AR7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AR7'].font = Font(name = 'Arial', size= 7)
    ws['AR7'].fill = plomo_claro_fill
    ws['AR7'].border = border_negro
    
    ws['AT7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AT7'].font = Font(name = 'Arial', size= 7)
    ws['AT7'].fill = plomo_claro_fill
    ws['AT7'].border = border_negro
    
    ws['AV7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AV7'].font = Font(name = 'Arial', size= 7)
    ws['AV7'].fill = plomo_claro_fill
    ws['AV7'].border = border_negro
    
    ws['AZ7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AZ7'].font = Font(name = 'Arial', size= 7)
    ws['AZ7'].fill = plomo_claro_fill
    ws['AZ7'].border = border_negro
    
    ws['BB7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BB7'].font = Font(name = 'Arial', size= 7)
    ws['BB7'].fill = plomo_claro_fill
    ws['BB7'].border = border_negro
    
    ws['BE7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BE7'].font = Font(name = 'Arial', size= 7)
    ws['BE7'].fill = plomo_claro_fill
    ws['BE7'].border = border_negro
    
    ws['BG7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BG7'].font = Font(name = 'Arial', size= 7)
    ws['BG7'].fill = plomo_claro_fill
    ws['BG7'].border = border_negro
    
    ws['BI7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BI7'].font = Font(name = 'Arial', size= 7)
    ws['BI7'].fill = plomo_claro_fill
    ws['BI7'].border = border_negro
    
    ws['BL7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BL7'].font = Font(name = 'Arial', size= 7)
    ws['BL7'].fill = plomo_claro_fill
    ws['BL7'].border = border_negro
    
    ws['BN7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BN7'].font = Font(name = 'Arial', size= 7)
    ws['BN7'].fill = plomo_claro_fill
    ws['BN7'].border = border_negro
    
    ws['BP7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BP7'].font = Font(name = 'Arial', size= 7)
    ws['BP7'].fill = plomo_claro_fill
    ws['BP7'].border = border_negro
    
    ws['BS7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BS7'].font = Font(name = 'Arial', size= 7)
    ws['BS7'].fill = plomo_claro_fill
    ws['BS7'].border = border_negro
    
    ws['BU7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BU7'].font = Font(name = 'Arial', size= 7)
    ws['BU7'].fill = plomo_claro_fill
    ws['BU7'].border = border_negro
    
    ws['BY7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BY7'].font = Font(name = 'Arial', size= 7)
    ws['BY7'].fill = plomo_claro_fill
    ws['BY7'].border = border_negro
    
    ws['CC7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CC7'].font = Font(name = 'Arial', size= 7)
    ws['CC7'].fill = plomo_claro_fill
    ws['CC7'].border = border_negro
    
    ws['CE7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CE7'].font = Font(name = 'Arial', size= 7)
    ws['CE7'].fill = plomo_claro_fill
    ws['CE7'].border = border_negro
    
    ws['CH7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CH7'].font = Font(name = 'Arial', size= 7)
    ws['CH7'].fill = plomo_claro_fill
    ws['CH7'].border = border_negro
    
    ws['CJ7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CJ7'].font = Font(name = 'Arial', size= 7)
    ws['CJ7'].fill = plomo_claro_fill
    ws['CJ7'].border = border_negro
    
    ws['CL7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CL7'].font = Font(name = 'Arial', size= 7)
    ws['CL7'].fill = plomo_claro_fill
    ws['CL7'].border = border_negro
    
    ws['CO7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CO7'].font = Font(name = 'Arial', size= 7)
    ws['CO7'].fill = plomo_claro_fill
    ws['CO7'].border = border_negro
    
    ws['CQ7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CQ7'].font = Font(name = 'Arial', size= 7)
    ws['CQ7'].fill = plomo_claro_fill
    ws['CQ7'].border = border_negro
    
    ws['CS7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CS7'].font = Font(name = 'Arial', size= 7)
    ws['CS7'].fill = plomo_claro_fill
    ws['CS7'].border = border_negro

    ws['CU7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CU7'].font = Font(name = 'Arial', size= 7)
    ws['CU7'].fill = plomo_claro_fill
    ws['CU7'].border = border_negro

    ws['CW7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CW7'].font = Font(name = 'Arial', size= 7)
    ws['CW7'].fill = plomo_claro_fill
    ws['CW7'].border = border_negro
    
    ws['CY7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CY7'].font = Font(name = 'Arial', size= 7)
    ws['CY7'].fill = plomo_claro_fill
    ws['CY7'].border = border_negro
    
    ws['DB7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DB7'].font = Font(name = 'Arial', size= 7)
    ws['DB7'].fill = plomo_claro_fill
    ws['DB7'].border = border_negro
    
    ws['DE7'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DE7'].font = Font(name = 'Arial', size= 7)
    ws['DE7'].fill = plomo_claro_fill
    ws['DE7'].border = border_negro

    # CODIGO HIS
    
    ws['D8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D8'].font = Font(name = 'Arial', size= 7)
    ws['D8'].fill = azul_claro_fill
    ws['D8'].border = border_negro
    
    ws['S8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S8'].font = Font(name = 'Arial', size= 7)
    ws['S8'].fill = azul_claro_fill
    ws['S8'].border = border_negro
    
    ws['U8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U8'].font = Font(name = 'Arial', size= 7)
    ws['U8'].fill = azul_claro_fill
    ws['U8'].border = border_negro
    
    ws['W8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W8'].font = Font(name = 'Arial', size= 7)
    ws['W8'].fill = azul_claro_fill
    ws['W8'].border = border_negro
    
    ws['Y8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y8'].font = Font(name = 'Arial', size= 7)
    ws['Y8'].fill = azul_claro_fill
    ws['Y8'].border = border_negro
    
    ws['Y8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y8'].font = Font(name = 'Arial', size= 7)
    ws['Y8'].fill = azul_claro_fill
    ws['Y8'].border = border_negro
    
    ws['AB8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB8'].font = Font(name = 'Arial', size= 7)
    ws['AB8'].fill = azul_claro_fill
    ws['AB8'].border = border_negro
    
    ws['AD8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD8'].font = Font(name = 'Arial', size= 7)
    ws['AD8'].fill = azul_claro_fill
    ws['AD8'].border = border_negro
    
    ws['AF8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF8'].font = Font(name = 'Arial', size= 7)
    ws['AF8'].fill = azul_claro_fill
    ws['AF8'].border = border_negro
    
    ws['AH8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH8'].font = Font(name = 'Arial', size= 7)
    ws['AH8'].fill = azul_claro_fill
    ws['AH8'].border = border_negro
    
    ws['AJ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ8'].font = Font(name = 'Arial', size= 7)
    ws['AJ8'].fill = azul_claro_fill
    ws['AJ8'].border = border_negro
    
    ws['AL8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL8'].font = Font(name = 'Arial', size= 7)
    ws['AL8'].fill = azul_claro_fill
    ws['AL8'].border = border_negro
    
    ws['AN8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN8'].font = Font(name = 'Arial', size= 7)
    ws['AN8'].fill = azul_claro_fill
    ws['AN8'].border = border_negro
    
    ws['AP8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP8'].font = Font(name = 'Arial', size= 7)
    ws['AP8'].fill = azul_claro_fill
    ws['AP8'].border = border_negro
    
    ws['AR8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AR8'].font = Font(name = 'Arial', size= 7)
    ws['AR8'].fill = azul_claro_fill
    ws['AR8'].border = border_negro
    
    ws['AT8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AT8'].font = Font(name = 'Arial', size= 7)
    ws['AT8'].fill = azul_claro_fill
    ws['AT8'].border = border_negro
    
    ws['AV8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AV8'].font = Font(name = 'Arial', size= 7)
    ws['AV8'].fill = azul_claro_fill
    ws['AV8'].border = border_negro
    
    ws['AZ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AZ8'].font = Font(name = 'Arial', size= 7)
    ws['AZ8'].fill = azul_claro_fill
    ws['AZ8'].border = border_negro
    
    ws['BB8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BB8'].font = Font(name = 'Arial', size= 7)
    ws['BB8'].fill = azul_claro_fill
    ws['BB8'].border = border_negro
    
    ws['BE8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BE8'].font = Font(name = 'Arial', size= 7)
    ws['BE8'].fill = azul_claro_fill
    ws['BE8'].border = border_negro
    
    ws['BG8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BG8'].font = Font(name = 'Arial', size= 7)
    ws['BG8'].fill = azul_claro_fill
    ws['BG8'].border = border_negro
    
    ws['BI8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BI8'].font = Font(name = 'Arial', size= 7)
    ws['BI8'].fill = azul_claro_fill
    ws['BI8'].border = border_negro
    
    ws['BL8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BL8'].font = Font(name = 'Arial', size= 7)
    ws['BL8'].fill = azul_claro_fill
    ws['BL8'].border = border_negro
    
    ws['BN8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BN8'].font = Font(name = 'Arial', size= 7)
    ws['BN8'].fill = azul_claro_fill
    ws['BN8'].border = border_negro
    
    ws['BP8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BP8'].font = Font(name = 'Arial', size= 7)
    ws['BP8'].fill = azul_claro_fill
    ws['BP8'].border = border_negro
    
    ws['BS8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BS8'].font = Font(name = 'Arial', size= 7)
    ws['BS8'].fill = azul_claro_fill
    ws['BS8'].border = border_negro
    
    ws['BU8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BU8'].font = Font(name = 'Arial', size= 7)
    ws['BU8'].fill = azul_claro_fill
    ws['BU8'].border = border_negro
    
    ws['BY8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BY8'].font = Font(name = 'Arial', size= 7)
    ws['BY8'].fill = azul_claro_fill
    ws['BY8'].border = border_negro
    
    ws['CC8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CC8'].font = Font(name = 'Arial', size= 7)
    ws['CC8'].fill = azul_claro_fill
    ws['CC8'].border = border_negro
    
    ws['CE8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CE8'].font = Font(name = 'Arial', size= 7)
    ws['CE8'].fill = azul_claro_fill
    ws['CE8'].border = border_negro
    
    ws['CH8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CH8'].font = Font(name = 'Arial', size= 7)
    ws['CH8'].fill = azul_claro_fill
    ws['CH8'].border = border_negro
    
    ws['CJ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CJ8'].font = Font(name = 'Arial', size= 7)
    ws['CJ8'].fill = azul_claro_fill
    ws['CJ8'].border = border_negro
    
    ws['CL8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CL8'].font = Font(name = 'Arial', size= 7)
    ws['CL8'].fill = azul_claro_fill
    ws['CL8'].border = border_negro
    
    ws['CO8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CO8'].font = Font(name = 'Arial', size= 7)
    ws['CO8'].fill = azul_claro_fill
    ws['CO8'].border = border_negro
    
    ws['CQ8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CQ8'].font = Font(name = 'Arial', size= 7)
    ws['CQ8'].fill = azul_claro_fill
    ws['CQ8'].border = border_negro
    
    ws['CS8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CS8'].font = Font(name = 'Arial', size= 7)
    ws['CS8'].fill = azul_claro_fill
    ws['CS8'].border = border_negro
    
    ws['CU8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CU8'].font = Font(name = 'Arial', size= 7)
    ws['CU8'].fill = azul_claro_fill
    ws['CU8'].border = border_negro
    
    ws['CW8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CW8'].font = Font(name = 'Arial', size= 7)
    ws['CW8'].fill = azul_claro_fill
    ws['CW8'].border = border_negro
    
    ws['CY8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CY8'].font = Font(name = 'Arial', size= 7)
    ws['CY8'].fill = azul_claro_fill
    ws['CY8'].border = border_negro
    
    ws['DB8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DB8'].font = Font(name = 'Arial', size= 7)
    ws['DB8'].fill = azul_claro_fill
    ws['DB8'].border = border_negro
    
    ws['DE8'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DE8'].font = Font(name = 'Arial', size= 7)
    ws['DE8'].fill = azul_claro_fill
    ws['DE8'].border = border_negro
    
    
    ws['B7'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B7'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B7'].fill = plomo_claro_fill
    ws['B7'].border = border_negro
    ws['B7'] = 'INTERVALO'
    
    ws['B8'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B8'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B8'].fill = azul_claro_fill
    ws['B8'].border = border_negro
    ws['B8'] = 'COD HIS'
    
    ### BORDE DE CELDAS CONBINADAS
    
    # NUM y DEN
    inicio_columna = 'B'
    fin_columna = 'DG'
    fila = 5
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    # NUM y DEN
    inicio_columna = 'B'
    fin_columna = 'DG'
    fila = 6
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
        
    # INTERVALO
    inicio_columna = 'B'
    fin_columna = 'DG'
    fila = 7
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
        
    # CODIGO HIS 
    inicio_columna = 'B'
    fin_columna = 'DG'
    fila = 8
    from openpyxl.utils import column_index_from_string
    # Convertir letras de columna a índices numéricos
    indice_inicio = column_index_from_string(inicio_columna)
    indice_fin = column_index_from_string(fin_columna)
    # Iterar sobre las columnas en el rango especificado
    for col in range(indice_inicio, indice_fin + 1):
        celda = ws.cell(row=fila, column=col)
        celda.border = border_negro
    
    ##### imprimer fecha y hora del reporte
    fecha_hora_actual = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    nombre_usuario = getpass.getuser()

    # Obtener el usuario actualmente autenticado
    try:
        user = User.objects.get(is_active=True)
    except User.DoesNotExist:
        user = None
    except User.MultipleObjectsReturned:
        # Manejar el caso donde hay múltiples usuarios activos
        user = User.objects.filter(is_active=True).first()  # Por ejemplo, obtener el primero
    
    # Asignar fecha y hora a la celda A1
    ws['V1'].value = 'Fecha y Hora:'
    ws['W1'].value = fecha_hora_actual

    # Asignar nombre de usuario a la celda A2
    ws['V2'].value = 'Usuario:'
    ws['W2'].value = nombre_usuario
    
    # Formatear las etiquetas en negrita
    etiqueta_font = Font(name='Arial', size=8)
    ws['V1'].font = etiqueta_font
    ws['W1'].font = etiqueta_font
    ws['V2'].font = etiqueta_font
    ws['W2'].font = etiqueta_font

    # Alinear el texto
    ws['V1'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W1'].alignment = Alignment(horizontal="left", vertical="center")
    ws['V2'].alignment = Alignment(horizontal="right", vertical="center")
    ws['W2'].alignment = Alignment(horizontal="left", vertical="center")
    
    ## crea titulo del reporte
    ws['B1'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B1'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B1'] = 'OFICINA DE TECNOLOGIAS DE LA INFORMACION'
    
    ws['B2'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B2'].font = Font(name = 'Arial', size= 7, bold = True)
    ws['B2'] = 'DIRECCION REGIONAL DE SALUD JUNIN'
    
    ws['B4'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B4'].font = Font(name = 'Arial', size= 12, bold = True)
    ws['B4'] = 'SEGUIMIENTO NOMINAL DEL INDICADOR MC-02. NIÑAS Y NIÑOS MENORES DE 12 MESES DE EDAD PROCEDENTES DE LOS QUINTILES 1 Y 2 DE POBREZA DEPARTAMENTAL QUE RECIBEN EL PAQUETE INTEGRADO DE SERVICIOS'
    
    ws['B3'].alignment = Alignment(horizontal= "left", vertical="center")
    ws['B3'].font = Font(name = 'Arial', size= 7, bold = True, color='0000CC')
    ws['B3'] ='El usuario se compromete a mantener la confidencialidad de los datos personales que conozca como resultado del reporte realizado, cumpliendo con lo establecido en la Ley N° 29733 - Ley de Protección de Datos Personales y sus normas complementarias.'
        
    ws['B9'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['B9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['B9'].fill = fill
    ws['B9'].border = border
    ws['B9'] = 'TD'
    
    ws['C9'].alignment = Alignment(horizontal= "center", vertical="center")
    ws['C9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['C9'].fill = fill
    ws['C9'].border = border
    ws['C9'] = 'NUM DOC'
    
    ws['D9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['D9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['D9'].fill = fill
    ws['D9'].border = border
    ws['D9'] = 'FECHA NAC'      
    
    ws['E9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['E9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['E9'].fill = fill
    ws['E9'].border = border
    ws['E9'] = 'SEXO' 
    
    ws['F9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['F9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['F9'].fill = fill
    ws['F9'].border = border
    ws['F9'] = 'SEGURO'     
    
    ws['G9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['G9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['G9'].fill = fill
    ws['G9'].border = border
    ws['G9'] = 'ED DIAS'    
    
    ws['H9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['H9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['H9'].fill = fill
    ws['H9'].border = border
    ws['H9'] = 'ED MES'    
    
    ws['I9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['I9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['I9'].fill = fill
    ws['I9'].border = border
    ws['I9'] = 'CNV'    
    
    ws['J9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['J9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['J9'].fill = fill
    ws['J9'].border = border
    ws['J9'] = 'PESO'  
    
    ws['K9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['K9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['K9'].fill = fill
    ws['K9'].border = border
    ws['K9'] = 'BPN'  
    
    ws['L9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['L9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['L9'].fill = fill
    ws['L9'].border = border
    ws['L9'] = 'SEM GEST'  
    
    ws['M9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['M9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['M9'].fill = fill
    ws['M9'].border = border
    ws['M9'] = 'PREM'  
    
    ws['N9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['N9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['N9'].fill = fill
    ws['N9'].border = border
    ws['N9'] = 'BPN PREM'  
    
    ws['O9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['O9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['O9'].fill = fill
    ws['O9'].border = border
    ws['O9'] = 'DENOM'  
    
    ws['P9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['P9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['P9'].fill = fill
    ws['P9'].border = border
    ws['P9'] = 'SIN DNI'  
    
    ws['Q9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Q9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Q9'].fill = green_fill_2
    ws['Q9'].border = border
    ws['Q9'] = 'CRED'    
    
    ws['R9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['R9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['R9'].fill = green_fill_2
    ws['R9'].border = border
    ws['R9'] = 'CRED RN' 
    
    ws['S9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['S9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['S9'].fill = green_fill_2
    ws['S9'].border = border
    ws['S9'] = '1° CRED RN' 
    
    ws['T9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['T9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['T9'].fill = green_fill_2
    ws['T9'].border = border
    ws['T9'] = 'V' 
    
    ws['U9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['U9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['U9'].fill = green_fill_2
    ws['U9'].border = border
    ws['U9'] = '2° CRED RN' 
    
    ws['V9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['V9'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['V9'].fill = green_fill_2
    ws['V9'].border = border
    ws['V9'] = 'V' 
    
    ws['W9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['W9'].font = Font(name = 'Arial', size= 7, bold = True, color='000000')
    ws['W9'].fill = green_fill_2
    ws['W9'].border = border
    ws['W9'] = '3° CRED RN'   
    
    ws['X9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['X9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['X9'].fill = green_fill_2
    ws['X9'].border = border
    ws['X9'] = 'V' 
    
    ws['Y9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Y9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Y9'].fill = green_fill_2
    ws['Y9'].border = border
    ws['Y9'] = '4° CRED RN' 
    
    ws['Z9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['Z9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['Z9'].fill = green_fill_2
    ws['Z9'].border = border
    ws['Z9'] = 'V' 
    
    ws['AA9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AA9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AA9'].fill = green_fill
    ws['AA9'].border = border
    ws['AA9'] = 'CRED MES' 
    
    ws['AB9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AB9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AB9'].fill = green_fill
    ws['AB9'].border = border
    ws['AB9'] = '1° CRED'     
    
    ws['AC9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AC9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AC9'].fill = green_fill
    ws['AC9'].border = border
    ws['AC9'] = 'V' 
    
    ws['AD9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AD9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AD9'].fill = green_fill
    ws['AD9'].border = border
    ws['AD9'] = '2° CRED' 
    
    ws['AE9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AE9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AE9'].fill = green_fill
    ws['AE9'].border = border
    ws['AE9'] = 'V' 
    
    ws['AF9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AF9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AF9'].fill = green_fill
    ws['AF9'].border = border
    ws['AF9'] = '3° CRED' 
    
    ws['AG9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AG9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AG9'].fill = green_fill
    ws['AG9'].border = border
    ws['AG9'] = 'V' 
    
    ws['AH9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AH9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AH9'].fill = green_fill
    ws['AH9'].border = border
    ws['AH9'] = '4° CRED' 
    
    ws['AI9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AI9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AI9'].fill = green_fill
    ws['AI9'].border = border
    ws['AI9'] = 'V' 
    
    ws['AJ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AJ9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AJ9'].fill = green_fill
    ws['AJ9'].border = border
    ws['AJ9'] = '5° CRED' 
    
    ws['AK9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AK9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AK9'].fill = green_fill
    ws['AK9'].border = border
    ws['AK9'] = 'V' 
    
    ws['AL9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AL9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AL9'].fill = green_fill
    ws['AL9'].border = border
    ws['AL9'] = '6° CRED' 
    
    ws['AM9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AM9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AM9'].fill = green_fill
    ws['AM9'].border = border
    ws['AM9'] = 'V' 
    
    ws['AN9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AN9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AN9'].fill = green_fill
    ws['AN9'].border = border
    ws['AN9'] = '7° CRED' 
    
    ws['AO9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AO9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AO9'].fill = green_fill
    ws['AO9'].border = border
    ws['AO9'] = 'V' 
    
    ws['AP9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AP9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AP9'].fill = green_fill
    ws['AP9'].border = border
    ws['AP9'] = '8° CRED' 
    
    ws['AQ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AQ9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AQ9'].fill = green_fill
    ws['AQ9'].border = border
    ws['AQ9'] = 'V' 
    
    ws['AR9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AR9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AR9'].fill = green_fill
    ws['AR9'].border = border
    ws['AR9'] = '9° CRED' 
    
    ws['AS9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AS9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AS9'].fill = green_fill
    ws['AS9'].border = border
    ws['AS9'] = 'V' 
    
    ws['AT9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AT9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AT9'].fill = green_fill
    ws['AT9'].border = border
    ws['AT9'] = '10° CRED' 
    
    ws['AU9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AU9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AU9'].fill = green_fill
    ws['AU9'].border = border
    ws['AU9'] = 'V' 
    
    ws['AV9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AV9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AV9'].fill = green_fill
    ws['AV9'].border = border
    ws['AV9'] = '11° CRED' 
    
    ws['AW9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AW9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AW9'].fill = green_fill
    ws['AW9'].border = border
    ws['AW9'] = 'V' 
    
    ws['AX9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AX9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AX9'].fill = yellow_fill
    ws['AX9'].border = border
    ws['AX9'] = 'NUM VAC' 
    
    ws['AY9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AY9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AY9'].fill = yellow_fill
    ws['AY9'].border = border
    ws['AY9'] = 'NUM NEUMO' 
    
    ws['AZ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['AZ9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['AZ9'].fill = yellow_fill
    ws['AZ9'].border = border
    ws['AZ9'] = '1° NEUMO' 
    
    ws['BA9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BA9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BA9'].fill = yellow_fill
    ws['BA9'].border = border
    ws['BA9'] = 'V' 
    
    ws['BB9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BB9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BB9'].fill = yellow_fill
    ws['BB9'].border = border
    ws['BB9'] = '2° NEUMO' 
    
    ws['BC9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BC9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BC9'].fill = yellow_fill
    ws['BC9'].border = border
    ws['BC9'] = 'V'     
    
    ws['BD9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BD9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BD9'].fill = yellow_fill
    ws['BD9'].border = border
    ws['BD9'] = 'NUM POLIO' 
    
    ws['BE9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BE9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BE9'].fill = yellow_fill
    ws['BE9'].border = border
    ws['BE9'] = '1° POLIO' 
    
    ws['BF9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BF9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BF9'].fill = yellow_fill
    ws['BF9'].border = border
    ws['BF9'] = 'V' 
    
    ws['BG9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BG9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BG9'].fill = yellow_fill
    ws['BG9'].border = border
    ws['BG9'] = '2° POLIO' 
    
    ws['BH9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BH9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BH9'].fill = yellow_fill
    ws['BH9'].border = border
    ws['BH9'] = 'V' 
    
    ws['BI9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BI9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BI9'].fill = yellow_fill
    ws['BI9'].border = border
    ws['BI9'] = '3° POLIO' 
    
    ws['BJ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BJ9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BJ9'].fill = yellow_fill
    ws['BJ9'].border = border
    ws['BJ9'] = 'V' 
    
    ws['BK9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BK9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BK9'].fill = yellow_fill
    ws['BK9'].border = border
    ws['BK9'] = 'NUM PENTA' 
    
    ws['BL9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BL9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BL9'].fill = yellow_fill
    ws['BL9'].border = border
    ws['BL9'] = '1° PENTA' 
    
    ws['BM9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BM9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BM9'].fill = yellow_fill
    ws['BM9'].border = border
    ws['BM9'] = 'V' 
    
    ws['BN9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BN9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BN9'].fill = yellow_fill
    ws['BN9'].border = border
    ws['BN9'] = '2° PENTA' 
    
    ws['BO9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BO9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BO9'].fill = yellow_fill
    ws['BO9'].border = border
    ws['BO9'] = 'V' 
    
    ws['BP9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BP9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BP9'].fill = yellow_fill
    ws['BP9'].border = border
    ws['BP9'] = '3° PENTA' 
    
    ws['BQ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BQ9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BQ9'].fill = yellow_fill
    ws['BQ9'].border = border
    ws['BQ9'] = 'V'  
    
    ws['BR9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BR9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BR9'].fill = yellow_fill
    ws['BR9'].border = border
    ws['BR9'] = 'NUM ROTA' 
    
    ws['BS9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BS9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BS9'].fill = yellow_fill
    ws['BS9'].border = border
    ws['BS9'] = '1° ROTA' 
    
    ws['BT9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BT9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BT9'].fill = yellow_fill
    ws['BT9'].border = border
    ws['BT9'] = 'V' 
    
    ws['BU9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BU9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BU9'].fill = yellow_fill
    ws['BU9'].border = border
    ws['BU9'] = '2° ROTA' 
    
    ws['BV9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BV9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BV9'].fill = yellow_fill
    ws['BV9'].border = border
    ws['BV9'] = 'V' 
    
    ws['BW9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BW9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BW9'].fill = blue_fill
    ws['BW9'].border = border
    ws['BW9'] = 'NUM ESQ'
    
    ws['BX9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BX9'].font = Font(name = 'Arial', size= 8, bold = True, color='000000')
    ws['BX9'].fill = blue_fill
    ws['BX9'].border = border
    ws['BX9'] = 'ESQ 4M' 
    
    ws['BY9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BY9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['BY9'].fill = blue_fill
    ws['BY9'].border = border
    ws['BY9'] = 'SUP 4M' 
    
    ws['BZ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['BZ9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['BZ9'].fill = blue_fill
    ws['BZ9'].border = border
    ws['BZ9'] = 'V' 
    
    ws['CA9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CA9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CA9'].fill = blue_fill
    ws['CA9'].border = border
    ws['CA9'] = 'ESQ 6M' 
    
    ws['CB9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CB9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CB9'].fill = blue_fill
    ws['CB9'].border = border
    ws['CB9'] = 'NUM SUP 6M' 
    
    ws['CC9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CC9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CC9'].fill = blue_fill
    ws['CC9'].border = border
    ws['CC9'] = '1° SUP 6M' 
    
    ws['CD9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CD9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CD9'].fill = blue_fill
    ws['CD9'].border = border
    ws['CD9'] = 'V' 
    
    ws['CE9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CE9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CE9'].fill = blue_fill
    ws['CE9'].border = border
    ws['CE9'] = '2° SUP 6M' 
    
    ws['CF9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CF9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CF9'].fill = blue_fill
    ws['CF9'].border = border
    ws['CF9'] = 'V' 
    
    ws['CG9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CG9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CG9'].fill = blue_fill
    ws['CG9'].border = border
    ws['CG9'] = 'NUM TOO 6M' 
    
    ws['CH9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CH9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CH9'].fill = blue_fill
    ws['CH9'].border = border
    ws['CH9'] = '1° TTO 6M' 
    
    ws['CI9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CI9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CI9'].fill = blue_fill
    ws['CI9'].border = border
    ws['CI9'] = 'V' 
    
    ws['CJ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CJ9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CJ9'].fill = blue_fill
    ws['CJ9'].border = border
    ws['CJ9'] = '2° TTO 6M' 
    
    ws['CK9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CK9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CK9'].fill = blue_fill
    ws['CK9'].border = border
    ws['CK9'] = 'V' 
    
    ws['CL9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CL9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CL9'].fill = blue_fill
    ws['CL9'].border = border
    ws['CL9'] = '3° TTO 6M' 
    
    ws['CM9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CM9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CM9'].fill = blue_fill
    ws['CM9'].border = border
    ws['CM9'] = 'V' 
    
    ws['CN9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CN9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CN9'].fill = blue_fill
    ws['CN9'].border = border
    ws['CN9'] = 'NUM MULT 6M' 
    
    ws['CO9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CO9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CO9'].fill = blue_fill
    ws['CO9'].border = border
    ws['CO9'] = '1° MULTI 6M' 
    
    ws['CP9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CP9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CP9'].fill = blue_fill
    ws['CP9'].border = border
    ws['CP9'] = 'V' 
    
    ws['CQ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CQ9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CQ9'].fill = blue_fill
    ws['CQ9'].border = border
    ws['CQ9'] = '2° MULTI 6M' 
    
    ws['CR9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CR9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CR9'].fill = blue_fill
    ws['CR9'].border = border
    ws['CR9'] = 'V' 
    
    ws['CS9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CS9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CS9'].fill = blue_fill
    ws['CS9'].border = border
    ws['CS9'] = '3° MULTI 6M' 
    
    ws['CT9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CT9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CT9'].fill = blue_fill
    ws['CT9'].border = border
    ws['CT9'] = 'V' 
    
    ws['CU9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CU9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CU9'].fill = blue_fill
    ws['CU9'].border = border
    ws['CU9'] = '4° MULTI 6M' 
    
    ws['CV9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CV9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CV9'].fill = blue_fill
    ws['CV9'].border = border
    ws['CV9'] = 'V' 
    
    ws['CW9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CW9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CW9'].fill = blue_fill
    ws['CW9'].border = border
    ws['CW9'] = '5° MULTI 6M' 
    
    ws['CX9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CX9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CX9'].fill = blue_fill
    ws['CX9'].border = border
    ws['CX9'] = 'V' 
    
    ws['CY9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CY9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CY9'].fill = blue_fill
    ws['CY9'].border = border
    ws['CY9'] = '6° MULTI 6M' 
    
    ws['CZ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['CZ9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['CZ9'].fill = blue_fill
    ws['CZ9'].border = border
    ws['CZ9'] = 'V' 
    
    ws['DA9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DA9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DA9'].fill = blue_fill
    ws['DA9'].border = border
    ws['DA9'] = 'NUM DOSAJE HB' 
    
    ws['DB9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DB9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DB9'].fill = blue_fill
    ws['DB9'].border = border
    ws['DB9'] = 'DOSAJE HB' 
    
    ws['DC9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DC9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DC9'].fill = blue_fill
    ws['DC9'].border = border
    ws['DC9'] = 'NUM HB' 
    
    ws['DD9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DD9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DD9'].fill = gray_fill
    ws['DD9'].border = border
    ws['DD9'] = 'NUM DNI EMISION' 
    
    ws['DE9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DE9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DE9'].fill = gray_fill
    ws['DE9'].border = border
    ws['DE9'] = 'EMISION' 
    
    ws['DF9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DF9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DF9'].fill = gray_fill
    ws['DF9'].border = border
    ws['DF9'] = 'DNI 30D' 
    
    ws['DG9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DG9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DG9'].fill = gray_fill
    ws['DG9'].border = border
    ws['DG9'] = 'DNI 60D' 
        
    ws['DH9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DH9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DH9'].fill = fill
    ws['DH9'].border = border
    ws['DH9'] = 'MES' 
    
    ws['DI9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DI9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DI9'].fill = gray_fill
    ws['DI9'].border = border
    ws['DI9'] = 'IND' 
    
    ws['DJ9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DJ9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DJ9'].fill = orange_fill
    ws['DJ9'].border = border
    ws['DJ9'] = 'UBIGEO'  
    
    ws['DK9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DK9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DK9'].fill = orange_fill
    ws['DK9'].border = border
    ws['DK9'] = 'PROVINCIA'       
    
    ws['DL9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DL9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DL9'].fill = orange_fill
    ws['DL9'].border = border
    ws['DL9'] = 'DISTRITO' 
    
    ws['DM9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DM9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DM9'].fill = orange_fill
    ws['DM9'].border = border
    ws['DM9'] = 'RED'  
    
    ws['DN9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DN9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DN9'].fill = orange_fill
    ws['DN9'].border = border
    ws['DN9'] = 'MICRORED'  
    
    ws['DO9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DO9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DO9'].fill = orange_fill
    ws['DO9'].border = border
    ws['DO9'] = 'COD EST'  
    
    ws['DP9'].alignment = Alignment(horizontal= "center", vertical="center", wrap_text=True)
    ws['DP9'].font = Font(name = 'Arial', size= 8, bold = True)
    ws['DP9'].fill = orange_fill
    ws['DP9'].border = border
    ws['DP9'] = 'ESTABLECIMIENTO'  
    
    # Definir estilos
    header_font = Font(name = 'Arial', size= 8, bold = True)
    centered_alignment = Alignment(horizontal='center')
    border = Border(left=Side(style='thin', color='A9A9A9'),
            right=Side(style='thin', color='A9A9A9'),
            top=Side(style='thin', color='A9A9A9'),
            bottom=Side(style='thin', color='A9A9A9'))
    header_fill = PatternFill(patternType='solid', fgColor='00B0F0')
    
    
    # Definir los caracteres especiales de check y X
    check_mark = '✓'  # Unicode para check
    x_mark = '✗'  # Unicode para X
    sub_cumple = 'CUMPLE'
    sub_no_cumple = 'NO CUMPLE'
    
    # Escribir datos
    for row, record in enumerate(results, start=10):
        for col, value in enumerate(record.values(), start=2):
            cell = ws.cell(row=row, column=col, value=value)

            # Alinear a la izquierda solo en las columnas 6,14,15,16
            if col in [116, 117, 118, 120]:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

            # Aplicar color en la columna 27
            if col == 113:
                if isinstance(value, str):
                    value_upper = value.strip().upper()
                    if value_upper == "NO CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')  # Fondo rojo
                        cell.font = Font(name='Arial', size=8, bold = True,color="FFFFFF")  # Letra blanca
                    elif value_upper == "CUMPLE":
                        cell.fill = PatternFill(patternType='solid', fgColor='00FF00')  # Fondo verde
                        cell.font = Font(name='Arial', size=8,  bold = True,color="FFFFFF")  # Letra blanca
                    else:
                        cell.font = Font(name='Arial', size=8, bold = True)
                else:
                    cell.font = Font(name='Arial', size=8,  bold = True)
            
            # Aplicar color de letra SUB INDICADORES
            elif col in [15, 18, 27, 51, 56, 63, 70, 76, 79, 80, 85, 92]:
                if value == 0:
                    cell.value = sub_no_cumple  # Insertar check
                    cell.font = Font(name='Arial', size=7, color="FF0000") 
                elif value == 1:
                    cell.value = sub_cumple # Insertar check
                    cell.font = Font(name='Arial', size=7, color="00B050")
                else:
                    cell.font = Font(name='Arial', size=7)
            # Fuente normal para otras columnas
            # Aplicar color de letra SUB GENERALIDADES
            elif col in [17, 50, 75, 105, 108]:
                if value == 0:
                    cell.value = sub_no_cumple  # Insertar check
                    cell.font = Font(name='Arial', size=7, color="FF0000") 
                    cell.fill = gray_fill # Letra roja
                elif value == 1:
                    cell.value = sub_cumple # Insertar check
                    cell.font = Font(name='Arial', size=7, color="00B050")
                    cell.fill = gray_fill# Letra verde
                else:
                    cell.font = Font(name='Arial', size=7)
            # Fuente normal para otras columnas
            else:
                cell.font = Font(name='Arial', size=8)  # Fuente normal para otras columnas
            
            # Aplicar caracteres especiales check y X
            if col in [9, 11, 13, 14, 16, 20, 22, 24, 26, 29, 31, 33, 35, 37, 39, 41, 43, 45, 47, 49, 53, 55, 58, 60, 62, 65, 67, 69, 72, 74, 78, 82, 84, 87, 89, 91, 94, 96, 98, 100, 102, 104, 107, 110, 111]:
                if value == 1:
                    cell.value = check_mark  # Insertar check
                    cell.font = Font(name='Arial', size=10, color='00B050')  # Letra verde
                elif value == 0:
                    cell.value = x_mark  # Insertar X
                    cell.font = Font(name='Arial', size=10, color='FF0000')  # Letra roja
                else:
                    cell.font = Font(name='Arial', size=8)  # Fuente normal si no es 1 o 0
            
            cell.border = border

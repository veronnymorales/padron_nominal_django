from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tempfile import NamedTemporaryFile

from django.http import HttpResponse

def fill_worksheet(ws, results):
    """
    Función que rellena la hoja de cálculo con los resultados de seguimiento.
    Aplica estilos, formatos, etc.
    """
    # Configuración de columnas, estilos, etc.
    ws.row_dimensions[1].height = 14
    # ... resto de configuraciones

    border = Border(
        left=Side(style='thin', color='A9A9A9'),
        right=Side(style='thin', color='A9A9A9'),
        top=Side(style='thin', color='A9A9A9'),
        bottom=Side(style='thin', color='A9A9A9')
    )
    # ... más estilos

    for row_num, record in enumerate(results, start=9):
        for col_num, value in enumerate(record, start=2):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
    # Resto de la lógica de formateo...

def fill_worksheet_cobertura_paquete_nino(ws, results):
    """
    Rellena la hoja de cálculo con resultados de cobertura.
    """
    ws['A1'] = 'Red'
    ws['B1'] = 'MicroRed'
    ws['C1'] = 'Establecimiento'
    # ... y más cabeceras

    # Escribir data
    for row_num, row_data in enumerate(results, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            # ... ajuste de estilos

def generar_excel_seguimiento(resultado_seguimiento):
    """
    Genera el archivo Excel para el seguimiento y retorna un HttpResponse con el archivo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Seguimiento'

    # Llenar la hoja
    fill_worksheet(ws, resultado_seguimiento)

    nombre_archivo = "rpt_paquete_nino_red.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f'attachment; filename={nombre_archivo}'
    wb.save(response)
    return response

def generar_excel_cobertura(resultado_cobertura):
    """
    Genera el archivo Excel para la cobertura y retorna un HttpResponse con el archivo.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Cobertura'

    fill_worksheet_cobertura_paquete_nino(ws, resultado_cobertura)

    nombre_archivo = "rpt_cobertura_paquete_nino.xlsx"
    with NamedTemporaryFile() as tmp_file:
        wb.save(tmp_file.name)
        tmp_file.seek(0)
        response = HttpResponse(
            tmp_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
        return response
